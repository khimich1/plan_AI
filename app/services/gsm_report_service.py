"""GSM usage report: fill blank → soffice → .xls → zip (+ waybills).

Reuses soffice helpers from ``gsm_export_service`` (isolated LO profile,
bare ``--convert-to xls``). Never modifies originals under ``ГСМ/**``.
"""

from __future__ import annotations

import io
import json
import shutil
import tempfile
import zipfile
from collections.abc import Sequence
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.repositories.gsm_repository import GsmRepository
from app.services.gsm_export_service import (
    DEFAULT_SOFFICE_TIMEOUT_SEC,
    GsmExportError,
    GsmExportService,
    convert_with_soffice,
)
from app.services.gsm_kit_gate import filter_kit_vehicle_ids
from core.gsm.season import parse_season_switches
from core.gsm.usage_report import (
    KIT_STATUSES,
    UsageMonthBlock,
    attach_transactions_to_rows,
    build_month_block,
    expand_tx_window,
    format_approval_date,
    group_waybills_by_month,
    report_filename,
    tx_day_qty,
)

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_TEMPLATE = PROJECT_ROOT / "core" / "gsm" / "templates" / "gsm_usage_report.xlsx"

# Template layout (1-based), from лист «Образец»
DATA_START_ROW = 18
# Columns A..O (1..15); P = destination (как на листе «2026»)
COL_SEQ = 1
COL_MARK = 2
COL_PLATE = 3
COL_DRIVER = 4
COL_GRADE = 5
COL_FUEL_START = 6
COL_ODO_START = 7
COL_ODO_END = 8
COL_KM = 9
COL_NORM = 10
COL_BURN_NORM = 11
COL_BURN_FACT = 12
COL_RECEIVED = 13
COL_FUEL_END = 14
COL_NOTE = 15
COL_DEST = 16


class GsmReportError(Exception):
    """Usage-report failure with stable machine code for HTTP mapping."""

    def __init__(self, message: str, *, code: str, details: dict[str, Any] | None = None) -> None:
        super().__init__(message)
        self.code = code
        self.details = details or {}


class GsmReportService:
    def __init__(
        self,
        repo: GsmRepository,
        *,
        template_path: Path | None = None,
        soffice_timeout_sec: int = DEFAULT_SOFFICE_TIMEOUT_SEC,
        export_service: GsmExportService | None = None,
    ) -> None:
        self._repo = repo
        self._template_path = Path(template_path) if template_path else DEFAULT_TEMPLATE
        self._timeout = int(soffice_timeout_sec)
        self._export = export_service or GsmExportService(
            repo=repo,
            soffice_timeout_sec=self._timeout,
        )

    def build_usage_zip(
        self,
        *,
        period_from: date,
        period_to: date,
        vehicle_ids: list[int] | None,
    ) -> tuple[bytes, str]:
        """Build zip of per-vehicle usage reports + waybills; return (bytes, filename)."""
        if period_to < period_from:
            raise GsmReportError(
                "period_to must be >= period_from",
                code="gsm_report_invalid_period",
            )

        ids = self._resolve_vehicle_ids(vehicle_ids)
        allowed, blocked = filter_kit_vehicle_ids(
            self._repo,
            ids,
            period_from=period_from,
            period_to=period_to,
            purpose="kit",
        )
        if not allowed:
            if len(ids) == 1 and blocked:
                first = blocked[0]
                raise GsmReportError(
                    first.message or "комплект запрещён",
                    code=first.code or "gsm_kit_red",
                )
            raise GsmReportError(
                "нет ПЛ к комплекту / все красные",
                code="gsm_report_no_data",
            )

        eligible: list[int] = []
        for vid in allowed:
            if not self._kit_waybills(vid, period_from, period_to):
                continue
            eligible.append(vid)

        if not eligible:
            raise GsmReportError(
                "нет ПЛ к комплекту / все красные",
                code="gsm_report_no_data",
            )

        blocks_by_vehicle: dict[int, list[UsageMonthBlock]] = {}
        for vid in eligible:
            blocks = self._build_vehicle_blocks(vid, period_from, period_to)
            if blocks:
                blocks_by_vehicle[vid] = blocks

        if not blocks_by_vehicle:
            raise GsmReportError(
                "нет ПЛ к комплекту / все красные",
                code="gsm_report_no_data",
            )

        if not self._template_path.exists():
            raise GsmReportError(
                f"usage report template not found: {self._template_path}",
                code="gsm_export_template_missing",
            )

        buf = io.BytesIO()
        with tempfile.TemporaryDirectory(prefix="gsm_usage_") as tmp:
            workdir = Path(tmp)
            outdir = workdir / "out"
            outdir.mkdir()
            filled_dir = workdir / "filled"
            filled_dir.mkdir()

            with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                for vid, blocks in blocks_by_vehicle.items():
                    vehicle = self._repo.get_vehicle(vid)
                    assert vehicle is not None
                    xlsx_path = filled_dir / f"usage_{vid}.xlsx"
                    self._write_report_xlsx(
                        blocks,
                        period_to=period_to,
                        out_path=xlsx_path,
                    )
                    try:
                        xls_path = convert_with_soffice(
                            xlsx_path, "xls", outdir, workdir, self._timeout
                        )
                    except GsmExportError as exc:
                        raise GsmReportError(
                            str(exc),
                            code=exc.code,
                            details=exc.details,
                        ) from exc

                    entry = report_filename(str(vehicle["plate_number"]))
                    zf.writestr(entry, xls_path.read_bytes())

                # Waybills for the same vehicles/period (existing export pipeline).
                try:
                    wb_zip_bytes, _ = self._export.export_zip(
                        vehicle_ids=sorted(eligible),
                        period_from=period_from,
                        period_to=period_to,
                    )
                except GsmExportError as exc:
                    if exc.code == "gsm_export_empty":
                        wb_zip_bytes = b""
                    else:
                        raise GsmReportError(
                            str(exc),
                            code=exc.code,
                            details=exc.details,
                        ) from exc

                if wb_zip_bytes:
                    with zipfile.ZipFile(io.BytesIO(wb_zip_bytes), "r") as wb_zf:
                        for name in wb_zf.namelist():
                            zf.writestr(name, wb_zf.read(name))

        filename = (
            f"gsm_usage_report_{period_from.isoformat()}_{period_to.isoformat()}.zip"
        )
        return buf.getvalue(), filename

    def build_vehicle_report_xlsx(
        self,
        *,
        vehicle_id: int,
        period_from: date,
        period_to: date,
        out_path: Path,
    ) -> Path:
        """Fill template for one vehicle (no soffice). Used by acceptance tests."""
        if period_to < period_from:
            raise GsmReportError(
                "period_to must be >= period_from",
                code="gsm_report_invalid_period",
            )
        if self._repo.get_vehicle(vehicle_id) is None:
            raise GsmReportError(
                f"vehicle #{vehicle_id} not found",
                code="gsm_vehicle_not_found",
            )
        blocks = self._build_vehicle_blocks(vehicle_id, period_from, period_to)
        if not blocks:
            raise GsmReportError(
                "нет ПЛ к комплекту / все красные",
                code="gsm_report_no_data",
            )
        return self._write_report_xlsx(blocks, period_to=period_to, out_path=out_path)

    # ------------------------------------------------------------------
    # Domain assembly
    # ------------------------------------------------------------------

    def _resolve_vehicle_ids(self, vehicle_ids: list[int] | None) -> list[int]:
        if vehicle_ids is None:
            return [int(v["id"]) for v in self._repo.list_vehicles(active_only=True)]
        ids = sorted({int(v) for v in vehicle_ids})
        if not ids:
            raise GsmReportError(
                "vehicle_ids must be non-empty or null",
                code="gsm_report_invalid_period",
            )
        for vid in ids:
            if self._repo.get_vehicle(vid) is None:
                raise GsmReportError(
                    f"vehicle #{vid} not found",
                    code="gsm_vehicle_not_found",
                )
        return ids

    def _kit_waybills(
        self,
        vehicle_id: int,
        period_from: date,
        period_to: date,
    ) -> list[dict[str, Any]]:
        return [
            wb
            for wb in self._repo.list_waybills(
                vehicle_id=vehicle_id,
                period_from=period_from,
                period_to=period_to,
            )
            if str(wb.get("status") or "") in KIT_STATUSES
        ]

    def _build_vehicle_blocks(
        self,
        vehicle_id: int,
        period_from: date,
        period_to: date,
    ) -> list[UsageMonthBlock]:
        vehicle = self._repo.get_vehicle(vehicle_id)
        assert vehicle is not None

        waybills = self._kit_waybills(vehicle_id, period_from, period_to)
        if not waybills:
            return []
        waybills.sort(key=lambda r: (str(r["date"]), int(r["id"])))

        switches_raw = self._repo.get_setting("season_switches")
        try:
            season_switches = parse_season_switches(switches_raw)
        except ValueError as exc:
            raise GsmReportError(
                f"invalid season_switches setting: {switches_raw!r}",
                code="gsm_settings_invalid",
            ) from exc

        drivers_by_id: dict[int, dict[str, Any]] = {}
        route_by_wb_id: dict[int, list[dict[str, Any]]] = {}
        for wb in waybills:
            did = int(wb["driver_id"])
            if did not in drivers_by_id:
                driver = self._repo.get_driver(did)
                if driver is None:
                    raise GsmReportError(
                        f"driver #{did} not found",
                        code="gsm_driver_not_found",
                    )
                drivers_by_id[did] = driver
            route_by_wb_id[int(wb["id"])] = _parse_route_list(wb.get("route_json"))

        tx_from, tx_to = expand_tx_window(period_from, period_to)
        txs = self._repo.list_transactions(
            period_from=tx_from,
            period_to=tx_to,
            vehicle_id=vehicle_id,
            service_type="fuel",
        )
        tx_pairs = [p for p in (tx_day_qty(t) for t in txs) if p is not None]
        row_dates = [_as_date(wb["date"]) for wb in waybills]
        received_list = attach_transactions_to_rows(row_dates, tx_pairs)
        received_by_date = {
            d: qty for d, qty in zip(row_dates, received_list, strict=True)
        }

        fuel_grade = _dominant_fuel_grade(txs) or "АИ-95"

        blocks: list[UsageMonthBlock] = []
        for year, month, month_wbs in group_waybills_by_month(waybills):
            block = build_month_block(
                year=year,
                month=month,
                waybills=month_wbs,
                vehicle=vehicle,
                drivers_by_id=drivers_by_id,
                route_by_wb_id=route_by_wb_id,
                received_by_date=received_by_date,
                season_switches=season_switches,
                fuel_grade=fuel_grade,
            )
            if block is not None:
                blocks.append(block)
        return blocks

    # ------------------------------------------------------------------
    # Template fill
    # ------------------------------------------------------------------

    def _write_report_xlsx(
        self,
        blocks: list[UsageMonthBlock],
        *,
        period_to: date,
        out_path: Path,
    ) -> Path:
        shutil.copy2(self._template_path, out_path)
        wb = load_workbook(out_path, data_only=False)
        ws = wb.active

        # Clear anything below the header (placeholder footer etc.)
        _clear_from_row(ws, DATA_START_ROW)

        ws["O5"] = format_approval_date(period_to)

        cursor = DATA_START_ROW
        for i, block in enumerate(blocks):
            if i == 0:
                # First block: reuse header title cell F9
                ws["F9"] = block.title
                cursor = self._write_month_rows(ws, block, start_row=DATA_START_ROW)
            else:
                # Stack subsequent month blocks downward with a small gap
                cursor += 2
                cursor = self._write_month_header_mini(ws, block, start_row=cursor)
                cursor = self._write_month_rows(ws, block, start_row=cursor)

            # ИТОГО + fuel-grade duplicate
            cursor = self._write_totals(ws, block, start_row=cursor)

        # Accountant footer
        cursor += 3
        ws.cell(cursor, 3).value = (
            "Бухгалтер _____________________________________ Никифорова Е.А."
        )

        wb.save(out_path)
        return out_path

    def _write_month_header_mini(
        self, ws: Worksheet, block: UsageMonthBlock, *, start_row: int
    ) -> int:
        """Compact month title for stacked blocks (not the first)."""
        ws.cell(start_row, 1).value = (
            "Отчет по использованию горюче-смазочных материалов "
        )
        ws.cell(start_row + 2, 6).value = block.title
        # Column headers abbreviated — reuse labels from row 15/16 of template
        hdr = start_row + 4
        labels = [
            (1, "№ п/п"),
            (2, "Марка а/м "),
            (3, "Государственный номер "),
            (4, "Ф.И.О. сотрудника, за которым закреплен а/м"),
            (5, "Марка бензина"),
            (6, "Остаток бензина в баке на начало месяца, л"),
            (7, "Показатели спидометра"),
            (9, "Пробег за месяц, км"),
            (10, "Норма по приказу, л/100км"),
            (11, "Расход бензина , л"),
            (13, "Получено ГСМ по смарт-карте, л"),
            (14, "Остаток бензина в баке на конец месяца, л"),
            (15, "Примечание"),
        ]
        for col, text in labels:
            ws.cell(hdr, col).value = text
        ws.cell(hdr + 1, 7).value = "на начало месяца"
        ws.cell(hdr + 1, 8).value = "на конец месяца"
        ws.cell(hdr + 1, 11).value = "по норме"
        ws.cell(hdr + 1, 12).value = "по факту"
        for col in range(1, 16):
            ws.cell(hdr + 2, col).value = col
        return hdr + 3

    def _write_month_rows(
        self, ws: Worksheet, block: UsageMonthBlock, *, start_row: int
    ) -> int:
        row_i = start_row
        for r in block.rows:
            ws.cell(row_i, COL_SEQ).value = r.seq
            # Mark/plate/grade only on first row of the block (as in образец)
            if r.seq == 1:
                ws.cell(row_i, COL_MARK).value = r.vehicle_mark
                ws.cell(row_i, COL_PLATE).value = r.plate.replace(" ", "")
                ws.cell(row_i, COL_GRADE).value = r.fuel_grade
            ws.cell(row_i, COL_DRIVER).value = r.driver_short
            ws.cell(row_i, COL_FUEL_START).value = r.fuel_start if r.seq == 1 else None
            ws.cell(row_i, COL_ODO_START).value = r.odometer_start
            ws.cell(row_i, COL_ODO_END).value = r.odometer_end
            ws.cell(row_i, COL_KM).value = r.km
            ws.cell(row_i, COL_NORM).value = r.norm_l_per_100
            ws.cell(row_i, COL_BURN_NORM).value = r.burn_norm
            ws.cell(row_i, COL_BURN_FACT).value = r.burn_fact
            ws.cell(row_i, COL_RECEIVED).value = r.received
            ws.cell(row_i, COL_FUEL_END).value = r.fuel_end
            ws.cell(row_i, COL_NOTE).value = r.note
            if r.destination:
                ws.cell(row_i, COL_DEST).value = r.destination
            row_i += 1
        return row_i

    def _write_totals(
        self, ws: Worksheet, block: UsageMonthBlock, *, start_row: int
    ) -> int:
        # ИТОГО
        ws.cell(start_row, COL_SEQ).value = "ИТОГО"
        ws.cell(start_row, COL_FUEL_START).value = block.fuel_start
        ws.cell(start_row, COL_BURN_NORM).value = block.burn_norm
        ws.cell(start_row, COL_BURN_FACT).value = block.burn_fact
        ws.cell(start_row, COL_RECEIVED).value = block.received
        ws.cell(start_row, COL_FUEL_END).value = block.fuel_end

        # Fuel-grade duplicate row (as in образец)
        dup = start_row + 1
        ws.cell(dup, COL_GRADE).value = block.fuel_grade
        ws.cell(dup, COL_FUEL_START).value = block.fuel_start
        ws.cell(dup, COL_NORM).value = block.fuel_grade
        ws.cell(dup, COL_BURN_NORM).value = block.burn_norm
        ws.cell(dup, COL_BURN_FACT).value = block.burn_fact
        ws.cell(dup, COL_RECEIVED).value = block.received
        ws.cell(dup, COL_FUEL_END).value = block.fuel_end
        return dup + 1


def _clear_from_row(ws: Worksheet, start_row: int) -> None:
    max_row = ws.max_row or start_row
    max_col = ws.max_column or 16
    for r in range(start_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).value = None


def _as_date(value: Any) -> date:
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value)[:10])


def _parse_route_list(raw: Any) -> list[dict[str, Any]]:
    if raw is None or raw == "":
        return []
    try:
        parsed = json.loads(raw) if isinstance(raw, str) else raw
    except (TypeError, ValueError, json.JSONDecodeError):
        return []
    if not isinstance(parsed, list):
        return []
    return [item for item in parsed if isinstance(item, dict)]


def _dominant_fuel_grade(txs: Sequence[dict[str, Any]]) -> str | None:
    counts: dict[str, float] = {}
    for tx in txs:
        grade = tx.get("fuel_grade")
        qty = tx.get("qty_liters")
        if not grade or qty is None:
            continue
        counts[str(grade)] = counts.get(str(grade), 0.0) + float(qty)
    if not counts:
        return None
    return max(counts.items(), key=lambda kv: kv[1])[0]

"""Shipment logistics service: реестр рейсов, propose→confirm, выезд, отмена.

Паттерн — SgpService: ``db_path``, ``ShipmentError(ValueError).code`` → 422,
одна транзакция на мутацию, audit рядом с UPDATE/INSERT.
"""

from __future__ import annotations

import json
import logging
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

from app.core.settings import get_settings
from app.domain.enums import (
    DeliveryType,
    KpStatus,
    ShipmentItemType,
    ShipmentStatus,
)
from app.repositories.shipment_repository import ShipmentRepository
from app.schemas.logistics import (
    LogisticsKpSearchItem,
    LogisticsKpSearchResponse,
    PileCatalogItem,
    PileCatalogResponse,
    ShipmentAvailableByKp,
    ShipmentCard,
    ShipmentItem,
    ShipmentItemInput,
    ShipmentLayoutMetadata,
    ShipmentLayoutStack,
    ShipmentLayoutTier,
    ShipmentLayoutUnit,
    ShipmentListItem,
    ShipmentListResponse,
    ShipmentLoadingStep,
    ShipmentMutationResponse,
    ShipmentOrderItem,
    ShipmentOrderPatch,
    ShipmentOrderRemainderItem,
    ShipmentProposeItem,
    ShipmentProposeResponse,
    ShipmentProposeWarning,
)
from app.services.shipment_completion_service import ShipmentCompletionService
from app.services.shipment_errors import ShipmentError
from core.config.settings import VehicleClassLimits
from core.kp_db_common import _connect
from core.kp_db_schema import ensure_schema
from core.kp_db_shipments import available_qty
from core.kp_plate_weight import resolve_kp_line_weight_kg
from core.shipment_packing import PlateCandidate, VehicleLimits, pack_shipment

logger = logging.getLogger(__name__)

ALLOWED_LOGISTICS_KP_STATUSES: tuple[str, ...] = (
    KpStatus.IN_WORK.value,
    KpStatus.ON_SGP.value,
)


_DATE_FORMAT = "%Y-%m-%d"


class ShipmentService:
    def __init__(
        self,
        *,
        db_path: str,
        repo: ShipmentRepository | None = None,
        completion: ShipmentCompletionService | None = None,
    ) -> None:
        self.db_path = db_path
        self._repo = repo or ShipmentRepository(db_path=db_path)
        self._completion = completion or ShipmentCompletionService(
            db_path=db_path,
            repo=self._repo,
            maybe_write_event=self._maybe_write_event,
        )

    # ------------------------------------------------------------------
    # CRUD (SHIP-200)
    # ------------------------------------------------------------------

    def create(
        self,
        *,
        shipment_date: str,
        delivery_type: str,
        kp_ids: list[int],
        actor: str | None = None,
    ) -> ShipmentCard:
        self._validate_date(shipment_date)
        kp_ids = self._validate_kp_ids(kp_ids)
        conn = self._repo.connect()
        try:
            cur = conn.cursor()
            self._assert_kp_exists(cur, kp_ids)
            shipment_id = self._insert_shipment_with_orders(
                cur,
                shipment_date=shipment_date,
                delivery_type=delivery_type,
                kp_ids=kp_ids,
                actor=actor,
            )
            conn.commit()
        except ShipmentError:
            conn.rollback()
            raise
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
        return self.get(shipment_id)

    def reuse_transport(
        self,
        source_id: int,
        *,
        shipment_date: str,
        delivery_type: str,
        kp_ids: list[int],
        actor: str | None = None,
    ) -> ShipmentCard:
        """Создать новый рейс, скопировав только транспортные поля из source."""
        self._validate_date(shipment_date)
        kp_ids = self._validate_kp_ids(kp_ids)
        conn = self._repo.connect()
        try:
            cur = conn.cursor()
            source = self._fetch_shipment_row(cur, source_id)
            self._assert_kp_exists(cur, kp_ids)
            shipment_id = self._insert_shipment_with_orders(
                cur,
                shipment_date=shipment_date,
                delivery_type=delivery_type,
                kp_ids=kp_ids,
                actor=actor,
            )
            self._repo.copy_transport_fields(
                cur, source=source, shipment_id=shipment_id
            )
            conn.commit()
        except ShipmentError:
            conn.rollback()
            raise
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
        return self.get(shipment_id)

    def get(self, shipment_id: int) -> ShipmentCard:
        conn = self._repo.connect()
        try:
            cur = conn.cursor()
            row = self._fetch_shipment_row(cur, shipment_id)
            orders = self._fetch_orders(cur, [shipment_id])
            items = self._fetch_items(cur, shipment_id)
            card = self._to_card(row, orders.get(shipment_id, []), items)
            card.available_by_kp = self._available_by_kp(cur, card.orders)
            return card
        finally:
            conn.close()

    def list_shipments(
        self,
        *,
        date_from: str | None = None,
        date_to: str | None = None,
        kp_id: int | None = None,
        carrier_id: int | None = None,
        delivery_type: str | None = None,
        status: str | None = None,
        no_upd: bool = False,
        attention: bool = False,
        limit: int = 200,
        offset: int = 0,
    ) -> ShipmentListResponse:
        conn = self._repo.connect()
        try:
            cur = conn.cursor()
            clauses: list[str] = []
            params: list[Any] = []
            if date_from:
                clauses.append("s.shipment_date >= ?")
                params.append(date_from)
            if date_to:
                clauses.append("s.shipment_date <= ?")
                params.append(date_to)
            if kp_id is not None:
                clauses.append(
                    "EXISTS (SELECT 1 FROM shipment_orders so "
                    "WHERE so.shipment_id = s.id AND so.kp_id = ?)"
                )
                params.append(int(kp_id))
            if carrier_id is not None:
                clauses.append("s.carrier_id = ?")
                params.append(int(carrier_id))
            if delivery_type:
                clauses.append("s.delivery_type = ?")
                params.append(delivery_type)
            if status:
                clauses.append("s.status = ?")
                params.append(status)
            if no_upd:
                clauses.append("(s.upd_no IS NULL OR s.upd_no = '')")
            if attention:
                clauses.append("s.attention = 1")
            rows = self._repo.list_shipment_rows(
                cur, clauses=clauses, params=params, limit=limit, offset=offset
            )
            shipment_ids = [int(r["id"]) for r in rows]
            orders_map = self._fetch_orders(cur, shipment_ids)
            weights = self._fetch_weight_totals(cur, shipment_ids)
            items = [
                self._to_list_item(row, orders_map.get(int(row["id"]), []), weights)
                for row in rows
            ]
            return ShipmentListResponse(items=items, count=len(items))
        finally:
            conn.close()

    def patch(
        self,
        shipment_id: int,
        *,
        fields: dict[str, Any],
        orders: list[ShipmentOrderPatch] | None = None,
        actor: str | None = None,
    ) -> ShipmentCard:
        conn = self._repo.connect()
        try:
            cur = conn.cursor()
            row = self._fetch_shipment_row(cur, shipment_id)
            self._assert_in_work(row, shipment_id)
            if "shipment_date" in fields and fields["shipment_date"] is not None:
                self._validate_date(str(fields["shipment_date"]))
            if "carrier_id" in fields and fields["carrier_id"] is not None:
                self._assert_carrier_exists(cur, int(fields["carrier_id"]))
            self._repo.update_shipment_fields(cur, shipment_id, fields)
            if orders is not None:
                self._replace_orders(cur, shipment_id, orders)
            conn.commit()
        except ShipmentError:
            conn.rollback()
            raise
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
        return self.get(shipment_id)

    # ------------------------------------------------------------------
    # Propose (SHIP-201)
    # ------------------------------------------------------------------

    def propose(
        self,
        shipment_id: int,
        *,
        vehicle_class: str | None = None,
    ) -> ShipmentProposeResponse:
        settings = get_settings()
        limits_kg = settings.vehicle_class_limits_kg
        conn = self._repo.connect()
        try:
            cur = conn.cursor()
            row = self._fetch_shipment_row(cur, shipment_id)
            self._assert_in_work(row, shipment_id)
            effective_class = vehicle_class or row["vehicle_class"] or "t20"
            if effective_class not in limits_kg:
                raise ShipmentError(
                    f"Неизвестный класс ТС «{effective_class}» (есть: {', '.join(sorted(limits_kg))})",
                    code="shipment_invalid_vehicle_class",
                )

            candidates = self._fetch_propose_candidates(cur, shipment_id)
            if effective_class == "t30plus":
                response = self._propose_legacy_weight_fifo(
                    candidates,
                    effective_class=effective_class,
                    limits_kg=limits_kg,
                )
            else:
                class_limits = settings.vehicle_class_limits.get(
                    effective_class, settings.vehicle_class_limits["t20"]
                )
                response = self._propose_v2_packing(
                    candidates,
                    effective_class=effective_class,
                    class_limits=class_limits,
                    limits_kg=limits_kg,
                )

            self._repo.update_propose_snapshot(
                cur,
                shipment_id,
                json.dumps(response.model_dump(), ensure_ascii=False),
            )
            conn.commit()
            return response
        except ShipmentError:
            conn.rollback()
            raise
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def _fetch_propose_candidates(
        self, cur: sqlite3.Cursor, shipment_id: int
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for order in self._fetch_orders(cur, [shipment_id]).get(shipment_id, []):
            if order.kp_id is None:
                continue
            cur.execute(
                """
                SELECT id, kp_id, plate_name, length_m, width_m, load_class,
                       qty, completed_date
                FROM completed_plates
                WHERE kp_id = ? AND qty > 0
                ORDER BY completed_date, id
                """,
                (order.kp_id,),
            )
            for cp in cur.fetchall():
                available = available_qty(cur, int(cp["id"]))
                if available <= 0:
                    continue
                unit, _ = resolve_kp_line_weight_kg(
                    {
                        "length_m": cp["length_m"],
                        "width_m": cp["width_m"],
                        "qty": 1,
                    }
                )
                rows.append(
                    {
                        "completed_plate_id": int(cp["id"]),
                        "kp_id": int(cp["kp_id"]),
                        "plate_name": str(cp["plate_name"] or ""),
                        "length_m": cp["length_m"],
                        "width_m": cp["width_m"],
                        "load_class": cp["load_class"],
                        "qty": available,
                        "unit_weight_kg": unit,
                        "completed_date": cp["completed_date"],
                    }
                )
        return rows

    @staticmethod
    def _propose_legacy_weight_fifo(
        candidates: list[dict[str, Any]],
        *,
        effective_class: str | None,
        limits_kg: dict[str, int],
    ) -> ShipmentProposeResponse:
        limit = limits_kg.get(effective_class) if effective_class else None
        fit: list[ShipmentProposeItem] = []
        not_fit: list[ShipmentProposeItem] = []
        accumulated = 0.0
        for row in candidates:
            available = int(row["qty"])
            unit = float(row["unit_weight_kg"])
            fit_qty, rest_qty = ShipmentService._split_by_limit(
                available, unit, limit, accumulated
            )
            accumulated += unit * fit_qty
            base = {
                "completed_plate_id": row["completed_plate_id"],
                "kp_id": row["kp_id"],
                "plate_name": row["plate_name"],
                "length_m": row["length_m"],
                "width_m": row["width_m"],
                "load_class": row["load_class"],
                "available_qty": available,
                "unit_weight_kg": unit,
                "completed_date": row["completed_date"],
            }
            if fit_qty > 0:
                fit.append(
                    ShipmentProposeItem(qty=fit_qty, weight_kg=unit * fit_qty, **base)
                )
            if rest_qty > 0:
                not_fit.append(
                    ShipmentProposeItem(
                        qty=rest_qty,
                        weight_kg=unit * rest_qty,
                        reason_code="weight_limit",
                        reason_text="Превышен лимит веса класса ТС",
                        **base,
                    )
                )

        total_weight = sum(item.weight_kg or 0.0 for item in fit)
        return ShipmentProposeResponse(
            items=fit,
            not_fit=not_fit,
            total_weight_kg=total_weight,
            overload=bool(limit is not None and total_weight > limit),
            vehicle_class=effective_class,
            vehicle_class_limits_kg=limits_kg,
        )

    @staticmethod
    def _propose_v2_packing(
        candidates: list[dict[str, Any]],
        *,
        effective_class: str,
        class_limits: VehicleClassLimits,
        limits_kg: dict[str, int],
    ) -> ShipmentProposeResponse:
        pack_candidates = [
            PlateCandidate(
                completed_plate_id=row["completed_plate_id"],
                kp_id=row["kp_id"],
                plate_name=row["plate_name"],
                length_m=row["length_m"],
                width_m=row["width_m"],
                load_class=row["load_class"],
                qty=int(row["qty"]),
                unit_weight_kg=float(row["unit_weight_kg"]),
                completed_date=row["completed_date"],
            )
            for row in candidates
        ]
        result = pack_shipment(
            pack_candidates,
            limits=VehicleLimits(
                max_weight_kg=class_limits.max_weight_kg,
                body_length_m=class_limits.body_length_m,
                max_tiers=class_limits.max_tiers,
            ),
        )

        def _to_item(line: Any) -> ShipmentProposeItem:
            return ShipmentProposeItem(
                completed_plate_id=line.completed_plate_id,
                kp_id=line.kp_id,
                plate_name=line.plate_name,
                length_m=line.length_m,
                width_m=line.width_m,
                load_class=line.load_class,
                qty=line.qty,
                available_qty=line.available_qty,
                unit_weight_kg=line.unit_weight_kg,
                weight_kg=line.weight_kg,
                completed_date=line.completed_date,
                reason_code=line.reason_code,
                reason_text=line.reason_text,
            )

        layout = None
        if result.layout is not None:
            layout = ShipmentLayoutMetadata(
                body_length_m=result.layout.body_length_m,
                body_used_m=result.layout.body_used_m,
                stacks=[
                    ShipmentLayoutStack(
                        index=stack.index,
                        marking_length_m=stack.marking_length_m,
                        tiers=[
                            ShipmentLayoutTier(
                                index=tier.index,
                                units=[
                                    ShipmentLayoutUnit(
                                        completed_plate_id=unit.completed_plate_id,
                                        kp_id=unit.kp_id,
                                        plate_name=unit.plate_name,
                                        width_m=unit.width_m,
                                    )
                                    for unit in tier.units
                                ],
                            )
                            for tier in stack.tiers
                        ],
                    )
                    for stack in result.layout.stacks
                ],
                loading_steps=[
                    ShipmentLoadingStep(
                        step=step.step,
                        stack_index=step.stack_index,
                        tier_index=step.tier_index,
                        description=step.description,
                    )
                    for step in result.layout.loading_steps
                ],
            )

        return ShipmentProposeResponse(
            items=[_to_item(line) for line in result.items],
            not_fit=[_to_item(line) for line in result.not_fit],
            order_remainder=[
                ShipmentOrderRemainderItem(
                    completed_plate_id=line.completed_plate_id,
                    kp_id=line.kp_id,
                    plate_name=line.plate_name,
                    qty_remaining=line.qty_remaining,
                )
                for line in result.order_remainder
            ],
            warnings=[
                ShipmentProposeWarning(
                    code=warning.code.value,
                    message=warning.message,
                    kp_ids=warning.kp_ids,
                )
                for warning in result.warnings
            ],
            total_weight_kg=result.total_weight_kg,
            overload=result.total_weight_kg > class_limits.max_weight_kg,
            vehicle_class=effective_class,
            vehicle_class_limits_kg=limits_kg,
            layout=layout,
        )

    @staticmethod
    def _split_by_limit(
        available: int,
        unit_weight: float,
        limit: float | None,
        accumulated: float,
    ) -> tuple[int, int]:
        """Сколько штук строки влезает в остаток лимита ТС (fit, rest)."""
        if limit is None or unit_weight <= 0:
            return available, 0
        remaining = limit - accumulated
        fit = min(available, int(remaining // unit_weight))
        fit = max(fit, 0)
        return fit, available - fit

    # ------------------------------------------------------------------
    # Confirm состава (SHIP-202)
    # ------------------------------------------------------------------

    # ------------------------------------------------------------------
    # Confirm / выезд / отмена — делегирование в ShipmentCompletionService
    # ------------------------------------------------------------------

    def put_items(
        self,
        shipment_id: int,
        items: list[ShipmentItemInput],
        *,
        actor: str | None = None,
    ) -> ShipmentCard:
        self._completion.put_items(shipment_id, items, actor=actor)
        return self.get(shipment_id)

    def complete(
        self,
        shipment_id: int,
        *,
        actor: str | None = None,
        events_enabled: bool | None = None,
        export_dir: str | None = None,
    ) -> ShipmentMutationResponse:
        return self._completion.complete(
            shipment_id,
            actor=actor,
            events_enabled=events_enabled,
            export_dir=export_dir,
        )

    def cancel(
        self,
        shipment_id: int,
        *,
        actor: str | None = None,
    ) -> ShipmentMutationResponse:
        return self._completion.cancel(shipment_id, actor=actor)


    def build_shipment_event_payload(self, shipment_id: int) -> dict[str, Any]:
        conn = self._repo.connect()
        try:
            cur = conn.cursor()
            row = self._fetch_shipment_row(cur, shipment_id)
            orders = self._fetch_orders(cur, [shipment_id]).get(shipment_id, [])
            items: list[dict[str, Any]] = []
            total_weight = 0.0
            for item in self._fetch_items(cur, shipment_id):
                weight = float(item.weight_kg or 0.0)
                total_weight += weight
                if item.item_type == ShipmentItemType.PLATE.value:
                    items.append(
                        {
                            "type": "plate",
                            "plate_name": item.plate_name,
                            "length_m": item.length_m,
                            "width_m": item.width_m,
                            "nomenclature_id": None,
                            "qty": item.qty,
                            "weight_kg": weight,
                        }
                    )
                else:
                    items.append(
                        {
                            "type": "free",
                            "mark": item.mark,
                            "qty": item.qty,
                            "weight_kg": weight,
                        }
                    )
            return {
                "event": "shipment_completed",
                "version": 1,
                "shipment_id": int(row["id"]),
                "shipment_date": row["shipment_date"],
                "completed_at": row["completed_at"],
                "delivery_type": row["delivery_type"],
                "orders": [
                    {"kp_id": order.kp_id, "ya_order_no": order.ya_order_no, "uid_kp": None}
                    for order in orders
                ],
                "items": items,
                "carrier": {"name": row["carrier_name"]} if row["carrier_name"] else None,
                "driver_name": row["driver_name"],
                "vehicle_text": row["vehicle_text"],
                "upd_no": row["upd_no"],
                "total_weight_kg": total_weight,
            }
        finally:
            conn.close()

    def _maybe_write_event(
        self,
        shipment_id: int,
        *,
        events_enabled: bool | None,
        export_dir: str | None,
    ) -> None:
        settings = get_settings()
        enabled = settings.shipment_events_enabled if events_enabled is None else events_enabled
        if not enabled:
            return
        try:
            payload = self.build_shipment_event_payload(shipment_id)
            directory = Path(export_dir) if export_dir else Path(settings.exchange_export_dir)
            directory.mkdir(parents=True, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%dT%H%M%S")
            path = directory / f"shipment_completed_{shipment_id}_{ts}.json"
            path.write_text(
                json.dumps(payload, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        except Exception:
            logger.exception(
                "Не удалось записать событие shipment_completed для рейса #%s",
                shipment_id,
            )

    # ------------------------------------------------------------------
    # Поиск КП для создания рейса (ACL B: в работе / На СГП)
    # ------------------------------------------------------------------

    def search_kp(
        self,
        *,
        kp_id: int | None = None,
        customer: str | None = None,
        limit: int = 50,
    ) -> LogisticsKpSearchResponse:
        from app.repositories.kp_offers_repository import KpOffersRepository

        repo = KpOffersRepository(db_path=self.db_path)
        if kp_id is not None:
            raw = repo.get_by_id(kp_id)
            status = (raw or {}).get("status") or ""
            rows = [raw] if raw and status in ALLOWED_LOGISTICS_KP_STATUSES else []
            return LogisticsKpSearchResponse(
                mode="number",
                items=[self._to_logistics_kp_search_item(r) for r in rows],
                total=len(rows),
                truncated=False,
            )

        name = (customer or "").strip()
        rows, total = repo.search_by_customer_name(
            name,
            limit=limit,
            readable_statuses=ALLOWED_LOGISTICS_KP_STATUSES,
        )
        return LogisticsKpSearchResponse(
            mode="customer",
            items=[self._to_logistics_kp_search_item(raw) for raw in rows],
            total=total,
            truncated=total > limit,
        )

    @staticmethod
    def _to_logistics_kp_search_item(raw: dict) -> LogisticsKpSearchItem:
        product = str(raw.get("product_type") or "plates").lower()
        if product not in ("plates", "piles"):
            product = "plates"
        return LogisticsKpSearchItem(
            kp_id=int(raw.get("kp_id") or 0),
            customer_name=raw.get("customer_name"),
            status=raw.get("status") or None,
            product_type=product,  # type: ignore[arg-type]
        )

    # ------------------------------------------------------------------
    # Каталог свай: автокомплит (SHIP-100/300)
    # ------------------------------------------------------------------

    def search_pile_catalog(self, q: str = "", *, limit: int = 20) -> PileCatalogResponse:
        query = (q or "").strip()
        ensure_schema(self.db_path)
        conn = _connect(self.db_path)
        try:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            if query:
                cur.execute(
                    """
                    SELECT id, mark, length_m, section_mm, volume_m3, weight_kg, pcs_per_20t
                    FROM pile_catalog
                    WHERE mark LIKE ?
                    ORDER BY mark
                    LIMIT ?
                    """,
                    (f"%{query}%", limit),
                )
            else:
                cur.execute(
                    """
                    SELECT id, mark, length_m, section_mm, volume_m3, weight_kg, pcs_per_20t
                    FROM pile_catalog
                    ORDER BY mark
                    LIMIT ?
                    """,
                    (limit,),
                )
            items = [
                PileCatalogItem(
                    id=int(row["id"]),
                    mark=row["mark"],
                    length_m=row["length_m"],
                    section_mm=row["section_mm"],
                    volume_m3=row["volume_m3"],
                    weight_kg=float(row["weight_kg"]),
                    pcs_per_20t=row["pcs_per_20t"],
                )
                for row in cur.fetchall()
            ]
        except sqlite3.Error as exc:
            logger.exception("Ошибка чтения каталога свай: %s", exc)
            raise ShipmentError(
                "Ошибка чтения каталога свай",
                code="pile_catalog_read_failed",
            ) from exc
        finally:
            conn.close()
        return PileCatalogResponse(items=items, count=len(items))

    # ------------------------------------------------------------------
    # Печатная форма «Лист отгрузки» (SHIP-500)
    # ------------------------------------------------------------------

    def export_shipment_sheet_xlsx(self, shipment_id: int) -> bytes:
        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Font

        card = self.get(shipment_id)
        wb = Workbook()
        ws = wb.active
        ws.title = "Лист отгрузки"

        bold = Font(bold=True)
        title = Font(bold=True, size=14)
        ws.append([f"Лист отгрузки № {card.id} от {card.shipment_date}"])
        ws["A1"].font = title
        delivery_label = (
            "Доставка" if card.delivery_type == DeliveryType.DELIVERY.value else "Самовывоз"
        )
        ya_orders = ", ".join(
            filter(
                None,
                [
                    f"{order.ya_order_no or 'без ЯР'} (КП #{order.kp_id})"
                    for order in card.orders
                ],
            )
        )
        customers = ", ".join(
            dict.fromkeys(
                order.customer_name for order in card.orders if order.customer_name
            )
        )
        carrier_line = card.carrier_name or ""
        if card.delivery_type == DeliveryType.PICKUP.value and card.proxy_no:
            carrier_line = f"Доверенность № {card.proxy_no}" + (
                f" ({carrier_line})" if carrier_line else ""
            )
        vehicle_line = card.vehicle_text or ""
        if card.vehicle_class:
            vehicle_line = f"{vehicle_line} [{card.vehicle_class}]".strip()

        ws.append([f"Тип выдачи: {delivery_label}"])
        ws.append([f"Заказы: {ya_orders}"])
        ws.append([f"Заказчик: {customers}"])
        ws.append([f"Перевозчик: {carrier_line}"])
        ws.append([f"Водитель: {card.driver_name or ''}"])
        ws.append([f"ТС: {vehicle_line}"])
        ws.append([])

        headers = ["№", "Марка / плита", "Размеры, м", "Кол-во, шт", "Вес, кг", "Примечание"]
        ws.append(headers)
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font = bold

        total_qty = 0
        total_weight = 0.0
        ordered_items = sorted(card.items, key=lambda item: (item.sort_order, item.id))
        for position, item in enumerate(ordered_items, start=1):
            if item.item_type == ShipmentItemType.PLATE.value:
                name = item.plate_name or f"СГП #{item.completed_plate_id}"
                dims = f"{item.length_m or 0:g} × {item.width_m or 0:g}"
            else:
                name = item.mark or ""
                dims = ""
            weight = float(item.weight_kg or 0.0)
            total_qty += item.qty
            total_weight += weight
            ws.append([position, name, dims, item.qty, weight, item.note or ""])

        totals_row = ws.max_row + 1
        ws.append(["", "Итого", "", total_qty, total_weight, ""])
        for cell in ws[totals_row]:
            cell.font = bold

        for column, width in zip("ABCDEF", (5, 28, 14, 12, 12, 30)):
            ws.column_dimensions[column].width = width

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ------------------------------------------------------------------
    # internals — SQL через ShipmentRepository; маппинг карточки здесь
    # ------------------------------------------------------------------

    def _insert_shipment_with_orders(
        self,
        cur: sqlite3.Cursor,
        *,
        shipment_date: str,
        delivery_type: str,
        kp_ids: list[int],
        actor: str | None,
    ) -> int:
        return self._repo.insert_shipment_with_orders(
            cur,
            shipment_date=shipment_date,
            delivery_type=delivery_type,
            kp_ids=kp_ids,
            actor=actor,
        )

    @staticmethod
    def _validate_date(value: str) -> None:
        try:
            datetime.strptime(value, _DATE_FORMAT)
        except ValueError:
            raise ShipmentError(
                f"Дата рейса должна быть в формате ГГГГ-ММ-ДД, получено «{value}»",
                code="shipment_invalid_date",
            ) from None

    @staticmethod
    def _validate_kp_ids(kp_ids: list[int]) -> list[int]:
        unique = list(dict.fromkeys(int(kp_id) for kp_id in kp_ids))
        if not unique:
            raise ShipmentError(
                "Рейс требует хотя бы один заказ (КП)",
                code="shipment_no_orders",
            )
        return unique

    def _assert_kp_exists(self, cur: sqlite3.Cursor, kp_ids: list[int]) -> None:
        self._repo.assert_kp_exists(cur, kp_ids)

    def _assert_carrier_exists(self, cur: sqlite3.Cursor, carrier_id: int) -> None:
        self._repo.assert_carrier_exists(cur, carrier_id)

    @staticmethod
    def _assert_in_work(row: sqlite3.Row, shipment_id: int) -> None:
        if row["status"] != ShipmentStatus.IN_WORK.value:
            raise ShipmentError(
                f"Рейс #{shipment_id} уже обработан — изменения недоступны",
                code="shipment_not_in_work",
            )

    def _fetch_shipment_row(self, cur: sqlite3.Cursor, shipment_id: int) -> sqlite3.Row:
        return self._repo.fetch_shipment_row(cur, shipment_id)

    def _fetch_orders(
        self, cur: sqlite3.Cursor, shipment_ids: list[int]
    ) -> dict[int, list[ShipmentOrderItem]]:
        return self._repo.fetch_orders(cur, shipment_ids)

    def _fetch_items(self, cur: sqlite3.Cursor, shipment_id: int) -> list[ShipmentItem]:
        return self._repo.fetch_items(cur, shipment_id)

    def _fetch_weight_totals(
        self, cur: sqlite3.Cursor, shipment_ids: list[int]
    ) -> dict[int, float]:
        return self._repo.fetch_weight_totals(cur, shipment_ids)

    def _available_by_kp(
        self, cur: sqlite3.Cursor, orders: list[ShipmentOrderItem]
    ) -> list[ShipmentAvailableByKp]:
        return self._repo.available_by_kp(cur, orders)

    def _replace_orders(
        self,
        cur: sqlite3.Cursor,
        shipment_id: int,
        orders: list[ShipmentOrderPatch],
    ) -> None:
        self._repo.replace_orders(cur, shipment_id, orders)

    def _prefill_ya_order_no(self, cur: sqlite3.Cursor, kp_id: int) -> str | None:
        return self._repo.prefill_ya_order_no(cur, kp_id)

    @staticmethod
    def _to_card(
        row: sqlite3.Row,
        orders: list[ShipmentOrderItem],
        items: list[ShipmentItem],
    ) -> ShipmentCard:
        total_weight = sum(float(item.weight_kg or 0.0) for item in items)
        return ShipmentCard(
            id=int(row["id"]),
            shipment_date=str(row["shipment_date"]),
            delivery_type=str(row["delivery_type"]),
            status=str(row["status"]),
            attention=bool(row["attention"]),
            attention_comment=row["attention_comment"],
            carrier_id=row["carrier_id"],
            carrier_name=row["carrier_name"],
            driver_name=row["driver_name"],
            vehicle_text=row["vehicle_text"],
            vehicle_class=row["vehicle_class"],
            proxy_no=row["proxy_no"],
            upd_no=row["upd_no"],
            freight_request_no=row["freight_request_no"],
            planned_cost=row["planned_cost"],
            time_slot=row["time_slot"],
            completed_at=row["completed_at"],
            actor=row["actor"],
            created_at=row["created_at"],
            orders=orders,
            items=items,
            total_weight_kg=total_weight,
        )

    @staticmethod
    def _to_list_item(
        row: sqlite3.Row,
        orders: list[ShipmentOrderItem],
        weights: dict[int, float],
    ) -> ShipmentListItem:
        shipment_id = int(row["id"])
        return ShipmentListItem(
            id=shipment_id,
            shipment_date=str(row["shipment_date"]),
            delivery_type=str(row["delivery_type"]),
            status=str(row["status"]),
            attention=bool(row["attention"]),
            attention_comment=row["attention_comment"],
            carrier_id=row["carrier_id"],
            carrier_name=row["carrier_name"],
            driver_name=row["driver_name"],
            vehicle_text=row["vehicle_text"],
            vehicle_class=row["vehicle_class"],
            proxy_no=row["proxy_no"],
            upd_no=row["upd_no"],
            freight_request_no=row["freight_request_no"],
            planned_cost=row["planned_cost"],
            time_slot=row["time_slot"],
            created_at=row["created_at"],
            orders=orders,
            total_weight_kg=float(weights.get(shipment_id, 0.0)),
        )

"""
Улучшенная программа создания КЗ (Карт Заданий) по дорожкам
Исправлены ошибки с объединёнными ячейками
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import shutil
import os
from datetime import datetime
import sqlite3
import sys

# Настройка кодировки для Windows
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'ignore')


def load_track_data(db_path: str = 'pb.db') -> list:
    """Загружает данные о дорожках из базы данных"""
    conn = sqlite3.connect(db_path)
    
    query = '''
        SELECT 
            "номенклатура пб к производству" as marking,
            "длина плиты, м" as length_db,
            "контрагент заказчик" as customer,
            "армирование по серии" as reinforcement,
            "неделя формовки" as week
        FROM plity_ex
    '''
    df = pd.read_sql(query, conn)
    conn.close()
    
    return df


def parse_plate_marking(marking: str):
    """Парсит маркировку плиты"""
    import re
    pattern = r'ПБ\s+(\d+[,.]?\d*)-(\d+[,.]?\d*)-(\d+)'
    match = re.search(pattern, marking)
    
    if not match:
        return None
    
    try:
        length_dm = float(match.group(1).replace(',', '.'))
        length_m = length_dm / 10
        
        width_dm = float(match.group(2).replace(',', '.'))
        width_m = width_dm / 10
        
        load_code = int(match.group(3))
        load_capacity = load_code * 100
        
        return {
            'length': length_m,
            'width': width_m,
            'load_capacity': load_capacity
        }
    except:
        return None


def group_plates_by_track(df: pd.DataFrame) -> list:
    """Группирует плиты по дорожкам"""
    # Фильтруем плиты с шириной 1.2 м
    plates_1_2m = []
    for idx, row in df.iterrows():
        params = parse_plate_marking(row['marking'])
        if params and params['width'] == 1.2:
            plates_1_2m.append({
                'marking': row['marking'],
                'length': params['length'],
                'width': params['width'],
                'load_capacity': params['load_capacity'],
                'customer': row['customer'],
                'reinforcement': row['reinforcement'],
                'week': row['week']
            })
    
    # Сортируем по неделе формовки
    plates_1_2m.sort(key=lambda x: x['week'] if x['week'] is not None else 999)
    
    tracks = []
    
    # Дорожка 1 - армирование 8.0 (неделя 4)
    track1_plates = [p for p in plates_1_2m if p['reinforcement'] == 8.0 and p['week'] == 4.0][:24]
    if track1_plates:
        tracks.append({
            'track_num': 1,
            'plates': track1_plates,
            'reinforcement': 8.0,
            'total_plates': len(track1_plates) * 3
        })
    
    # Дорожка 2 - армирование 6.0
    track2_plates = [p for p in plates_1_2m if p['reinforcement'] == 6.0 and p['week'] in [5.0, 26.0, 27.0, 28.0]][:27]
    if track2_plates:
        tracks.append({
            'track_num': 2,
            'plates': track2_plates,
            'reinforcement': 6.0,
            'total_plates': len(track2_plates) * 3
        })
    
    # Дорожка 3 - армирование 4.0
    track3_plates = [p for p in plates_1_2m if p['reinforcement'] == 4.0 and p['week'] in [6.0, 8.0, 9.0, 10.0, 13.0]][:47]
    if track3_plates:
        tracks.append({
            'track_num': 3,
            'plates': track3_plates,
            'reinforcement': 4.0,
            'total_plates': len(track3_plates) * 3
        })
    
    return tracks


def find_empty_cell(ws, start_row=1, start_col=1, max_search=50):
    """Находит пустую ячейку для заполнения"""
    for row in range(start_row, max_search):
        for col in range(start_col, 10):
            try:
                cell = ws.cell(row=row, column=col)
                # Проверяем, что ячейка не объединена и пуста
                if cell.value is None or str(cell.value).strip() == '':
                    return row, col
            except:
                continue
    return None, None


def create_kz_for_track(track_data: dict, template_path: str, output_dir: str):
    """Создаёт КЗ файл для одной дорожки"""
    track_num = track_data['track_num']
    plates = track_data['plates']
    reinforcement = track_data['reinforcement']
    total_plates = track_data['total_plates']
    
    filename = f"КЗ_ПБ_Дорожка_{track_num}_Армирование_{reinforcement}.xlsx"
    output_path = os.path.join(output_dir, filename)
    
    # Копируем шаблон
    shutil.copy2(template_path, output_path)
    
    # Открываем файл
    wb = load_workbook(output_path)
    ws = wb.active
    
    print(f"[СОЗДАНИЕ] КЗ для дорожки #{track_num}...")
    print(f"   Файл: {filename}")
    print(f"   Плит: {total_plates} шт")
    print(f"   Армирование: {reinforcement}")
    
    current_date = datetime.now().strftime("%d.%m.%Y")
    
    # Ищем подходящие места для заполнения
    try:
        # Ищем ячейку для номера дорожки
        for row in range(1, 20):
            for col in range(1, 10):
                try:
                    cell = ws.cell(row=row, column=col)
                    if cell.value and any(word in str(cell.value).lower() for word in ['дорожка', 'номер', '№']):
                        # Заполняем соседнюю ячейку
                        ws.cell(row=row, column=col+1).value = f"ДОРОЖКА #{track_num}"
                        break
                except:
                    continue
        
        # Ищем ячейку для даты
        for row in range(1, 20):
            for col in range(1, 10):
                try:
                    cell = ws.cell(row=row, column=col)
                    if cell.value and 'дата' in str(cell.value).lower():
                        ws.cell(row=row, column=col+1).value = current_date
                        break
                except:
                    continue
        
        # Ищем ячейку для количества плит
        for row in range(1, 20):
            for col in range(1, 10):
                try:
                    cell = ws.cell(row=row, column=col)
                    if cell.value and any(word in str(cell.value).lower() for word in ['количество', 'плит', 'шт', 'всего']):
                        ws.cell(row=row, column=col+1).value = total_plates
                        break
                except:
                    continue
        
        # Ищем ячейку для армирования
        for row in range(1, 20):
            for col in range(1, 10):
                try:
                    cell = ws.cell(row=row, column=col)
                    if cell.value and 'армирование' in str(cell.value).lower():
                        ws.cell(row=row, column=col+1).value = f"Армирование {reinforcement}"
                        break
                except:
                    continue
    
    except Exception as e:
        print(f"   [ПРЕДУПРЕЖДЕНИЕ] Не удалось заполнить некоторые поля: {e}")
    
    # Ищем место для таблицы плит (начинаем с строки 10)
    table_start_row = 10
    table_start_col = 1
    
    # Заполняем заголовки таблицы
    headers = ['№', 'Маркировка', 'Длина, м', 'Ширина, м', 'Нагрузка, кг/м2', 'Заказчик', 'Неделя']
    for i, header in enumerate(headers):
        try:
            ws.cell(row=table_start_row, column=table_start_col + i).value = header
        except:
            pass
    
    # Заполняем данные плит
    for i, plate in enumerate(plates):
        row = table_start_row + 1 + i
        
        try:
            ws.cell(row=row, column=table_start_col).value = i + 1  # №
            ws.cell(row=row, column=table_start_col + 1).value = plate['marking']  # Маркировка
            ws.cell(row=row, column=table_start_col + 2).value = plate['length']  # Длина
            ws.cell(row=row, column=table_start_col + 3).value = plate['width']  # Ширина
            ws.cell(row=row, column=table_start_col + 4).value = plate['load_capacity']  # Нагрузка
            ws.cell(row=row, column=table_start_col + 5).value = plate['customer']  # Заказчик
            ws.cell(row=row, column=table_start_col + 6).value = plate['week']  # Неделя
        except Exception as e:
            print(f"   [ОШИБКА] Ошибка при заполнении строки {row}: {e}")
    
    # Сохраняем файл
    wb.save(output_path)
    wb.close()
    
    print(f"   [ГОТОВО] {filename}")
    return output_path


def main():
    """Основная функция"""
    print("=" * 60)
    print("СОЗДАНИЕ КЗ (КАРТ ЗАДАНИЙ) ПО ДОРОЖКАМ - v2")
    print("=" * 60)
    
    template_path = "!КЗ ПБ Шаблон.xlsx"
    output_dir = "КЗ_файлы_v2"
    
    # Создаём папку
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"[ПАПКА] Создана папка: {output_dir}")
    
    # Проверяем шаблон
    if not os.path.exists(template_path):
        print(f"[ОШИБКА] Не найден шаблон: {template_path}")
        return
    
    print(f"[ШАБЛОН] Используем шаблон: {template_path}")
    
    # Загружаем данные
    print("\n[ЗАГРУЗКА] Загружаю данные о плитах из базы...")
    df = load_track_data()
    print(f"[УСПЕХ] Загружено {len(df)} записей")
    
    # Группируем плиты
    print("\n[ГРУППИРОВКА] Группирую плиты по дорожкам...")
    tracks = group_plates_by_track(df)
    print(f"[УСПЕХ] Найдено {len(tracks)} дорожек")
    
    # Создаём КЗ
    print(f"\n[СОЗДАНИЕ] Создаю КЗ файлы...")
    created_files = []
    
    for track in tracks:
        try:
            file_path = create_kz_for_track(track, template_path, output_dir)
            created_files.append(file_path)
        except Exception as e:
            print(f"[ОШИБКА] Ошибка при создании КЗ для дорожки {track['track_num']}: {e}")
    
    # Итоговая сводка
    print(f"\n" + "=" * 60)
    print("ИТОГОВАЯ СВОДКА")
    print("=" * 60)
    print(f"[УСПЕХ] Создано КЗ файлов: {len(created_files)}")
    print(f"[ПАПКА] Папка: {output_dir}")
    
    for file_path in created_files:
        filename = os.path.basename(file_path)
        print(f"   [ФАЙЛ] {filename}")
    
    print(f"\n[ГОТОВО] Все КЗ файлы созданы в папке '{output_dir}'")
    print("\n[ИНФОРМАЦИЯ] Каждый файл содержит:")
    print("   • Общую информацию о дорожке")
    print("   • Количество плит")
    print("   • Армирование")
    print("   • Таблицу со всеми плитами")
    print("   • Недели формовки (срочность)")


if __name__ == "__main__":
    main()




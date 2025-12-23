#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Чтение данных о себестоимости из Excel-файла

Работает с листами:
- "Стоимость": прямые затраты на материалы
- "Себестоимость": КЭФ и общие затраты

ВАЖНО: НЕ парсит размеры плит! Использует существующий парсер из config_and_data.py
"""

import re
from typing import Optional, List, Dict, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def find_header_row(sheet: Worksheet, keyword: str = 'Наименование') -> Optional[int]:
    """
    Находит строку с заголовками таблицы.
    
    Args:
        sheet: Лист Excel
        keyword: Ключевое слово для поиска (например, "Наименование")
    
    Returns:
        Номер строки (1-indexed) или None
    """
    for row_idx in range(1, min(20, sheet.max_row + 1)):
        for cell in sheet[row_idx]:
            if cell.value and keyword.lower() in str(cell.value).lower():
                return row_idx
    return None


def find_column_by_keywords(
    sheet: Worksheet,
    header_row: int,
    keywords: List[str],
    max_rows: int = 3
) -> Optional[int]:
    """
    Находит колонку по ключевым словам в многострочном заголовке.
    
    Args:
        sheet: Лист Excel
        header_row: Начальная строка заголовка
        keywords: Список ключевых слов (все должны встретиться)
        max_rows: Максимум строк для поиска в заголовке
    
    Returns:
        Номер колонки (1-indexed) или None
    """
    for col_idx in range(1, sheet.max_column + 1):
        # Собираем текст из нескольких строк заголовка
        text_parts = []
        for row_offset in range(max_rows):
            cell = sheet.cell(header_row + row_offset, col_idx)
            if cell.value:
                text_parts.append(str(cell.value).strip())
        
        combined_text = ' '.join(text_parts).lower()
        
        # Проверяем, что ВСЕ ключевые слова встречаются
        if all(kw.lower() in combined_text for kw in keywords):
            return col_idx
    
    return None


def read_cost_sheet(sheet: Worksheet) -> List[Dict]:
    """
    Читает лист "Стоимость" с прямыми затратами.
    
    Формат листа:
    - Заголовок многострочный (обычно 3 строки)
    - Колонки: Наименование ЖБИ, Длина, Бетон|марка, Бетон|объем,
      Армирование|Стоимость, Бетон|Стоимость, Петли|Стоимость,
      Изоформ-Б, Итого стоимость
    
    Returns:
        Список словарей с данными о себестоимости
    """
    # Находим заголовок
    header_row = find_header_row(sheet, 'Наименование')
    if not header_row:
        print("[EXCEL] ⚠️ Не найден заголовок на листе 'Стоимость'")
        return []
    
    print(f"[EXCEL] Заголовок найден в строке {header_row}")
    
    # Находим нужные колонки
    col_name = find_column_by_keywords(sheet, header_row, ['Наименование', 'ЖБИ'])
    col_length = find_column_by_keywords(sheet, header_row, ['Длина', 'дм'])
    col_concrete_grade = find_column_by_keywords(sheet, header_row, ['Бетон', 'марка'])
    col_concrete_volume = find_column_by_keywords(sheet, header_row, ['Бетон', 'объем', 'м'])
    col_reinforcement_cost = find_column_by_keywords(sheet, header_row, ['Армирование', 'Стоимость'])
    col_concrete_cost = find_column_by_keywords(sheet, header_row, ['Бетон', 'Стоимость'])
    col_loops_cost = find_column_by_keywords(sheet, header_row, ['Петли', 'Стоимость'])
    col_izoform_cost = find_column_by_keywords(sheet, header_row, ['Изоформ'])
    col_total_cost = find_column_by_keywords(sheet, header_row, ['Итого', 'стоимость'])
    
    # Проверяем, что нашли все критичные колонки
    if not all([col_name, col_length, col_total_cost]):
        print("[EXCEL] ⚠️ Не найдены критичные колонки (Наименование, Длина, Итого)")
        return []
    
    print(f"[EXCEL] Колонки найдены: name={col_name}, length={col_length}, "
          f"total={col_total_cost}")
    
    # Читаем данные (начинаем со строки после заголовка + 2 для пропуска подзаголовков)
    data_start_row = header_row + 3
    results = []
    
    for row_idx in range(data_start_row, sheet.max_row + 1):
        # Читаем наименование
        plate_name_cell = sheet.cell(row_idx, col_name)
        plate_name = str(plate_name_cell.value or '').strip()
        
        # Пропускаем пустые строки и заголовки
        if not plate_name or len(plate_name) < 5:
            continue
        
        # Пропускаем строки, не похожие на плиты
        if not any(x in plate_name.upper() for x in ['ПБ', 'ПК', 'ПЛИТ']):
            continue
        
        # Читаем длину (для валидации)
        length_dm = None
        if col_length:
            length_val = sheet.cell(row_idx, col_length).value
            try:
                length_dm = int(float(length_val)) if length_val else None
            except (ValueError, TypeError):
                pass
        
        # Читаем марку бетона
        concrete_grade = None
        if col_concrete_grade:
            grade_val = sheet.cell(row_idx, col_concrete_grade).value
            concrete_grade = str(grade_val).strip() if grade_val else None
        
        # Читаем объём бетона
        volume_m3 = None
        if col_concrete_volume:
            vol_val = sheet.cell(row_idx, col_concrete_volume).value
            try:
                volume_m3 = float(vol_val) if vol_val else None
            except (ValueError, TypeError):
                pass
        
        # Читаем стоимости компонентов
        reinforcement_cost = 0.0
        if col_reinforcement_cost:
            val = sheet.cell(row_idx, col_reinforcement_cost).value
            try:
                reinforcement_cost = float(val) if val else 0.0
            except (ValueError, TypeError):
                pass
        
        concrete_cost = 0.0
        if col_concrete_cost:
            val = sheet.cell(row_idx, col_concrete_cost).value
            try:
                concrete_cost = float(val) if val else 0.0
            except (ValueError, TypeError):
                pass
        
        loops_cost = 0.0
        if col_loops_cost:
            val = sheet.cell(row_idx, col_loops_cost).value
            try:
                loops_cost = float(val) if val else 0.0
            except (ValueError, TypeError):
                pass
        
        izoform_cost = 0.0
        if col_izoform_cost:
            val = sheet.cell(row_idx, col_izoform_cost).value
            try:
                izoform_cost = float(val) if val else 0.0
            except (ValueError, TypeError):
                pass
        
        # Читаем итоговую стоимость
        total_cost = None
        if col_total_cost:
            val = sheet.cell(row_idx, col_total_cost).value
            try:
                total_cost = float(val) if val else None
            except (ValueError, TypeError):
                pass
        
        # Если нет итоговой стоимости, пропускаем
        if total_cost is None or total_cost <= 0:
            continue
        
        # Формируем запись
        results.append({
            'plate_name_excel': plate_name,
            'length_dm': length_dm,
            'concrete_grade': concrete_grade,
            'volume_m3': volume_m3,
            'reinforcement_cost': reinforcement_cost,
            'concrete_cost': concrete_cost,
            'loops_cost': loops_cost,
            'izoform_cost': izoform_cost,
            'direct_cost': total_cost,
            'source_row': row_idx,
        })
    
    print(f"[EXCEL] ✓ Прочитано {len(results)} записей с листа 'Стоимость'")
    return results


def read_kef_from_costing_sheet(sheet: Worksheet) -> Optional[float]:
    """
    Читает КЭФ (коэффициент) с листа "Себестоимость".
    
    КЭФ используется для добавления общих затрат к прямым затратам.
    
    Returns:
        Значение КЭФ или None
    """
    # Ищем ячейку с текстом "КЭФ" или "K.Э.Ф" или "коэффициент"
    for row_idx in range(1, min(100, sheet.max_row + 1)):
        for col_idx in range(1, min(20, sheet.max_column + 1)):
            cell = sheet.cell(row_idx, col_idx)
            cell_text = str(cell.value or '').strip().upper()
            
            # Расширяем поиск: КЭФ, коэффициент, КОЭФ, накладные
            if any(kw in cell_text for kw in ['КЭФ', 'K.Э.Ф', 'KEF', 'КОЭФ', 'НАКЛАДН', 'ОБЩИХ ЗАТРАТ']):
                # Ищем значение справа, снизу или в той же ячейке (через =)
                candidates = [
                    (row_idx, col_idx),      # та же ячейка (может быть формат "КЭФ: 1.25")
                    (row_idx, col_idx + 1),  # справа
                    (row_idx, col_idx + 2),  # справа через 1
                    (row_idx + 1, col_idx),  # снизу
                    (row_idx + 1, col_idx + 1),  # по диагонали
                ]
                
                for r, c in candidates:
                    try:
                        val_cell = sheet.cell(r, c)
                        cell_str = str(val_cell.value or '').strip()
                        
                        # Пытаемся извлечь число из строки типа "КЭФ: 1.25" или "1,25"
                        # Заменяем запятую на точку для float()
                        cell_str = cell_str.replace(',', '.')
                        
                        # Ищем первое число в строке
                        import re
                        match = re.search(r'(\d+[.,]?\d*)', cell_str)
                        if match:
                            kef = float(match.group(1).replace(',', '.'))
                            if 1.0 <= kef <= 3.0:  # Разумный диапазон для КЭФ
                                print(f"[EXCEL] ✓ КЭФ найден: {kef} (строка {r}, колонка {c})")
                                return kef
                    except (ValueError, TypeError, AttributeError):
                        continue
    
    print("[EXCEL] ⚠️ КЭФ не найден на листе 'Себестоимость'")
    print("[EXCEL] ℹ️  Будет использован КЭФ = 1.0 (без накладных)")
    return 1.0  # Дефолтное значение


def read_factory_costs_from_excel(xlsx_path: str) -> Tuple[List[Dict], Optional[float]]:
    """
    Главная функция: читает себестоимость из Excel.
    
    Args:
        xlsx_path: Путь к файлу Excel
    
    Returns:
        (список записей о себестоимости, значение КЭФ)
    """
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"[EXCEL] ❌ Не удалось открыть файл {xlsx_path}: {e}")
        return [], None
    
    # Читаем лист "Стоимость"
    cost_sheet = None
    for name in ['Стоимость', 'стоимость', 'Cost']:
        if name in wb.sheetnames:
            cost_sheet = wb[name]
            break
    
    if not cost_sheet:
        print("[EXCEL] ❌ Лист 'Стоимость' не найден")
        return [], None
    
    cost_data = read_cost_sheet(cost_sheet)
    
    # Читаем КЭФ с листа "Себестоимость"
    kef = None
    costing_sheet = None
    for name in ['Себестоимость', 'себестоимость', 'Costing']:
        if name in wb.sheetnames:
            costing_sheet = wb[name]
            break
    
    if costing_sheet:
        kef = read_kef_from_costing_sheet(costing_sheet)
    else:
        print("[EXCEL] ⚠️ Лист 'Себестоимость' не найден, КЭФ не загружен")
    
    wb.close()
    
    return cost_data, kef


if __name__ == '__main__':
    # Тест чтения
    import os
    test_path = os.path.join(
        os.path.dirname(os.path.dirname(__file__)),
        'банк знаний',
        'Расчет новых цен на ПБ 10.09.2025 (1).xlsx'
    )
    
    if os.path.exists(test_path):
        print(f"[TEST] Чтение файла: {test_path}")
        data, kef = read_factory_costs_from_excel(test_path)
        print(f"\n=== Результат ===")
        print(f"Записей: {len(data)}")
        print(f"КЭФ: {kef}")
        if data:
            print(f"\nПример первой записи:")
            for k, v in list(data[0].items())[:5]:
                print(f"  {k}: {v}")
    else:
        print(f"[TEST] ❌ Файл не найден: {test_path}")


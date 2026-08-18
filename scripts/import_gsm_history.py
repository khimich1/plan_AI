#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Одноразовый импорт истории ГСМ → gsm_* таблицы.

Источники (только чтение):
  • Роману.xlsx — машины и топливные карты
  • пул_поездок.xlsx / лист routes_ab — библиотека маршрутов
  • geo_cache/stations.geojson — АЗС с координатами
  • заполненные ПЛ *.xls — водители + крайний confirmed waybill на машину

Полная история ПЛ не импортируется (D9): только крайний confirmed PL
на каждую машину как seed одометра/остатка.

Пример:
  python scripts/import_gsm_history.py --db plita.db --gsm-dir "ГСМ"
"""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Mapping

import xlrd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.repositories.gsm_repository import GsmRepository
from core.kp_db_schema import ensure_schema
from scripts.build_gsm_trip_pool import (
    VehicleInfo,
    _as_str,
    _cell,
    iter_waybill_files,
    load_registry,
    normalize_plate,
    parse_date_from_name,
    parse_waybill,
)

DEFAULT_GSM_DIR = PROJECT_ROOT / "ГСМ"
ROMANU_NAME = "Роману.xlsx"
POOL_NAME = "пул_поездок.xlsx"
ROUTES_SHEET = "routes_ab"
STATIONS_REL = Path("geo_cache") / "stations.geojson"

# Ячейки бланка ОКУД 0345001 (0-based, xlrd; проверено на реальных ПЛ).
_CELL_DRIVER_FIO = (11, 17)
_CELL_LICENSE = (13, 23)
_CELL_SNILS_LINE = (16, 5)
_CELL_PERSONNEL = (11, 72)
_CELL_ODO_START = (19, 77)  # BZ20
_CELL_ODO_END = (44, 76)  # BY45
_CELL_FUEL_START = (33, 70)  # BS34
_CELL_FUEL_ISSUED = (37, 70)  # BS38
_CELL_FUEL_END = (38, 70)  # BS39

_LATIN_TO_CYR = str.maketrans(
    {
        "C": "С",  # U+0043 → U+0421
        "c": "с",
    }
)

_LICENSE_RE = re.compile(
    r"(?P<num>[\d\s]+?)\s*,?\s*выдано\s*(?P<issued>[\d.]+)",
    re.IGNORECASE,
)
_SNILS_RE = re.compile(r"СНИЛС\s*([\d\- ]+)", re.IGNORECASE)
_FLOAT_RE = re.compile(r"[\d]+(?:[.,]\d+)?")


@dataclass
class ImportReport:
    """Сводка импорта; conflicts — для ручного разбора (Лоншакова и т.п.)."""

    vehicles: int = 0
    cards: int = 0
    drivers: int = 0
    routes: int = 0
    waybills: int = 0
    stations: int = 0
    conflicts: list[dict[str, Any]] = field(default_factory=list)


def normalize_driver_name(name: str) -> str:
    """Нормализация ФИО: Latin C→С, trim, схлопывание пробелов."""
    s = (name or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s.translate(_LATIN_TO_CYR)


def _record_get(record: Any, *keys: str) -> Any:
    if isinstance(record, Mapping):
        for key in keys:
            if key in record and record[key] not in (None, ""):
                return record[key]
        return None
    for key in keys:
        if hasattr(record, key):
            val = getattr(record, key)
            if val not in (None, ""):
                return val
    return None


def build_driver_conflicts(records: Iterable[Any]) -> list[dict[str, Any]]:
    """Конфликты реквизитов водителей (разные ВУ при одном ФИО и т.п.)."""
    by_name: dict[str, list[Any]] = defaultdict(list)
    for rec in records:
        full_name = _record_get(rec, "full_name", "name", "driver")
        if not full_name:
            continue
        by_name[normalize_driver_name(str(full_name))].append(rec)

    conflicts: list[dict[str, Any]] = []
    for name, items in sorted(by_name.items()):
        licenses = sorted(
            {
                str(_record_get(item, "license_number", "license")).strip()
                for item in items
                if _record_get(item, "license_number", "license")
            }
        )
        if len(licenses) > 1:
            conflicts.append(
                {
                    "type": "license_conflict",
                    "full_name": name,
                    "license_numbers": licenses,
                    "message": (
                        f"Удостоверение {name}: "
                        + " vs ".join(licenses)
                    ),
                }
            )

        snils_values = [
            _record_get(item, "snils")
            for item in items
            if _record_get(item, "snils")
        ]
        if not snils_values:
            conflicts.append(
                {
                    "type": "missing_snils",
                    "full_name": name,
                    "message": f"Нет СНИЛС у {name}",
                }
            )

        personnel_values = [
            _record_get(item, "personnel_number", "personnel")
            for item in items
            if _record_get(item, "personnel_number", "personnel")
        ]
        if not personnel_values:
            conflicts.append(
                {
                    "type": "missing_personnel",
                    "full_name": name,
                    "message": f"Нет табельного номера у {name}",
                }
            )

    return conflicts


collect_driver_conflicts = build_driver_conflicts


def _parse_liters(raw: str) -> float:
    s = (raw or "").strip().replace(",", ".")
    m = _FLOAT_RE.search(s)
    if not m:
        return 0.0
    return float(m.group(0).replace(",", "."))


def _parse_float_cell(value: Any) -> float | None:
    if value in ("", None):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _parse_int_cell(value: Any) -> int | None:
    num = _parse_float_cell(value)
    if num is None:
        return None
    return int(round(num))


def _parse_license(raw: str) -> tuple[str, str | None]:
    s = _as_str(raw)
    if not s:
        return "", None
    m = _LICENSE_RE.search(s)
    if m:
        num = re.sub(r"\s+", " ", m.group("num").strip())
        issued = m.group("issued").rstrip("г.").rstrip(".")
        return num, issued
    return s, None


def _parse_snils(raw: str) -> str | None:
    s = _as_str(raw)
    m = _SNILS_RE.search(s)
    if not m:
        return None
    digits = re.sub(r"[^\d]", "", m.group(1))
    if len(digits) == 11:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:9]}-{digits[9:]}"
    return re.sub(r"\s+", "", m.group(1).strip())


def _split_card_numbers(raw: str) -> list[str]:
    parts = re.split(r"[,;/]+", raw or "")
    out: list[str] = []
    for part in parts:
        s = part.strip()
        if not s:
            continue
        if isinstance(s, str) and s.endswith(".0") and s[:-2].isdigit():
            s = s[:-2]
        # openpyxl may give int-like floats already stringified by registry
        try:
            if "." not in s and float(s) == int(float(s)):
                s = str(int(float(s)))
        except ValueError:
            pass
        out.append(s)
    return out


def _extract_driver_from_sheet(sh: xlrd.sheet.Sheet) -> dict[str, Any]:
    fio = _as_str(_cell(sh, *_CELL_DRIVER_FIO))
    license_raw = _as_str(_cell(sh, *_CELL_LICENSE))
    license_number, license_issued_at = _parse_license(license_raw)
    snils = _parse_snils(_as_str(_cell(sh, *_CELL_SNILS_LINE)))
    personnel = _as_str(_cell(sh, *_CELL_PERSONNEL))
    return {
        "full_name": normalize_driver_name(fio) if fio else "",
        "raw_name": fio,
        "license_number": license_number,
        "license_issued_at": license_issued_at,
        "snils": snils,
        "personnel_number": personnel or None,
    }


def _extract_balance_from_sheet(sh: xlrd.sheet.Sheet) -> dict[str, Any]:
    return {
        "odometer_start": _parse_int_cell(_cell(sh, *_CELL_ODO_START)),
        "odometer_end": _parse_int_cell(_cell(sh, *_CELL_ODO_END)),
        "fuel_start": _parse_float_cell(_cell(sh, *_CELL_FUEL_START)),
        "fuel_issued": _parse_float_cell(_cell(sh, *_CELL_FUEL_ISSUED)),
        "fuel_end": _parse_float_cell(_cell(sh, *_CELL_FUEL_END)),
    }


def _pick_most_common(values: Iterable[Any]) -> Any | None:
    cleaned = [v for v in values if v not in (None, "")]
    if not cleaned:
        return None
    return Counter(cleaned).most_common(1)[0][0]


def _merge_driver_records(records: list[dict[str, Any]]) -> dict[str, Any]:
    name = _pick_most_common(r["full_name"] for r in records) or ""
    return {
        "full_name": name,
        "license_number": _pick_most_common(r["license_number"] for r in records) or "",
        "license_issued_at": _pick_most_common(
            r.get("license_issued_at") for r in records
        ),
        "snils": _pick_most_common(r.get("snils") for r in records),
        "personnel_number": _pick_most_common(
            r.get("personnel_number") for r in records
        ),
    }


def _import_vehicles_and_cards(
    repo: GsmRepository,
    registry: dict[str, VehicleInfo],
    assigned_at: str,
) -> dict[str, int]:
    """Идемпотентно создаёт машины и карты. Возвращает plate_norm → vehicle_id."""
    existing_by_plate = {
        normalize_plate(v["plate_number"]): int(v["id"])
        for v in repo.list_vehicles(active_only=False)
    }
    existing_cards = {
        str(c["card_number"]): int(c["id"])
        for c in repo.list_cards(include_archived=True)
    }
    plate_to_id: dict[str, int] = dict(existing_by_plate)

    for plate_norm, info in registry.items():
        vehicle_id = plate_to_id.get(plate_norm)
        if vehicle_id is None:
            vehicle_id = repo.create_vehicle(
                name=info.mark or info.plate,
                plate_number=info.plate,
                tank_volume_liters=_parse_liters(info.tank_l),
                norm_summer=_parse_liters(info.norm_summer),
                norm_winter=_parse_liters(info.norm_winter),
            )
            plate_to_id[plate_norm] = vehicle_id

        for card_number in _split_card_numbers(info.fuel_card):
            if card_number in existing_cards:
                continue
            card_id = repo.create_card(
                card_number=card_number,
                vehicle_id=vehicle_id,
                assigned_at=assigned_at,
            )
            existing_cards[card_number] = card_id

    return plate_to_id


def _import_stations(repo: GsmRepository, geojson_path: Path) -> int:
    if not geojson_path.is_file():
        return 0
    data = json.loads(geojson_path.read_text(encoding="utf-8"))
    existing = {
        str(s["address"]).strip().lower(): int(s["id"])
        for s in repo.list_stations()
    }
    created = 0
    for feature in data.get("features") or []:
        props = feature.get("properties") or {}
        address = _as_str(props.get("адрес") or props.get("address"))
        if not address:
            continue
        key = address.strip().lower()
        if key in existing:
            continue
        geometry = feature.get("geometry") or {}
        coords = geometry.get("coordinates") or []
        lon = float(coords[0]) if len(coords) >= 2 else None
        lat = float(coords[1]) if len(coords) >= 2 else None
        brand = _as_str(props.get("бренд") or props.get("brand")) or None
        source = (
            _as_str(props.get("источник_координат") or props.get("geocode_source"))
            or None
        )
        station_id = repo.create_station(
            address=address,
            brand=brand,
            lat=lat,
            lon=lon,
            geocode_source=source,
        )
        existing[key] = station_id
        created += 1
    return created


def _load_routes_ab(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        return []
    wb = load_workbook(path, data_only=True, read_only=True)
    if ROUTES_SHEET not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[ROUTES_SHEET]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        return []
    headers = [(_as_str(h) if h is not None else "") for h in header_row]
    out: list[dict[str, Any]] = []
    for row in rows_iter:
        cells = list(row)
        if not any(c is not None and str(c).strip() for c in cells):
            continue
        item = {
            headers[i]: cells[i] if i < len(cells) else None
            for i in range(len(headers))
            if headers[i]
        }
        out.append(item)
    wb.close()
    return out


def _import_routes(
    repo: GsmRepository,
    routes_rows: list[dict[str, Any]],
    plate_to_id: dict[str, int],
) -> int:
    existing_keys: set[tuple[int, str, str, int]] = set()
    for route in repo.list_routes():
        existing_keys.add(
            (
                int(route["vehicle_id"]),
                str(route["addr_a"]),
                str(route["addr_b"]),
                int(route["km"]),
            )
        )

    created = 0
    for row in routes_rows:
        plate = normalize_plate(_as_str(row.get("гос_номер")))
        vehicle_id = plate_to_id.get(plate)
        if vehicle_id is None:
            continue
        addr_a = _as_str(row.get("адрес_A"))
        addr_b = _as_str(row.get("адрес_B"))
        km_raw = row.get("км")
        km_val = _parse_float_cell(km_raw)
        if not addr_a or not addr_b or km_val is None:
            continue
        km = int(round(km_val))
        key = (vehicle_id, addr_a, addr_b, km)
        if key in existing_keys:
            continue
        freq_raw = _parse_float_cell(row.get("частота"))
        frequency = int(freq_raw) if freq_raw is not None else 1
        repo.create_route(
            vehicle_id=vehicle_id,
            addr_a=addr_a,
            addr_b=addr_b,
            km=km,
            frequency=max(frequency, 1),
        )
        existing_keys.add(key)
        created += 1
    return created


def _collect_pl_data(
    gsm_dir: Path,
) -> tuple[list[dict[str, Any]], dict[str, tuple[date, Path, dict[str, Any]]]]:
    """Собирает записи водителей и крайний ПЛ на каждую папку машины."""
    driver_records: list[dict[str, Any]] = []
    latest_by_folder: dict[str, tuple[date, Path, dict[str, Any]]] = {}

    for vehicle_folder, path in iter_waybill_files(gsm_dir):
        pl_date = parse_date_from_name(path.name)
        if pl_date is None:
            continue
        book = xlrd.open_workbook(str(path), formatting_info=False)
        sh = book.sheet_by_index(0)
        driver = _extract_driver_from_sheet(sh)
        if driver["full_name"] and driver["license_number"]:
            driver_records.append(driver)

        balance = _extract_balance_from_sheet(sh)
        meta, trips = parse_waybill(path, vehicle_folder)
        payload = {
            "vehicle_folder": vehicle_folder,
            "plate": meta.get("plate") or "",
            "driver": driver,
            "balance": balance,
            "trips": trips,
            "path": str(path),
        }
        prev = latest_by_folder.get(vehicle_folder)
        if prev is None or pl_date > prev[0]:
            latest_by_folder[vehicle_folder] = (pl_date, path, payload)

    return driver_records, latest_by_folder


def _import_drivers(
    repo: GsmRepository,
    driver_records: list[dict[str, Any]],
) -> dict[str, int]:
    """Идемпотентно создаёт водителей. Ключ — normalize_driver_name."""
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for rec in driver_records:
        by_name[normalize_driver_name(rec["full_name"])].append(rec)

    existing_by_name = {
        normalize_driver_name(d["full_name"]): d
        for d in repo.list_drivers(active_only=False)
    }
    name_to_id: dict[str, int] = {
        name: int(row["id"]) for name, row in existing_by_name.items()
    }

    for name, records in by_name.items():
        merged = _merge_driver_records(records)
        if not merged["full_name"] or not merged["license_number"]:
            continue
        existing = existing_by_name.get(name)
        if existing is None:
            driver_id = repo.create_driver(
                full_name=merged["full_name"],
                license_number=merged["license_number"],
                license_issued_at=merged.get("license_issued_at"),
                personnel_number=merged.get("personnel_number"),
                snils=merged.get("snils"),
            )
            name_to_id[name] = driver_id
            existing_by_name[name] = {"id": driver_id, **merged}
            continue

        updates: dict[str, Any] = {}
        if not existing.get("snils") and merged.get("snils"):
            updates["snils"] = merged["snils"]
        if not existing.get("personnel_number") and merged.get("personnel_number"):
            updates["personnel_number"] = merged["personnel_number"]
        if not existing.get("license_issued_at") and merged.get("license_issued_at"):
            updates["license_issued_at"] = merged["license_issued_at"]
        if updates:
            repo.update_driver(int(existing["id"]), **updates)

    return name_to_id


def _resolve_vehicle_id_for_pl(
    payload: dict[str, Any],
    plate_to_id: dict[str, int],
    registry: dict[str, VehicleInfo],
) -> int | None:
    plate = normalize_plate(payload.get("plate") or "")
    if plate and plate in plate_to_id:
        return plate_to_id[plate]
    folder = (payload.get("vehicle_folder") or "").lower()
    folder_digits = re.findall(r"\d{3}", folder)
    for plate_norm, info in registry.items():
        mark = (info.mark or "").lower()
        if not mark or mark not in folder:
            continue
        plate_digits = re.findall(r"\d{3}", plate_norm)
        if folder_digits and plate_digits and plate_digits[0] not in folder:
            continue
        return plate_to_id.get(plate_norm)
    return None


def _import_latest_waybills(
    repo: GsmRepository,
    latest_by_folder: dict[str, tuple[date, Path, dict[str, Any]]],
    plate_to_id: dict[str, int],
    driver_ids: dict[str, int],
    registry: dict[str, VehicleInfo],
) -> int:
    imported = 0
    for pl_date, _path, payload in latest_by_folder.values():
        vehicle_id = _resolve_vehicle_id_for_pl(payload, plate_to_id, registry)
        if vehicle_id is None:
            continue
        driver = payload.get("driver") or {}
        driver_name = normalize_driver_name(driver.get("full_name") or "")
        driver_id = driver_ids.get(driver_name)
        if driver_id is None:
            continue
        balance = payload.get("balance") or {}
        trips = payload.get("trips") or []
        route_json = json.dumps(
            [
                {
                    "seq": t.seq,
                    "addr_from": t.addr_from,
                    "addr_to": t.addr_to,
                    "km": t.km,
                    "time_dep": t.time_dep,
                    "time_ret": t.time_ret,
                }
                for t in trips
            ],
            ensure_ascii=False,
        )
        repo.upsert_waybill(
            vehicle_id=vehicle_id,
            date=pl_date,
            driver_id=driver_id,
            status="confirmed",
            source="imported",
            odometer_start=balance.get("odometer_start"),
            odometer_end=balance.get("odometer_end"),
            fuel_start=balance.get("fuel_start"),
            fuel_issued=balance.get("fuel_issued"),
            fuel_end=balance.get("fuel_end"),
            route_json=route_json,
        )
        imported += 1
    return imported


def run_import(db_path: str | Path, gsm_dir: str | Path) -> ImportReport:
    """Импорт справочников и крайних ПЛ из ``gsm_dir`` в БД ``db_path``."""
    db = str(db_path)
    root = Path(gsm_dir)
    ensure_schema(db)
    repo = GsmRepository(db_path=db)
    report = ImportReport()

    assigned_at = datetime.now().date().isoformat()
    registry = load_registry(root / ROMANU_NAME)
    plate_to_id = _import_vehicles_and_cards(repo, registry, assigned_at)

    _import_stations(repo, root / STATIONS_REL)

    routes_rows = _load_routes_ab(root / POOL_NAME)
    _import_routes(repo, routes_rows, plate_to_id)

    driver_records, latest_by_folder = _collect_pl_data(root)
    report.conflicts = build_driver_conflicts(driver_records)
    driver_ids = _import_drivers(repo, driver_records)
    _import_latest_waybills(
        repo, latest_by_folder, plate_to_id, driver_ids, registry
    )

    report.vehicles = len(repo.list_vehicles(active_only=False))
    report.cards = len(repo.list_cards(include_archived=True))
    report.drivers = len(repo.list_drivers(active_only=False))
    report.routes = len(repo.list_routes())
    report.stations = len(repo.list_stations())
    with sqlite3.connect(db) as conn:
        row = conn.execute(
            "SELECT COUNT(*) FROM gsm_waybill "
            "WHERE status = 'confirmed' AND source = 'imported'"
        ).fetchone()
    report.waybills = int(row[0])
    return report


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Импорт истории ГСМ (Роману.xlsx, routes_ab, stations, крайние ПЛ)."
    )
    parser.add_argument(
        "--db",
        required=True,
        help="Путь к SQLite БД (например plita.db).",
    )
    parser.add_argument(
        "--gsm-dir",
        default=str(DEFAULT_GSM_DIR),
        help='Каталог с исходниками ГСМ (по умолчанию "ГСМ").',
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    report = run_import(args.db, args.gsm_dir)
    print(
        "GSM import done:",
        f"vehicles={report.vehicles}",
        f"cards={report.cards}",
        f"drivers={report.drivers}",
        f"routes={report.routes}",
        f"stations={report.stations}",
        f"waybills={report.waybills}",
        f"conflicts={len(report.conflicts)}",
    )
    for conflict in report.conflicts:
        print("  conflict:", conflict.get("message") or conflict)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

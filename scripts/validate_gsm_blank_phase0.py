#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Phase 0 (блокер) модуля ГСМ: round-trip бланка путевого листа.

Проверяет производственный пайплайн экспорта ПЛ до начала основных работ:
  1. .xls (ОКУД 0345001, формулы) → .xlsx через LibreOffice headless;
  2. .xlsx открывается openpyxl, формулы сохранены (не превратились в значения);
  3. Сценарий A (passthrough): openpyxl load/save без изменений → экспорт в .xls
     через soffice → все ячейки совпадают с оригиналом (±0,01);
  4. Сценарий B (modification): меняем входную ячейку-константу, на которую
     ссылается максимум формул (напр. км плеча на стр.2) → экспорт →
     зависимые формулы пересчитались (итого км, расход, остаток, одометр);
  5. PDF финального бланка для визуальной проверки вёрстки.

Вердикт:
  PASS    — оба сценария зелёные, пайплайн экспорта по спеке возможен;
  PARTIAL — значения сохраняются, но пересчёт формул не работает → нужен
            макрос/настройка recalc в soffice;
  FAIL    — ячейки расходятся уже в сценарии A → пересборка бланка.

См. ai_docs/specs/gsm-module-putevye-listy.md (SC-0).

Пример:
  python scripts/validate_gsm_blank_phase0.py \
      --template "ГСМ/Geely Monjaro/2025 год/Апрель 2025/ПЛ 03.04.25.xls"
"""

from __future__ import annotations

import argparse
import re
import subprocess
import sys
import warnings
from dataclasses import dataclass, field
from pathlib import Path

import xlrd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]

DEFAULT_REPORT = (
    PROJECT_ROOT / "ai_docs" / "develop" / "reports" / "2026-08-14-gsm-blank-phase0.md"
)
FLOAT_TOL = 0.01
MODIFY_DELTA = 10.0              # сценарий B: +10 к входной ячейке (км/л)
MAX_RANGE_CELLS = 100            # предел развёртывания диапазонов вида I5:J7
REF_RE = re.compile(
    r"(?:'([^']+)'!)?(\$?[A-Z]{1,3}\$?\d+)(?::(\$?[A-Z]{1,3}\$?\d+))?"
)


@dataclass
class CellDiff:
    sheet: str
    row: int
    col: int
    before: object
    after: object
    kind: str  # 'numeric' | 'text' | 'empty_to_value' | 'value_to_empty'


@dataclass
class Phase0Result:
    formula_count_xlsx: int = 0
    openpyxl_warnings: list[str] = field(default_factory=list)
    cells_compared: int = 0
    diffs_a: list[CellDiff] = field(default_factory=list)
    modified_cell: str | None = None
    modified_from: float | None = None
    modified_to: float | None = None
    changed_after_modify: list[CellDiff] = field(default_factory=list)
    pdf_path: Path | None = None
    verdict: str = "FAIL"
    notes: list[str] = field(default_factory=list)


def run_soffice(args: list[str], workdir: Path, timeout: int) -> None:
    """Единая точка вызова LibreOffice с изолированным профилем."""
    profile = workdir / "lo_profile"
    cmd = [
        "soffice", "--headless", "--norestore",
        f"-env:UserInstallation=file://{profile}",
        *args,
    ]
    proc = subprocess.run(
        cmd, capture_output=True, text=True, timeout=timeout, cwd=workdir,
    )
    if proc.returncode != 0:
        raise RuntimeError(f"soffice failed: {' '.join(cmd)}\n{proc.stderr}")


def convert(src: Path, fmt: str, outdir: Path, workdir: Path, timeout: int) -> Path:
    """soffice --convert-to с проверкой результата.

    NB: явный фильтр 'xls:"MS Excel 97"' даёт impl_store 0x81a на этой версии
    LibreOffice; голый 'xls' использует тот же фильтр, но работает.
    """
    run_soffice(
        ["--convert-to", fmt, "--outdir", str(outdir), str(src)],
        workdir, timeout,
    )
    out = outdir / f"{src.stem}.{fmt}"
    if not out.exists():
        raise RuntimeError(f"soffice не создал {out}")
    return out


def snapshot_xls(path: Path) -> dict[tuple[str, int, int], object]:
    """Все непустые ячейки .xls: {(лист, строка, колонка): значение}."""
    wb = xlrd.open_workbook(str(path))
    snap: dict[tuple[str, int, int], object] = {}
    for sh in wb.sheets():
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                v = sh.cell_value(r, c)
                if isinstance(v, str):
                    v = v.strip()
                if v != "" and v is not None:
                    snap[(sh.name, r, c)] = v
    return snap


def diff_snapshots(
    before: dict[tuple[str, int, int], object],
    after: dict[tuple[str, int, int], object],
) -> tuple[int, list[CellDiff]]:
    """Сравнение снапшотов с допуском для float. Возвращает (всего сравнено, расхождения)."""
    diffs: list[CellDiff] = []
    keys = set(before) | set(after)
    for sheet, r, c in sorted(keys):
        b = before.get((sheet, r, c), "")
        a = after.get((sheet, r, c), "")
        if isinstance(b, str):
            b = b.strip()
        if isinstance(a, str):
            a = a.strip()
        if b == a:
            continue
        if isinstance(b, float) and isinstance(a, float):
            if abs(b - a) <= FLOAT_TOL:
                continue
            diffs.append(CellDiff(sheet, r, c, b, a, "numeric"))
        elif b == "":
            diffs.append(CellDiff(sheet, r, c, b, a, "empty_to_value"))
        elif a == "":
            diffs.append(CellDiff(sheet, r, c, b, a, "value_to_empty"))
        else:
            diffs.append(CellDiff(sheet, r, c, b, a, "text"))
    return len(keys), diffs


def count_formulas_xlsx(path: Path) -> int:
    wb = load_workbook(path, data_only=False)
    return sum(
        1
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.data_type == "f"
    )


def _expand_ref(
    sheet: str, start: str, end: str | None
) -> list[tuple[str, str]]:
    """Разворачивает ссылку/диапазон в список (лист, координата)."""
    from openpyxl.utils import column_index_from_string, get_column_letter

    def split(coord: str) -> tuple[int, int]:
        m = re.match(r"\$?([A-Z]{1,3})\$?(\d+)", coord)
        assert m is not None
        return column_index_from_string(m.group(1)), int(m.group(2))

    c1, r1 = split(start)
    if end is None:
        return [(sheet, f"{get_column_letter(c1)}{r1}")]
    c2, r2 = split(end)
    if (c2 - c1 + 1) * (r2 - r1 + 1) > MAX_RANGE_CELLS:
        return []
    return [
        (sheet, f"{get_column_letter(c)}{r}")
        for r in range(min(r1, r2), max(r1, r2) + 1)
        for c in range(min(c1, c2), max(c1, c2) + 1)
    ]


def find_modification_target(
    path: Path,
) -> tuple[str, int, int, float, list[str]] | None:
    """Ищет входную ячейку-константу, от которой зависит максимум формул.

    Парсит ссылки формул (включая cross-sheet и диапазоны), выбирает
    числовой прецедент с наибольшим числом зависимых формул.
    Возвращает (лист, row_1based, col_1based, значение, зависимые_формулы).
    """
    wb = load_workbook(path, data_only=False)
    formula_cells: set[tuple[str, str]] = set()
    direct: dict[tuple[str, str], set[tuple[str, str]]] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type != "f" or not isinstance(cell.value, str):
                    continue
                owner = (ws.title, cell.coordinate)
                formula_cells.add(owner)
                for m in REF_RE.finditer(cell.value):
                    ref_sheet = m.group(1) or ws.title
                    for target in _expand_ref(ref_sheet, m.group(2), m.group(3)):
                        direct.setdefault(target, set()).add(owner)

    def chain_size(seed: tuple[str, str]) -> int:
        """Сколько уникальных формул прямо/транзитивно зависит от ячейки."""
        seen: set[tuple[str, str]] = set()
        queue = [seed]
        while queue:
            cur = queue.pop()
            for owner in direct.get(cur, ()):
                if owner not in seen:
                    seen.add(owner)
                    queue.append(owner)
        return len(seen)

    best: tuple[str, str] | None = None
    best_chain = 0
    for target in direct:
        sheet, coord = target
        v = wb[sheet][coord].value
        if isinstance(v, (int, float)) and not isinstance(v, bool) and v > 0:
            score = chain_size(target)
            if score > best_chain:
                best, best_chain = target, score
    if best is None:
        return None
    sheet, coord = best
    cell = wb[sheet][coord]
    owners = sorted(f"{s}!{c}" for s, c in direct[best])
    return sheet, cell.row, cell.column, float(cell.value), owners


def modify_cell(src: Path, dst: Path, sheet: str, row: int, col: int, value: float) -> None:
    wb = load_workbook(src, data_only=False)
    wb[sheet].cell(row=row, column=col).value = value
    wb.save(dst)


def fmt_cell(diff: CellDiff) -> str:
    return (
        f"| {diff.sheet} | R{diff.row + 1}C{diff.col + 1} | "
        f"{diff.before} | {diff.after} | {diff.kind} |"
    )


def write_report(result: Phase0Result, template: Path, report_path: Path) -> None:
    lines = [
        "# Phase 0: round-trip бланка ПЛ (SC-0)",
        "",
        f"Дата прогона: 2026-08-14. Шаблон: `{template}`",
        "",
        f"## Вердикт: **{result.verdict}**",
        "",
        "## Сводка",
        "",
        f"- Формул в .xlsx после конвертации: **{result.formula_count_xlsx}**",
        f"- Ячеек сравнено (сценарий A): {result.cells_compared}",
        f"- Расхождений A (passthrough): **{len(result.diffs_a)}**",
        f"- Входная ячейка сценария B: {result.modified_cell or 'не найдена'}",
    ]
    if result.modified_cell:
        lines.append(
            f"- Изменение: {result.modified_from} → {result.modified_to}; "
            f"пересчитанных ячеек: **{len(result.changed_after_modify)}**"
        )
    if result.pdf_path:
        lines.append(f"- PDF для визуальной проверки: `{result.pdf_path}`")
    if result.openpyxl_warnings:
        lines.append(f"- Предупреждений openpyxl: {len(result.openpyxl_warnings)}")
    if result.notes:
        lines += ["", "## Заметки", ""] + [f"- {n}" for n in result.notes]
    if result.diffs_a:
        lines += [
            "", "## Расхождения сценария A (топ-20)", "",
            "| Лист | Ячейка | Оригинал | После | Тип |",
            "|---|---|---|---|---|",
            *[fmt_cell(d) for d in result.diffs_a[:20]],
        ]
    if result.changed_after_modify:
        lines += [
            "", "## Пересчитанные ячейки сценария B", "",
            "| Лист | Ячейка | Было | Стало | Тип |",
            "|---|---|---|---|---|",
            *[fmt_cell(d) for d in result.changed_after_modify[:20]],
        ]
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--template", required=True, type=Path, help="Путь к реальному ПЛ .xls")
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--workdir", type=Path, default=PROJECT_ROOT / "tmp" / "gsm_blank_phase0")
    parser.add_argument("--timeout", type=int, default=120, help="Таймаут вызова soffice, сек")
    args = parser.parse_args()

    template = args.template.resolve()
    if not template.exists():
        print(f"Шаблон не найден: {template}", file=sys.stderr)
        return 1
    workdir = args.workdir.resolve()
    workdir.mkdir(parents=True, exist_ok=True)

    result = Phase0Result()
    print(f"[1/6] Снапшот оригинала: {template.name}")
    original = snapshot_xls(template)

    print("[2/6] Конвертация .xls → .xlsx (soffice)")
    step1 = convert(template, "xlsx", workdir, workdir, args.timeout)

    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        result.formula_count_xlsx = count_formulas_xlsx(step1)
        passthrough = workdir / "step2a_passthrough.xlsx"
        wb = load_workbook(step1, data_only=False)
        wb.save(passthrough)
    result.openpyxl_warnings = [str(w.message) for w in caught]
    print(f"       Формул в .xlsx: {result.formula_count_xlsx}")
    if result.formula_count_xlsx == 0:
        result.notes.append("Формулы не пережили конвертацию .xls→.xlsx — стоп.")
        write_report(result, template, args.report)
        print("Вердикт: FAIL (формулы потеряны при конвертации)")
        return 1

    print("[3/6] Сценарий A: passthrough → .xls")
    final_a = convert(passthrough, "xls", workdir, workdir, args.timeout)
    snap_a = snapshot_xls(final_a)
    result.cells_compared, result.diffs_a = diff_snapshots(original, snap_a)

    print("[4/6] Сценарий B: модификация входной ячейки → .xls")
    found = find_modification_target(step1)
    changed: list[CellDiff] = []
    if found is None:
        result.notes.append("Числовой прецедент формул не найден — сценарий B пропущен.")
    else:
        sheet, row, col, old_value, owners = found
        result.notes.append(
            f"Сценарий B: входная ячейка {sheet}!R{row}C{col} "
            f"(значение {old_value}), зависимых формул: {len(owners)} "
            f"({'; '.join(owners[:5])})"
        )
        new_value = old_value + MODIFY_DELTA
        result.modified_cell = f"{sheet}!R{row}C{col}"
        result.modified_from, result.modified_to = old_value, new_value
        modified = workdir / "step2b_modified.xlsx"
        modify_cell(step1, modified, sheet, row, col, new_value)
        final_b = convert(modified, "xls", workdir, workdir, args.timeout)
        snap_b = snapshot_xls(final_b)
        _, changed = diff_snapshots(snap_a, snap_b)
        # Сама изменённая ячейка — не «пересчёт», убираем из списка
        result.changed_after_modify = [
            d for d in changed
            if not (d.sheet == sheet and d.row == row - 1 and d.col == col - 1)
        ]
        print(f"       Пересчиталось ячеек: {len(result.changed_after_modify)}")

    print("[5/6] PDF для визуальной проверки")
    result.pdf_path = convert(final_a, "pdf", workdir, workdir, args.timeout)

    print("[6/6] Вердикт")
    hard_diffs = [d for d in result.diffs_a if d.kind in ("numeric", "text")]
    if hard_diffs:
        result.verdict = "FAIL"
        result.notes.append(
            f"Сценарий A: {len(hard_diffs)} расхождений значений/текста — "
            "конвертация ломает данные, нужна пересборка бланка."
        )
    elif found is not None and not result.changed_after_modify:
        result.verdict = "PARTIAL"
        result.notes.append(
            "Значения сохраняются, но ни одна зависимая ячейка не пересчиталась "
            "после изменения входной — нужен recalc-макрос или настройка "
            "OOXMLRecalcMode в профиле soffice."
        )
    elif found is None:
        result.verdict = "PARTIAL"
        result.notes.append("Сценарий B не выполнен (нет якорной ячейки) — пересчёт не подтверждён.")
    else:
        result.verdict = "PASS"
        result.notes.append(
            "Пайплайн экспорта по спеке работоспособен: шаблон .xlsx + openpyxl fill "
            "+ soffice export с пересчётом формул."
        )

    write_report(result, template, args.report)
    print(f"Вердикт: {result.verdict}")
    print(f"Отчёт: {args.report}")
    print(f"PDF:   {result.pdf_path}")
    return 0 if result.verdict == "PASS" else 1


if __name__ == "__main__":
    sys.exit(main())

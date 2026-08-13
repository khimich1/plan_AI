#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Сборка пула поездок из архива путевых листов ГСМ → один Excel.

Листы: facts, routes_ab, rounds_aba, unpaired, summary.
См. ai_docs/ideas/gsm-trip-pool.md
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Iterable

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_GSM_DIR = PROJECT_ROOT / "ГСМ"
DEFAULT_REGISTRY = DEFAULT_GSM_DIR / "Роману.xlsx"
DEFAULT_OUT = DEFAULT_GSM_DIR / "пул_поездок.xlsx"

BLANK_DIR_MARKERS = ("не заполн", "не заполненые")
SKIP_NAME_PREFIXES = (".~", "~$")

DATE_RE = re.compile(
    r"(?P<d>\d{1,2})\.(?P<m>\d{1,2})\.(?P<y>\d{2,4})"
)


@dataclass(frozen=True)
class VehicleInfo:
    mark: str
    plate: str
    drivers_registry: str
    fuel_card: str
    tank_l: str
    norm_summer: str
    norm_winter: str


@dataclass
class TripFact:
    vehicle_folder: str
    mark: str
    plate: str
    pl_date: date | None
    seq: int
    addr_from: str
    addr_to: str
    addr_from_norm: str
    addr_to_norm: str
    time_dep: str
    time_ret: str
    km: float | None
    driver: str
    fuel: str
    source_path: str
    paired: bool = False


@dataclass
class RoundTrip:
    vehicle_folder: str
    mark: str
    plate: str
    pl_date: date | None
    addr_a: str
    addr_b: str
    km_ab: float | None
    km_ba: float | None
    km_sum: float | None
    time_dep_1: str
    time_ret_1: str
    time_dep_2: str
    time_ret_2: str
    driver: str
    fuel: str
    source_path: str


def _cell(sh: xlrd.sheet.Sheet, row: int, col: int) -> Any:
    if row >= sh.nrows or col >= sh.ncols:
        return ""
    return sh.cell_value(row, col)


def _as_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value).strip()


def normalize_address(raw: str) -> str:
    """Грубая нормализация адреса для ключа уникальности."""
    s = (raw or "").strip().lower().replace("ё", "е")
    s = s.replace("«", '"').replace("»", '"').replace("“", '"').replace("”", '"')
    s = s.replace("№", " ")
    # Даты в хвосте адреса (многодневные ПЛ) ломают зеркальную склейку
    s = re.sub(r"[,;\s]*\d{1,2}\.\d{1,2}\.\d{2,4}\s*г?\.?\s*$", "", s)
    replacements = (
        (r"\bг\.\s*", "г "),
        (r"\bгород\s+", "г "),
        (r"\bул\.\s*", "ул "),
        (r"\bулица\s+", "ул "),
        (r"\bпр-д\s*", "проезд "),
        (r"\bпроезд\s+", "проезд "),
        (r"\bпр\.\s*", "пр "),
        (r"\bд\.\s*", "д "),
        (r"\bдом\s+", "д "),
        (r"\bстр\.\s*", "стр "),
        (r"\bстроение\s+", "стр "),
        (r"\bпом\.\s*", "пом "),
        (r"\bпомещение\s+", "пом "),
        (r"\booо\b", "ооо"),
        (r"\booo\b", "ооо"),
    )
    for pattern, repl in replacements:
        s = re.sub(pattern, repl, s, flags=re.IGNORECASE)
    s = re.sub(r"[,\.;]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def format_hm(h: Any, m: Any) -> str:
    hs = _as_str(h)
    ms = _as_str(m)
    if not hs and not ms:
        return ""
    try:
        hi = int(float(hs)) if hs else 0
        mi = int(float(ms)) if ms else 0
        return f"{hi:02d}:{mi:02d}"
    except ValueError:
        return f"{hs}:{ms}".strip(":")


def parse_km(value: Any) -> float | None:
    if value in ("", None):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_date_from_name(name: str) -> date | None:
    m = DATE_RE.search(name)
    if not m:
        return None
    d = int(m.group("d"))
    mo = int(m.group("m"))
    y = int(m.group("y"))
    if y < 100:
        y += 2000
    try:
        return date(y, mo, d)
    except ValueError:
        return None


def is_blank_dir(path: Path) -> bool:
    return any(marker in part.lower() for part in path.parts for marker in BLANK_DIR_MARKERS)


def iter_waybill_files(gsm_dir: Path) -> Iterable[tuple[str, Path]]:
    """Yield (vehicle_folder, path) for filled waybills only."""
    for car_dir in sorted(p for p in gsm_dir.iterdir() if p.is_dir()):
        for path in sorted(car_dir.rglob("*.xls")):
            if path.name.startswith(SKIP_NAME_PREFIXES):
                continue
            if is_blank_dir(path.relative_to(gsm_dir)):
                continue
            yield car_dir.name, path


def load_registry(path: Path) -> dict[str, VehicleInfo]:
    """Ключ — нормализованный гос.номер."""
    from openpyxl import load_workbook

    if not path.is_file():
        return {}

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    by_plate: dict[str, VehicleInfo] = {}
    rows = ws.iter_rows(values_only=True)
    header = None
    for row in rows:
        cells = ["" if c is None else str(c).strip() for c in row]
        if not any(cells):
            continue
        joined = " ".join(cells).lower()
        if header is None and "гос" in joined and "марка" in joined:
            header = [c.lower() for c in cells]
            continue
        if header is None:
            continue

        def col(*names: str) -> str:
            for i, h in enumerate(header):
                if any(n in h for n in names):
                    return cells[i] if i < len(cells) else ""
            return ""

        mark = col("марка")
        plate = col("гос")
        if not plate:
            continue
        info = VehicleInfo(
            mark=mark,
            plate=plate,
            drivers_registry=col("водитель"),
            fuel_card=col("топливн", "карт"),
            tank_l=col("объем", "бак"),
            norm_summer=col("лето"),
            norm_winter=col("зима"),
        )
        by_plate[normalize_plate(plate)] = info
    wb.close()
    return by_plate


def normalize_plate(plate: str) -> str:
    s = (plate or "").upper().replace("Ё", "Е")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def parse_waybill(path: Path, vehicle_folder: str) -> tuple[dict[str, str], list[TripFact]]:
    book = xlrd.open_workbook(str(path), formatting_info=False)
    sh1 = book.sheet_by_index(0)
    sh2 = book.sheet_by_index(1) if book.nsheets > 1 else None

    mark = _as_str(_cell(sh1, 9, 26))
    plate = _as_str(_cell(sh1, 10, 39))
    driver = _as_str(_cell(sh1, 11, 17))
    fuel = _as_str(_cell(sh1, 25, 1)) or _as_str(_cell(sh1, 25, 2))
    pl_date = parse_date_from_name(path.name)

    meta = {
        "mark": mark,
        "plate": plate,
        "driver": driver,
        "fuel": fuel,
    }

    trips: list[TripFact] = []
    if sh2 is None:
        return meta, trips

    for r in range(4, sh2.nrows):
        label0 = _as_str(_cell(sh2, r, 0)).lower()
        if "пройдено" in label0:
            break

        seq_raw = _cell(sh2, r, 1)
        addr_from = _as_str(_cell(sh2, r, 2))
        addr_to = _as_str(_cell(sh2, r, 3))
        # Итог «пройдено, км» иногда лежит в col2 без подписи в col0
        if not addr_to and addr_from.replace(".", "", 1).isdigit():
            break
        if not addr_from or not addr_to:
            continue

        try:
            seq = int(float(seq_raw)) if seq_raw not in ("", None) else len(trips) + 1
        except (TypeError, ValueError):
            seq = len(trips) + 1

        km = parse_km(_cell(sh2, r, 8))
        trips.append(
            TripFact(
                vehicle_folder=vehicle_folder,
                mark=mark,
                plate=plate,
                pl_date=pl_date,
                seq=seq,
                addr_from=addr_from,
                addr_to=addr_to,
                addr_from_norm=normalize_address(addr_from),
                addr_to_norm=normalize_address(addr_to),
                time_dep=format_hm(_cell(sh2, r, 4), _cell(sh2, r, 5)),
                time_ret=format_hm(_cell(sh2, r, 6), _cell(sh2, r, 7)),
                km=km,
                driver=driver,
                fuel=fuel,
                source_path=str(path),
            )
        )
    return meta, trips


def pair_rounds(trips: list[TripFact]) -> tuple[list[RoundTrip], list[TripFact]]:
    """Склейка соседних зеркальных плеч внутри одного ПЛ."""
    by_file: dict[str, list[TripFact]] = defaultdict(list)
    for t in trips:
        by_file[t.source_path].append(t)

    rounds: list[RoundTrip] = []
    unpaired: list[TripFact] = []

    for _path, legs in by_file.items():
        legs_sorted = sorted(legs, key=lambda x: x.seq)
        used: set[int] = set()
        i = 0
        while i < len(legs_sorted):
            if i in used:
                i += 1
                continue
            a = legs_sorted[i]
            matched = False
            if i + 1 < len(legs_sorted) and (i + 1) not in used:
                b = legs_sorted[i + 1]
                if (
                    a.addr_from_norm
                    and a.addr_to_norm
                    and a.addr_from_norm == b.addr_to_norm
                    and a.addr_to_norm == b.addr_from_norm
                ):
                    km_ab = a.km
                    km_ba = b.km
                    km_sum = None
                    if km_ab is not None or km_ba is not None:
                        km_sum = (km_ab or 0.0) + (km_ba or 0.0)
                    rounds.append(
                        RoundTrip(
                            vehicle_folder=a.vehicle_folder,
                            mark=a.mark,
                            plate=a.plate,
                            pl_date=a.pl_date,
                            addr_a=a.addr_from,
                            addr_b=a.addr_to,
                            km_ab=km_ab,
                            km_ba=km_ba,
                            km_sum=km_sum,
                            time_dep_1=a.time_dep,
                            time_ret_1=a.time_ret,
                            time_dep_2=b.time_dep,
                            time_ret_2=b.time_ret,
                            driver=a.driver,
                            fuel=a.fuel,
                            source_path=a.source_path,
                        )
                    )
                    a.paired = True
                    b.paired = True
                    used.add(i)
                    used.add(i + 1)
                    matched = True
                    i += 2
                    continue
            if not matched:
                unpaired.append(a)
                used.add(i)
                i += 1
    return rounds, unpaired


def enrich(plate: str, registry: dict[str, VehicleInfo]) -> VehicleInfo | None:
    return registry.get(normalize_plate(plate))


def build_routes(trips: list[TripFact], registry: dict[str, VehicleInfo]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str, str, float | None], list[TripFact]] = defaultdict(list)
    for t in trips:
        key = (t.vehicle_folder, t.addr_from_norm, t.addr_to_norm, t.km)
        groups[key].append(t)

    rows: list[dict[str, Any]] = []
    for (vehicle, a_norm, b_norm, km), items in sorted(
        groups.items(), key=lambda x: (-len(x[1]), x[0][0], x[0][1], x[0][2])
    ):
        if not a_norm and not b_norm:
            continue
        sample = items[0]
        times = [x.time_dep for x in items if x.time_dep]
        typical_dep = Counter(times).most_common(1)[0][0] if times else ""
        dates = sorted({x.pl_date.isoformat() for x in items if x.pl_date})
        drivers = sorted({x.driver for x in items if x.driver})
        fuels = sorted({x.fuel for x in items if x.fuel})
        info = enrich(sample.plate, registry)
        rows.append(
            {
                "машина": vehicle,
                "марка": sample.mark or (info.mark if info else ""),
                "гос_номер": sample.plate or (info.plate if info else ""),
                "адрес_A": sample.addr_from,
                "адрес_B": sample.addr_to,
                "адрес_A_норм": a_norm,
                "адрес_B_норм": b_norm,
                "км": km,
                "частота": len(items),
                "типичное_время_выезда": typical_dep,
                "водители": "; ".join(drivers),
                "топливо": "; ".join(fuels),
                "топливная_карта": info.fuel_card if info else "",
                "объем_бака": info.tank_l if info else "",
                "норма_лето": info.norm_summer if info else "",
                "норма_зима": info.norm_winter if info else "",
                "водители_реестр": info.drivers_registry if info else "",
                "даты_примеры": ", ".join(dates[:8]) + ("…" if len(dates) > 8 else ""),
                "дат_всего": len(dates),
            }
        )
    return rows


def _write_sheet(ws, headers: list[str], rows: list[dict[str, Any] | Any]) -> None:
    bold = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = bold
    for r_idx, row in enumerate(rows, 2):
        if isinstance(row, dict):
            for c_idx, h in enumerate(headers, 1):
                ws.cell(r_idx, c_idx, row.get(h, ""))
        else:
            for c_idx, h in enumerate(headers, 1):
                ws.cell(r_idx, c_idx, getattr(row, h, ""))
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = len(headers[col - 1])
        for r_idx in range(2, min(len(rows) + 2, 200)):
            val = ws.cell(r_idx, col).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[letter].width = max_len + 2
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"


def fact_to_row(t: TripFact, registry: dict[str, VehicleInfo]) -> dict[str, Any]:
    info = enrich(t.plate, registry)
    return {
        "машина": t.vehicle_folder,
        "марка": t.mark or (info.mark if info else ""),
        "гос_номер": t.plate or (info.plate if info else ""),
        "дата_пл": t.pl_date.isoformat() if t.pl_date else "",
        "номер_строки": t.seq,
        "адрес_отправления": t.addr_from,
        "адрес_назначения": t.addr_to,
        "время_выезда": t.time_dep,
        "время_возвращения": t.time_ret,
        "км": t.km if t.km is not None else "",
        "водитель": t.driver,
        "топливо": t.fuel,
        "топливная_карта": info.fuel_card if info else "",
        "объем_бака": info.tank_l if info else "",
        "норма_лето": info.norm_summer if info else "",
        "норма_зима": info.norm_winter if info else "",
        "путь_к_файлу": t.source_path,
    }


def round_to_row(r: RoundTrip, registry: dict[str, VehicleInfo]) -> dict[str, Any]:
    info = enrich(r.plate, registry)
    return {
        "машина": r.vehicle_folder,
        "марка": r.mark or (info.mark if info else ""),
        "гос_номер": r.plate or (info.plate if info else ""),
        "дата_пл": r.pl_date.isoformat() if r.pl_date else "",
        "адрес_A": r.addr_a,
        "адрес_B": r.addr_b,
        "км_A_B": r.km_ab if r.km_ab is not None else "",
        "км_B_A": r.km_ba if r.km_ba is not None else "",
        "км_сумма": r.km_sum if r.km_sum is not None else "",
        "время_выезда_1": r.time_dep_1,
        "время_возвращения_1": r.time_ret_1,
        "время_выезда_2": r.time_dep_2,
        "время_возвращения_2": r.time_ret_2,
        "водитель": r.driver,
        "топливо": r.fuel,
        "топливная_карта": info.fuel_card if info else "",
        "объем_бака": info.tank_l if info else "",
        "норма_лето": info.norm_summer if info else "",
        "норма_зима": info.norm_winter if info else "",
        "путь_к_файлу": r.source_path,
    }


def build_summary(
    trips: list[TripFact],
    routes: list[dict[str, Any]],
    rounds: list[RoundTrip],
    unpaired: list[TripFact],
) -> list[dict[str, Any]]:
    by_vehicle = sorted({t.vehicle_folder for t in trips})
    rows = []
    for v in by_vehicle:
        v_trips = [t for t in trips if t.vehicle_folder == v]
        v_routes = [r for r in routes if r["машина"] == v]
        v_rounds = [r for r in rounds if r.vehicle_folder == v]
        v_unp = [t for t in unpaired if t.vehicle_folder == v]
        files = {t.source_path for t in v_trips}
        paired_ratio = (
            (2 * len(v_rounds) / len(v_trips)) if v_trips else 0.0
        )
        rows.append(
            {
                "машина": v,
                "файлов_пл": len(files),
                "фактов_плеч": len(v_trips),
                "уник_маршрутов_ab": len(v_routes),
                "рейсов_aba": len(v_rounds),
                "непарных_плеч": len(v_unp),
                "доля_плеч_в_парах": round(paired_ratio, 3),
            }
        )
    rows.append(
        {
            "машина": "ИТОГО",
            "файлов_пл": len({t.source_path for t in trips}),
            "фактов_плеч": len(trips),
            "уник_маршрутов_ab": len(routes),
            "рейсов_aba": len(rounds),
            "непарных_плеч": len(unpaired),
            "доля_плеч_в_парах": round((2 * len(rounds) / len(trips)) if trips else 0.0, 3),
        }
    )
    return rows


def build_workbook(
    trips: list[TripFact],
    rounds: list[RoundTrip],
    unpaired: list[TripFact],
    routes: list[dict[str, Any]],
    summary: list[dict[str, Any]],
    registry: dict[str, VehicleInfo],
) -> Workbook:
    wb = Workbook()

    ws = wb.active
    ws.title = "summary"
    _write_sheet(
        ws,
        [
            "машина",
            "файлов_пл",
            "фактов_плеч",
            "уник_маршрутов_ab",
            "рейсов_aba",
            "непарных_плеч",
            "доля_плеч_в_парах",
        ],
        summary,
    )

    ws = wb.create_sheet("facts")
    fact_headers = [
        "машина",
        "марка",
        "гос_номер",
        "дата_пл",
        "номер_строки",
        "адрес_отправления",
        "адрес_назначения",
        "время_выезда",
        "время_возвращения",
        "км",
        "водитель",
        "топливо",
        "топливная_карта",
        "объем_бака",
        "норма_лето",
        "норма_зима",
        "путь_к_файлу",
    ]
    _write_sheet(ws, fact_headers, [fact_to_row(t, registry) for t in trips])

    ws = wb.create_sheet("routes_ab")
    route_headers = [
        "машина",
        "марка",
        "гос_номер",
        "адрес_A",
        "адрес_B",
        "адрес_A_норм",
        "адрес_B_норм",
        "км",
        "частота",
        "типичное_время_выезда",
        "водители",
        "топливо",
        "топливная_карта",
        "объем_бака",
        "норма_лето",
        "норма_зима",
        "водители_реестр",
        "даты_примеры",
        "дат_всего",
    ]
    _write_sheet(ws, route_headers, routes)

    ws = wb.create_sheet("rounds_aba")
    round_headers = [
        "машина",
        "марка",
        "гос_номер",
        "дата_пл",
        "адрес_A",
        "адрес_B",
        "км_A_B",
        "км_B_A",
        "км_сумма",
        "время_выезда_1",
        "время_возвращения_1",
        "время_выезда_2",
        "время_возвращения_2",
        "водитель",
        "топливо",
        "топливная_карта",
        "объем_бака",
        "норма_лето",
        "норма_зима",
        "путь_к_файлу",
    ]
    _write_sheet(ws, round_headers, [round_to_row(r, registry) for r in rounds])

    ws = wb.create_sheet("unpaired")
    _write_sheet(ws, fact_headers, [fact_to_row(t, registry) for t in unpaired])
    return wb


def run(gsm_dir: Path, registry_path: Path, out_path: Path) -> dict[str, Any]:
    if not gsm_dir.is_dir():
        raise FileNotFoundError(f"Нет каталога ГСМ: {gsm_dir}")

    registry = load_registry(registry_path)
    trips: list[TripFact] = []
    errors: list[str] = []
    files_ok = 0

    for vehicle_folder, path in iter_waybill_files(gsm_dir):
        try:
            _meta, file_trips = parse_waybill(path, vehicle_folder)
            trips.extend(file_trips)
            files_ok += 1
        except Exception as exc:  # noqa: BLE001 — собираем отчёт по битым файлам
            errors.append(f"{path}: {exc}")

    trips.sort(
        key=lambda t: (
            t.vehicle_folder,
            t.pl_date or date.min,
            t.seq,
            t.source_path,
        )
    )
    rounds, unpaired = pair_rounds(trips)
    routes = build_routes(trips, registry)
    summary = build_summary(trips, routes, rounds, unpaired)

    wb = build_workbook(trips, rounds, unpaired, routes, summary, registry)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    return {
        "out": str(out_path),
        "files_ok": files_ok,
        "facts": len(trips),
        "routes": len(routes),
        "rounds": len(rounds),
        "unpaired": len(unpaired),
        "errors": errors,
        "registry": len(registry),
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Пул поездок ГСМ → Excel")
    parser.add_argument("--gsm-dir", type=Path, default=DEFAULT_GSM_DIR)
    parser.add_argument("--registry", type=Path, default=DEFAULT_REGISTRY)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args(argv)

    try:
        report = run(args.gsm_dir, args.registry, args.out)
    except FileNotFoundError as exc:
        print(f"Ошибка: {exc}", file=sys.stderr)
        return 1

    print(f"OK → {report['out']}")
    print(
        f"ПЛ: {report['files_ok']}, facts: {report['facts']}, "
        f"routes_ab: {report['routes']}, rounds_aba: {report['rounds']}, "
        f"unpaired: {report['unpaired']}, registry: {report['registry']}"
    )
    if report["errors"]:
        print(f"Ошибок чтения: {len(report['errors'])}")
        for line in report["errors"][:20]:
            print(f"  - {line}")
        if len(report["errors"]) > 20:
            print(f"  … ещё {len(report['errors']) - 20}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

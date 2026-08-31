#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Событийная лента «поездки + заправки + мойки» → один Excel.

Читает ГСМ/транзакции/*.xls (выгрузки ЛК топливного оператора) и
ГСМ/пул_поездок.xlsx (листы facts, routes_ab), матчит транзакции карт
на дни поездок по ключу (машина, дата), геокодит адреса АЗС/моек
через общий кэш geo_cache/addresses.json.

Листы результата:
- лента       — хронология дня машины: выезд/приезд/заправка/мойка;
- рейсы_дня   — агрегат дня: км, нормативный расход, заправлено, дельта;
- маршруты    — уникальные A↔B + типовые АЗС в дни маршрута;
- вне_поездок — транзакции без поездки в тот же день (деньги не теряются).

См. ai_docs/ideas/poezdki-zapravki-moyki-lenta.md
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from scripts.build_gsm_routes_map import (  # noqa: E402
    DEFAULT_GEO_CACHE,
    NOMINATIM_RATE_LIMIT_SEC,
    coords_for_norm,
    geocode_addresses,
    load_address_cache,
    nominatim_geocode,
    save_address_cache,
    simplify_address_for_geocode,
)
from scripts.build_gsm_trip_pool import normalize_address  # noqa: E402

DEFAULT_TX_DIR = PROJECT_ROOT / "ГСМ" / "транзакции"
DEFAULT_POOL = PROJECT_ROOT / "ГСМ" / "пул_поездок.xlsx"
DEFAULT_OUT = PROJECT_ROOT / "ГСМ" / "поездки_заправки.xlsx"
DEFAULT_STATIONS = PROJECT_ROOT / "ГСМ" / "geo_cache" / "stations.geojson"
DEFAULT_ROUTE_STATIONS = (
    PROJECT_ROOT / "ГСМ" / "geo_cache" / "route_stations.json"
)

WASH_SERVICE = "Мойка"
WINTER_MONTHS = (11, 12, 1, 2, 3)

# Порядок типов при равном времени внутри дня.
EVENT_TYPE_ORDER = {"выезд": 0, "заправка": 1, "мойка": 1, "приезд": 2}


@dataclass(frozen=True)
class FuelTx:
    """Одна транзакция топливной карты (заправка или мойка)."""

    dt: datetime
    card: str
    service: str
    qty: float | None
    unit: str
    price: float | None
    amount: float | None
    vendor: str
    brand: str
    city: str
    address: str
    addr_norm: str
    source: str

    @property
    def is_wash(self) -> bool:
        return self.service.strip().lower() == WASH_SERVICE.lower()


@dataclass(frozen=True)
class TripLeg:
    """Одно плечо поездки из листа facts пула."""

    vehicle: str
    plate: str
    day: date
    seq: int
    addr_from: str
    addr_to: str
    time_dep: str
    time_ret: str
    km: float | None
    driver: str
    cards: str
    norm_summer: float | None
    norm_winter: float | None
    source: str


def _as_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value).strip()


def _parse_float(value: Any) -> float | None:
    if value in ("", None):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(_as_str(value).replace(",", "."))
    except ValueError:
        return None


def norm_card(value: Any) -> str:
    """Номер карты → строка без «.0» и пробелов."""
    return _as_str(value).removesuffix(".0").strip()


def split_cards(value: Any) -> list[str]:
    """«3005454266, 3005454268» → [3005454266, 3005454268]."""
    return [c for c in (norm_card(p) for p in _as_str(value).split(",")) if c]


def hm_to_min(value: Any) -> int | None:
    """«07:10» → 430; пусто/мусор → None."""
    text = _as_str(value)
    if not text:
        return None
    parts = text.split(":")
    try:
        return int(parts[0]) * 60 + int(parts[1])
    except (ValueError, IndexError):
        return None


def parse_tx_datetime(cell: xlrd.cell.Cell, datemode: int) -> datetime | None:
    """Дата транзакции: xldate или текст «2026-05-12 16:20:46»."""
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate.xldate_as_datetime(cell.value, datemode)
    text = _as_str(cell.value)
    if not text:
        return None
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def load_transactions(tx_dir: Path) -> tuple[list[FuelTx], list[str]]:
    """Все транзакции из *.xls каталога + предупреждения сверки с «Итоги:»."""
    txs: list[FuelTx] = []
    warnings: list[str] = []
    for path in sorted(tx_dir.glob("*.xls")):
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(0)
        header_row = None
        for r in range(min(5, sheet.nrows)):
            if _as_str(sheet.cell_value(r, 0)) == "Дата трн.":
                header_row = r
                break
        if header_row is None:
            warnings.append(f"{path.name}: не найдена строка шапки «Дата трн.»")
            continue
        headers = [_as_str(sheet.cell_value(header_row, c)) for c in range(sheet.ncols)]
        col = {name: i for i, name in enumerate(headers) if name}

        def value(r: int, name: str, fallback: int) -> Any:
            i = col.get(name, fallback)
            if i is None or i >= sheet.ncols:
                return None
            return sheet.cell_value(r, i)

        file_qty = 0.0
        file_amount = 0.0
        total_qty: float | None = None
        total_amount: float | None = None
        for r in range(header_row + 1, sheet.nrows):
            first = _as_str(sheet.cell_value(r, 0))
            if not first:
                continue
            if first.startswith("Итоги"):
                total_qty = _parse_float(value(r, "Кол-во", 3))
                total_amount = _parse_float(value(r, "Сумма с налогом, всего", 7))
                continue
            dt_col = col.get("Дата трн.", 0)
            dt = parse_tx_datetime(sheet.cell(r, dt_col), book.datemode)
            if dt is None:
                continue
            qty = _parse_float(value(r, "Кол-во", 3))
            amount = _parse_float(value(r, "Сумма с налогом, всего", 7))
            file_qty += qty or 0.0
            file_amount += amount or 0.0
            address = _as_str(value(r, "Адрес ТО", 11))
            txs.append(
                FuelTx(
                    dt=dt,
                    card=norm_card(value(r, "Карта", 1)),
                    service=_as_str(value(r, "Услуга", 2)),
                    qty=qty,
                    unit=_as_str(value(r, "Ед. изм.", 4)),
                    price=_parse_float(value(r, "Цена", 5)),
                    amount=amount,
                    vendor=_as_str(value(r, "ТО", 8)),
                    brand=_as_str(value(r, "Бренд", 9)),
                    city=_as_str(value(r, "Город", 10)),
                    address=address,
                    addr_norm=normalize_address(address),
                    source=path.name,
                )
            )
        if total_qty is not None and abs(file_qty - total_qty) > 0.01:
            warnings.append(
                f"{path.name}: Кол-во {file_qty:.2f} ≠ Итоги {total_qty:.2f}"
            )
        if total_amount is not None and abs(file_amount - total_amount) > 0.01:
            warnings.append(
                f"{path.name}: Сумма {file_amount:.2f} ≠ Итоги {total_amount:.2f}"
            )
    return txs, warnings


def _load_pool_sheet(
    pool_path: Path, sheet_name: str
) -> tuple[list[str], list[tuple[Any, ...]]]:
    wb = load_workbook(pool_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = [_as_str(h) for h in next(rows)]
        return header, [r for r in rows if r and any(v not in ("", None) for v in r)]
    finally:
        wb.close()


def _pool_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _as_str(value)
    if not text:
        return None
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def load_legs(pool_path: Path) -> list[TripLeg]:
    """Плечи поездок с листа facts."""
    header, rows = _load_pool_sheet(pool_path, "facts")
    col = {name: i for i, name in enumerate(header) if name}

    def cell(row: tuple[Any, ...], name: str) -> Any:
        i = col.get(name)
        return row[i] if i is not None and i < len(row) else None

    legs: list[TripLeg] = []
    for row in rows:
        day = _pool_date(cell(row, "дата_пл"))
        if day is None:
            continue
        legs.append(
            TripLeg(
                vehicle=_as_str(cell(row, "машина")),
                plate=_as_str(cell(row, "гос_номер")),
                day=day,
                seq=int(_parse_float(cell(row, "номер_строки")) or 0),
                addr_from=_as_str(cell(row, "адрес_отправления")),
                addr_to=_as_str(cell(row, "адрес_назначения")),
                time_dep=_as_str(cell(row, "время_выезда")),
                time_ret=_as_str(cell(row, "время_возвращения")),
                km=_parse_float(cell(row, "км")),
                driver=_as_str(cell(row, "водитель")),
                cards=_as_str(cell(row, "топливная_карта")),
                norm_summer=_parse_float(cell(row, "норма_лето")),
                norm_winter=_parse_float(cell(row, "норма_зима")),
                source=_as_str(cell(row, "путь_к_файлу")),
            )
        )
    return legs


def day_intervals(legs: list[TripLeg]) -> dict[tuple[str, date], tuple[int, int]]:
    """(машина, дата) → (минуты первого выезда, минуты последнего возвращения)."""
    intervals: dict[tuple[str, date], tuple[int, int]] = {}
    for leg in legs:
        dep = hm_to_min(leg.time_dep)
        ret = hm_to_min(leg.time_ret)
        if dep is None or ret is None:
            continue
        key = (leg.vehicle, leg.day)
        if key in intervals:
            lo, hi = intervals[key]
            intervals[key] = (min(lo, dep), max(hi, ret))
        else:
            intervals[key] = (dep, ret)
    return intervals


def tx_position(
    tx: FuelTx, interval: tuple[int, int] | None
) -> str:
    """Позиция транзакции внутри дня поездки."""
    if interval is None:
        return ""
    t = tx.dt.hour * 60 + tx.dt.minute
    if t < interval[0]:
        return "до_выезда"
    if t > interval[1]:
        return "после_возвращения"
    return "в_рейсе"


def norm_consumption(leg: TripLeg) -> float | None:
    """Нормативный расход плеча, л: км × норма(сезон) / 100."""
    norm = leg.norm_winter if leg.day.month in WINTER_MONTHS else leg.norm_summer
    if leg.km is None or norm is None:
        return None
    return leg.km * norm / 100.0


# Сегменты адреса транзакции, которые мешают геокоду: «АЗС №2156», «ТД НМ № 011»…
_STATION_JUNK_SEGMENT_RE = re.compile(r"^(?:АЗС\b.*|ТД\s+.*)$", re.IGNORECASE)
_STATION_JUNK_INLINE_RE = re.compile(r"территория", re.IGNORECASE)
# Open Location Code в адресе: «PVGW+3P Кострома», «5J49+28 Сусанино».
PLUS_CODE_RE = re.compile(
    r"\b[23456789CFGHJMPQRVWXcfghjmpqrvwx]{4,8}"
    r"\+[23456789CFGHJMPQRVWXcfghjmpqrvwx]{2,3}\b"
)
# Хвост-дом в части адреса: « 20 б», « 108 корпус 2», « 15/1».
_HOUSE_TAIL_RE = re.compile(
    r"\s+\d+[А-Яа-яA-Za-z]?(?:/\d+[А-Яа-яA-Za-z]?)?(?:\s+(?:к|корпус)\s*\d+)?$"
)


def clean_station_display(address: str, city: str = "") -> str:
    """Адрес АЗС/мойки без служебных хвостов («АЗС №4», «ТД НМ № 011»).

    Если города нет в тексте адреса — дописываем (помогает адресам
    вида «д.Кузнечиха, ул.Луговая, д.12» без населённого пункта выше).
    """
    parts = [p.strip() for p in (address or "").split(",")]
    kept = [
        p
        for p in parts
        if p
        and not _STATION_JUNK_SEGMENT_RE.match(p)
        and not _STATION_JUNK_INLINE_RE.search(p)
        and len(p) > 1
    ]
    cleaned = ", ".join(kept) or (address or "").strip()
    if (
        city
        and city.casefold() not in cleaned.casefold()
        and not city.endswith("р-н")
    ):
        cleaned = f"{cleaned}, {city}"
    return cleaned


def _extract_plus_code(address: str) -> tuple[str, str, str] | None:
    """(код, город, полный хвост) из адреса с Open Location Code."""
    match = PLUS_CODE_RE.search(address or "")
    if not match:
        return None
    code = match.group(0).upper()
    tail = (address or "")[match.end():].strip(" ,")
    city = tail.split(",")[0].strip()
    return code, city, tail


def resolve_plus_codes(
    txs: list[FuelTx], addresses: dict[str, Any], *, offline: bool
) -> int:
    """Плюс-коды декодируются локально (openlocationcode), без Nominatim.

    Короткие коды («PVGW+3P») восстанавливаются относительно центра города
    (один запрос Nominatim на город). Возвращает число разрешённых адресов.
    """
    if offline:
        return 0
    try:
        from openlocationcode import openlocationcode as olc
    except ImportError:
        return 0
    resolved = 0
    city_ref: dict[str, tuple[float, float] | None] = {}
    for tx in txs:
        if not tx.addr_norm:
            continue
        existing = addresses.get(tx.addr_norm)
        if isinstance(existing, dict) and existing.get("lat") is not None:
            continue
        found = _extract_plus_code(tx.address)
        if not found:
            continue
        code, city, tail = found
        lat: float | None = None
        lon: float | None = None
        if olc.isFull(code):
            area = olc.decode(code)
            lat, lon = area.latitudeCenter, area.longitudeCenter
        elif city:
            # референс — весь хвост с регионом: «Сусанино, Костромская обл.»,
            # иначе Nominatim может выбрать одноимённый населённый пункт
            ref_query = tail or f"{city}, Россия"
            if ref_query not in city_ref:
                hit = nominatim_geocode(ref_query)
                city_ref[ref_query] = (hit[0], hit[1]) if hit else None
                time.sleep(NOMINATIM_RATE_LIMIT_SEC)
            ref = city_ref[ref_query]
            if ref is not None:
                # recoverNearest возвращает полный код, его нужно декодировать
                area = olc.decode(olc.recoverNearest(code, ref[0], ref[1]))
                lat, lon = area.latitudeCenter, area.longitudeCenter
        if lat is None:
            continue
        addresses[tx.addr_norm] = {
            "lat": lat,
            "lon": lon,
            "display": tx.address,
            "query": code,
            "source": "plus_code",
            "error": None,
        }
        resolved += 1
    return resolved


def mark_failed_for_retry(norms: set[str], addresses: dict[str, Any]) -> bool:
    """Пометить упавшие записи tx-адресов retry=True (перезапрос при прогоне)."""
    changed = False
    for norm in norms:
        entry = addresses.get(norm)
        if (
            isinstance(entry, dict)
            and entry.get("lat") is None
            and entry.get("error")
        ):
            entry["retry"] = True
            changed = True
    return changed


_HOUSE_PART_RE = re.compile(r"^\d+[А-Яа-яA-Za-z]?(?:/\d+[А-Яа-яA-Za-z]?)?$")


def retry_houseless(
    norms: set[str], display_map: dict[str, str], cache_path: Path
) -> int:
    """Для не найденных с домом — запрос без номера дома (точность до улицы)."""
    cache = load_address_cache(cache_path)
    addresses = cache.setdefault("addresses", {})
    resolved = 0
    dirty = False
    for norm in sorted(norms):
        entry = addresses.get(norm)
        if isinstance(entry, dict) and entry.get("lat") is not None:
            continue
        simplified = simplify_address_for_geocode(display_map.get(norm) or norm)
        parts = [p.strip() for p in simplified.split(",") if p.strip()]
        candidates: list[str] = []
        without_house_parts = [p for p in parts if not _HOUSE_PART_RE.fullmatch(p)]
        if len(without_house_parts) < len(parts) and without_house_parts:
            candidates.append(", ".join(without_house_parts))
        for i, part in enumerate(parts):
            stripped = _HOUSE_TAIL_RE.sub("", part)
            if stripped != part and stripped:
                candidates.append(", ".join([*parts[:i], stripped, *parts[i + 1:]]))
        for query in candidates[:2]:
            hit = nominatim_geocode(query)
            time.sleep(NOMINATIM_RATE_LIMIT_SEC)
            if hit is None:
                continue
            lat, lon, nom_display = hit
            addresses[norm] = {
                "lat": lat,
                "lon": lon,
                "display": nom_display,
                "query": query,
                "source": "nominatim_street",
                "error": None,
            }
            resolved += 1
            dirty = True
            break
    if dirty:
        save_address_cache(cache_path, cache)
    return resolved


def geocode_tx_addresses(
    txs: list[FuelTx],
    legs: list[TripLeg],
    cache_path: Path,
    *,
    offline: bool,
    force_geocode: bool,
) -> dict[str, Any]:
    """Геокод адресов АЗС/моек (+ добор адресов плеч) → addresses из кэша."""
    norms: set[str] = set()
    tx_norms: set[str] = set()
    display_map: dict[str, str] = {}
    for tx in txs:
        if not tx.addr_norm:
            continue
        norms.add(tx.addr_norm)
        tx_norms.add(tx.addr_norm)
        display_map.setdefault(
            tx.addr_norm, clean_station_display(tx.address, tx.city)
        )
    for leg in legs:
        for addr in (leg.addr_from, leg.addr_to):
            norm = normalize_address(addr)
            if norm:
                norms.add(norm)
                display_map.setdefault(norm, addr)

    cache = load_address_cache(cache_path)
    addresses = cache.setdefault("addresses", {})
    if resolve_plus_codes(txs, addresses, offline=offline):
        save_address_cache(cache_path, cache)
    if not offline and mark_failed_for_retry(tx_norms, addresses):
        save_address_cache(cache_path, cache)

    geocode_addresses(
        sorted(norms),
        display_map,
        cache_path,
        offline=offline,
        force_geocode=force_geocode,
    )
    if not offline:
        retry_houseless(tx_norms, display_map, cache_path)
    cache = load_address_cache(cache_path)
    return cache.get("addresses", {})


def _coords(addresses: dict[str, Any], norm: str) -> tuple[Any, Any]:
    coords = coords_for_norm(addresses, norm)
    if coords is None:
        return None, None
    return coords[0], coords[1]


def build_feed_rows(
    legs: list[TripLeg],
    txs: list[FuelTx],
    card2vehicle: dict[str, str],
    intervals: dict[tuple[str, date], tuple[int, int]],
    addresses: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[FuelTx], dict[tuple[str, date], list[FuelTx]]]:
    """События ленты + несопоставленные транзакции + матч (машина,дата)→tx."""
    trip_days = {(leg.vehicle, leg.day) for leg in legs}

    matched: dict[tuple[str, date], list[FuelTx]] = {}
    unmatched: list[FuelTx] = []
    for tx in txs:
        vehicle = card2vehicle.get(tx.card, "")
        key = (vehicle, tx.dt.date())
        if vehicle and key in trip_days:
            matched.setdefault(key, []).append(tx)
        else:
            unmatched.append(tx)

    rows: list[dict[str, Any]] = []

    def emit(row: dict[str, Any], sort_key: tuple[Any, ...]) -> None:
        row["_sort"] = sort_key
        rows.append(row)

    for leg in legs:
        base = {
            "дата": leg.day.isoformat(),
            "машина": leg.vehicle,
            "гос_номер": leg.plate,
            "км": leg.km,
            "водитель": leg.driver,
            "топливная_карта": "",
            "источник": Path(leg.source).name if leg.source else "",
        }
        lat, lon = _coords(addresses, normalize_address(leg.addr_from))
        dep_min = hm_to_min(leg.time_dep)
        emit(
            {
                **base,
                "время": leg.time_dep,
                "тип": "выезд",
                "адрес": leg.addr_from,
                "широта": lat,
                "долгота": lon,
                "куда": leg.addr_to,
            },
            (leg.vehicle, leg.day, dep_min if dep_min is not None else -1, 0, leg.seq),
        )
        lat, lon = _coords(addresses, normalize_address(leg.addr_to))
        ret_min = hm_to_min(leg.time_ret)
        emit(
            {
                **base,
                "время": leg.time_ret,
                "тип": "приезд",
                "адрес": leg.addr_to,
                "широта": lat,
                "долгота": lon,
                "куда": "",
            },
            (leg.vehicle, leg.day, ret_min if ret_min is not None else 9999, 2, leg.seq),
        )

    for (vehicle, day), day_txs in matched.items():
        interval = intervals.get((vehicle, day))
        for tx in day_txs:
            lat, lon = _coords(addresses, tx.addr_norm)
            t_min = tx.dt.hour * 60 + tx.dt.minute
            emit(
                {
                    "дата": day.isoformat(),
                    "машина": vehicle,
                    "гос_номер": "",
                    "время": tx.dt.strftime("%H:%M:%S"),
                    "тип": "мойка" if tx.is_wash else "заправка",
                    "адрес": tx.address,
                    "широта": lat,
                    "долгота": lon,
                    "куда": "",
                    "км": None,
                    "водитель": "",
                    "услуга": tx.service,
                    "кол-во": tx.qty,
                    "ед_изм": tx.unit,
                    "цена": tx.price,
                    "сумма": tx.amount,
                    "бренд": tx.brand,
                    "точка_обслуживания": tx.vendor,
                    "город": tx.city,
                    "позиция_в_дне": tx_position(tx, interval),
                    "топливная_карта": tx.card,
                    "источник": tx.source,
                },
                (vehicle, day, t_min, 1, 0),
            )

    rows.sort(key=lambda r: r["_sort"])
    per_day_counter: dict[tuple[str, str], int] = {}
    for row in rows:
        key = (row["машина"], row["дата"])
        per_day_counter[key] = per_day_counter.get(key, 0) + 1
        row["№_события"] = per_day_counter[key]
        del row["_sort"]
    return rows, unmatched, matched


def build_day_rows(
    legs: list[TripLeg],
    matched: dict[tuple[str, date], list[FuelTx]],
) -> list[dict[str, Any]]:
    """Агрегат дня машины: км, норма, заправлено, мойки, дельта."""
    days: dict[tuple[str, date], dict[str, Any]] = {}
    for leg in legs:
        key = (leg.vehicle, leg.day)
        d = days.setdefault(
            key,
            {
                "дата": leg.day.isoformat(),
                "машина": leg.vehicle,
                "гос_номер": leg.plate,
                "плеч": 0,
                "км": 0.0,
                "расход_норма_л": 0.0,
            },
        )
        d["плеч"] += 1
        d["км"] += leg.km or 0.0
        d["расход_норма_л"] += norm_consumption(leg) or 0.0
    for key, day_txs in matched.items():
        if key not in days:
            continue
        d = days[key]
        fuel = [t for t in day_txs if not t.is_wash]
        wash = [t for t in day_txs if t.is_wash]
        d["заправок"] = len(fuel)
        d["литров"] = round(sum(t.qty or 0.0 for t in fuel), 2)
        d["сумма_заправок"] = round(sum(t.amount or 0.0 for t in fuel), 2)
        d["моек"] = len(wash)
        d["сумма_моек"] = round(sum(t.amount or 0.0 for t in wash), 2)
    rows = []
    for key in sorted(days, key=lambda k: (k[0], k[1])):
        d = days[key]
        d.setdefault("заправок", 0)
        d.setdefault("литров", 0.0)
        d.setdefault("сумма_заправок", 0.0)
        d.setdefault("моек", 0)
        d.setdefault("сумма_моек", 0.0)
        d["км"] = round(d["км"], 1)
        d["расход_норма_л"] = round(d["расход_норма_л"], 1)
        d["дельта_л"] = round(d["литров"] - d["расход_норма_л"], 1)
        rows.append(d)
    return rows


_GOROD_TOKEN_RE = re.compile(r"(?<!\S)г\s+")


def _station_group_key(tx: FuelTx, addresses: dict[str, Any]) -> str:
    """Ключ группировки АЗС: координаты (~10 м), иначе норма без токена «г»."""
    coords = coords_for_norm(addresses, tx.addr_norm)
    if coords is not None:
        return f"{coords[0]:.4f},{coords[1]:.4f}"
    return _GOROD_TOKEN_RE.sub("", tx.addr_norm) or tx.address or tx.city


def _mode(values: list[str]) -> str:
    """Самое частое непустое значение."""
    counts: dict[str, int] = {}
    for v in values:
        if v:
            counts[v] = counts.get(v, 0) + 1
    return max(counts, key=lambda k: counts[k]) if counts else ""


def build_stations_geojson(
    txs: list[FuelTx],
    matched: dict[tuple[str, date], list[FuelTx]],
    unmatched: list[FuelTx],
    addresses: dict[str, Any],
    card2vehicle: dict[str, str],
) -> dict[str, Any]:
    """Точки АЗС/моек с агрегатами → GeoJSON FeatureCollection.

    Группировка по координатам (~10 м); точки без координат пропускаются.
    """
    matched_ids = {id(t) for day_txs in matched.values() for t in day_txs}
    groups: dict[str, list[FuelTx]] = {}
    for tx in txs:
        groups.setdefault(_station_group_key(tx, addresses), []).append(tx)

    features: list[dict[str, Any]] = []
    for group_txs in groups.values():
        coords = coords_for_norm(addresses, group_txs[0].addr_norm)
        if coords is None:
            continue
        lat, lon = coords
        fuel = [t for t in group_txs if not t.is_wash]
        wash = [t for t in group_txs if t.is_wash]
        has_fuel = bool(fuel)
        has_wash = bool(wash)
        station_type = (
            "азс+мойка" if has_fuel and has_wash else ("мойка" if has_wash else "азс")
        )
        unmatched_n = sum(1 for t in group_txs if id(t) not in matched_ids)
        entry = addresses.get(group_txs[0].addr_norm) or {}
        features.append(
            {
                "type": "Feature",
                "geometry": {"type": "Point", "coordinates": [lon, lat]},
                "properties": {
                    "тип": station_type,
                    "коорд": f"{lat:.4f},{lon:.4f}",
                    "бренд": _mode([t.brand for t in group_txs]),
                    "адрес": _mode([t.address for t in group_txs]),
                    "город": _mode([t.city for t in group_txs]),
                    "транзакций": len(group_txs),
                    "заправок": len(fuel),
                    "моек": len(wash),
                    "литров": round(sum(t.qty or 0.0 for t in fuel), 1),
                    "сумма_руб": round(sum(t.amount or 0.0 for t in group_txs), 2),
                    "машины": "; ".join(
                        sorted(
                            {
                                card2vehicle.get(t.card, "")
                                for t in group_txs
                                if card2vehicle.get(t.card)
                            }
                        )
                    ),
                    "первая_дата": min(t.dt.date().isoformat() for t in group_txs),
                    "последняя_дата": max(t.dt.date().isoformat() for t in group_txs),
                    "вне_поездок": unmatched_n,
                    "источник_координат": entry.get("source", ""),
                },
            }
        )
    features.sort(
        key=lambda f: -float(f["properties"].get("транзакций") or 0)
    )
    return {"type": "FeatureCollection", "features": features}


# Ручные координаты трассовых АЗС (не находятся свободным геокодом).
DEFAULT_MANUAL_STATIONS = (
    PROJECT_ROOT / "ГСМ" / "geo_cache" / "stations_manual.json"
)


def load_manual_stations(path: Path) -> dict[str, Any]:
    """stations_manual.json: {addr_norm: {lat, lon, comment}} → entries кэша."""
    path = Path(path)
    if not path.is_file():
        return {}

    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}
    stations = data.get("stations", {}) if isinstance(data, dict) else {}
    out: dict[str, Any] = {}
    for norm, entry in stations.items():
        if not isinstance(entry, dict):
            continue
        lat = entry.get("lat")
        lon = entry.get("lon")
        if lat is None or lon is None:
            continue
        out[norm] = {
            "lat": float(lat),
            "lon": float(lon),
            "display": entry.get("display") or norm,
            "query": "manual",
            "source": "manual",
            "error": None,
        }
    return out


def save_manual_template(
    path: Path, missing_norms: list[str], display_map: dict[str, str]
) -> int:
    """Дописать в stations_manual.json ненайденные адреса с пустыми координатами.

    Существующие записи (в т.ч. заполненные вручную) не трогаем.
    Возвращает число добавленных строк-заготовок.
    """
    path = Path(path)
    data: dict[str, Any] = {"version": 1, "stations": {}}
    if path.is_file():
        try:
            loaded = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(loaded, dict) and isinstance(loaded.get("stations"), dict):
                data = loaded
        except (json.JSONDecodeError, OSError):
            pass
    stations = data.setdefault("stations", {})
    added = 0
    for norm in sorted(missing_norms):
        if norm in stations:
            continue
        stations[norm] = {
            "lat": None,
            "lon": None,
            "display": display_map.get(norm, norm),
            "comment": "",
        }
        added += 1
    if added or not path.is_file():
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2, sort_keys=True),
            encoding="utf-8",
        )
    return added


@dataclass(frozen=True)
class RouteStationLink:
    """Связь «маршрут ↔ типовая АЗС» (топ-3 станции маршрута по заправкам)."""

    vehicle: str
    addr_a: str
    addr_b: str
    a_norm: str
    b_norm: str
    station_norm: str
    station_address: str
    заправок: int


def build_route_station_links(
    pool_path: Path,
    legs: list[TripLeg],
    matched: dict[tuple[str, date], list[FuelTx]],
    addresses: dict[str, Any],
) -> list[RouteStationLink]:
    """Структурные связи маршрут↔АЗС для карты (routes_via.geojson).

    Заправка дня относится ко всем маршрутам, которые машина ехала в этот день
    (обычно день = один круг ABA, так что неоднозначность редка).
    """
    header, rows = _load_pool_sheet(pool_path, "routes_ab")
    col = {name: i for i, name in enumerate(header) if name}

    def cell(row: tuple[Any, ...], name: str) -> Any:
        i = col.get(name)
        return row[i] if i is not None and i < len(row) else None

    route_days: dict[tuple[str, str, str], set[date]] = {}
    for leg in legs:
        a = normalize_address(leg.addr_from)
        b = normalize_address(leg.addr_to)
        route_days.setdefault((leg.vehicle, a, b), set()).add(leg.day)

    links: list[RouteStationLink] = []
    for row in rows:
        vehicle = _as_str(cell(row, "машина"))
        addr_a = _as_str(cell(row, "адрес_A"))
        addr_b = _as_str(cell(row, "адрес_B"))
        a_norm = normalize_address(addr_a) or _as_str(cell(row, "адрес_A_норм"))
        b_norm = normalize_address(addr_b) or _as_str(cell(row, "адрес_B_норм"))
        days = route_days.get((vehicle, a_norm, b_norm), set())
        stations: dict[str, int] = {}
        station_names: dict[str, str] = {}
        station_norms: dict[str, str] = {}
        for day in days:
            for tx in matched.get((vehicle, day), []):
                if tx.is_wash:
                    continue
                key = _station_group_key(tx, addresses)
                stations[key] = stations.get(key, 0) + 1
                station_names.setdefault(key, tx.address or tx.city)
                # норму станции берём с координатами, если она есть у части транзакций
                prev_norm = station_norms.get(key)
                if prev_norm is None or (
                    coords_for_norm(addresses, prev_norm) is None
                    and coords_for_norm(addresses, tx.addr_norm) is not None
                ):
                    station_norms[key] = tx.addr_norm
        for key, n in sorted(stations.items(), key=lambda kv: -kv[1])[:3]:
            links.append(
                RouteStationLink(
                    vehicle=vehicle,
                    addr_a=addr_a,
                    addr_b=addr_b,
                    a_norm=a_norm,
                    b_norm=b_norm,
                    station_norm=station_norms[key],
                    station_address=station_names[key],
                    заправок=n,
                )
            )
    return links


def save_route_station_links(path: Path, links: list[RouteStationLink]) -> None:
    """route_stations.json — контракт для build_gsm_routes_map.py."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": 1,
        "links": [
            {
                "машина": lk.vehicle,
                "адрес_A": lk.addr_a,
                "адрес_B": lk.addr_b,
                "address_a_norm": lk.a_norm,
                "address_b_norm": lk.b_norm,
                "station_norm": lk.station_norm,
                "станция": lk.station_address,
                "заправок": lk.заправок,
            }
            for lk in links
        ],
    }
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8"
    )


def build_route_rows(
    pool_path: Path,
    legs: list[TripLeg],
    matched: dict[tuple[str, date], list[FuelTx]],
    addresses: dict[str, Any],
) -> list[dict[str, Any]]:
    """Уникальные маршруты routes_ab + типовые АЗС в дни этого маршрута."""
    links = build_route_station_links(pool_path, legs, matched, addresses)
    by_route: dict[tuple[str, str, str], list[RouteStationLink]] = {}
    for lk in links:
        by_route.setdefault((lk.vehicle, lk.a_norm, lk.b_norm), []).append(lk)

    route_days: dict[tuple[str, str, str], set[date]] = {}
    for leg in legs:
        a = normalize_address(leg.addr_from)
        b = normalize_address(leg.addr_to)
        route_days.setdefault((leg.vehicle, a, b), set()).add(leg.day)

    header, rows = _load_pool_sheet(pool_path, "routes_ab")
    col = {name: i for i, name in enumerate(header) if name}

    def cell(row: tuple[Any, ...], name: str) -> Any:
        i = col.get(name)
        return row[i] if i is not None and i < len(row) else None

    out: list[dict[str, Any]] = []
    for row in rows:
        vehicle = _as_str(cell(row, "машина"))
        addr_a = _as_str(cell(row, "адрес_A"))
        addr_b = _as_str(cell(row, "адрес_B"))
        a_norm = normalize_address(addr_a) or _as_str(cell(row, "адрес_A_норм"))
        b_norm = normalize_address(addr_b) or _as_str(cell(row, "адрес_B_норм"))
        count = 0
        liters = 0.0
        for day in route_days.get((vehicle, a_norm, b_norm), set()):
            for tx in matched.get((vehicle, day), []):
                if tx.is_wash:
                    continue
                count += 1
                liters += tx.qty or 0.0
        route_links = by_route.get((vehicle, a_norm, b_norm), [])
        out.append(
            {
                "машина": vehicle,
                "адрес_A": addr_a,
                "адрес_B": addr_b,
                "км": _parse_float(cell(row, "км")),
                "частота": _parse_float(cell(row, "частота")),
                "дат_всего": _parse_float(cell(row, "дат_всего")),
                "заправок_в_дни": count,
                "литров_в_дни": round(liters, 1),
                "типовые_АЗС": "; ".join(
                    f"{lk.station_address} (×{lk.заправок})" for lk in route_links
                ),
            }
        )
    return out


def build_unmatched_rows(
    unmatched: list[FuelTx],
    card2vehicle: dict[str, str],
    addresses: dict[str, Any],
) -> list[dict[str, Any]]:
    """Транзакции без поездки в тот же день."""
    rows = []
    for tx in sorted(unmatched, key=lambda t: (card2vehicle.get(t.card, ""), t.dt)):
        lat, lon = _coords(addresses, tx.addr_norm)
        rows.append(
            {
                "дата": tx.dt.date().isoformat(),
                "машина": card2vehicle.get(tx.card, "карта не в пуле"),
                "топливная_карта": tx.card,
                "время": tx.dt.strftime("%H:%M:%S"),
                "тип": "мойка" if tx.is_wash else "заправка",
                "услуга": tx.service,
                "кол-во": tx.qty,
                "ед_изм": tx.unit,
                "цена": tx.price,
                "сумма": tx.amount,
                "бренд": tx.brand,
                "точка_обслуживания": tx.vendor,
                "город": tx.city,
                "адрес": tx.address,
                "широта": lat,
                "долгота": lon,
                "источник": tx.source,
            }
        )
    return rows


def _write_sheet(
    ws, headers: list[str], rows: list[dict[str, Any]], max_width: int = 60
) -> None:
    bold = Font(bold=True)
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h).font = bold
    for r_idx, row in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            value = row.get(h)
            if value is not None:
                ws.cell(r_idx, c, value)
    for c, h in enumerate(headers, 1):
        letter = get_column_letter(c)
        max_len = len(str(h))
        for r_idx in range(2, min(len(rows) + 2, 200)):
            val = ws.cell(r_idx, c).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), max_width))
        ws.column_dimensions[letter].width = max_len + 2
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"
    ws.freeze_panes = "A2"


FEED_HEADERS = [
    "дата", "машина", "гос_номер", "№_события", "время", "тип",
    "адрес", "широта", "долгота", "куда", "км", "водитель",
    "услуга", "кол-во", "ед_изм", "цена", "сумма", "бренд",
    "точка_обслуживания", "город", "позиция_в_дне", "топливная_карта",
    "источник",
]
DAY_HEADERS = [
    "дата", "машина", "гос_номер", "плеч", "км", "расход_норма_л",
    "заправок", "литров", "сумма_заправок", "моек", "сумма_моек", "дельта_л",
]
ROUTE_HEADERS = [
    "машина", "адрес_A", "адрес_B", "км", "частота", "дат_всего",
    "заправок_в_дни", "литров_в_дни", "типовые_АЗС",
]
UNMATCHED_HEADERS = [
    "дата", "машина", "топливная_карта", "время", "тип", "услуга",
    "кол-во", "ед_изм", "цена", "сумма", "бренд", "точка_обслуживания",
    "город", "адрес", "широта", "долгота", "источник",
]


def run(
    tx_dir: Path,
    pool_path: Path,
    out_path: Path,
    cache_path: Path,
    *,
    offline: bool = False,
    force_geocode: bool = False,
    stations_out: Path | None = None,
) -> dict[str, Any]:
    stations_out = Path(stations_out) if stations_out else DEFAULT_STATIONS
    txs, tx_warnings = load_transactions(tx_dir)
    legs = load_legs(pool_path)
    if not txs:
        raise ValueError(f"Нет транзакций в {tx_dir}")
    if not legs:
        raise ValueError(f"Нет поездок на листе facts в {pool_path}")

    card2vehicle: dict[str, str] = {}
    for leg in legs:
        for card in split_cards(leg.cards):
            card2vehicle.setdefault(card, leg.vehicle)

    intervals = day_intervals(legs)
    addresses = geocode_tx_addresses(
        txs, legs, cache_path, offline=offline, force_geocode=force_geocode
    )
    manual = load_manual_stations(DEFAULT_MANUAL_STATIONS)
    addresses.update(manual)

    feed_rows, unmatched, matched = build_feed_rows(
        legs, txs, card2vehicle, intervals, addresses
    )
    day_rows = build_day_rows(legs, matched)
    route_rows = build_route_rows(pool_path, legs, matched, addresses)
    unmatched_rows = build_unmatched_rows(unmatched, card2vehicle, addresses)

    wb = Workbook()
    _write_sheet(wb.active, FEED_HEADERS, feed_rows)
    wb.active.title = "лента"
    _write_sheet(wb.create_sheet("рейсы_дня"), DAY_HEADERS, day_rows)
    _write_sheet(wb.create_sheet("маршруты"), ROUTE_HEADERS, route_rows)
    _write_sheet(wb.create_sheet("вне_поездок"), UNMATCHED_HEADERS, unmatched_rows)
    wb.save(out_path)

    stations = build_stations_geojson(txs, matched, unmatched, addresses, card2vehicle)
    Path(stations_out).parent.mkdir(parents=True, exist_ok=True)
    Path(stations_out).write_text(
        json.dumps(stations, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    links = build_route_station_links(pool_path, legs, matched, addresses)
    save_route_station_links(DEFAULT_ROUTE_STATIONS, links)

    missing_norms = sorted(
        {
            t.addr_norm
            for t in txs
            if coords_for_norm(addresses, t.addr_norm) is None
        }
    )
    display_by_norm = {t.addr_norm: t.address for t in txs}
    template_added = save_manual_template(
        DEFAULT_MANUAL_STATIONS, missing_norms, display_by_norm
    )

    geocoded_tx = sum(
        1 for t in txs if coords_for_norm(addresses, t.addr_norm) is not None
    )
    return {
        "out": str(out_path),
        "stations_geojson": str(stations_out),
        "транзакций": len(txs),
        "сопоставлено": len(txs) - len(unmatched),
        "вне_поездок": len(unmatched),
        "плеч": len(legs),
        "событий_ленты": len(feed_rows),
        "дней": len(day_rows),
        "маршрутов": len(route_rows),
        "tx_с_координатами": geocoded_tx,
        "станций_на_карте": len(stations["features"]),
        "route_station_связей": len(links),
        "manual_применено": len(manual),
        "manual_добавлено_заготовок": template_added,
        "без_координат_адресов": len(missing_norms),
        "предупреждения": tx_warnings,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Лента «поездки+заправки+мойки» → поездки_заправки.xlsx"
    )
    parser.add_argument("--tx-dir", type=Path, default=DEFAULT_TX_DIR,
                        help="Каталог с transactions_excel*.xls")
    parser.add_argument("--pool", type=Path, default=DEFAULT_POOL,
                        help="пул_поездок.xlsx")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT,
                        help="Куда писать результат")
    parser.add_argument("--geo-cache", type=Path, default=DEFAULT_GEO_CACHE,
                        help="Кэш геокода addresses.json")
    parser.add_argument("--offline", action="store_true",
                        help="Только кэш геокода, без сети")
    parser.add_argument("--force-geocode", action="store_true",
                        help="Перезапросить Nominatim даже для кэшированных")
    parser.add_argument("--stations-out", type=Path, default=DEFAULT_STATIONS,
                        help="GeoJSON точек АЗС/моек для карты")
    args = parser.parse_args(argv)

    report = run(
        args.tx_dir,
        args.pool,
        args.out,
        args.geo_cache,
        offline=args.offline,
        force_geocode=args.force_geocode,
        stations_out=args.stations_out,
    )
    print(f"Записано: {report['out']}")
    print(
        "транзакций: {транзакций} (сопоставлено {сопоставлено}, "
        "вне поездок {вне_поездок})".format(**report)
    )
    print(
        "плеч: {плеч}, событий ленты: {событий_ленты}, дней: {дней}, "
        "маршрутов: {маршрутов}".format(**report)
    )
    print(
        "координаты у транзакций: {tx_с_координатами}/{транзакций}".format(**report)
    )
    print(
        "станций на карте: {станций_на_карте} → {stations_geojson} "
        "(без координат адресов: {без_координат_адресов}, "
        "manual заготовок: +{manual_добавлено_заготовок})".format(**report)
    )
    for warning in report["предупреждения"]:
        print(f"ВНИМАНИЕ: {warning}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

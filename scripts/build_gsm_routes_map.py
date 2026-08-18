#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Сборка HTML-карты маршрутов ГСМ из пула поездок.

Читает лист routes_ab → геокод → OSRM → Leaflet HTML.
См. ai_docs/specs/gsm-routes-map.md
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

import requests
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from scripts.build_gsm_trip_pool import normalize_address

DEFAULT_XLSX = PROJECT_ROOT / "ГСМ" / "пул_поездок.xlsx"
DEFAULT_OUT = PROJECT_ROOT / "ГСМ" / "карта_маршрутов.html"
DEFAULT_GEO_CACHE = PROJECT_ROOT / "ГСМ" / "geo_cache" / "addresses.json"
DEFAULT_ROUTES_CACHE = PROJECT_ROOT / "ГСМ" / "geo_cache" / "routes.geojson"
DEFAULT_STATIONS = PROJECT_ROOT / "ГСМ" / "geo_cache" / "stations.geojson"
SHEET_ROUTES_AB = "routes_ab"

REQUIRED_HEADERS = ("машина", "адрес_A", "адрес_B", "км", "частота")

GEO_CACHE_VERSION = 1
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
NOMINATIM_USER_AGENT = "ShishovGSM/1.0 (local; contact: factory)"
NOMINATIM_RATE_LIMIT_SEC = 1.1
NOMINATIM_TIMEOUT_SEC = 30

OSRM_BASE_URL = "https://router.project-osrm.org/route/v1/driving"
OSRM_RATE_LIMIT_SEC = 0.2
OSRM_TIMEOUT_SEC = 60
# overview=simplified + Douglas-Peucker → обычно <200 точек на маршрут.
OSRM_OVERVIEW = "simplified"
SIMPLIFY_TOLERANCE_M = 30.0
# Если после упрощения GeoJSON всё ещё слишком большой для <script> — sibling + fetch.
EMBED_GEOJSON_MAX_BYTES = 5_000_000
# Ключ кэша геометрии: округлённые lon,lat A→B (см. route_osrm_cache_key).
# Один трек переиспользуется для разных машин с теми же координатами.

# Фиксированная палитра для 4 машин завода (Leaflet polylines).
VEHICLE_COLORS: dict[str, str] = {
    "Geely Monjaro": "#d62728",
    "Hyundai Palisade": "#1f77b4",
    "Geely Tugella 848": "#2ca02c",
    "Geely Tugella 952": "#ff7f0e",
}
DEFAULT_VEHICLE_COLOR = "#7f7f7f"
LEAFLET_CSS_CDN = "https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"
LEAFLET_JS_CDN = "https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"
# Кострома — запасной центр, если FeatureCollection пуст.
DEFAULT_MAP_CENTER = (57.77, 40.93)
DEFAULT_MAP_ZOOM = 8

EARTH_RADIUS_KM = 6371.0088
EARTH_RADIUS_M = 6_371_008.8
# Photon (Komoot) — CORS-friendly geocoder для браузерного поиска.
PHOTON_API_URL = "https://photon.komoot.io/api/"


@dataclass(frozen=True)
class RouteAb:
    """Одна уникальная пара A→B с листа routes_ab."""

    vehicle: str
    addr_a: str
    addr_b: str
    km: float | None
    frequency: int
    addr_a_norm: str
    addr_b_norm: str
    mark: str = ""
    plate: str = ""
    typical_dep: str = ""
    drivers: str = ""
    fuel: str = ""


@dataclass(frozen=True)
class GeocodeStats:
    """Сводка геокодирования уникальных адресов."""

    geocoded: int = 0
    failed: int = 0
    from_cache: int = 0
    missing: tuple[str, ...] = ()


@dataclass(frozen=True)
class OsrmStats:
    """Сводка построения треков OSRM → routes.geojson."""

    routed: int = 0
    from_cache: int = 0
    skipped_missing_coords: int = 0
    skipped_osrm_error: int = 0
    skipped_offline: int = 0
    features: int = 0
    skips: tuple[str, ...] = ()


def _as_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value).strip()


def _parse_float(value: Any) -> float | None:
    if value in ("", None):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _as_str(value).replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _parse_int(value: Any) -> int:
    if value in ("", None):
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = _as_str(value)
    try:
        return int(float(text.replace(",", ".")))
    except ValueError:
        return 0


def load_routes_ab(xlsx_path: Path, *, sheet_name: str = SHEET_ROUTES_AB) -> list[RouteAb]:
    """Прочитать лист routes_ab → список маршрутов."""
    path = Path(xlsx_path)
    if not path.is_file():
        raise FileNotFoundError(f"Нет файла пула: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"В {path} нет листа {sheet_name!r}. Есть: {', '.join(wb.sheetnames)}"
            )
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise ValueError(f"Лист {sheet_name!r} пуст") from exc

        headers = [_as_str(h) for h in header_row]
        index = {name: i for i, name in enumerate(headers) if name}
        missing = [h for h in REQUIRED_HEADERS if h not in index]
        if missing:
            raise ValueError(
                f"На листе {sheet_name!r} нет колонок: {', '.join(missing)}"
            )

        def cell(row: tuple[Any, ...], name: str) -> Any:
            i = index.get(name)
            if i is None or i >= len(row):
                return None
            return row[i]

        routes: list[RouteAb] = []
        for row in rows:
            if row is None or all(v in ("", None) for v in row):
                continue
            addr_a = _as_str(cell(row, "адрес_A"))
            addr_b = _as_str(cell(row, "адрес_B"))
            if not addr_a and not addr_b:
                continue

            a_norm_sheet = _as_str(cell(row, "адрес_A_норм"))
            b_norm_sheet = _as_str(cell(row, "адрес_B_норм"))
            addr_a_norm = normalize_address(addr_a) or a_norm_sheet
            addr_b_norm = normalize_address(addr_b) or b_norm_sheet

            routes.append(
                RouteAb(
                    vehicle=_as_str(cell(row, "машина")),
                    addr_a=addr_a,
                    addr_b=addr_b,
                    km=_parse_float(cell(row, "км")),
                    frequency=_parse_int(cell(row, "частота")),
                    addr_a_norm=addr_a_norm,
                    addr_b_norm=addr_b_norm,
                    mark=_as_str(cell(row, "марка")),
                    plate=_as_str(cell(row, "гос_номер")),
                    typical_dep=_as_str(cell(row, "типичное_время_выезда")),
                    drivers=_as_str(cell(row, "водители")),
                    fuel=_as_str(cell(row, "топливо")),
                )
            )
        return routes
    finally:
        wb.close()


def collect_unique_addresses(routes: list[RouteAb]) -> list[str]:
    """Уникальные нормализованные адреса из A и B (отсортированный список)."""
    seen: set[str] = set()
    for r in routes:
        if r.addr_a_norm:
            seen.add(r.addr_a_norm)
        if r.addr_b_norm:
            seen.add(r.addr_b_norm)
    return sorted(seen)


def address_display_map(routes: list[RouteAb]) -> dict[str, str]:
    """norm → первый встреченный «сырой» адрес (для геокода/подписей)."""
    mapping: dict[str, str] = {}
    for r in routes:
        if r.addr_a_norm and r.addr_a_norm not in mapping:
            mapping[r.addr_a_norm] = r.addr_a or r.addr_a_norm
        if r.addr_b_norm and r.addr_b_norm not in mapping:
            mapping[r.addr_b_norm] = r.addr_b or r.addr_b_norm
    return mapping


def empty_address_cache() -> dict[str, Any]:
    return {"version": GEO_CACHE_VERSION, "addresses": {}}


def load_address_cache(path: Path) -> dict[str, Any]:
    """Загрузить кэш адресов; при отсутствии файла — пустой кэш."""
    path = Path(path)
    if not path.is_file():
        return empty_address_cache()
    raw = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise ValueError(f"Кэш адресов должен быть объектом JSON: {path}")
    addresses = raw.get("addresses")
    if addresses is None:
        addresses = {}
    if not isinstance(addresses, dict):
        raise ValueError(f"Поле addresses в кэше должно быть объектом: {path}")
    return {
        "version": int(raw.get("version") or GEO_CACHE_VERSION),
        "addresses": addresses,
    }


def save_address_cache(path: Path, cache: dict[str, Any]) -> None:
    """Записать кэш адресов (атомарно через .tmp)."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": int(cache.get("version") or GEO_CACHE_VERSION),
        "addresses": cache.get("addresses") or {},
    }
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    tmp.replace(path)


def _coord_as_float(value: Any) -> float | None:
    """Числовой lat/lon; строки вроде \"57.7\" допускаются (ручные правки кэша)."""
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        return _parse_float(value)
    return None


def cache_entry_has_coords(entry: Any) -> bool:
    """True, если в записи есть числовые lat/lon (в т.ч. строки-числа)."""
    if not isinstance(entry, dict):
        return False
    return (
        _coord_as_float(entry.get("lat")) is not None
        and _coord_as_float(entry.get("lon")) is not None
    )


def coerce_cache_entry_coords(entry: dict[str, Any]) -> tuple[bool, bool]:
    """Привести lat/lon к float на месте.

    Returns:
        (has_coords, changed) — оба валидны; запись изменена.
    """
    lat = _coord_as_float(entry.get("lat"))
    lon = _coord_as_float(entry.get("lon"))
    if lat is None or lon is None:
        return False, False
    changed = entry.get("lat") != lat or entry.get("lon") != lon
    entry["lat"] = lat
    entry["lon"] = lon
    return True, changed


def cache_entry_should_skip_network(
    entry: Any, *, force_geocode: bool = False
) -> bool:
    """True: не ходить в Nominatim (есть ключ в кэше, в т.ч. not_found).

    Повторный запрос только при force_geocode или entry.retry is True.
    """
    if not isinstance(entry, dict):
        return False
    if force_geocode:
        return False
    if entry.get("retry") is True:
        return False
    return True


_COUNTRY_TAIL_RE = re.compile(r",?\s*(?:Россия|Russia|РФ)\s*$", re.IGNORECASE)
# Почтовый индекс в хвосте: «…, Россия, 150032» / «…, Россия 150032»
_POSTAL_TAIL_RE = re.compile(r",?\s*\d{6}\s*$")
# ООО/СК/ИП и т.п. — хвост до конца строки (в т.ч. ООО СЗ "…", Завод ЖБИ "…").
# \b слева обязателен: иначе «СК» съедает окончание «Приволжск» → «Приволж».
_ORG_TAIL_RE = re.compile(
    r",?\s*\b(?:"
    r"(?:ООО|OOO|АО|ЗАО|ПАО|ОАО|ИП|СК|ГК|СЗ|НКО|ТСЖ|СНТ|МБУ|МУП|ГУП)\b"
    r"|Завод\b"
    r").*$",
    re.IGNORECASE,
)
# стр./корп./лит./пом./к. + значение — только как отдельные метки (не внутри слов)
_STRUCT_DROP_RE = re.compile(
    r",?\s*\b(?:стр|корп|лит|пом)\.?\s*[А-Яа-яA-Za-z0-9]+\b"
    r"|,\s*\bк\.\s*[А-Яа-яA-Za-z0-9]+\b",
    re.IGNORECASE,
)
# Однобуквенные/короткие аббревиатуры: справа обязательны точка, пробел или конец
# строки — иначе «Переславль» → «переулок еславль», «поселение» → «посёлок еление».
_ABBREV_EXPAND: tuple[tuple[re.Pattern[str], str], ...] = (
    (re.compile(r"\bпр-д\.?\s*", re.IGNORECASE), "проезд "),
    (re.compile(r"\bпр-кт\.?\s*", re.IGNORECASE), "проспект "),
    (re.compile(r"\bпр-т\.?\s*", re.IGNORECASE), "проспект "),
    (re.compile(r"\bпл(?:\.\s*|\s+|$)", re.IGNORECASE), "площадь "),
    (re.compile(r"\bобл(?:\.\s*|\s+|$)", re.IGNORECASE), "область "),
    (re.compile(r"\bпер(?:\.\s*|\s+|$)", re.IGNORECASE), "переулок "),
    (re.compile(r"\bпос(?:\.\s*|\s+|$)", re.IGNORECASE), "посёлок "),
    (re.compile(r"\bш\.\s*", re.IGNORECASE), "шоссе "),
    (re.compile(r"\bпр(?:ос)?(?:\.\s*|\s+|$)", re.IGNORECASE), "проспект "),
)
# г./ул./д./дом — точка или пробел обязательны (не трогать «готово»)
_ADDR_LABEL_RE = re.compile(
    r"(?<![А-Яа-яA-Za-z0-9])"
    r"(?:г|ул|д|дом)"
    r"(?:\.\s*|\s+)"
    r"(?=[А-Яа-яA-Z0-9«\"'])",
    re.IGNORECASE,
)
_HOUSE_TOKEN_RE = re.compile(
    r"^\d+[А-Яа-яA-Za-z]?(?:/\d+[А-Яа-яA-Za-z]?)?$"
)


def _is_house_token(token: str) -> bool:
    return bool(_HOUSE_TOKEN_RE.fullmatch(token.strip()))


_REGION_PART_RE = re.compile(r"область|район|округ|Россия", re.IGNORECASE)


def _merge_street_house(parts: list[str]) -> list[str]:
    """«…, Кузнецкая, 18Б» → «…, Кузнецкая 18Б» (дом склеиваем с улицей,
    где бы он ни стоял; не склеиваем с регионом/районом)."""
    out: list[str] = []
    for part in parts:
        if out and _is_house_token(part) and not _REGION_PART_RE.search(out[-1]):
            out[-1] = f"{out[-1]} {part}"
        else:
            out.append(part)
    return out


def simplify_address_for_geocode(display: str) -> str:
    """Упростить адрес для Nominatim: без ООО/меток г./ул./д., форма city, street house."""
    text = (display or "").strip()
    if not text:
        return ""
    text = _POSTAL_TAIL_RE.sub("", text)
    text = _COUNTRY_TAIL_RE.sub("", text).strip().rstrip(",").strip()
    text = _POSTAL_TAIL_RE.sub("", text)
    text = _COUNTRY_TAIL_RE.sub("", text).strip().rstrip(",").strip()
    text = _ORG_TAIL_RE.sub("", text).strip().rstrip(",").strip()
    text = _STRUCT_DROP_RE.sub(" ", text)
    for pat, repl in _ABBREV_EXPAND:
        text = pat.sub(repl, text)
    text = _ADDR_LABEL_RE.sub(" ", text)
    text = re.sub(r"\s*,\s*", ", ", text)
    text = re.sub(r"\s+", " ", text).strip(" ,")
    parts = [p.strip() for p in text.split(",") if p.strip()]
    parts = _merge_street_house(parts)
    return ", ".join(parts)


def nominatim_query_variants(display: str) -> list[str]:
    """2–3 варианта запроса; первый — предпочтительный упрощённый."""
    simplified = simplify_address_for_geocode(display)
    variants: list[str] = []

    def _add(query: str) -> None:
        q = query.strip()
        if q and q not in variants:
            variants.append(q)

    _add(simplified)
    if simplified:
        _add(f"{simplified}, Россия")
    parts = [p.strip() for p in simplified.split(",") if p.strip()] if simplified else []
    if len(parts) >= 3 and "область" in parts[0].casefold():
        _add(", ".join(parts[1:]))
    if not variants:
        raw = (display or "").strip()
        raw = _COUNTRY_TAIL_RE.sub("", raw).strip().rstrip(",").strip()
        _add(raw)
    if not variants:
        _add("Россия")
    return variants


def nominatim_query(display: str) -> str:
    """Основной запрос к Nominatim (упрощённый адрес)."""
    return nominatim_query_variants(display)[0]


def nominatim_geocode(
    query: str,
    *,
    session: requests.Session | None = None,
    timeout: float = NOMINATIM_TIMEOUT_SEC,
) -> tuple[float, float, str] | None:
    """Один запрос к Nominatim. None, если ничего не найдено."""
    http = session or requests.Session()
    headers = {
        "User-Agent": NOMINATIM_USER_AGENT,
        "Accept-Language": "ru",
    }
    params = {
        "q": query,
        "format": "json",
        "limit": 1,
        "countrycodes": "ru",
    }
    resp = http.get(
        NOMINATIM_URL,
        params=params,
        headers=headers,
        timeout=timeout,
    )
    resp.raise_for_status()
    data = resp.json()
    if not isinstance(data, list) or not data:
        return None
    hit = data[0]
    lat = float(hit["lat"])
    lon = float(hit["lon"])
    display = _as_str(hit.get("display_name")) or query
    return lat, lon, display


def geocode_addresses(
    norms: list[str],
    display_map: dict[str, str],
    cache_path: Path,
    *,
    offline: bool = False,
    force_geocode: bool = False,
    sleep_sec: float = NOMINATIM_RATE_LIMIT_SEC,
    sleep_fn: Callable[[float], None] = time.sleep,
    geocode_fn: Callable[[str], tuple[float, float, str] | None] | None = None,
    session: requests.Session | None = None,
) -> GeocodeStats:
    """Геокодировать уникальные адреса с кэшем на диске.

    --offline: только кэш, без сети; отсутствующие попадают в missing.
    Отрицательный кэш (lat/lon null, error not_found и т.п.) не перезапрашивается
    онлайн, пока нет force_geocode / entry.retry.
    """
    cache = load_address_cache(cache_path)
    addresses: dict[str, Any] = cache.setdefault("addresses", {})
    geocoded = 0
    failed = 0
    from_cache = 0
    missing: list[str] = []
    network_calls = 0
    dirty = False

    def _do_geocode(query: str) -> tuple[float, float, str] | None:
        if geocode_fn is not None:
            return geocode_fn(query)
        return nominatim_geocode(query, session=session)

    for norm in norms:
        if not norm:
            continue
        entry = addresses.get(norm)
        if isinstance(entry, dict):
            _, coords_changed = coerce_cache_entry_coords(entry)
            if coords_changed:
                dirty = True
            if cache_entry_should_skip_network(
                entry, force_geocode=force_geocode and not offline
            ):
                from_cache += 1
                continue

        if offline:
            missing.append(norm)
            continue

        display = display_map.get(norm) or norm
        variants = nominatim_query_variants(display)
        query = variants[0]
        result: tuple[float, float, str] | None = None

        try:
            for query in variants:
                if network_calls > 0 and sleep_sec > 0:
                    sleep_fn(sleep_sec)
                network_calls += 1
                result = _do_geocode(query)
                if result is not None:
                    break
        except (requests.RequestException, ValueError, KeyError, TypeError) as exc:
            addresses[norm] = {
                "lat": None,
                "lon": None,
                "display": display,
                "query": query,
                "source": "nominatim",
                "error": str(exc),
            }
            failed += 1
            dirty = True
            continue

        if result is None:
            addresses[norm] = {
                "lat": None,
                "lon": None,
                "display": display,
                "query": query,
                "source": "nominatim",
                "error": "not_found",
            }
            failed += 1
            dirty = True
            continue

        lat, lon, nom_display = result
        addresses[norm] = {
            "lat": lat,
            "lon": lon,
            "display": nom_display,
            "query": query,
            "source": "nominatim",
            "error": None,
        }
        geocoded += 1
        dirty = True

    if dirty:
        save_address_cache(cache_path, cache)

    return GeocodeStats(
        geocoded=geocoded,
        failed=failed,
        from_cache=from_cache,
        missing=tuple(missing),
    )


def route_osrm_cache_key(
    lon1: float, lat1: float, lon2: float, lat2: float
) -> str:
    """Ключ кэша трека: округлённые координаты A→B (5 знаков ≈ 1 м).

    Документация: кэш геометрии привязан к координатам, не к машине/нормам.
    Разные машины с теми же lat/lon переиспользуют один OSRM-ответ.
    """
    return f"{lon1:.5f},{lat1:.5f}->{lon2:.5f},{lat2:.5f}"


def empty_routes_geojson() -> dict[str, Any]:
    return {"type": "FeatureCollection", "features": []}


def load_routes_geojson(path: Path) -> dict[str, Any]:
    """Загрузить кэш треков; при отсутствии файла — пустая FeatureCollection."""
    path = Path(path)
    if not path.is_file():
        return empty_routes_geojson()
    raw = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise ValueError(f"routes.geojson должен быть объектом JSON: {path}")
    features = raw.get("features")
    if features is None:
        features = []
    if not isinstance(features, list):
        raise ValueError(f"Поле features должно быть массивом: {path}")
    return {"type": "FeatureCollection", "features": features}


def load_stations(path: Path) -> dict[str, Any] | None:
    """Точки АЗС/моек для слоёв карты; None, если файла нет (слои не рисуем)."""
    path = Path(path)
    if not path.is_file():
        return None
    raw = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict) or not isinstance(raw.get("features"), list):
        raise ValueError(f"stations.geojson: ожидалась FeatureCollection: {path}")
    return {"type": "FeatureCollection", "features": raw["features"]}


def save_routes_geojson(path: Path, collection: dict[str, Any]) -> None:
    """Записать FeatureCollection (атомарно через .tmp, компактный JSON)."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "type": "FeatureCollection",
        "features": collection.get("features") or [],
    }
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n",
        encoding="utf-8",
    )
    tmp.replace(path)


def coords_for_norm(
    addresses: dict[str, Any], norm: str
) -> tuple[float, float] | None:
    """(lat, lon) из кэша адресов или None."""
    if not norm:
        return None
    entry = addresses.get(norm)
    if not isinstance(entry, dict):
        return None
    lat = _coord_as_float(entry.get("lat"))
    lon = _coord_as_float(entry.get("lon"))
    if lat is None or lon is None:
        return None
    return lat, lon


def _feature_identity(props: dict[str, Any]) -> tuple[str, str, str]:
    return (
        _as_str(props.get("машина")),
        _as_str(props.get("address_a_norm")),
        _as_str(props.get("address_b_norm")),
    )


def _index_routes_cache(
    features: list[Any],
) -> tuple[
    dict[str, dict[str, Any]],
    dict[tuple[str, str, str], dict[str, Any]],
]:
    """Индексы: cache_key → geometry-feature; (машина, a_norm, b_norm) → feature."""
    by_geom: dict[str, dict[str, Any]] = {}
    by_route: dict[tuple[str, str, str], dict[str, Any]] = {}
    for feat in features:
        if not isinstance(feat, dict):
            continue
        props = feat.get("properties") or {}
        if not isinstance(props, dict):
            continue
        geom = feat.get("geometry")
        key = _as_str(props.get("cache_key"))
        if key and isinstance(geom, dict) and geom.get("type") == "LineString":
            by_geom.setdefault(key, feat)
        ident = _feature_identity(props)
        if any(ident):
            by_route[ident] = feat
    return by_geom, by_route


def make_route_feature(
    route: RouteAb,
    *,
    cache_key: str,
    coordinates: list[list[float]],
    osrm_distance_m: float | None = None,
) -> dict[str, Any]:
    """GeoJSON Feature LineString для одного маршрута A→B."""
    props: dict[str, Any] = {
        "cache_key": cache_key,
        "машина": route.vehicle,
        "адрес_A": route.addr_a,
        "адрес_B": route.addr_b,
        "км": route.km,
        "частота": route.frequency,
        "address_a_norm": route.addr_a_norm,
        "address_b_norm": route.addr_b_norm,
    }
    if osrm_distance_m is not None:
        props["osrm_distance_m"] = osrm_distance_m
    return {
        "type": "Feature",
        "properties": props,
        "geometry": {"type": "LineString", "coordinates": coordinates},
    }


def _perp_dist_m(
    point: tuple[float, float],
    start: tuple[float, float],
    end: tuple[float, float],
) -> float:
    """Расстояние от точки до отрезка start–end в метрах (локальная проекция)."""
    lon0, lat0 = point
    ax, ay = _local_xy_m(start[0], start[1], lon0, lat0)
    bx, by = _local_xy_m(end[0], end[1], lon0, lat0)
    abx, aby = bx - ax, by - ay
    ab2 = abx * abx + aby * aby
    if ab2 <= 0.0:
        return math.hypot(ax, ay)
    t = ((-ax) * abx + (-ay) * aby) / ab2
    t = max(0.0, min(1.0, t))
    return math.hypot(ax + t * abx, ay + t * aby)


def simplify_linestring_coords(
    coordinates: list[Any],
    *,
    tolerance_m: float = SIMPLIFY_TOLERANCE_M,
) -> list[list[float]]:
    """Douglas-Peucker: упростить LineString, всегда сохраняя первую и последнюю точки.

    Цель — обычно <200 точек на маршрут при tolerance ~20–50 м.
    """
    pts: list[list[float]] = []
    for raw in coordinates:
        if not isinstance(raw, (list, tuple)) or len(raw) < 2:
            continue
        try:
            lon, lat = float(raw[0]), float(raw[1])
        except (TypeError, ValueError):
            continue
        pts.append([lon, lat])
    if len(pts) <= 2:
        return pts
    if tolerance_m <= 0:
        return pts

    keep = [False] * len(pts)
    keep[0] = True
    keep[-1] = True
    stack: list[tuple[int, int]] = [(0, len(pts) - 1)]
    while stack:
        i, j = stack.pop()
        if j <= i + 1:
            continue
        start = (pts[i][0], pts[i][1])
        end = (pts[j][0], pts[j][1])
        max_d = -1.0
        max_idx = i
        for k in range(i + 1, j):
            d = _perp_dist_m((pts[k][0], pts[k][1]), start, end)
            if d > max_d:
                max_d = d
                max_idx = k
        if max_d > tolerance_m:
            keep[max_idx] = True
            stack.append((i, max_idx))
            stack.append((max_idx, j))

    simplified = [pts[i] for i, flag in enumerate(keep) if flag]
    # Запасной даунсэмпл, если DP оставил слишком много точек (длинные трассы).
    max_points = 200
    if len(simplified) > max_points:
        step = max(1, (len(simplified) - 1) // (max_points - 1))
        sampled = simplified[::step]
        if sampled[-1] != simplified[-1]:
            sampled.append(simplified[-1])
        simplified = sampled
    return simplified


def osrm_route(
    lon1: float,
    lat1: float,
    lon2: float,
    lat2: float,
    *,
    session: requests.Session | None = None,
    timeout: float = OSRM_TIMEOUT_SEC,
    base_url: str = OSRM_BASE_URL,
) -> tuple[list[list[float]], float] | None:
    """Запрос к публичному OSRM. (coordinates GeoJSON, distance_m) или None."""
    return osrm_route_multi(
        [(lon1, lat1), (lon2, lat2)],
        session=session,
        timeout=timeout,
        base_url=base_url,
    )


def build_routes_geojson(
    routes: list[RouteAb],
    address_cache: dict[str, Any],
    routes_cache_path: Path,
    *,
    offline: bool = False,
    sleep_sec: float = OSRM_RATE_LIMIT_SEC,
    sleep_fn: Callable[[float], None] = time.sleep,
    osrm_fn: (
        Callable[[float, float, float, float], tuple[list[list[float]], float] | None]
        | None
    ) = None,
    session: requests.Session | None = None,
) -> OsrmStats:
    """Построить/обновить кэш треков OSRM для маршрутов с координатами A и B.

    --offline: только кэш, без сети.
    Кэш-ключ геометрии: route_osrm_cache_key(lonA, latA, lonB, latB).
    """
    addresses = address_cache.get("addresses") or {}
    if not isinstance(addresses, dict):
        addresses = {}

    collection = load_routes_geojson(routes_cache_path)
    features: list[Any] = list(collection.get("features") or [])
    dirty = False

    # Упростить устаревшие dense-треки (overview=full) без повторного OSRM.
    for feat in features:
        if not isinstance(feat, dict):
            continue
        geom = feat.get("geometry")
        if not isinstance(geom, dict) or geom.get("type") != "LineString":
            continue
        coords = geom.get("coordinates")
        if not isinstance(coords, list) or len(coords) < 3:
            continue
        simplified = simplify_linestring_coords(coords)
        if len(simplified) < len(coords):
            geom["coordinates"] = simplified
            dirty = True

    by_geom, by_route = _index_routes_cache(features)

    routed = 0
    from_cache = 0
    skipped_missing_coords = 0
    skipped_osrm_error = 0
    skipped_offline = 0
    skip_msgs: list[str] = []
    network_calls = 0

    def _do_osrm(
        lon1: float, lat1: float, lon2: float, lat2: float
    ) -> tuple[list[list[float]], float] | None:
        if osrm_fn is not None:
            result = osrm_fn(lon1, lat1, lon2, lat2)
            if result is None:
                return None
            coords, dist_m = result
            return simplify_linestring_coords(coords), dist_m
        return osrm_route(lon1, lat1, lon2, lat2, session=session)

    for route in routes:
        coords_a = coords_for_norm(addresses, route.addr_a_norm)
        coords_b = coords_for_norm(addresses, route.addr_b_norm)
        if coords_a is None or coords_b is None:
            skipped_missing_coords += 1
            missing_parts: list[str] = []
            if coords_a is None:
                missing_parts.append(f"A={route.addr_a_norm or route.addr_a!r}")
            if coords_b is None:
                missing_parts.append(f"B={route.addr_b_norm or route.addr_b!r}")
            msg = (
                f"skip missing coords: {route.vehicle!r} "
                f"{' '.join(missing_parts)}"
            )
            skip_msgs.append(msg)
            continue

        lat1, lon1 = coords_a
        lat2, lon2 = coords_b
        cache_key = route_osrm_cache_key(lon1, lat1, lon2, lat2)
        ident = (route.vehicle, route.addr_a_norm, route.addr_b_norm)

        existing = by_route.get(ident)
        existing_props = (
            existing.get("properties")
            if isinstance(existing, dict)
            else None
        )
        existing_cache_key = (
            _as_str(existing_props.get("cache_key"))
            if isinstance(existing_props, dict)
            else ""
        )
        if (
            existing is not None
            and isinstance(existing.get("geometry"), dict)
            and existing["geometry"].get("type") == "LineString"
            and existing_cache_key == cache_key
        ):
            from_cache += 1
            continue

        geom_src = by_geom.get(cache_key)
        if geom_src is not None:
            src_geom = geom_src["geometry"]
            src_props = geom_src.get("properties") or {}
            dist = src_props.get("osrm_distance_m")
            dist_m = float(dist) if isinstance(dist, (int, float)) else None
            feat = make_route_feature(
                route,
                cache_key=cache_key,
                coordinates=list(src_geom["coordinates"]),
                osrm_distance_m=dist_m,
            )
            features.append(feat)
            by_route[ident] = feat
            from_cache += 1
            dirty = True
            continue

        if offline:
            skipped_offline += 1
            msg = f"skip offline (no cached track): {route.vehicle!r} {cache_key}"
            skip_msgs.append(msg)
            continue

        if network_calls > 0 and sleep_sec > 0:
            sleep_fn(sleep_sec)
        network_calls += 1

        try:
            result = _do_osrm(lon1, lat1, lon2, lat2)
        except (requests.RequestException, ValueError, KeyError, TypeError) as exc:
            skipped_osrm_error += 1
            msg = f"skip OSRM error: {route.vehicle!r} {cache_key}: {exc}"
            skip_msgs.append(msg)
            continue

        if result is None:
            skipped_osrm_error += 1
            msg = f"skip OSRM empty/no route: {route.vehicle!r} {cache_key}"
            skip_msgs.append(msg)
            continue

        coordinates, dist_m = result
        feat = make_route_feature(
            route,
            cache_key=cache_key,
            coordinates=coordinates,
            osrm_distance_m=dist_m,
        )
        features.append(feat)
        by_geom[cache_key] = feat
        by_route[ident] = feat
        routed += 1
        dirty = True

    if dirty:
        save_routes_geojson(
            routes_cache_path,
            {"type": "FeatureCollection", "features": features},
        )

    return OsrmStats(
        routed=routed,
        from_cache=from_cache,
        skipped_missing_coords=skipped_missing_coords,
        skipped_osrm_error=skipped_osrm_error,
        skipped_offline=skipped_offline,
        features=len(features),
        skips=tuple(skip_msgs),
    )


# ——— треки «маршрут через типовую АЗС» (waypoint OSRM) ———

DEFAULT_ROUTE_STATIONS = (
    PROJECT_ROOT / "ГСМ" / "geo_cache" / "route_stations.json"
)
DEFAULT_VIA_CACHE = PROJECT_ROOT / "ГСМ" / "geo_cache" / "routes_via.geojson"


@dataclass
class ViaStats:
    """Счётчики построения via-треков A→АЗС→B."""

    links: int = 0
    routed: int = 0
    from_cache: int = 0
    skipped_missing_coords: int = 0
    skipped_osrm_error: int = 0
    skipped_offline: int = 0
    features: int = 0
    skips: tuple[str, ...] = ()


def load_route_station_links(path: Path) -> list[dict[str, Any]]:
    """Связи маршрут↔АЗС из route_stations.json (пишет build_gsm_trip_feed)."""
    path = Path(path)
    if not path.is_file():
        return []
    raw = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict) or not isinstance(raw.get("links"), list):
        raise ValueError(f"route_stations.json: ожидался объект с links: {path}")
    return [lk for lk in raw["links"] if isinstance(lk, dict)]


def osrm_route_multi(
    points: list[tuple[float, float]],
    *,
    session: requests.Session | None = None,
    timeout: float = OSRM_TIMEOUT_SEC,
    base_url: str = OSRM_BASE_URL,
) -> tuple[list[list[float]], float] | None:
    """OSRM-трек через произвольное число точек (lon, lat)."""
    if len(points) < 2:
        return None
    http = session or requests.Session()
    url = f"{base_url}/" + ";".join(f"{lon},{lat}" for lon, lat in points)
    params = {"overview": OSRM_OVERVIEW, "geometries": "geojson"}
    headers = {"User-Agent": NOMINATIM_USER_AGENT}
    resp = http.get(url, params=params, headers=headers, timeout=timeout)
    resp.raise_for_status()
    data = resp.json()
    if not isinstance(data, dict) or data.get("code") != "Ok":
        return None
    routes = data.get("routes")
    if not isinstance(routes, list) or not routes:
        return None
    first = routes[0]
    if not isinstance(first, dict):
        return None
    geom = first.get("geometry")
    if not isinstance(geom, dict) or geom.get("type") != "LineString":
        return None
    coords = geom.get("coordinates")
    if not isinstance(coords, list) or len(coords) < 2:
        return None
    distance = first.get("distance")
    dist_m = float(distance) if isinstance(distance, (int, float)) else 0.0
    return simplify_linestring_coords(coords), dist_m


def via_osrm_cache_key(
    lon1: float, lat1: float, slon: float, slat: float, lon2: float, lat2: float
) -> str:
    """Ключ кэша via-трека: A→АЗС→B (координаты, 5 знаков)."""
    return (
        f"{lon1:.5f},{lat1:.5f}|{slon:.5f},{slat:.5f}|{lon2:.5f},{lat2:.5f}"
    )


def build_via_geojson(
    links: list[dict[str, Any]],
    address_cache: dict[str, Any],
    routes_collection: dict[str, Any],
    via_cache_path: Path,
    *,
    offline: bool = False,
    sleep_sec: float = OSRM_RATE_LIMIT_SEC,
    sleep_fn: Callable[[float], None] = time.sleep,
    osrm_multi_fn: (
        Callable[[list[tuple[float, float]]], tuple[list[list[float]], float] | None]
        | None
    ) = None,
    session: requests.Session | None = None,
) -> ViaStats:
    """Треки A→АЗС→B для связей маршрут↔типовая АЗС → routes_via.geojson.

    «Крюк» считается против OSRM-дистанции прямого трека (тот же измеритель).
    Кэш геометрии — по координатам тройки точек; свойства маршрута/станции
    обновляются в кэше при каждом прогоне (геометрия переиспользуется).
    """
    addresses = address_cache.get("addresses") or {}
    if not isinstance(addresses, dict):
        addresses = {}

    _, by_route = _index_routes_cache(
        list(routes_collection.get("features") or [])
    )

    collection = load_routes_geojson(via_cache_path)
    features: list[dict[str, Any]] = [
        f
        for f in (collection.get("features") or [])
        if isinstance(f, dict)
    ]
    by_key: dict[str, dict[str, Any]] = {}
    for feat in features:
        props = feat.get("properties") or {}
        key = _as_str(props.get("cache_key"))
        if key and (feat.get("geometry") or {}).get("type") == "LineString":
            by_key.setdefault(key, feat)

    stats_links = 0
    routed = from_cache = 0
    skipped_missing = skipped_error = skipped_offline = 0
    skip_msgs: list[str] = []
    network_calls = 0
    dirty = False
    seen: set[tuple[str, str, str, str]] = set()
    referenced_keys: set[str] = set()

    def _do_osrm(
        points: list[tuple[float, float]],
    ) -> tuple[list[list[float]], float] | None:
        if osrm_multi_fn is not None:
            result = osrm_multi_fn(points)
            if result is None:
                return None
            coords, dist_m = result
            return simplify_linestring_coords(coords), dist_m
        return osrm_route_multi(points, session=session)

    for link in links:
        vehicle = _as_str(link.get("машина"))
        a_norm = _as_str(link.get("address_a_norm"))
        b_norm = _as_str(link.get("address_b_norm"))
        station_norm = _as_str(link.get("station_norm"))
        dedup_key = (vehicle, a_norm, b_norm, station_norm)
        if dedup_key in seen:
            continue
        seen.add(dedup_key)
        stats_links += 1

        coords_a = coords_for_norm(addresses, a_norm)
        coords_b = coords_for_norm(addresses, b_norm)
        coords_s = coords_for_norm(addresses, station_norm)
        if coords_a is None or coords_b is None or coords_s is None:
            skipped_missing += 1
            missing_parts = []
            if coords_a is None:
                missing_parts.append(f"A={a_norm or link.get('адрес_A')!r}")
            if coords_b is None:
                missing_parts.append(f"B={b_norm or link.get('адрес_B')!r}")
            if coords_s is None:
                missing_parts.append(
                    f"station={station_norm or link.get('станция')!r}"
                )
            skip_msgs.append(
                f"via skip missing coords: {vehicle!r} {' '.join(missing_parts)}"
            )
            continue

        lat1, lon1 = coords_a
        lat2, lon2 = coords_b
        slat, slon = coords_s
        cache_key = via_osrm_cache_key(lon1, lat1, slon, slat, lon2, lat2)
        referenced_keys.add(cache_key)

        direct = by_route.get((vehicle, a_norm, b_norm))
        direct_props = (direct or {}).get("properties") or {}
        direct_m = direct_props.get("osrm_distance_m")
        direct_km = (
            round(float(direct_m) / 1000.0, 2)
            if isinstance(direct_m, (int, float))
            else None
        )

        props: dict[str, Any] = {
            "cache_key": cache_key,
            "машина": vehicle,
            "адрес_A": _as_str(link.get("адрес_A")),
            "адрес_B": _as_str(link.get("адрес_B")),
            "address_a_norm": a_norm,
            "address_b_norm": b_norm,
            "станция": _as_str(link.get("станция")),
            "station_norm": station_norm,
            "station_ll": f"{slat:.4f},{slon:.4f}",
            "заправок": link.get("заправок"),
            "км_напрямую": direct_km,
        }

        existing = by_key.get(cache_key)
        if existing is not None:
            geom = existing.get("geometry") or {}
            coords_exist = geom.get("coordinates")
            if isinstance(coords_exist, list) and len(coords_exist) >= 2:
                merged = {**(existing.get("properties") or {}), **props}
                via_m = merged.get("osrm_distance_m")
                _finalize_via_props(merged, via_m)
                if merged != existing.get("properties"):
                    existing["properties"] = merged
                    dirty = True
                from_cache += 1
                continue

        if offline:
            skipped_offline += 1
            skip_msgs.append(f"via skip offline: {vehicle!r} {cache_key}")
            continue

        if network_calls > 0 and sleep_sec > 0:
            sleep_fn(sleep_sec)
        network_calls += 1
        try:
            result = _do_osrm([(lon1, lat1), (slon, slat), (lon2, lat2)])
        except (requests.RequestException, ValueError, KeyError, TypeError) as exc:
            skipped_error += 1
            skip_msgs.append(f"via OSRM error: {vehicle!r} {cache_key}: {exc}")
            continue
        if result is None:
            skipped_error += 1
            skip_msgs.append(f"via OSRM empty/no route: {vehicle!r} {cache_key}")
            continue

        coordinates, dist_m = result
        feat = {
            "type": "Feature",
            "properties": props,
            "geometry": {"type": "LineString", "coordinates": coordinates},
        }
        _finalize_via_props(feat["properties"], dist_m)
        features.append(feat)
        by_key[cache_key] = feat
        routed += 1
        dirty = True

    # Убрать фичи устаревших геометрий (координаты станции/точек изменились
    # или связь удалена): их ключи больше не запрошены текущими связями.
    if links:
        pruned = [
            f
            for f in features
            if (f.get("properties") or {}).get("cache_key") in referenced_keys
        ]
        if len(pruned) != len(features):
            features = pruned
            dirty = True

    if dirty:
        save_routes_geojson(
            via_cache_path,
            {"type": "FeatureCollection", "features": features},
        )

    return ViaStats(
        links=stats_links,
        routed=routed,
        from_cache=from_cache,
        skipped_missing_coords=skipped_missing,
        skipped_osrm_error=skipped_error,
        skipped_offline=skipped_offline,
        features=len(features),
        skips=tuple(skip_msgs),
    )


def _finalize_via_props(props: dict[str, Any], via_m: Any) -> None:
    """км_с_заездом/крюк_км из OSRM-дистанции via-трека."""
    if not isinstance(via_m, (int, float)):
        return
    props["osrm_distance_m"] = via_m
    via_km = round(float(via_m) / 1000.0, 2)
    props["км_с_заездом"] = via_km
    direct_km = props.get("км_напрямую")
    if isinstance(direct_km, (int, float)):
        props["крюк_км"] = round(via_km - float(direct_km), 2)


def recommend_via_threshold_km(detours: list[float]) -> float:
    """Порог «заправка по пути»: минимум из 2/3/5/8 км, покрывающий ≥85% крюков.

    Смотрим только на правдоподобные значения (≤ 20 км); большие крюки —
    аномалии геокода или реальные заезды, их порог определять не должен.
    """
    plausible = sorted(d for d in detours if 0.0 <= d <= 20.0)
    if not plausible:
        return 5.0
    for threshold in (2.0, 3.0, 5.0, 8.0):
        covered = sum(1 for d in plausible if d <= threshold) / len(plausible)
        if covered >= 0.85:
            return threshold
    return 13.0


def via_detour_histogram(
    features: list[dict[str, Any]],
) -> list[tuple[str, int]]:
    """Распределение крюков по корзинам для отчёта калибровки."""
    bins = (1.0, 2.0, 3.0, 5.0, 8.0, 13.0, 21.0, 34.0, 55.0)
    counts = [0] * (len(bins) + 1)
    for feat in features:
        detour = (feat.get("properties") or {}).get("крюк_км")
        if not isinstance(detour, (int, float)):
            continue
        idx = 0
        while idx < len(bins) and detour > bins[idx]:
            idx += 1
        counts[idx] += 1
    labels = [f"≤{bins[0]:g}"]
    labels += [f"{bins[i]:g}–{bins[i + 1]:g}" for i in range(len(bins) - 1)]
    labels.append(f">{bins[-1]:g}")
    return list(zip(labels, counts))


def apply_via_threshold(
    features: list[dict[str, Any]], threshold_km: float
) -> None:
    """Флаг «крюк_по_пути» в properties по порогу (в HTML, не в кэш)."""
    for feat in features:
        props = feat.get("properties")
        if not isinstance(props, dict):
            props = {}
            feat["properties"] = props
        detour = props.get("крюк_км")
        if isinstance(detour, (int, float)):
            props["крюк_по_пути"] = bool(detour <= threshold_km)
        else:
            props["крюк_по_пути"] = None


def count_geocoded_addresses(address_cache: dict[str, Any]) -> int:
    """Сколько адресов в кэше имеют валидные координаты."""
    addresses = address_cache.get("addresses") or {}
    if not isinstance(addresses, dict):
        return 0
    return sum(1 for e in addresses.values() if cache_entry_has_coords(e))


def known_addresses_from_cache(address_cache: dict[str, Any]) -> list[dict[str, Any]]:
    """Адрес→координаты для встраивания в HTML (офлайн-поиск известных точек)."""
    addresses = address_cache.get("addresses") or {}
    if not isinstance(addresses, dict):
        return []
    out: list[dict[str, Any]] = []
    for norm, entry in addresses.items():
        if not isinstance(entry, dict) or not cache_entry_has_coords(entry):
            continue
        lat = _coord_as_float(entry.get("lat"))
        lon = _coord_as_float(entry.get("lon"))
        if lat is None or lon is None:
            continue
        display = _as_str(entry.get("display")) or _as_str(norm)
        out.append(
            {
                "norm": _as_str(norm),
                "display": display,
                "lat": lat,
                "lon": lon,
            }
        )
    out.sort(key=lambda a: a["display"].lower())
    return out


def haversine_km(
    lon1: float, lat1: float, lon2: float, lat2: float
) -> float:
    """Большой круг между двумя точками (lon, lat), км."""
    rlat1, rlat2 = math.radians(lat1), math.radians(lat2)
    dlat = rlat2 - rlat1
    dlon = math.radians(lon2 - lon1)
    a = (
        math.sin(dlat / 2) ** 2
        + math.cos(rlat1) * math.cos(rlat2) * math.sin(dlon / 2) ** 2
    )
    return 2 * EARTH_RADIUS_KM * math.asin(min(1.0, math.sqrt(a)))


def _local_xy_m(
    lon: float, lat: float, lon0: float, lat0: float
) -> tuple[float, float]:
    """Эквидистантная проекция в метры относительно (lon0, lat0)."""
    x = math.radians(lon - lon0) * math.cos(math.radians(lat0)) * EARTH_RADIUS_M
    y = math.radians(lat - lat0) * EARTH_RADIUS_M
    return x, y


def point_to_segment_km(
    point_lonlat: tuple[float, float],
    a_lonlat: tuple[float, float],
    b_lonlat: tuple[float, float],
) -> float:
    """Кратчайшее расстояние от точки до отрезка A–B, км (локальная проекция)."""
    lon0, lat0 = point_lonlat
    ax, ay = _local_xy_m(a_lonlat[0], a_lonlat[1], lon0, lat0)
    bx, by = _local_xy_m(b_lonlat[0], b_lonlat[1], lon0, lat0)
    abx, aby = bx - ax, by - ay
    ab2 = abx * abx + aby * aby
    if ab2 <= 0.0:
        dist_m = math.hypot(ax, ay)
    else:
        # P в начале координат; вектор A→P = (-ax, -ay)
        t = ((-ax) * abx + (-ay) * aby) / ab2
        t = max(0.0, min(1.0, t))
        cx = ax + t * abx
        cy = ay + t * aby
        dist_m = math.hypot(cx, cy)
    return dist_m / 1000.0


def distance_point_to_linestring_km(
    point_lonlat: tuple[float, float],
    coordinates: list[Any],
) -> float | None:
    """Мин. расстояние точки до GeoJSON LineString (координаты [lon, lat])."""
    if not isinstance(coordinates, list) or len(coordinates) < 2:
        return None
    best: float | None = None
    prev: tuple[float, float] | None = None
    for raw in coordinates:
        if not isinstance(raw, (list, tuple)) or len(raw) < 2:
            prev = None
            continue
        try:
            lon, lat = float(raw[0]), float(raw[1])
        except (TypeError, ValueError):
            prev = None
            continue
        cur = (lon, lat)
        if prev is not None:
            d = point_to_segment_km(point_lonlat, prev, cur)
            if best is None or d < best:
                best = d
        prev = cur
    return best


def nearest_routes(
    point_lonlat: tuple[float, float],
    features: list[Any],
    *,
    k: int = 3,
) -> list[dict[str, Any]]:
    """Топ-k маршрутов по расстоянию от точки до линии (км), без порога.

    Каждый элемент: машина, адрес_A, адрес_B, distance_km + прочие свойства фичи.
    """
    if k <= 0:
        return []
    scored: list[tuple[float, dict[str, Any]]] = []
    for feat in features:
        if not isinstance(feat, dict):
            continue
        geom = feat.get("geometry")
        if not isinstance(geom, dict) or geom.get("type") != "LineString":
            continue
        coords = geom.get("coordinates")
        if not isinstance(coords, list):
            continue
        dist = distance_point_to_linestring_km(point_lonlat, coords)
        if dist is None:
            continue
        props = feat.get("properties") or {}
        if not isinstance(props, dict):
            props = {}
        row: dict[str, Any] = {
            "машина": props.get("машина", ""),
            "адрес_A": props.get("адрес_A", ""),
            "адрес_B": props.get("адрес_B", ""),
            "distance_km": round(dist, 4),
        }
        for key in ("км", "частота", "cache_key", "osrm_distance_m"):
            if key in props:
                row[key] = props[key]
        scored.append((dist, row))
    scored.sort(key=lambda item: item[0])
    return [row for _, row in scored[:k]]


def _json_for_script(value: Any) -> str:
    """JSON, безопасный для вставки в <script> (экранирует </)."""
    return json.dumps(value, ensure_ascii=False).replace("</", "<\\/")


def _vehicles_from_geojson(routes_geojson: dict[str, Any]) -> list[str]:
    """Уникальные машины из features, порядок: палитра, затем остальные."""
    seen: set[str] = set()
    for feat in routes_geojson.get("features") or []:
        if not isinstance(feat, dict):
            continue
        props = feat.get("properties") or {}
        if not isinstance(props, dict):
            continue
        name = _as_str(props.get("машина"))
        if name:
            seen.add(name)
    ordered = [v for v in VEHICLE_COLORS if v in seen]
    ordered.extend(sorted(seen - set(VEHICLE_COLORS)))
    return ordered


def write_map_html(
    out_path: Path,
    routes_geojson: dict[str, Any],
    vehicle_colors: dict[str, str] | None = None,
    known_addresses: list[dict[str, Any]] | None = None,
    *,
    embed_max_bytes: int = EMBED_GEOJSON_MAX_BYTES,
    stations: dict[str, Any] | None = None,
    via: dict[str, Any] | None = None,
    via_threshold_km: float | None = None,
) -> None:
    """Собрать Leaflet HTML: встроенный GeoJSON или sibling routes.geojson + fetch.

    Пустая FeatureCollection допустима — карта открывается с запасным центром.
    known_addresses — список {{norm, display, lat, lon}} для офлайн-поиска.
    Если сериализованный GeoJSON > embed_max_bytes — пишем sibling routes.geojson
    рядом с HTML и подгружаем через fetch (нужен http.server, не file://).
    stations — FeatureCollection точек АЗС/моек (опционально): слои с чекбоксами.
    via — FeatureCollection треков A→АЗС→B с крюк_км (опционально).
    """
    colors = dict(vehicle_colors) if vehicle_colors is not None else dict(VEHICLE_COLORS)
    collection = {
        "type": "FeatureCollection",
        "features": list(routes_geojson.get("features") or []),
    }
    feature_count = len(collection["features"])
    vehicles = _vehicles_from_geojson(collection)
    for name in vehicles:
        colors.setdefault(name, DEFAULT_VEHICLE_COLOR)

    addresses = list(known_addresses) if known_addresses is not None else []
    out = Path(out_path)
    geojson_bytes = json.dumps(
        collection, ensure_ascii=False, separators=(",", ":")
    ).encode("utf-8")
    embed_inline = len(geojson_bytes) <= embed_max_bytes
    sibling_name = "routes.geojson"
    if embed_inline:
        geojson_js = _json_for_script(collection)
        routes_url_js = _json_for_script(None)
    else:
        sibling = out.with_name(sibling_name)
        save_routes_geojson(sibling, collection)
        geojson_js = "null"
        routes_url_js = _json_for_script(sibling_name)
    colors_js = _json_for_script(colors)
    vehicles_js = _json_for_script(vehicles)
    addresses_js = _json_for_script(addresses)
    photon_js = _json_for_script(PHOTON_API_URL)
    stations_fc = (
        {"type": "FeatureCollection", "features": list(stations.get("features") or [])}
        if stations
        else {"type": "FeatureCollection", "features": []}
    )
    stations_js = _json_for_script(stations_fc)
    azs_count = sum(
        1
        for f in stations_fc["features"]
        if (f.get("properties") or {}).get("тип") != "мойка"
    )
    wash_count = len(stations_fc["features"]) - azs_count
    via_fc = (
        {"type": "FeatureCollection", "features": list(via.get("features") or [])}
        if via
        else {"type": "FeatureCollection", "features": []}
    )
    via_js = _json_for_script(via_fc)
    via_count = len(via_fc["features"])
    via_threshold_js = _json_for_script(via_threshold_km)
    via_threshold_label = (
        f"{via_threshold_km:g}" if via_threshold_km is not None else "—"
    )
    center_lat, center_lon = DEFAULT_MAP_CENTER
    addr_count = len(addresses)
    load_hint = (
        ""
        if embed_inline
        else (
            f"Маршруты: {sibling_name} (fetch). "
            "Если карта пуста — из папки ГСМ/: python -m http.server"
        )
    )

    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>Карта маршрутов ГСМ</title>
  <link rel="stylesheet" href="{LEAFLET_CSS_CDN}"/>
  <script src="{LEAFLET_JS_CDN}"></script>
  <style>
    html, body {{ height: 100%; margin: 0; }}
    #map {{ height: 100%; width: 100%; }}
    .panel {{
      position: absolute; z-index: 1000; top: 10px; right: 10px;
      background: #fff; padding: 10px 12px; border-radius: 4px;
      box-shadow: 0 1px 5px rgba(0,0,0,0.35); font: 13px/1.35 sans-serif;
      max-width: 300px; max-height: calc(100% - 24px); overflow: auto;
    }}
    .panel h1 {{ font-size: 14px; margin: 0 0 6px; }}
    .panel .meta {{ color: #555; margin: 0 0 8px; }}
    .panel label {{ display: flex; align-items: center; gap: 6px; margin: 3px 0; cursor: pointer; }}
    .swatch {{ width: 12px; height: 12px; border-radius: 2px; flex-shrink: 0; border: 1px solid #333; }}
    .search-box {{
      position: absolute; z-index: 1000; top: 10px; left: 50px;
      background: #fff; padding: 8px 10px; border-radius: 4px;
      box-shadow: 0 1px 5px rgba(0,0,0,0.35); font: 13px sans-serif;
      width: min(360px, calc(100vw - 80px));
    }}
    .search-box .row {{ display: flex; gap: 6px; align-items: center; }}
    .search-box input[type="text"] {{ flex: 1; min-width: 0; padding: 5px 6px; }}
    .search-box button {{ padding: 5px 10px; cursor: pointer; }}
    .search-box .hint {{ color: #666; font-size: 11px; margin-top: 5px; }}
    .search-box .status {{ font-size: 12px; margin-top: 6px; color: #333; }}
    .search-box .status.err {{ color: #a33; }}
    .nearest-list {{ margin: 8px 0 0; padding: 0; list-style: none; }}
    .nearest-list li {{
      margin: 0 0 6px; padding: 6px 8px; background: #f6f6f6;
      border-left: 3px solid #999; font-size: 12px; line-height: 1.35;
    }}
    .nearest-list li.top {{ border-left-color: #c0392b; background: #fff5f5; }}
    .nearest-list .dist {{ font-weight: 600; }}
  </style>
</head>
<body>
  <div id="map"></div>
  <div class="search-box" id="search-box">
    <form id="search-form" autocomplete="off">
      <div class="row">
        <input type="text" id="search-input" list="known-addresses"
               placeholder="Адрес (заправка, склад…)" />
        <button type="submit" id="search-btn">Найти</button>
      </div>
      <datalist id="known-addresses"></datalist>
    </form>
    <div class="hint">
      Известных адресов: {addr_count}. Выберите из списка (офлайн)
      или свободный текст через Photon (нужен интернет / не file://).
    </div>
    <div class="status" id="search-status"></div>
    <ol class="nearest-list" id="nearest-list"></ol>
  </div>
  <div class="panel" id="layers-panel">
    <h1>Маршруты ГСМ</h1>
    <p class="meta">маршрутов: <span id="feature-count">{feature_count}</span></p>
    <p class="meta" id="load-hint">{load_hint}</p>
    <label><input type="checkbox" id="chk-all" checked/> Все</label>
    <div id="vehicle-checks"></div>
    <hr style="border: none; border-top: 1px solid #ddd; margin: 8px 0;"/>
    <label>
      <input type="checkbox" id="chk-azs" checked/>
      <span class="swatch" style="border-radius: 50%; background: #0079c2;"></span>
      АЗС ({azs_count})
    </label>
    <label>
      <input type="checkbox" id="chk-wash" checked/>
      <span class="swatch" style="border-radius: 50%; background: #8e44ad;"></span>
      Мойки ({wash_count})
    </label>
    <label title="Треки A→АЗС→B. Зелёный — крюк ≤ порога, красный — больше, серый — нет данных">
      <input type="checkbox" id="chk-via" checked/>
      <span class="swatch" style="background: repeating-linear-gradient(90deg,#27ae60 0 4px,#fff 4px 6px);"></span>
      Через АЗС ({via_count}), порог {via_threshold_label} км
    </label>
  </div>
  <script>
    let ROUTES_GEOJSON = {geojson_js};
    const ROUTES_GEOJSON_URL = {routes_url_js};
    const VEHICLE_COLORS = {colors_js};
    const VEHICLES = {vehicles_js};
    const KNOWN_ADDRESSES = {addresses_js};
    const PHOTON_API = {photon_js};
    const STATIONS = {stations_js};
    const VIA = {via_js};
    const VIA_THRESHOLD_KM = {via_threshold_js};
    const BRAND_COLORS = {{
      "Газпромнефть": "#0079c2", "ГПН": "#0079c2",
      "TATNEFT": "#d9261c", "Татнефть": "#d9261c",
      "Роснефть": "#f5a800", "ЛУКОЙЛ": "#c8102e",
      "КТК": "#27ae60", "ТНК": "#16a085", "Shell": "#fbce07"
    }};
    const STATION_DEFAULT_COLOR = "#e67e22";
    const WASH_COLOR = "#8e44ad";
    let FEATURE_COUNT = {feature_count};
    const DEFAULT_CENTER = [{center_lat}, {center_lon}];
    const DEFAULT_ZOOM = {DEFAULT_MAP_ZOOM};
    const EARTH_R_M = 6371008.8;

    const map = L.map("map").setView(DEFAULT_CENTER, DEFAULT_ZOOM);
    L.tileLayer("https://tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png", {{
      maxZoom: 19,
      attribution: "&copy; OpenStreetMap"
    }}).addTo(map);

    function colorFor(vehicle) {{
      return VEHICLE_COLORS[vehicle] || "{DEFAULT_VEHICLE_COLOR}";
    }}

    function escapeHtml(s) {{
      return String(s ?? "")
        .replace(/&/g, "&amp;")
        .replace(/</g, "&lt;")
        .replace(/>/g, "&gt;")
        .replace(/"/g, "&quot;")
        .replace(/'/g, "&#39;");
    }}

    function popupHtml(props) {{
      const p = props || {{}};
      const km = p["км"] == null ? "—" : p["км"];
      const freq = p["частота"] == null ? "—" : p["частота"];
      return (
        "<b>" + escapeHtml(p["машина"] || "—") + "</b><br/>" +
        "A: " + escapeHtml(p["адрес_A"] || "—") + "<br/>" +
        "B: " + escapeHtml(p["адрес_B"] || "—") + "<br/>" +
        "км: " + escapeHtml(km) + "<br/>" +
        "частота: " + escapeHtml(freq)
      );
    }}

    const layersByVehicle = {{}};
    VEHICLES.forEach(function (v) {{
      layersByVehicle[v] = L.layerGroup().addTo(map);
    }});

    const boundsLayer = L.featureGroup();
    const routeLayers = [];
    const checksRoot = document.getElementById("vehicle-checks");
    const chkAll = document.getElementById("chk-all");
    const vehicleChecks = {{}};
    const loadHintEl = document.getElementById("load-hint");

    function baseStyle(feature) {{
      const v = (feature.properties && feature.properties["машина"]) || "";
      return {{ color: colorFor(v), weight: 4, opacity: 0.85 }};
    }}

    function syncAllCheckbox() {{
      const names = Object.keys(vehicleChecks);
      if (!names.length) {{
        chkAll.checked = true;
        return;
      }}
      chkAll.checked = names.every(function (v) {{ return vehicleChecks[v].checked; }});
    }}

    function setVehicleVisible(v, on) {{
      const group = layersByVehicle[v];
      if (!group) return;
      if (on) {{
        if (!map.hasLayer(group)) map.addLayer(group);
      }} else if (map.hasLayer(group)) {{
        map.removeLayer(group);
      }}
    }}

    function buildVehicleChecks() {{
      checksRoot.innerHTML = "";
      VEHICLES.forEach(function (v) {{
        const label = document.createElement("label");
        const input = document.createElement("input");
        input.type = "checkbox";
        input.checked = true;
        const swatch = document.createElement("span");
        swatch.className = "swatch";
        swatch.style.background = colorFor(v);
        label.appendChild(input);
        label.appendChild(swatch);
        label.appendChild(document.createTextNode(v));
        checksRoot.appendChild(label);
        vehicleChecks[v] = input;
        input.addEventListener("change", function () {{
          setVehicleVisible(v, input.checked);
          syncAllCheckbox();
        }});
      }});
    }}

    chkAll.addEventListener("change", function () {{
      const on = chkAll.checked;
      Object.keys(vehicleChecks).forEach(function (v) {{
        vehicleChecks[v].checked = on;
        setVehicleVisible(v, on);
      }});
    }});

    function mountRoutes(collection) {{
      ROUTES_GEOJSON = collection || {{ type: "FeatureCollection", features: [] }};
      FEATURE_COUNT = (ROUTES_GEOJSON.features || []).length;
      const countEl = document.getElementById("feature-count");
      if (countEl) countEl.textContent = String(FEATURE_COUNT);
      L.geoJSON(ROUTES_GEOJSON, {{
        style: baseStyle,
        onEachFeature: function (feature, layer) {{
          const props = feature.properties || {{}};
          layer.bindPopup(popupHtml(props));
          layer._gsmFeature = feature;
          layer._gsmBaseStyle = baseStyle(feature);
          routeLayers.push(layer);
          const v = props["машина"] || "";
          if (layersByVehicle[v]) {{
            layersByVehicle[v].addLayer(layer);
          }} else {{
            layer.addTo(map);
          }}
          boundsLayer.addLayer(layer);
        }}
      }});
      if (FEATURE_COUNT > 0) {{
        try {{
          const bounds = boundsLayer.getBounds();
          if (bounds.isValid()) {{
            map.fitBounds(bounds, {{ padding: [30, 30] }});
          }}
        }} catch (e) {{ /* empty / invalid */ }}
      }}
      buildVehicleChecks();
    }}

    function showRoutesLoadError(err) {{
      const msg =
        "Не удалось загрузить routes.geojson. " +
        "Запустите из папки ГСМ/: python -m http.server";
      if (loadHintEl) loadHintEl.textContent = msg;
      if (typeof setStatus === "function") setStatus(msg, true);
      console.warn(err);
      mountRoutes({{ type: "FeatureCollection", features: [] }});
    }}

    if (ROUTES_GEOJSON && ROUTES_GEOJSON.features) {{
      mountRoutes(ROUTES_GEOJSON);
    }} else if (ROUTES_GEOJSON_URL) {{
      if (loadHintEl && !loadHintEl.textContent) {{
        loadHintEl.textContent = "Загрузка " + ROUTES_GEOJSON_URL + "…";
      }}
      fetch(ROUTES_GEOJSON_URL)
        .then(function (resp) {{
          if (!resp.ok) throw new Error("HTTP " + resp.status);
          return resp.json();
        }})
        .then(function (data) {{
          mountRoutes(data);
          if (loadHintEl) {{
            loadHintEl.textContent =
              "Маршруты из " + ROUTES_GEOJSON_URL +
              ". При file:// откройте через: python -m http.server";
          }}
        }})
        .catch(showRoutesLoadError);
    }} else {{
      mountRoutes({{ type: "FeatureCollection", features: [] }});
    }}

    /* —— точки АЗС и моек —— */
    const azsGroup = L.layerGroup();
    const washGroup = L.layerGroup();
    const stationMarkers = [];
    const viaLayers = [];
    let activeStationKey = null;

    function stationColor(props) {{
      if (props["тип"] === "мойка") return WASH_COLOR;
      return BRAND_COLORS[props["бренд"]] || STATION_DEFAULT_COLOR;
    }}

    function stationRadius(props) {{
      const liters = Number(props["литров"]) || 0;
      if (props["тип"] === "мойка") return 6;
      return Math.max(5, Math.min(14, 5 + Math.sqrt(liters) / 12));
    }}

    function stationPopup(props) {{
      const p = props || {{}};
      const kind =
        p["тип"] === "мойка" ? "Мойка" :
        (p["тип"] === "азс+мойка" ? "АЗС + мойка" : "АЗС");
      let html =
        "<b>" + kind +
        (p["бренд"] ? " · " + escapeHtml(p["бренд"]) : "") + "</b><br/>" +
        escapeHtml(p["адрес"] || "—") + "<br/>";
      if (p["тип"] !== "мойка") {{
        html +=
          "заправок: " + escapeHtml(p["заправок"]) +
          " · литров: " + escapeHtml(p["литров"]) +
          " · сумма: " + escapeHtml(p["сумма_руб"]) + " ₽<br/>";
      }} else {{
        html +=
          "моек: " + escapeHtml(p["моек"]) +
          " · сумма: " + escapeHtml(p["сумма_руб"]) + " ₽<br/>";
      }}
      if (p["машины"]) html += "машины: " + escapeHtml(p["машины"]) + "<br/>";
      html +=
        "период: " + escapeHtml(p["первая_дата"]) +
        " → " + escapeHtml(p["последняя_дата"]);
      if (Number(p["вне_поездок"]) > 0) {{
        html += "<br/><span style='color:#a33'>вне поездок: " +
          escapeHtml(p["вне_поездок"]) + "</span>";
      }}
      return html;
    }}

    (STATIONS.features || []).forEach(function (feat) {{
      const geom = feat && feat.geometry;
      if (!geom || geom.type !== "Point") return;
      const p = feat.properties || {{}};
      const allUnmatched =
        Number(p["вне_поездок"]) > 0 &&
        Number(p["вне_поездок"]) === Number(p["транзакций"]);
      const marker = L.circleMarker(
        [geom.coordinates[1], geom.coordinates[0]],
        {{
          radius: stationRadius(p),
          color: "#222",
          weight: 1,
          fillColor: stationColor(p),
          fillOpacity: allUnmatched ? 0.35 : 0.9
        }}
      );
      marker.bindPopup(stationPopup(p));
      marker._stationKey = p["коорд"] || null;
      marker.on("click", function () {{ highlightStationRoutes(marker._stationKey); }});
      stationMarkers.push(marker);
      (p["тип"] === "мойка" ? washGroup : azsGroup).addLayer(marker);
    }});
    azsGroup.addTo(map);
    washGroup.addTo(map);

    /* —— via-треки A→АЗС→B —— */
    const viaGroup = L.layerGroup();

    function viaColor(props) {{
      if (props["крюк_по_пути"] === true) return "#27ae60";
      if (props["крюк_по_пути"] === false) return "#c0392b";
      return "#7f8c8d";
    }}

    function viaPopup(props) {{
      const p = props || {{}};
      const detour = (typeof p["крюк_км"] === "number") ? p["крюк_км"] : null;
      let html =
        "<b>" + escapeHtml(p["машина"] || "—") + "</b> · через АЗС<br/>" +
        "АЗС: " + escapeHtml(p["станция"] || "—") + "<br/>" +
        "заправок: " + escapeHtml(p["заправок"] ?? "—") + "<br/>";
      if (typeof p["км_напрямую"] === "number") {{
        html += "напрямую: " + escapeHtml(p["км_напрямую"]) + " км";
      }}
      if (typeof p["км_с_заездом"] === "number") {{
        html += " · с заездом: " + escapeHtml(p["км_с_заездом"]) + " км";
      }}
      if (detour !== null) {{
        const verdict = p["крюк_по_пути"] === true
          ? "по пути"
          : (p["крюк_по_пути"] === false ? "КРЮК" : "");
        html += "<br/>крюк: <b>" + escapeHtml(detour) + " км</b>";
        if (verdict) {{
          html += " <span style='color:" +
            (p["крюк_по_пути"] === false ? "#c0392b" : "#27ae60") +
            "'>" + verdict + "</span>";
        }}
      }}
      return html;
    }}

    (VIA.features || []).forEach(function (feat) {{
      const geom = feat && feat.geometry;
      if (!geom || geom.type !== "LineString") return;
      const p = feat.properties || {{}};
      const latlngs = (geom.coordinates || []).map(function (c) {{
        return [c[1], c[0]];
      }});
      if (latlngs.length < 2) return;
      const line = L.polyline(latlngs, {{
        color: viaColor(p),
        weight: 3,
        opacity: 0.75,
        dashArray: "7 6"
      }});
      line.bindPopup(viaPopup(p));
      line._stationKey = p["station_ll"] || null;
      line._viaBaseStyle = {{ color: viaColor(p), weight: 3, opacity: 0.75 }};
      viaLayers.push(line);
      viaGroup.addLayer(line);
    }});
    viaGroup.addTo(map);

    function highlightStationRoutes(key) {{
      if (!key || activeStationKey === key) {{
        activeStationKey = null;
        viaLayers.forEach(function (l) {{ l.setStyle(l._viaBaseStyle); }});
        return;
      }}
      activeStationKey = key;
      viaLayers.forEach(function (l) {{
        if (l._stationKey === key) {{
          l.setStyle({{ weight: 6, opacity: 1 }});
          if (l.bringToFront) l.bringToFront();
        }} else {{
          l.setStyle({{ weight: 2, opacity: 0.15 }});
        }}
      }});
    }}
    map.on("click", function () {{
      if (activeStationKey) highlightStationRoutes(null);
    }});

    function bindStationToggle(id, group) {{
      const chk = document.getElementById(id);
      if (!chk) return;
      chk.addEventListener("change", function () {{
        if (chk.checked) {{
          if (!map.hasLayer(group)) map.addLayer(group);
        }} else if (map.hasLayer(group)) {{
          map.removeLayer(group);
        }}
      }});
    }}
    bindStationToggle("chk-azs", azsGroup);
    bindStationToggle("chk-wash", washGroup);
    bindStationToggle("chk-via", viaGroup);

    /* —— поиск адреса → топ-3 ближайших маршрута —— */
    const datalist = document.getElementById("known-addresses");
    KNOWN_ADDRESSES.forEach(function (a) {{
      const opt = document.createElement("option");
      opt.value = a.display || a.norm || "";
      if (a.norm && a.display && a.norm !== a.display) {{
        opt.label = a.norm;
      }}
      datalist.appendChild(opt);
    }});

    const searchInput = document.getElementById("search-input");
    const searchStatus = document.getElementById("search-status");
    const nearestList = document.getElementById("nearest-list");
    let searchMarker = null;
    let highlightedLayer = null;

    function setStatus(text, isErr) {{
      searchStatus.textContent = text || "";
      searchStatus.className = "status" + (isErr ? " err" : "");
    }}

    function localXy(lon, lat, lon0, lat0) {{
      const x = (lon - lon0) * Math.PI / 180 * Math.cos(lat0 * Math.PI / 180) * EARTH_R_M;
      const y = (lat - lat0) * Math.PI / 180 * EARTH_R_M;
      return [x, y];
    }}

    function pointToSegmentKm(plon, plat, aLon, aLat, bLon, bLat) {{
      const a = localXy(aLon, aLat, plon, plat);
      const b = localXy(bLon, bLat, plon, plat);
      const abx = b[0] - a[0], aby = b[1] - a[1];
      const ab2 = abx * abx + aby * aby;
      let distM;
      if (ab2 <= 0) {{
        distM = Math.hypot(a[0], a[1]);
      }} else {{
        let t = ((-a[0]) * abx + (-a[1]) * aby) / ab2;
        t = Math.max(0, Math.min(1, t));
        const cx = a[0] + t * abx;
        const cy = a[1] + t * aby;
        distM = Math.hypot(cx, cy);
      }}
      return distM / 1000;
    }}

    function distanceToLineStringKm(plon, plat, coordinates) {{
      if (!coordinates || coordinates.length < 2) return null;
      let best = null;
      for (let i = 1; i < coordinates.length; i++) {{
        const a = coordinates[i - 1], b = coordinates[i];
        if (!a || !b || a.length < 2 || b.length < 2) continue;
        const d = pointToSegmentKm(plon, plat, a[0], a[1], b[0], b[1]);
        if (best === null || d < best) best = d;
      }}
      return best;
    }}

    function nearestRoutes(plon, plat, k) {{
      k = k || 3;
      const scored = [];
      (ROUTES_GEOJSON.features || []).forEach(function (feat, idx) {{
        const geom = feat && feat.geometry;
        if (!geom || geom.type !== "LineString") return;
        const dist = distanceToLineStringKm(plon, plat, geom.coordinates);
        if (dist === null) return;
        const p = feat.properties || {{}};
        scored.push({{
          idx: idx,
          машина: p["машина"] || "",
          адрес_A: p["адрес_A"] || "",
          адрес_B: p["адрес_B"] || "",
          км: p["км"],
          частота: p["частота"],
          distance_km: Math.round(dist * 10000) / 10000
        }});
      }});
      scored.sort(function (a, b) {{ return a.distance_km - b.distance_km; }});
      return scored.slice(0, k);
    }}

    function clearHighlight() {{
      if (highlightedLayer) {{
        highlightedLayer.setStyle(highlightedLayer._gsmBaseStyle || {{ weight: 4 }});
        highlightedLayer = null;
      }}
      routeLayers.forEach(function (layer) {{
        if (layer.setStyle && layer._gsmBaseStyle) {{
          layer.setStyle(layer._gsmBaseStyle);
        }}
      }});
    }}

    function highlightTop(result) {{
      clearHighlight();
      if (!result) return;
      const layer = routeLayers[result.idx];
      if (!layer || !layer.setStyle) return;
      layer.setStyle({{
        color: "#c0392b",
        weight: 8,
        opacity: 1
      }});
      if (layer.bringToFront) layer.bringToFront();
      highlightedLayer = layer;
      if (layer.openPopup) layer.openPopup();
    }}

    function renderNearest(results) {{
      nearestList.innerHTML = "";
      results.forEach(function (r, i) {{
        const li = document.createElement("li");
        if (i === 0) li.className = "top";
        const dist = (typeof r.distance_km === "number")
          ? r.distance_km.toFixed(2) : "—";
        li.innerHTML =
          "<div class=\\"dist\\">#" + (i + 1) + " · " + dist + " км до линии</div>" +
          "<div><b>" + escapeHtml(r["машина"] || "—") + "</b></div>" +
          "<div>A: " + escapeHtml(r["адрес_A"] || "—") + "</div>" +
          "<div>B: " + escapeHtml(r["адрес_B"] || "—") + "</div>";
        nearestList.appendChild(li);
      }});
    }}

    function showPoint(lat, lon, label) {{
      if (searchMarker) {{
        map.removeLayer(searchMarker);
        searchMarker = null;
      }}
      searchMarker = L.marker([lat, lon]).addTo(map);
      if (label) searchMarker.bindPopup(escapeHtml(label)).openPopup();
      map.setView([lat, lon], Math.max(map.getZoom(), 11));
      const top = nearestRoutes(lon, lat, 3);
      renderNearest(top);
      highlightTop(top[0] || null);
      if (!top.length) {{
        setStatus("Маршруты не найдены (пустая карта).", true);
      }} else {{
        setStatus("Топ-" + top.length + " ближайших к «" + (label || "") + "»");
      }}
    }}

    function normalizeQuery(s) {{
      return (s || "").trim().toLowerCase().replace(/\\s+/g, " ");
    }}

    function findKnownAddress(query) {{
      const q = normalizeQuery(query);
      if (!q) return null;
      let exact = null, partial = null;
      for (let i = 0; i < KNOWN_ADDRESSES.length; i++) {{
        const a = KNOWN_ADDRESSES[i];
        const d = normalizeQuery(a.display);
        const n = normalizeQuery(a.norm);
        if (d === q || n === q) {{ exact = a; break; }}
        if (!partial && (d.indexOf(q) >= 0 || n.indexOf(q) >= 0 ||
            q.indexOf(d) >= 0 || (n && q.indexOf(n) >= 0))) {{
          partial = a;
        }}
      }}
      return exact || partial;
    }}

    function photonGeocode(query) {{
      const url = PHOTON_API + "?q=" + encodeURIComponent(query) +
        "&lang=ru&limit=1&lat=57.77&lon=40.93";
      return fetch(url).then(function (resp) {{
        if (!resp.ok) throw new Error("HTTP " + resp.status);
        return resp.json();
      }}).then(function (data) {{
        const feats = (data && data.features) || [];
        if (!feats.length) return null;
        const f = feats[0];
        const coords = f.geometry && f.geometry.coordinates;
        if (!coords || coords.length < 2) return null;
        const props = f.properties || {{}};
        const parts = [props.name, props.street, props.city, props.state, props.country]
          .filter(Boolean);
        const label = parts.length ? parts.join(", ") : query;
        return {{ lon: coords[0], lat: coords[1], label: label }};
      }});
    }}

    document.getElementById("search-form").addEventListener("submit", function (ev) {{
      ev.preventDefault();
      const q = (searchInput.value || "").trim();
      if (!q) {{
        setStatus("Введите адрес.", true);
        return;
      }}
      const known = findKnownAddress(q);
      if (known) {{
        showPoint(known.lat, known.lon, known.display || known.norm || q);
        return;
      }}
      setStatus("Геокодинг через Photon…");
      nearestList.innerHTML = "";
      photonGeocode(q).then(function (hit) {{
        if (!hit) {{
          setStatus(
            "Адрес не найден. Выберите из списка известных (" +
            KNOWN_ADDRESSES.length + ") или уточните запрос.",
            true
          );
          return;
        }}
        showPoint(hit.lat, hit.lon, hit.label);
      }}).catch(function (err) {{
        const hint = (location.protocol === "file:")
          ? " file:// часто блокирует fetch — откройте через python -m http.server"
          : " (CORS/сеть)";
        setStatus(
          "Не удалось геокодировать" + hint +
          ". Выберите адрес из списка известных (" +
          KNOWN_ADDRESSES.length + ").",
          true
        );
        console.warn(err);
      }});
    }});
  </script>
</body>
</html>
"""
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    tmp = out.with_suffix(out.suffix + ".tmp")
    tmp.write_text(html, encoding="utf-8")
    tmp.replace(out)


def run(
    xlsx_path: Path,
    out_path: Path,
    *,
    offline: bool = False,
    force_geocode: bool = False,
    geo_cache_path: Path | None = None,
    routes_cache_path: Path | None = None,
    stations_path: Path | None = None,
    route_stations_path: Path | None = None,
    via_cache_path: Path | None = None,
    via_threshold_km: float | None = None,
    sleep_sec: float = NOMINATIM_RATE_LIMIT_SEC,
    osrm_sleep_sec: float = OSRM_RATE_LIMIT_SEC,
    sleep_fn: Callable[[float], None] = time.sleep,
    geocode_fn: Callable[[str], tuple[float, float, str] | None] | None = None,
    osrm_fn: (
        Callable[[float, float, float, float], tuple[list[list[float]], float] | None]
        | None
    ) = None,
    osrm_multi_fn: (
        Callable[[list[tuple[float, float]]], tuple[list[list[float]], float] | None]
        | None
    ) = None,
) -> dict[str, Any]:
    """Загрузить маршруты, геокодировать, OSRM → routes.geojson → Leaflet HTML."""
    cache_path = Path(geo_cache_path) if geo_cache_path else DEFAULT_GEO_CACHE
    routes_path = (
        Path(routes_cache_path) if routes_cache_path else DEFAULT_ROUTES_CACHE
    )
    links_path = (
        Path(route_stations_path)
        if route_stations_path
        else DEFAULT_ROUTE_STATIONS
    )
    via_path = Path(via_cache_path) if via_cache_path else DEFAULT_VIA_CACHE
    routes = load_routes_ab(xlsx_path)
    addresses = collect_unique_addresses(routes)
    display_map = address_display_map(routes)
    vehicles = sorted({r.vehicle for r in routes if r.vehicle})

    with requests.Session() as session:
        stats = geocode_addresses(
            addresses,
            display_map,
            cache_path,
            offline=offline,
            force_geocode=force_geocode,
            sleep_sec=sleep_sec,
            sleep_fn=sleep_fn,
            geocode_fn=geocode_fn,
            session=session,
        )
        address_cache = load_address_cache(cache_path)
        geocoded_count = count_geocoded_addresses(address_cache)
        if geocoded_count == 0:
            osrm_stats = OsrmStats(
                skips=(
                    "нет координат в addresses.json — "
                    "сначала геокодируйте адреса (без --offline) "
                    "или заполните кэш вручную",
                ),
            )
        else:
            osrm_stats = build_routes_geojson(
                routes,
                address_cache,
                routes_path,
                offline=offline,
                sleep_sec=osrm_sleep_sec,
                sleep_fn=sleep_fn,
                osrm_fn=osrm_fn,
                session=session,
            )

    routes_collection = load_routes_geojson(routes_path)
    known = known_addresses_from_cache(address_cache)
    stations = load_stations(stations_path or DEFAULT_STATIONS)

    via_stats: ViaStats | None = None
    via_collection: dict[str, Any] | None = None
    via_threshold = via_threshold_km
    via_histogram: list[tuple[str, int]] = []
    via_anomalies: list[Any] = []
    links = load_route_station_links(links_path)
    if links:
        with requests.Session() as session:
            via_stats = build_via_geojson(
                links,
                address_cache,
                routes_collection,
                via_path,
                offline=offline,
                sleep_sec=osrm_sleep_sec,
                sleep_fn=sleep_fn,
                osrm_multi_fn=osrm_multi_fn,
                session=session,
            )
        via_collection = load_routes_geojson(via_path)
        via_features = via_collection["features"]
        detours = [
            (f.get("properties") or {}).get("крюк_км")
            for f in via_features
        ]
        detours = [d for d in detours if isinstance(d, (int, float))]
        if via_threshold is None:
            via_threshold = recommend_via_threshold_km(detours)
        apply_via_threshold(via_features, via_threshold)
        via_histogram = via_detour_histogram(via_features)
        via_anomalies = sorted(
            (
                f
                for f in via_features
                if isinstance((f.get("properties") or {}).get("крюк_км"), (int, float))
            ),
            key=lambda f: -(f["properties"]["крюк_км"]),
        )[:10]

    write_map_html(
        out_path,
        routes_collection,
        VEHICLE_COLORS,
        known_addresses=known,
        stations=stations,
        via=via_collection,
        via_threshold_km=via_threshold,
    )
    html_features = len(routes_collection.get("features") or [])

    return {
        "xlsx": str(xlsx_path),
        "out": str(out_path),
        "geo_cache": str(cache_path),
        "routes_cache": str(routes_path),
        "routes": len(routes),
        "addresses": len(addresses),
        "geocoded_addresses": geocoded_count,
        "vehicles": vehicles,
        "offline": offline,
        "force_geocode": force_geocode,
        "geocoded": stats.geocoded,
        "failed": stats.failed,
        "from_cache": stats.from_cache,
        "missing": list(stats.missing),
        "osrm_routed": osrm_stats.routed,
        "osrm_from_cache": osrm_stats.from_cache,
        "osrm_skipped_missing_coords": osrm_stats.skipped_missing_coords,
        "osrm_skipped_osrm_error": osrm_stats.skipped_osrm_error,
        "osrm_skipped_offline": osrm_stats.skipped_offline,
        "osrm_features": osrm_stats.features,
        "osrm_skips": list(osrm_stats.skips),
        "html_features": html_features,
        "stations": len(stations["features"]) if stations else 0,
        "stations_path": str(stations_path or DEFAULT_STATIONS),
        "via_links": via_stats.links if via_stats else 0,
        "via_features": via_stats.features if via_stats else 0,
        "via_routed": via_stats.routed if via_stats else 0,
        "via_from_cache": via_stats.from_cache if via_stats else 0,
        "via_skipped_missing_coords": (
            via_stats.skipped_missing_coords if via_stats else 0
        ),
        "via_skipped_osrm_error": via_stats.skipped_osrm_error if via_stats else 0,
        "via_skipped_offline": via_stats.skipped_offline if via_stats else 0,
        "via_skips": list(via_stats.skips) if via_stats else [],
        "via_threshold_km": via_threshold,
        "via_histogram": via_histogram,
        "via_anomalies": [
            {
                "машина": (f.get("properties") or {}).get("машина"),
                "маршрут": (
                    f"{(f.get('properties') or {}).get('адрес_A', '')} → "
                    f"{(f.get('properties') or {}).get('адрес_B', '')}"
                ),
                "станция": (f.get("properties") or {}).get("станция"),
                "крюк_км": (f.get("properties") or {}).get("крюк_км"),
            }
            for f in via_anomalies
        ],
        "via_cache": str(via_path),
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Карта маршрутов ГСМ → HTML (routes_ab → Leaflet)"
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX,
        help=f"Путь к пул_поездок.xlsx (по умолчанию: {DEFAULT_XLSX})",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=DEFAULT_OUT,
        help=f"Путь к HTML (по умолчанию: {DEFAULT_OUT})",
    )
    parser.add_argument(
        "--geo-cache",
        type=Path,
        default=DEFAULT_GEO_CACHE,
        help=f"Кэш геокода addresses.json (по умолчанию: {DEFAULT_GEO_CACHE})",
    )
    parser.add_argument(
        "--routes-cache",
        type=Path,
        default=DEFAULT_ROUTES_CACHE,
        help=f"Кэш треков routes.geojson (по умолчанию: {DEFAULT_ROUTES_CACHE})",
    )
    parser.add_argument(
        "--offline",
        action="store_true",
        help="Не ходить в сеть (только кэш геокода и треков)",
    )
    parser.add_argument(
        "--force-geocode",
        action="store_true",
        help="Перезапросить Nominatim даже для записей уже в кэше (в т.ч. not_found)",
    )
    parser.add_argument(
        "--stations",
        type=Path,
        default=DEFAULT_STATIONS,
        help=f"GeoJSON точек АЗС/моек (по умолчанию: {DEFAULT_STATIONS}); "
        "если файла нет — карта строится без слоёв станций",
    )
    parser.add_argument(
        "--route-stations",
        type=Path,
        default=DEFAULT_ROUTE_STATIONS,
        help="Связи маршрут↔АЗС route_stations.json "
        f"(по умолчанию: {DEFAULT_ROUTE_STATIONS})",
    )
    parser.add_argument(
        "--via-cache",
        type=Path,
        default=DEFAULT_VIA_CACHE,
        help=f"Кэш via-треков routes_via.geojson (по умолчанию: {DEFAULT_VIA_CACHE})",
    )
    parser.add_argument(
        "--via-threshold-km",
        type=float,
        default=None,
        help="Порог «заправка по пути», км (по умолчанию: авто по распределению)",
    )
    args = parser.parse_args(argv)

    try:
        report = run(
            args.xlsx,
            args.out,
            offline=args.offline,
            force_geocode=args.force_geocode,
            geo_cache_path=args.geo_cache,
            routes_cache_path=args.routes_cache,
            stations_path=args.stations,
            route_stations_path=args.route_stations,
            via_cache_path=args.via_cache,
            via_threshold_km=args.via_threshold_km,
        )
    except (FileNotFoundError, ValueError, json.JSONDecodeError) as exc:
        print(f"Ошибка: {exc}", file=sys.stderr)
        return 1

    print(f"OK ← {report['xlsx']}")
    print(
        f"routes_ab: {report['routes']}, "
        f"уник. адресов: {report['addresses']}, "
        f"машин: {len(report['vehicles'])}"
    )
    if report["vehicles"]:
        print("машины: " + ", ".join(report["vehicles"]))
    print(
        f"геокод: geocoded={report['geocoded']}, "
        f"failed={report['failed']}, "
        f"from_cache={report['from_cache']}, "
        f"с координатами: {report['geocoded_addresses']}"
    )
    if report["missing"]:
        print(f"offline missing ({len(report['missing'])}):")
        for addr in report["missing"][:50]:
            print(f"  - {addr}")
        if len(report["missing"]) > 50:
            print(f"  … ещё {len(report['missing']) - 50}")
    print(f"кэш адресов → {report['geo_cache']}")
    print(
        f"OSRM: routed={report['osrm_routed']}, "
        f"from_cache={report['osrm_from_cache']}, "
        f"features={report['osrm_features']}, "
        f"skip_coords={report['osrm_skipped_missing_coords']}, "
        f"skip_error={report['osrm_skipped_osrm_error']}, "
        f"skip_offline={report['osrm_skipped_offline']}"
    )
    if report["osrm_skips"]:
        print(f"OSRM skips ({len(report['osrm_skips'])}):")
        for line in report["osrm_skips"][:40]:
            print(f"  - {line}")
        if len(report["osrm_skips"]) > 40:
            print(f"  … ещё {len(report['osrm_skips']) - 40}")
    print(f"кэш треков → {report['routes_cache']}")
    print(
        f"станций (АЗС+мойки): {report['stations']} ← {report['stations_path']}"
    )
    if report["via_links"]:
        print(
            f"via-треки: связей={report['via_links']}, "
            f"features={report['via_features']}, "
            f"routed={report['via_routed']}, "
            f"from_cache={report['via_from_cache']}, "
            f"skip_coords={report['via_skipped_missing_coords']}, "
            f"skip_error={report['via_skipped_osrm_error']}, "
            f"skip_offline={report['via_skipped_offline']}"
        )
        print(f"крюк_км распределение (порог {report['via_threshold_km']:g} км):")
        for label, count in report["via_histogram"]:
            if count:
                print(f"  {label:>8} км: {count}")
        if report["via_anomalies"]:
            print("топ аномалий (проверить координаты станции / атрибуцию дня):")
            for a in report["via_anomalies"]:
                print(
                    f"  {a['крюк_км']:>7.1f} км | {a['машина']} | "
                    f"{a['маршрут'][:50]} | {str(a['станция'])[:50]}"
                )
        for line in report["via_skips"][:20]:
            print(f"  - {line}")
        if len(report["via_skips"]) > 20:
            print(f"  … ещё {len(report['via_skips']) - 20}")
    print(
        f"HTML → {report['out']} "
        f"(features={report['html_features']}, offline={report['offline']})"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

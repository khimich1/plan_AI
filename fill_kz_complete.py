"""
Программа ПОЛНОГО заполнения КЗ (Карт Заданий) по дорожкам
Заполняет все поля шаблона данными из базы
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import shutil
import os
from datetime import datetime, timedelta
import sqlite3
import sys

# Настройка кодировки для Windows
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'ignore')


def load_track_data(db_path: str = 'pb.db') -> list:
    """Загружает данные о дорожках из базы данных"""
    conn = sqlite3.connect(db_path)
    
    query = '''
        SELECT 
            "номенклатура пб к производству" as marking,
            "длина плиты, м" as length_db,
            "контрагент заказчик" as customer,
            "армирование по серии" as reinforcement,
            "неделя формовки" as week
        FROM plity_ex
    '''
    df = pd.read_sql(query, conn)
    conn.close()
    
    return df


def parse_plate_marking(marking: str):
    """Парсит маркировку плиты"""
    import re
    pattern = r'ПБ\s+(\d+[,.]?\d*)-(\d+[,.]?\d*)-(\d+)'
    match = re.search(pattern, marking)
    
    if not match:
        return None
    
    try:
        length_dm = float(match.group(1).replace(',', '.'))
        length_m = length_dm / 10
        
        width_dm = float(match.group(2).replace(',', '.'))
        width_m = width_dm / 10
        
        load_code = int(match.group(3))
        load_capacity = load_code * 100
        
        return {
            'length': length_m,
            'width': width_m,
            'load_capacity': load_capacity
        }
    except:
        return None


def get_track_data():
    """Получает данные для каждой дорожки"""
    df = load_track_data()
    
    # Фильтруем плиты с шириной 1.2 м
    plates_1_2m = []
    for idx, row in df.iterrows():
        params = parse_plate_marking(row['marking'])
        if params and params['width'] == 1.2:
            plates_1_2m.append({
                'marking': row['marking'],
                'length': params['length'],
                'width': params['width'],
                'load_capacity': params['load_capacity'],
                'customer': row['customer'],
                'reinforcement': row['reinforcement'],
                'week': row['week']
            })
    
    # Сортируем по неделе формовки
    plates_1_2m.sort(key=lambda x: x['week'] if x['week'] is not None else 999)
    
    tracks = []
    
    # Дорожка 1 - армирование 8.0 (все недели от меньшей к большей)
    track1_plates = [p for p in plates_1_2m if p['reinforcement'] == 8.0]
    # Сортируем по неделям (от меньшей к большей, NaN в конец)
    track1_plates.sort(key=lambda x: x['week'] if x['week'] is not None else 999)
    
    if track1_plates:
        # Берём плиты рядами до тех пор, пока не наберём ~101 м длины дорожки
        selected_plates = []
        current_track_length = 0  # Длина дорожки (вдоль)
        target_track_length = 101.0  # Целевая длина дорожки
        plates_per_row = 3  # 3 плиты поперёк (по ширине)
        
        for plate in track1_plates:
            # Проверяем, поместится ли этот ряд плит в дорожку
            if current_track_length + plate['length'] <= target_track_length:
                # Добавляем плиту в ряд (3 плиты поперёк)
                selected_plates.append(plate)
                current_track_length += plate['length']  # Длина дорожки увеличивается на длину плиты
            else:
                break  # Больше не помещается
        
        # Суммарная длина всех плит (для расчётов)
        total_plates_length = sum(p['length'] for p in selected_plates)
        
        tracks.append({
            'track_num': 1,
            'plates': selected_plates,
            'reinforcement': 8.0,
            'total_plates': len(selected_plates) * plates_per_row,  # 3 плиты в каждом ряду
            'total_length': current_track_length,  # Длина дорожки
            'total_plates_length': total_plates_length  # Суммарная длина плит
        })
    
    # Дорожка 2 - армирование 6.0 (все недели от меньшей к большей)
    track2_plates = [p for p in plates_1_2m if p['reinforcement'] == 6.0]
    # Сортируем по неделям (от меньшей к большей, NaN в конец)
    track2_plates.sort(key=lambda x: x['week'] if x['week'] is not None else 999)
    
    if track2_plates:
        # Берём плиты рядами до тех пор, пока не наберём ~101 м длины дорожки
        selected_plates = []
        current_track_length = 0  # Длина дорожки (вдоль)
        target_track_length = 101.0  # Целевая длина дорожки
        plates_per_row = 3  # 3 плиты поперёк (по ширине)
        
        for plate in track2_plates:
            # Проверяем, поместится ли этот ряд плит в дорожку
            if current_track_length + plate['length'] <= target_track_length:
                # Добавляем плиту в ряд (3 плиты поперёк)
                selected_plates.append(plate)
                current_track_length += plate['length']  # Длина дорожки увеличивается на длину плиты
            else:
                break  # Больше не помещается
        
        # Суммарная длина всех плит (для расчётов)
        total_plates_length = sum(p['length'] for p in selected_plates)
        
        tracks.append({
            'track_num': 2,
            'plates': selected_plates,
            'reinforcement': 6.0,
            'total_plates': len(selected_plates) * plates_per_row,  # 3 плиты в каждом ряду
            'total_length': current_track_length,  # Длина дорожки
            'total_plates_length': total_plates_length  # Суммарная длина плит
        })
    
    # Дорожка 3 - армирование 4.0 (все недели от меньшей к большей)
    track3_plates = [p for p in plates_1_2m if p['reinforcement'] == 4.0]
    # Сортируем по неделям (от меньшей к большей, NaN в конец)
    track3_plates.sort(key=lambda x: x['week'] if x['week'] is not None else 999)
    
    if track3_plates:
        # Берём плиты рядами до тех пор, пока не наберём ~101 м длины дорожки
        selected_plates = []
        current_track_length = 0  # Длина дорожки (вдоль)
        target_track_length = 101.0  # Целевая длина дорожки
        plates_per_row = 3  # 3 плиты поперёк (по ширине)
        
        for plate in track3_plates:
            # Проверяем, поместится ли этот ряд плит в дорожку
            if current_track_length + plate['length'] <= target_track_length:
                # Добавляем плиту в ряд (3 плиты поперёк)
                selected_plates.append(plate)
                current_track_length += plate['length']  # Длина дорожки увеличивается на длину плиты
            else:
                break  # Больше не помещается
        
        # Суммарная длина всех плит (для расчётов)
        total_plates_length = sum(p['length'] for p in selected_plates)
        
        tracks.append({
            'track_num': 3,
            'plates': selected_plates,
            'reinforcement': 4.0,
            'total_plates': len(selected_plates) * plates_per_row,  # 3 плиты в каждом ряду
            'total_length': current_track_length,  # Длина дорожки
            'total_plates_length': total_plates_length  # Суммарная длина плит
        })
    
    return tracks


def calculate_plate_weight(length_m: float, width_m: float, thickness_m: float = 0.22) -> float:
    """Рассчитывает вес плиты в кг"""
    # Плотность бетона ~2400 кг/м³
    volume = length_m * width_m * thickness_m
    weight = volume * 2400
    return round(weight, 1)


def fill_kz_complete(track_data: dict, template_path: str, output_dir: str):
    """ПОЛНОСТЬЮ заполняет КЗ файл для дорожки"""
    track_num = track_data['track_num']
    plates = track_data['plates']
    reinforcement = track_data['reinforcement']
    total_plates = track_data['total_plates']
    total_length = track_data['total_length']
    
    filename = f"КЗ_ПБ_Дорожка_{track_num}_Полный.xlsx"
    output_path = os.path.join(output_dir, filename)
    
    # Копируем шаблон
    shutil.copy2(template_path, output_path)
    
    # Открываем файл
    wb = load_workbook(output_path)
    ws = wb.active
    
    print(f"[ЗАПОЛНЕНИЕ] КЗ для дорожки #{track_num}...")
    print(f"   Файл: {filename}")
    print(f"   Плит: {total_plates} шт")
    print(f"   Длина дорожки: {track_data.get('total_length', total_length):.2f} м")
    if 'total_plates_length' in track_data:
        print(f"   Суммарная длина плит: {track_data['total_plates_length']:.2f} м")
    print(f"   Армирование: {reinforcement}")
    
    # Заполняем заголовочные поля
    current_date = datetime.now()
    molding_date = current_date.strftime("%d.%m.%Y")
    acceptance_date = (current_date + timedelta(days=1)).strftime("%d.%m.%Y")
    
    try:
        # Строка 1: Дата формовки и № дорожки
        ws['B1'] = molding_date  # Дата формовки
        ws['D1'] = f"Дорожка #{track_num}"  # № дорожки
        
        # Строка 6: Дата приемки ОТК
        ws['B6'] = acceptance_date  # Дата приемки ОТК
        
        print(f"   [ЗАПОЛНЕНО] Даты: формовки {molding_date}, приемки {acceptance_date}")
        
    except Exception as e:
        print(f"   [ОШИБКА] Ошибка при заполнении дат: {e}")
    
    # Заполняем таблицу плит (строки 13-27)
    table_start_row = 13
    total_weight = 0
    
    try:
        for i, plate in enumerate(plates):
            if i >= 15:  # Максимум 15 строк в таблице
                break
                
            row = table_start_row + i
            
            # Армирование
            ws.cell(row=row, column=1).value = f"Арм. {reinforcement}"
            
            # Заказ (заказчик)
            ws.cell(row=row, column=2).value = plate['customer'][:20] if plate['customer'] else "Заказчик"
            
            # Номенклатура
            ws.cell(row=row, column=3).value = plate['marking'][:30] if plate['marking'] else "Плита ПБ"
            
            # Количество в формовку (3 плиты поперёк)
            ws.cell(row=row, column=4).value = 3
            
            # Метраж в формовку ПЛ (длина плиты, так как 3 плиты идут поперёк)
            metrage = plate['length']  # Длина плиты (3 плиты поперёк образуют ряд)
            ws.cell(row=row, column=5).value = round(metrage, 2)
            
            # Вес изделия (рассчитываем)
            weight_per_plate = calculate_plate_weight(plate['length'], plate['width'])
            total_weight_this_row = weight_per_plate * 3
            ws.cell(row=row, column=7).value = round(total_weight_this_row, 1)
            
            total_weight += total_weight_this_row
            
            # Примечание (неделя формовки)
            week_text = f"Неделя {plate['week']}" if plate['week'] else "Срочно"
            ws.cell(row=row, column=10).value = week_text
        
        print(f"   [ЗАПОЛНЕНО] Таблица плит: {len(plates)} позиций")
        print(f"   [РАСЧЕТ] Общий вес: {total_weight:.1f} кг")
        
    except Exception as e:
        print(f"   [ОШИБКА] Ошибка при заполнении таблицы: {e}")
    
    # Заполняем итоговые строки
    try:
        # Строка 12: Итого количество
        ws['D12'] = total_plates
        
        # Строка 9: Общий метраж в дорожке (длина дорожки)
        track_length = track_data.get('total_length', total_length)
        ws['G9'] = round(track_length, 2)
        
        print(f"   [ЗАПОЛНЕНО] Итоги: {total_plates} плит, {total_length:.2f} м")
        
    except Exception as e:
        print(f"   [ОШИБКА] Ошибка при заполнении итогов: {e}")
    
    # Сохраняем файл
    wb.save(output_path)
    wb.close()
    
    print(f"   [ГОТОВО] {filename}")
    return output_path


def main():
    """Основная функция"""
    print("=" * 70)
    print("ПОЛНОЕ ЗАПОЛНЕНИЕ КЗ (КАРТ ЗАДАНИЙ) ПО ДОРОЖКАМ")
    print("=" * 70)
    
    template_path = "!КЗ ПБ Шаблон.xlsx"
    output_dir = "КЗ_файлы_ПОЛНЫЕ"
    
    # Создаём папку
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"[ПАПКА] Создана папка: {output_dir}")
    
    # Проверяем шаблон
    if not os.path.exists(template_path):
        print(f"[ОШИБКА] Не найден шаблон: {template_path}")
        return
    
    print(f"[ШАБЛОН] Используем шаблон: {template_path}")
    
    # Получаем данные дорожек
    print("\n[ДАННЫЕ] Получаю данные о дорожках...")
    tracks = get_track_data()
    print(f"[УСПЕХ] Найдено {len(tracks)} дорожек")
    
    # Заполняем КЗ для каждой дорожки
    print(f"\n[ЗАПОЛНЕНИЕ] Создаю ПОЛНЫЕ КЗ файлы...")
    created_files = []
    
    for track in tracks:
        try:
            file_path = fill_kz_complete(track, template_path, output_dir)
            created_files.append(file_path)
        except Exception as e:
            print(f"[ОШИБКА] Ошибка при создании КЗ для дорожки {track['track_num']}: {e}")
    
    # Итоговая сводка
    print(f"\n" + "=" * 70)
    print("ИТОГОВАЯ СВОДКА")
    print("=" * 70)
    print(f"[УСПЕХ] Создано ПОЛНЫХ КЗ файлов: {len(created_files)}")
    print(f"[ПАПКА] Папка: {output_dir}")
    
    for file_path in created_files:
        filename = os.path.basename(file_path)
        print(f"   [ФАЙЛ] {filename}")
    
    print(f"\n[ГОТОВО] Все КЗ файлы ПОЛНОСТЬЮ заполнены!")
    print("\n[ИНФОРМАЦИЯ] Каждый файл содержит:")
    print("   • Даты формовки и приемки ОТК")
    print("   • Номер дорожки")
    print("   • Полную таблицу плит с расчетами")
    print("   • Метраж и количество плит")
    print("   • Вес изделий")
    print("   • Недели формовки (срочность)")
    print("   • Заказчиков")


if __name__ == "__main__":
    main()

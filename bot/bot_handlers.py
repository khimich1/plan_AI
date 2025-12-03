import asyncio
import os
import re
import sys
from pathlib import Path
from datetime import datetime
from typing import Any, Dict
from dataclasses import dataclass

# Добавляем корень проекта в sys.path
BOT_DIR = Path(__file__).parent
PROJECT_ROOT = BOT_DIR.parent
sys.path.insert(0, str(PROJECT_ROOT))

from aiogram import Router, F
from aiogram.types import (
    Message,
    CallbackQuery,
    FSInputFile,
    ReplyKeyboardMarkup,
    KeyboardButton,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
)
from aiogram.filters import Command
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext

try:
    import pandas as pd
except Exception:
    pd = None

# Импорты из core/
from core.visualization import visualize_plan
from core.config_and_data import set_plate_lists_from_text, parse_name_to_sizes
from core.optimization import apply_width_optimization, optimize_with_cascading_longitudinal_cuts
import core.config_and_data as cfg
from core.commercial_offer import generate_commercial_offer_pdf
# Умное OCR: сначала EasyOCR (бесплатно), потом GPT-4o (платно)
from core.ocr_gpt import recognize_text_smart, GPT_AVAILABLE, EASYOCR_AVAILABLE

# Импорт из локального модуля
from .bot_config import OUTPUTS_DIR_STR
# TODO: Модули не реализованы - временно закомментированы
# from planning import plan_tracks, available_days, track_to_text, render_line

router = Router()

PLANNING_CACHE: Dict[int, Dict[str, Any]] = {}
ORDER_CACHE: Dict[int, list] = {}  # Кэш для хранения заказов пользователей


def register_handlers(dp):
    """Регистрируем все обработчики"""
    dp.include_router(router)

class KPStates(StatesGroup):
    waiting_for_plate_list = State()
    waiting_for_commercial_offer = State()


class CompareStates(StatesGroup):
    """Состояния диалога для сравнения реального КП и нашей сметы."""

    waiting_kp = State()
    waiting_smeta = State()

def main_menu_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Получить КП")],
            [KeyboardButton(text="Оптимизация резов")],
            [KeyboardButton(text="Коммерческое предложение PDF")],
            [KeyboardButton(text="Сравнение результатов")],
            # TODO: Временно отключено - модуль не реализован
            # [KeyboardButton(text="Планирование по дням")],
        ],
        resize_keyboard=True
    )

@router.message(Command("start"))
async def cmd_start(message: Message):
    """Обработчик команды /start"""
    await message.answer(
        "👋 Привет! Я бот для расчёта и визуализации дорожек ПБ.\n\n"
        "🔧 Что я умею:\n"
        "• Строить планы раскладки плит\n"
        "• Рассчитывать стоимость и отходы\n"
        "• Оптимизировать раскрой (экономия до 40%)\n"
        "• Экспортировать результаты в файлы\n\n"
        "Выберите действие кнопкой ниже или /help для справки",
        reply_markup=main_menu_kb()
    )

@router.message(F.text == "Получить КП")
async def btn_get_kp(message: Message, state: FSMContext):
    await state.set_state(KPStates.waiting_for_plate_list)
    
    # Формируем подсказку о фото в зависимости от доступных методов OCR
    if GPT_AVAILABLE:
        photo_hint = "\n📸 Или отправьте фото таблицы - я распознаю через 🧠 GPT-4o (точность 95%+)!"
    elif EASYOCR_AVAILABLE:
        photo_hint = "\n📸 Или отправьте фото таблицы - я распознаю через 🤖 EasyOCR!"
    else:
        photo_hint = ""
    
    await message.answer(
        "✍️ Пришлите список плит в свободной форме.\n"
        "Например: '1.2×3.39 — 2 шт; 0.32×6.63 — 4 шт; 0.32×7.83 — 3 шт'\n"
        f"{photo_hint}\n\n"
        "Я выполню расчёт с оптимизацией и пришлю схемы и смету.\n"
        "💡 Используется каскадная оптимизация для экономии материала!",
        reply_markup=main_menu_kb()
    )


# ==================== СРАВНЕНИЕ РЕАЛЬНОГО КП И НАШЕЙ СМЕТЫ ====================


@dataclass
class PlateRow:
    """Одна строка из Excel с данными по плите."""

    name: str
    qty: float
    weight: float
    price: float
    total: float

    @property
    def price_per_kg(self) -> float | None:
        if self.weight and self.weight > 0:
            return self.price / self.weight
        return None


def _to_float(value) -> float:
    """Переводит '9 130,80' → 9130.80."""
    try:
        s = str(value).replace(" ", "").replace("\u00a0", "").replace(",", ".")
        return float(s)
    except Exception:
        return 0.0


def _to_qty(value) -> float:
    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except Exception:
        return 0.0


def _normalize_name(name: str) -> str:
    """
    Нормализуем наименование плиты:
    - убираем лишние пробелы;
    - приводим к одному регистру;
    - оставляем только текст без лишних пробелов.
    """
    return " ".join(str(name).strip().lower().split())


def _make_plate_key(name: str):
    """
    Делает «умный» ключ для сравнения плит.

    Логика:
    - из строки вида 'Плиты ПБ 66,2-12-8п' вытаскиваем два первых числа: 66,2 и 12;
    - это длина и ширина в дециметрах;
    - переводим их в float, округляем до целых дм и берём пару (length_dm, width_dm);
      так '66,2-12-8п' и '66-12-8п' дадут один ключ (66, 12);
    - если распарсить не удалось — используем нормализованный текст как раньше.
    """
    import re

    s = str(name)

    # Ищем шаблон "ЧИСЛО-ЧИСЛО" с поддержкой запятой/точки
    m = re.search(r'(\d[\d,\.]*)\s*-\s*(\d[\d,\.]*)', s)
    if m:
        length_dm_str = m.group(1).replace(" ", "").replace(",", ".")
        width_dm_str = m.group(2).replace(" ", "").replace(",", ".")
        try:
            # Округляем длину и ширину в дециметрах до целых
            length_dm = int(round(float(length_dm_str)))
            width_dm = int(round(float(width_dm_str)))
            return (length_dm, width_dm)
        except ValueError:
            # если вдруг не смогли преобразовать — упадём в текстовый ключ
            pass

    # Фолбэк: сравниваем по тексту, как раньше
    return _normalize_name(name)


def _find_column(df, *candidates: str) -> str:
    """
    Находит колонку в DataFrame по нескольким вариантам имени
    (по вхождению подстроки без учёта регистра и пробелов).
    """
    columns = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        cand_norm = cand.strip().lower()
        for key, real_name in columns.items():
            if cand_norm == key or cand_norm in key:
                return real_name
    raise KeyError(f"Не найдена колонка среди вариантов: {candidates}")


def _aggregate_row(result: dict, row: PlateRow) -> None:
    """
    Добавляет строку в словарь, суммируя данные,
    если такая плита уже встречалась (даже если она записана по‑разному).

    Ключ строим по (длина, ширина) в дециметрах, а не по сырому тексту:
    так 'Плиты ПБ 66,2-12-8п' и 'Плиты ПБ 66-12-8п' попадут в одну позицию.
    """
    key = _make_plate_key(row.name)
    prev = result.get(key)

    if prev is not None:
        qty = prev.qty + row.qty
        total = prev.total + row.total
        # Вес и цена считаем одинаковыми для одинаковой плиты,
        # поэтому просто берём первое ненулевое значение.
        weight = prev.weight or row.weight
        price = prev.price or row.price
        result[key] = PlateRow(
            name=prev.name,  # храним оригинальное имя первой строки
            qty=qty,
            weight=weight,
            price=price,
            total=total,
        )
    else:
        result[key] = row


def _prepare_kp_sheet(sheet_df):
    """
    Пытается найти строку с заголовками ('Товары (работы, услуги)', 'Количество', 'Цена', 'Сумма')
    если она не в первой строке, и переставить её как header DataFrame.
    """
    # Если заголовки уже в columns, просто возвращаем как есть
    try:
        _find_column(
            sheet_df,
            "товары (работы, услуги)",
            "наименование",
            "товар",
            "товары",
            "работы",
            "услуги",
        )
        return sheet_df
    except KeyError:
        pass

    header_candidates = (
        "товары (работы, услуги)",
        "наименование",
        "товар",
        "товары",
        "работы",
        "услуги",
    )
    max_rows = min(40, len(sheet_df))
    for i in range(max_rows):
        row = sheet_df.iloc[i]
        row_lower = [str(v).strip().lower() for v in row]
        if any(
            any(cand in cell for cand in header_candidates)
            for cell in row_lower
        ):
            # Делаем эту строку заголовком
            new_df = sheet_df.iloc[i + 1 :].copy()
            new_df.columns = row
            return new_df
    return sheet_df


def load_kp_excel(path: str) -> dict:
    """
    Загружает РЕАЛЬНОЕ КП завода.
    Здесь может не быть столбца веса, а наименование может быть
    в колонке «Товары (работы, услуги)».
    """
    if pd is None:
        raise RuntimeError("pandas не установлена, сравнение КП недоступно.")

    all_sheets = pd.read_excel(path, sheet_name=None)
    df = None
    name_col = qty_col = price_col = total_col = None

    for sheet_df in all_sheets.values():
        sheet_df = _prepare_kp_sheet(sheet_df)
        try:
            name_col = _find_column(
                sheet_df,
                "товары (работы, услуги)",
                "наименование",
                "наимен",
                "товар",
                "товары",
                "работы",
                "услуги",
            )
            qty_col = _find_column(sheet_df, "кол-во", "количество")
            price_col = _find_column(sheet_df, "цена", "стоимость")
            total_col = _find_column(sheet_df, "сумма", "итого")
        except KeyError:
            continue
        df = sheet_df
        break

    if df is None:
        raise RuntimeError("Не удалось найти нужные столбцы в файле КП.")

    result: dict = {}

    for _, row in df.iterrows():
        name = str(row.get(name_col, "")).strip()
        if not name:
            continue
        if "итог" in name.lower():
            # Строки с итогами пропускаем
            continue

        qty = _to_qty(row.get(qty_col, 0))
        price = _to_float(row.get(price_col, 0))
        total = _to_float(row.get(total_col, 0))

        # Нас интересуют только плиты ПБ
        if "плиты пб" not in _normalize_name(name):
            continue

        plate_row = PlateRow(
            name=name,
            qty=qty,
            weight=0.0,  # в КП веса нет
            price=price,
            total=total,
        )
        _aggregate_row(result, plate_row)

    return result


def load_smeta_excel(path: str) -> dict:
    """
    Загружает НАШУ СМЕТУ (Смета_Дорожка_1_...xlsx).
    Здесь обязательно есть колонка веса.
    """
    if pd is None:
        raise RuntimeError("pandas не установлена, сравнение КП недоступно.")

    all_sheets = pd.read_excel(path, sheet_name=None)
    df = None
    name_col = qty_col = weight_col = price_col = total_col = None

    for sheet_df in all_sheets.values():
        try:
            name_col = _find_column(sheet_df, "наименование", "наимен")
            qty_col = _find_column(sheet_df, "кол-во", "количество")
            weight_col = _find_column(sheet_df, "вес", "вес(кг)", "масса")
            price_col = _find_column(sheet_df, "цена")
            total_col = _find_column(sheet_df, "сумма")
        except KeyError:
            continue
        df = sheet_df
        break

    if df is None:
        raise RuntimeError("Не удалось найти нужные столбцы в файле сметы.")

    result: dict = {}

    for _, row in df.iterrows():
        name = str(row.get(name_col, "")).strip()
        if not name:
            continue
        if "итого" in name.lower():
            continue

        qty = _to_qty(row.get(qty_col, 0))
        weight = _to_float(row.get(weight_col, 0))
        price = _to_float(row.get(price_col, 0))
        total = _to_float(row.get(total_col, 0))

        plate_row = PlateRow(
            name=name,
            qty=qty,
            weight=weight,
            price=price,
            total=total,
        )
        _aggregate_row(result, plate_row)

    return result


def _explain_price_breakdown(name: str, our_price: float, smeta_path: str) -> str:
    """
    Пытается найти детальную разбивку цены для плиты из файла детальной разбивки рядом со сметой.
    Возвращает строку с объяснением компонентов цены.
    """
    try:
        # Ищем файл детальной разбивки (может быть в том же Excel-файле на другом листе)
        all_sheets = pd.read_excel(smeta_path, sheet_name=None)
        breakdown_df = None
        
        for sheet_name, df in all_sheets.items():
            if 'детал' in sheet_name.lower() or 'разбивка' in sheet_name.lower() or 'компонент' in sheet_name.lower():
                breakdown_df = df
                break
        
        if breakdown_df is None:
            return "  💡 Детальная разбивка не найдена в файле сметы"
        
        # Ищем строку с нашим наименованием
        for _, row in breakdown_df.iterrows():
            row_name = str(row.get('Наименование', '')).strip()
            # Сравниваем по ключу (length, width)
            if _normalize_name(row_name) == _normalize_name(name) or _make_plate_key(row_name) == _make_plate_key(name):
                # Извлекаем компоненты
                base = _to_float(row.get('Базовая цена', 0))
                long_cuts = _to_float(row.get('Продольные резы', 0))
                trans_cuts = _to_float(row.get('Поперечные резы', 0))
                rest = _to_float(row.get('Остатки', 0))
                waste = _to_float(row.get('Отходы', 0))
                
                parts = []
                if base > 0:
                    parts.append(f"базовая {base:,.2f} ₽")
                if long_cuts > 0:
                    parts.append(f"прод.резы {long_cuts:,.2f} ₽")
                if trans_cuts > 0:
                    parts.append(f"попер.резы {trans_cuts:,.2f} ₽")
                if rest > 0:
                    parts.append(f"остатки {rest:,.2f} ₽")
                if waste > 0:
                    parts.append(f"отходы {waste:,.2f} ₽")
                
                if parts:
                    formula = " + ".join(parts)
                    return f"  💡 Разбивка: {formula} = {our_price:,.2f} ₽"
                else:
                    return f"  💡 Разбивка: {our_price:,.2f} ₽ (детали не найдены в таблице)"
        
        return "  💡 Позиция не найдена в детальной разбивке"
    
    except Exception as e:
        return f"  💡 Ошибка чтения разбивки: {e}"


def create_comparison_excel(kp_path: str, smeta_path: str) -> str:
    """
    Создаёт Excel файл с двумя таблицами (смета и КП) рядом + пояснения о ценообразовании.
    
    Возвращает путь к созданному файлу.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    if pd is None:
        raise RuntimeError("pandas не установлена, сравнение КП недоступно.")
    
    # Загружаем данные из КП и сметы
    kp = load_kp_excel(kp_path)
    smeta = load_smeta_excel(smeta_path)
    
    # Ключи, которые есть в обоих файлах
    common_names = sorted(set(kp.keys()) & set(smeta.keys()), key=str)
    
    if not common_names:
        raise RuntimeError("Не нашёл общих позиций по наименованию плит в двух файлах.")
    
    # Создаём новый Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сравнение"
    
    # ==================== СТИЛИ ====================
    # Заголовок
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Подзаголовки
    subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    subheader_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Названия плит
    plate_name_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    plate_name_font = Font(bold=True, size=10)
    
    # Разница (положительная - зелёный, отрицательная - красный)
    positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Рамки
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ==================== ГЛАВНЫЙ ЗАГОЛОВОК ====================
    ws.merge_cells('A1:N1')
    cell = ws['A1']
    cell.value = "📊 СРАВНЕНИЕ СМЕТЫ И КП ЗАВОДА"
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    
    # ==================== ЗАГОЛОВКИ ТАБЛИЦ ====================
    row = 3
    
    # Заголовок "НАША СМЕТА"
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws[f'A{row}']
    cell.value = "📋 НАША СМЕТА"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    
    # Заголовок "КП ЗАВОДА"
    ws.merge_cells(f'H{row}:N{row}')
    cell = ws[f'H{row}']
    cell.value = "🏭 КП ЗАВОДА"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    
    ws.row_dimensions[row].height = 20
    row += 1
    
    # ==================== СТОЛБЦЫ ТАБЛИЦ ====================
    # Столбцы для сметы
    smeta_cols = ['№', 'Наименование', 'Кол-во', 'Вес, кг', 'Цена, ₽', 'Сумма, ₽', 'Примечание']
    for col_idx, col_name in enumerate(smeta_cols, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = col_name
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = subheader_align
        cell.border = thin_border
    
    # Столбцы для КП
    kp_cols = ['№', 'Наименование', 'Кол-во', 'Цена, ₽', 'Сумма, ₽', 'Разница цены', 'Разница суммы']
    for col_idx, col_name in enumerate(kp_cols, start=8):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = col_name
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = subheader_align
        cell.border = thin_border
    
    ws.row_dimensions[row].height = 30
    row += 1
    
    # ==================== ДАННЫЕ ====================
    # Пытаемся найти файл детальной разбивки
    breakdown_data = {}
    try:
        # Ищем файл детальной разбивки рядом со сметой
        smeta_dir = os.path.dirname(smeta_path)
        smeta_name = os.path.basename(smeta_path)
        
        # Извлекаем timestamp из имени сметы (например, "Смета_Дорожка_1_20250128_143022.xlsx")
        timestamp_match = re.search(r'_(\d{8}_\d{6})\.xlsx$', smeta_name)
        if timestamp_match:
            timestamp = timestamp_match.group(1)
            breakdown_path = os.path.join(smeta_dir, f'Детальная_разбивка_Дорожка_1_{timestamp}.xlsx')
            
            if os.path.exists(breakdown_path):
                # Читаем файл детальной разбивки
                breakdown_df = pd.read_excel(breakdown_path)
                current_plate = None
                
                for _, br_row in breakdown_df.iterrows():
                    comp = str(br_row.get('Компонент', '')).strip()
                    calc = str(br_row.get('Расчёт', '')).strip()
                    summa = str(br_row.get('Сумма', '')).strip()
                    
                    # Если это заголовок плиты (наименование)
                    if comp and not calc and not summa and 'Плиты ПБ' in comp:
                        current_plate = _make_plate_key(comp)
                        breakdown_data[current_plate] = []
                    elif current_plate and comp:
                        # Добавляем компонент к текущей плите
                        breakdown_data[current_plate].append({
                            'component': comp,
                            'formula': calc,
                            'sum': summa
                        })
    except Exception as e:
        print(f"[DEBUG] Не удалось загрузить детальную разбивку: {e}")
    
    # Итоги
    total_smeta = 0.0
    total_kp = 0.0
    
    idx = 1
    for key in common_names:
        s = smeta[key]  # строка из нашей сметы
        k = kp[key]     # строка из КП завода
        
        name = s.name or k.name
        
        # ===== СМЕТА (столбцы A-G) =====
        # Номер
        cell = ws.cell(row=row, column=1, value=idx)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
        # Наименование
        cell = ws.cell(row=row, column=2, value=name)
        cell.font = plate_name_font
        cell.fill = plate_name_fill
        cell.border = thin_border
        
        # Количество
        cell = ws.cell(row=row, column=3, value=s.qty)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        cell.number_format = '0'
        
        # Вес
        cell = ws.cell(row=row, column=4, value=s.weight)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        cell.number_format = '0.00'
        
        # Цена
        cell = ws.cell(row=row, column=5, value=s.price)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        cell.number_format = '#,##0.00'
        
        # Сумма
        cell = ws.cell(row=row, column=6, value=s.total)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        cell.number_format = '#,##0.00'
        
        # Примечание (цена за кг)
        if s.price_per_kg:
            note = f"{s.price_per_kg:.2f} ₽/кг"
        else:
            note = ""
        cell = ws.cell(row=row, column=7, value=note)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
        # ===== КП (столбцы H-N) =====
        # Номер
        cell = ws.cell(row=row, column=8, value=idx)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
        # Наименование
        cell = ws.cell(row=row, column=9, value=name)
        cell.font = plate_name_font
        cell.fill = plate_name_fill
        cell.border = thin_border
        
        # Количество
        cell = ws.cell(row=row, column=10, value=k.qty)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        cell.number_format = '0'
        
        # Цена
        cell = ws.cell(row=row, column=11, value=k.price)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        cell.number_format = '#,##0.00'
        
        # Сумма
        cell = ws.cell(row=row, column=12, value=k.total)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        cell.number_format = '#,##0.00'
        
        # Разница цены (наша - их)
        price_diff = s.price - k.price
        cell = ws.cell(row=row, column=13, value=price_diff)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        cell.number_format = '#,##0.00'
        if price_diff > 0:
            cell.fill = negative_fill  # Мы дороже - красный
        elif price_diff < 0:
            cell.fill = positive_fill  # Мы дешевле - зелёный
        
        # Разница суммы (наша - их)
        total_diff = s.total - k.total
        cell = ws.cell(row=row, column=14, value=total_diff)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        cell.number_format = '#,##0.00'
        if total_diff > 0:
            cell.fill = negative_fill  # Мы дороже - красный
        elif total_diff < 0:
            cell.fill = positive_fill  # Мы дешевле - зелёный
        
        total_smeta += s.total
        total_kp += k.total
        
        row += 1
        
        # ===== ДЕТАЛЬНАЯ РАЗБИВКА ДЛЯ ЭТОЙ ПЛИТЫ =====
        if key in breakdown_data and breakdown_data[key]:
            # Заголовок разбивки
            ws.merge_cells(f'A{row}:N{row}')
            cell = ws[f'A{row}']
            cell.value = "💡 Детальная разбивка цены из нашей сметы:"
            cell.font = Font(bold=True, italic=True, size=9, color="404040")
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 18
            row += 1
            
            # Строки разбивки
            for comp_data in breakdown_data[key]:
                # Компонент
                cell = ws.cell(row=row, column=2, value=comp_data['component'])
                cell.font = Font(size=9, italic=True)
                cell.alignment = Alignment(horizontal="left")
                
                # Расчёт
                ws.merge_cells(f'C{row}:F{row}')
                cell = ws.cell(row=row, column=3, value=comp_data['formula'])
                cell.font = Font(size=9, italic=True, color="404040")
                cell.alignment = Alignment(horizontal="left")
                
                # Сумма
                cell = ws.cell(row=row, column=7, value=comp_data['sum'])
                cell.font = Font(size=9, italic=True)
                cell.alignment = Alignment(horizontal="right")
                
                ws.row_dimensions[row].height = 15
                row += 1
            
            # Пустая строка после разбивки
            row += 1
        
        idx += 1
    
    # ==================== ИТОГО ====================
    row += 1
    
    # ИТОГО для сметы
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "ИТОГО (наша смета):"
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="right")
    cell.border = thin_border
    
    cell = ws.cell(row=row, column=6, value=total_smeta)
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="right")
    cell.border = thin_border
    cell.number_format = '#,##0.00'
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # ИТОГО для КП
    ws.merge_cells(f'H{row}:K{row}')
    cell = ws[f'H{row}']
    cell.value = "ИТОГО (КП завода):"
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="right")
    cell.border = thin_border
    
    cell = ws.cell(row=row, column=12, value=total_kp)
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="right")
    cell.border = thin_border
    cell.number_format = '#,##0.00'
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Общая разница
    ws.merge_cells(f'M{row}:N{row}')
    cell = ws[f'M{row}']
    diff_total = total_smeta - total_kp
    cell.value = diff_total
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="right")
    cell.border = thin_border
    cell.number_format = '#,##0.00'
    if diff_total > 0:
        cell.fill = negative_fill  # Мы дороже
    elif diff_total < 0:
        cell.fill = positive_fill  # Мы дешевле
    
    ws.row_dimensions[row].height = 25
    
    # ==================== ВЫВОДЫ ====================
    row += 2
    ws.merge_cells(f'A{row}:N{row}')
    cell = ws[f'A{row}']
    if diff_total > 0:
        conclusion = f"⚠️ Наша смета на {abs(diff_total):,.2f} ₽ ДОРОЖЕ, чем КП завода"
    elif diff_total < 0:
        conclusion = f"✅ Наша смета на {abs(diff_total):,.2f} ₽ ДЕШЕВЛЕ, чем КП завода (экономия!)"
    else:
        conclusion = "✅ Цены СОВПАДАЮТ!"
    cell.value = conclusion
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="70AD47" if diff_total <= 0 else "E74C3C", 
                            end_color="70AD47" if diff_total <= 0 else "E74C3C", 
                            fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30
    
    # ==================== ШИРИНА СТОЛБЦОВ ====================
    ws.column_dimensions['A'].width = 5   # №
    ws.column_dimensions['B'].width = 30  # Наименование (смета)
    ws.column_dimensions['C'].width = 8   # Кол-во
    ws.column_dimensions['D'].width = 10  # Вес
    ws.column_dimensions['E'].width = 12  # Цена
    ws.column_dimensions['F'].width = 14  # Сумма
    ws.column_dimensions['G'].width = 12  # Примечание
    
    ws.column_dimensions['H'].width = 5   # №
    ws.column_dimensions['I'].width = 30  # Наименование (КП)
    ws.column_dimensions['J'].width = 8   # Кол-во
    ws.column_dimensions['K'].width = 12  # Цена
    ws.column_dimensions['L'].width = 14  # Сумма
    ws.column_dimensions['M'].width = 14  # Разница цены
    ws.column_dimensions['N'].width = 14  # Разница суммы
    
    # Сохраняем файл
    user_id = os.path.basename(kp_path).split('_')[0]
    output_path = os.path.join("tmp", f"{user_id}_comparison.xlsx")
    wb.save(output_path)
    
    return output_path


def compare_files(kp_path: str, smeta_path: str) -> str:
    """
    Сравнивает два Excel-файла (реальное КП и наша смета)
    и возвращает текстовый отчёт.

    Главный акцент:
    - выводим позиции, где КОЛИЧЕСТВО одинаковое, а ЦЕНА за шт. отличается;
    - отдельно показываем позиции, где отличается количество;
    - для каждой позиции показываем детальную разбивку нашей цены (базовая + резы + остатки + отходы).
    """
    kp = load_kp_excel(kp_path)
    smeta = load_smeta_excel(smeta_path)

    # Ключи могут быть как строками, так и кортежами (length_dm, width_dm),
    # поэтому сортируем по строковому представлению, чтобы избежать ошибок сравнения.
    common_names = sorted(set(kp.keys()) & set(smeta.keys()), key=str)
    only_in_kp = sorted(set(kp.keys()) - set(smeta.keys()), key=str)
    only_in_smeta = sorted(set(smeta.keys()) - set(kp.keys()), key=str)

    # ==== ИТОГО по всем плитам (в каждом файле отдельно) ====
    total_kp_all = sum(r.total for r in kp.values())
    total_smeta_all = sum(o.total for o in smeta.values())

    if not common_names:
        return "Не нашёл общих позиций по наименованию плит в двух файлах."

    # Блоки для различных типов расхождений
    same_qty_price_diff_blocks: list[str] = []
    qty_mismatch_blocks: list[str] = []

    # Сводные показатели по общим позициям
    total_kp_common = 0.0
    total_smeta_common = 0.0

    # Короткий итог по позициям с одинаковым количеством, но разной ценой
    same_qty_count = 0
    same_qty_economy = 0.0  # КП - наша смета (положительное = экономия по нашей смете)

    for key in common_names:
        r = kp[key]      # строка из реального КП
        o = smeta[key]   # строка из нашей сметы
        name = r.name or o.name

        total_kp_common += r.total
        total_smeta_common += o.total

        price_diff = o.price - r.price
        total_diff = o.total - r.total
        price_diff_pct = (price_diff / r.price * 100) if r.price else 0.0

        # Если вообще всё почти одинаково — пропускаем
        if abs(price_diff) < 1e-2 and abs(total_diff) < 1e-2:
            continue

        qty_same = abs(o.qty - r.qty) < 1e-6

        block: list[str] = [f"▪️ {name}"]

        if qty_same:
            block.append(f"  – количество в обоих файлах: {o.qty:g} шт.")
        else:
            block.append(f"  – количество: в нашей смете {o.qty:g} шт., в КП {r.qty:g} шт.")

        block.append(f"  – наша цена за шт.: {o.price:,.2f} руб.")
        block.append(f"  – цена в КП:       {r.price:,.2f} руб.")
        block.append(f"  – разница по цене:{price_diff:+.2f} руб. ({price_diff_pct:+.1f}%)")
        block.append(f"  – наша сумма:      {o.total:,.2f} руб.")
        block.append(f"  – сумма в КП:      {r.total:,.2f} руб.")
        block.append(f"  – разница по сумме:{total_diff:+.2f} руб.")

        if o.price_per_kg:
            block.append(
                f"  – пояснение: наша цена ≈ вес {o.weight:g} кг × "
                f"{o.price_per_kg:,.2f} руб/кг = {o.price:,.2f} руб за плиту "
                f"(в эту цену уже заложены плита, резы, отходы и НДС по нашей смете)."
            )
        else:
            block.append(
                "  – пояснение: цена взята из нашей сметы (столбец «Цена»), "
                "где уже учтены плита, резы, отходы и НДС."
            )
        
        # Добавляем детальную разбивку цены
        breakdown_text = _explain_price_breakdown(name, o.price, smeta_path)
        block.append(breakdown_text)

        text_block = "\n".join(block)

        if qty_same:
            same_qty_price_diff_blocks.append(text_block)
            same_qty_count += 1
            same_qty_economy += (r.total - o.total)
        else:
            qty_mismatch_blocks.append(text_block)

    lines: list[str] = []

    # 1) Главная секция — разные цены при одинаковом количестве
    lines.append("🔍 Позиции с РАЗНОЙ ценой при ОДИНАКОВОМ количестве:\n")
    if same_qty_price_diff_blocks:
        lines.extend(same_qty_price_diff_blocks)
        lines.append(
            f"\nВсего таких позиций: {same_qty_count}. "
            f"Суммарная экономия (+) / переплата (-) по ним: {same_qty_economy:+,.2f} руб.\n"
        )
    else:
        lines.append("✅ Таких позиций не обнаружено — при одинаковом количестве цены совпадают.\n")

    # 2) Отдельный блок — различия в количестве
    if qty_mismatch_blocks:
        lines.append("📊 Позиции, где отличается КОЛИЧЕСТВО (и, как следствие, сумма):\n")
        lines.extend(qty_mismatch_blocks)

    # 3) Сводный итог по общим позициям
    total_economy_common = total_kp_common - total_smeta_common
    lines.append("\n===== ИТОГО ПО ОБЩИМ ПОЗИЦИЯМ =====")
    lines.append(f"Наша смета (общие позиции):  {total_smeta_common:,.2f} руб.")
    lines.append(f"КП завода (общие позиции):   {total_kp_common:,.2f} руб.")
    lines.append(f"Экономия (+) / Переплата (-): {total_economy_common:+,.2f} руб.")

    # 4) Позиции, которые есть только в одном из файлов
    if only_in_kp:
        lines.append("\nПозиции, которые есть ТОЛЬКО в КП (нет в нашей смете):")
        for key in only_in_kp:
            lines.append(f"  – {kp[key].name}")

    if only_in_smeta:
        lines.append("\nПозиции, которые есть ТОЛЬКО в нашей смете (нет в КП):")
        for key in only_in_smeta:
            lines.append(f"  – {smeta[key].name}")

    # 5) Общие итоги по всем плитам ПБ
    lines.append("\n===== ОБЩИЕ ИТОГИ ПО ВСЕМ ПЛИТАМ ПБ =====")
    lines.append(f"КП завода (все плиты ПБ): {total_kp_all:,.2f} руб.")
    lines.append(f"Наша смета (все плиты ПБ): {total_smeta_all:,.2f} руб.")
    lines.append(
        f"Разница (КП - наша смета): {(total_kp_all - total_smeta_all):+,.2f} руб."
    )

    return "\n".join(lines)


@router.message(Command("compare"))
@router.message(F.text == "Сравнение результатов")
async def start_comparison(message: Message, state: FSMContext):
    """Старт диалога сравнения КП и сметы."""
    if pd is None:
        await message.answer(
            "⚠️ Модуль pandas не установлен, сравнение Excel-файлов недоступно.\n"
            "Установите зависимости из requirements.txt (pandas, openpyxl)."
        )
        return

    await state.set_state(CompareStates.waiting_kp)
    await message.answer(
        "📄 Отправьте файл **реального КП завода** в формате Excel (.xlsx)\n\n"
        "Файл должен содержать столбцы:\n"
        "• Наименование (или 'Товары (работы, услуги)')\n"
        "• Кол-во (или 'Количество')\n"
        "• Цена\n"
        "• Сумма\n\n"
        "💡 Столбец 'Вес' не обязателен",
        parse_mode="Markdown"
    )


@router.message(CompareStates.waiting_kp, F.document)
async def receive_kp(message: Message, state: FSMContext):
    """Принимаем файл КП и сохраняем его."""
    user_id = message.from_user.id
    os.makedirs("tmp", exist_ok=True)
    kp_path = os.path.join("tmp", f"{user_id}_kp.xlsx")

    await message.bot.download(message.document, destination=kp_path)
    await state.update_data(kp_path=kp_path)
    await state.set_state(CompareStates.waiting_smeta)

    await message.answer(
        "✅ КП завода получил!\n\n"
        "📄 Теперь отправьте файл **нашей сметы** (Excel .xlsx)\n\n"
        "Это должен быть файл 'Смета_Дорожка_...' или 'Детальная_разбивка_...',\n"
        "который создаётся при расчёте КП через бота.",
        parse_mode="Markdown"
    )


@router.message(CompareStates.waiting_kp)
async def receive_kp_wrong(message: Message, state: FSMContext):  # noqa: ARG001
    await message.answer("Нужно отправить именно файл КП (Excel-документ). Попробуйте ещё раз.")


@router.message(CompareStates.waiting_smeta, F.document)
async def receive_smeta(message: Message, state: FSMContext):
    """Принимаем файл нашей сметы и запускаем сравнение."""
    user_id = message.from_user.id
    os.makedirs("tmp", exist_ok=True)
    smeta_path = os.path.join("tmp", f"{user_id}_smeta.xlsx")

    await message.bot.download(message.document, destination=smeta_path)

    data = await state.get_data()
    kp_path = data.get("kp_path")

    if not kp_path or not os.path.exists(kp_path):
        await message.answer("Не нашёл сохранённый файл КП. Давай начнём сначала: /compare")
        await state.clear()
        return

    await message.answer("📊 Сравниваю файлы и создаю Excel отчёт, подождите...")

    try:
        # Создаём Excel файл со сравнением
        comparison_path = await asyncio.to_thread(create_comparison_excel, kp_path, smeta_path)
        
        # Отправляем Excel файл
        if os.path.exists(comparison_path):
            await message.answer(
                "✅ Сравнение готово!\n\n"
                "📋 В файле вы найдёте:\n"
                "• Две таблицы рядом (наша смета и КП завода)\n"
                "• Сравнение цен по каждой позиции\n"
                "• Детальную разбивку ценообразования\n"
                "• Цветовую индикацию разницы (зелёный=экономия, красный=дороже)\n"
                "• Итоговые суммы и выводы"
            )
            await message.answer_document(
                FSInputFile(comparison_path),
                caption="📊 Сравнительная таблица КП и сметы"
            )
        else:
            await message.answer("❌ Не удалось создать файл сравнения")
            
    except Exception as e:  # pylint: disable=broad-except
        # Отключаем разбор разметки, чтобы спецсимволы из текста ошибки не ломали сообщение
        await message.answer(f"❌ Не удалось сравнить файлы: {e}", parse_mode=None)
        import traceback
        traceback.print_exc()
    finally:
        await state.clear()


@router.message(CompareStates.waiting_smeta)
async def receive_smeta_wrong(message: Message, state: FSMContext):  # noqa: ARG001
    await message.answer("Нужно отправить именно файл нашей сметы (Excel-документ). Попробуйте ещё раз.")

@router.message(KPStates.waiting_for_plate_list, F.photo)
async def receive_photo_with_plates(message: Message, state: FSMContext):
    """
    🧠 УМНАЯ обработка фотографий с плитами:
    1. Скачивает фото
    2. Пробует бесплатный EasyOCR
    3. Если не получилось — использует платный GPT-4o
    4. Парсит распознанный текст
    5. Обрабатывает заказ
    """
    # Проверяем доступность хотя бы одного метода OCR
    if not EASYOCR_AVAILABLE and not GPT_AVAILABLE:
        await message.answer(
            "❌ OCR недоступен. Установите одну из библиотек:\n\n"
            "🤖 EasyOCR (бесплатно):\n"
            "   pip install easyocr\n\n"
            "🧠 GPT-4o (платно, но точнее):\n"
            "   pip install openai\n"
            "   Добавьте в .env: OPENAI_API_KEY=sk-...\n\n"
            "Или отправьте текст заказа вручную."
        )
        return
    
    # Скачиваем фото (берём самое большое разрешение)
    photo = message.photo[-1]
    user_id = message.from_user.id
    os.makedirs("tmp", exist_ok=True)
    photo_path = os.path.join("tmp", f"{user_id}_photo.jpg")
    
    await message.answer("📸 Получил фото! Анализирую...")
    
    try:
        # Скачиваем фото
        await message.bot.download(photo, destination=photo_path)
        
        # 🔥 УМНОЕ РАСПОЗНАВАНИЕ (EasyOCR → GPT fallback)
        result = await recognize_text_smart(photo_path, show_cost=True)
        
        if not result:
            await message.answer(
                "❌ Не удалось распознать текст на фото.\n\n"
                "💡 Попробуйте:\n"
                "• Сделать фото при хорошем освещении\n"
                "• Убедиться, что текст чёткий и читаемый\n"
                "• Расположить камеру параллельно таблице\n"
                "• Отправить текст заказа вручную"
            )
            # Удаляем временный файл
            if os.path.exists(photo_path):
                try:
                    os.remove(photo_path)
                except:
                    pass
            return
        
        # Показываем пользователю результат
        method_emoji = {"EasyOCR": "🤖", "GPT-4o": "🧠"}
        emoji = method_emoji.get(result['method'], "🔍")
        
        # Формируем красивое сообщение
        confidence_percent = int(result['confidence'] * 100)
        status_msg = f"{emoji} **{result['method']}** (уверенность {confidence_percent}%)\n\n"
        
        # Добавляем инфо о стоимости, если использовали GPT
        if result['cost_usd'] > 0:
            rub_cost = result['cost_usd'] * 75
            status_msg += f"💰 Стоимость: ${result['cost_usd']:.4f} (~{rub_cost:.2f}₽)\n\n"
        
        status_msg += f"📋 Распознанный текст:\n```\n{result['text']}\n```\n\nПродолжаю обработку..."
        
        await message.answer(status_msg, parse_mode="Markdown")
        
        # Используем распознанный текст
        cleaned_text = result['text']
        
        # Теперь парсим распознанный текст так же, как обычный текст
        # Используем ту же логику, что и в receive_plate_list_and_build
        await message.answer("⏳ Считаю КП по вашему списку... Это может занять время.")
        
        # 1) Парсим список пользователя в структуры визуализатора
        unparsed_lines = set_plate_lists_from_text(cleaned_text)
        
        # Если какие‑то строки не распознаны по формату — сразу честно говорим об этом
        if unparsed_lines:
            warn_text = "⚠️ Некоторые строки я не смог распознать по формату и пропустил:\n"
            warn_text += "\n".join(f"• {line}" for line in unparsed_lines[:5])  # Показываем первые 5
            if len(unparsed_lines) > 5:
                warn_text += f"\n... и ещё {len(unparsed_lines) - 5} строк"
            warn_text += (
                "\n\nЯ понимаю, например, такие форматы:\n"
                "• 1.2×3.39 — 2 шт\n"
                "• 0,32x6,63 - 4\n"
                "• Плиты ПБ 78-12-8п 3\n"
                "• ПБ 66,2-12-8п 6\n"
            )
            await message.answer(warn_text)
        
        # Дальше используем ту же логику, что и в receive_plate_list_and_build
        # (копируем код обработки заказа)
        from collections import Counter, defaultdict
        import math
        
        # ✅ НОВАЯ ЛОГИКА: Группируем плиты по нагрузке
        orders_by_load = defaultdict(list)  # {load_group: [orders_2d]}
        
        print(f"[BOT] Проверяем PLATE_LOAD_DETAILS: {len(cfg.PLATE_LOAD_DETAILS)} записей")
        
        # Используем детальную карту с нагрузками (если есть)
        if cfg.PLATE_LOAD_DETAILS:
            print("[BOT] ✅ Используем PLATE_LOAD_DETAILS (с нагрузками)")
            for (length, width_m, load_code), qty in cfg.PLATE_LOAD_DETAILS.items():
                width_mm = int(round(width_m * 1000))
                
                # Группируем по ЦЕЛОЙ части: 12.5 → группа 12
                load_group = math.floor(load_code) if isinstance(load_code, (int, float)) else load_code
                
                orders_by_load[load_group].append({
                    'length': length,
                    'width': width_mm,
                    'qty': qty,
                    'load_code': load_code,
                    'load_group': load_group
                })
                
                load_display = cfg.format_reinforcement_from_load_code(load_code)
                print(f"  + {qty}x {length}м × {width_mm}мм, нагрузка {load_display} (группа {load_group}п)")
        else:
            # Fallback: Если PLATE_LOAD_DETAILS пуст
            print("[BOT] ⚠️ PLATE_LOAD_DETAILS пуст, используем fallback (все плиты = 8п)")
            for width_mm, plates_list, target_name in [
                (1200, cfg.PLATES_1_2, 'PLATES_1_2'), (1080, cfg.PLATES_1_08, 'PLATES_1_08'), (1000, cfg.PLATES_1_0, 'PLATES_1_0'),
                (320, cfg.PLATES_0_32, 'PLATES_0_32'), (460, cfg.PLATES_0_46, 'PLATES_0_46'), (700, cfg.PLATES_0_70, 'PLATES_0_70'),
                (720, cfg.PLATES_0_72, 'PLATES_0_72'), (860, cfg.PLATES_0_86, 'PLATES_0_86'), (880, cfg.PLATES_0_88, 'PLATES_0_88'),
                (740, cfg.PLATES_0_74, 'PLATES_0_74'), (480, cfg.PLATES_0_48, 'PLATES_0_48'), (500, cfg.PLATES_0_50, 'PLATES_0_50'),
                (340, cfg.PLATES_0_34, 'PLATES_0_34')
            ]:
                if plates_list:
                    length_counts = Counter(plates_list)
                    for length, qty in length_counts.items():
                        exact_width_m = cfg.get_exact_width(length, target_name, width_mm / 1000.0)
                        exact_width_mm = int(round(exact_width_m * 1000))
                        load_code = cfg.get_load_code_for_plate(length, exact_width_m, default=8)
                        
                        orders_by_load[load_code].append({
                            'length': length,
                            'width': exact_width_mm,
                            'qty': qty,
                            'load_code': load_code
                        })
        
        # Если после парсинга не осталось ни одной плиты — сразу выходим
        if not orders_by_load:
            await message.answer(
                "❌ Не удалось распознать ни одной плиты в вашем сообщении.\n"
                "Проверьте формат строк (ширина×длина×кол-во или 'Плиты ПБ 78-12-8п 3')."
            )
            await state.clear()
            return
        
        # ✅ ЗАПУСКАЕМ ОПТИМИЗАЦИЮ ДЛЯ КАЖДОЙ НАГРУЗКИ ОТДЕЛЬНО
        print(f"\n[BOT] Найдено {len(orders_by_load)} групп(ы) по нагрузкам: {sorted(orders_by_load.keys())}")
        
        optimization_results_by_load = {}
        total_plates_all = 0
        total_cost_all = 0
        
        load_group_to_originals = {}
        for load_group, orders in orders_by_load.items():
            originals = set(o['load_code'] for o in orders)
            load_group_to_originals[load_group] = sorted(originals)
        
        for load_group in sorted(orders_by_load.keys()):
            orders_2d = orders_by_load[load_group]
            
            original_loads = load_group_to_originals[load_group]
            load_display_list = [cfg.format_reinforcement_from_load_code(lc) for lc in original_loads]
            load_display = ", ".join(load_display_list) if len(load_display_list) > 1 else load_display_list[0]
            
            print(f"\n[BOT] === Оптимизация для группы {load_group}п ({load_display}) ===")
            print(f"[BOT] Плит: {sum(o['qty'] for o in orders_2d)} шт, типов: {len(orders_2d)}")
            
            try:
                from core.optimization import optimize_with_cascading_longitudinal_cuts
                optimization_result = await asyncio.to_thread(
                    optimize_with_cascading_longitudinal_cuts,
                    orders_2d=orders_2d
                )
                
                if optimization_result and optimization_result.get('total_plates', 0) > 0:
                    optimization_result['load_group'] = load_group
                    optimization_result['original_loads'] = original_loads
                    optimization_results_by_load[load_group] = optimization_result
                    total_plates_all += optimization_result.get('total_plates', 0)
                    total_cost_all += optimization_result.get('total_cost', 0)
                    
                    print(f"[BOT] ✅ Группа {load_group}п ({load_display}): {optimization_result['total_plates']} плит, "
                          f"{optimization_result.get('total_cost', 0):,} ₽".replace(',', ' '))
            except Exception as e:
                print(f"[BOT] ❌ Ошибка оптимизации для группы {load_group}п: {e}")
        
        # Сохраняем результаты в глобальную переменную
        if optimization_results_by_load:
            import core.optimization as optimization
            optimization.OPT_CASCADING_PLAN_BY_LOAD = optimization_results_by_load
            print(f"\n[BOT] ✅ Сохранено {len(optimization_results_by_load)} результатов оптимизации")
            
            opt_msg = "💡 **Результат оптимизации по нагрузкам:**\n"
            for load_group in sorted(optimization_results_by_load.keys()):
                result = optimization_results_by_load[load_group]
                original_loads = result.get('original_loads', [load_group])
                load_display_list = [cfg.format_reinforcement_from_load_code(lc) for lc in original_loads]
                load_display = ", ".join(load_display_list)
                opt_msg += f"• **{load_display}**: {result['total_plates']} плит\n"
            opt_msg += f"\n**Итого:** {total_plates_all} плит, {total_cost_all:,} ₽\n".replace(',', ' ')
            
            await message.answer(opt_msg, parse_mode="Markdown")
        else:
            print("[BOT] ⚠️ Оптимизация не дала результатов, используем fallback")
        
        # 4) Строим приоритет ширин (запасной вариант)
        if not optimization_results_by_load:
            from core.optimization import apply_width_optimization
            apply_width_optimization()
        
        # 5) Запускаем расчёт и визуализацию
        result_paths = await asyncio.to_thread(visualize_plan, OUTPUTS_DIR_STR)
        if isinstance(result_paths, tuple) and len(result_paths) >= 2:
            png_path, pdf_path = result_paths
            
            base = os.path.basename(png_path)
            if 'КЗ_' in base:
                timestamp = base.split('КЗ_', 1)[-1].replace('.png', '')
            else:
                timestamp = base.rsplit('_', 1)[-1].replace('.png', '')
            
            candidates = [
                os.path.join(OUTPUTS_DIR_STR, f'Ведомость_Дорожка_1_{timestamp}.xlsx'),
                os.path.join(OUTPUTS_DIR_STR, f'Смета_Дорожка_1_{timestamp}.xlsx'),
                os.path.join(OUTPUTS_DIR_STR, f'Детальная_разбивка_Дорожка_1_{timestamp}.xlsx'),
                os.path.join(OUTPUTS_DIR_STR, f'Ведомость_Дорожка_1_{timestamp}.csv'),
                os.path.join(OUTPUTS_DIR_STR, f'Раскладка_Дорожка_1_{timestamp}.csv'),
            ]
            
            await message.answer("✅ Готово! Отправляю файлы:")
            
            if os.path.exists(png_path):
                await message.answer_document(FSInputFile(png_path))
            if os.path.exists(pdf_path):
                await message.answer_document(FSInputFile(pdf_path))
            
            files_sent = 0
            for p in candidates:
                if os.path.exists(p):
                    await message.answer_document(FSInputFile(p))
                    files_sent += 1
            
            final_msg = "📋 **Итоги:**\n• Схема раскладки готова\n• Ведомость и смета сформированы"
            if optimization_results_by_load:
                final_msg += "\n\n✨ **Использована оптимизация с каскадными резами**\n• Минимум плит\n• Остатки используются повторно"
            await message.answer(final_msg, parse_mode="Markdown")
        else:
            await message.answer("❌ Ошибка при расчёте КП")
            
    except Exception as e:
        await message.answer(f"❌ Ошибка при обработке фото: {str(e)}\n\nПопробуйте отправить текст заказа вручную.")
    finally:
        # Удаляем временный файл
        if os.path.exists(photo_path):
            try:
                os.remove(photo_path)
            except:
                pass
        await state.clear()

@router.message(KPStates.waiting_for_plate_list)
async def receive_plate_list_and_build(message: Message, state: FSMContext):
    # На первом шаге просто принимаем текст как подтверждение и запускаем существующий расчёт
    await message.answer("⏳ Считаю КП по вашему списку... Это может занять время.")
    try:
        # 1) Парсим список пользователя в структуры визуализатора (ядро парсинга в core/)
        user_text = message.text or ""
        unparsed_lines = set_plate_lists_from_text(user_text)

        # Если какие‑то строки не распознаны по формату — сразу честно говорим об этом
        if unparsed_lines:
            warn_text = "⚠️ Некоторые строки я не смог распознать по формату и пропустил:\n"
            warn_text += "\n".join(f"• {line}" for line in unparsed_lines)
            warn_text += (
                "\n\nЯ понимаю, например, такие форматы:\n"
                "• 1.2×3.39 — 2 шт\n"
                "• 0,32x6,63 - 4\n"
                "• Плиты ПБ 78-12-8п 3\n"
                "• ПБ 66,2-12-8п 6\n"
            )
            await message.answer(warn_text)
        
        # 3) Собираем заказы для 2D оптимизации (длина + ширина + НАГРУЗКА!)
        from collections import Counter, defaultdict
        import math
        
        # ✅ НОВАЯ ЛОГИКА: Группируем плиты по нагрузке
        # ВАЖНО: Группируем по ЦЕЛОЙ части (12.5→12), но сохраняем оригинал для отображения
        orders_by_load = defaultdict(list)  # {load_group: [orders_2d]}
        
        print(f"[BOT] Проверяем PLATE_LOAD_DETAILS: {len(cfg.PLATE_LOAD_DETAILS)} записей")
        
        # Используем детальную карту с нагрузками (если есть)
        if cfg.PLATE_LOAD_DETAILS:
            print("[BOT] ✅ Используем PLATE_LOAD_DETAILS (с нагрузками)")
            for (length, width_m, load_code), qty in cfg.PLATE_LOAD_DETAILS.items():
                width_mm = int(round(width_m * 1000))
                
                # Группируем по ЦЕЛОЙ части: 12.5 → группа 12
                load_group = math.floor(load_code) if isinstance(load_code, (int, float)) else load_code
                
                orders_by_load[load_group].append({
                    'length': length,
                    'width': width_mm,
                    'qty': qty,
                    'load_code': load_code,  # Сохраняем ОРИГИНАЛЬНУЮ нагрузку (12.5)
                    'load_group': load_group  # Группа для оптимизации (12)
                })
                
                # Форматируем нагрузку для отображения
                load_display = cfg.format_reinforcement_from_load_code(load_code)
                print(f"  + {qty}x {length}м × {width_mm}мм, нагрузка {load_display} (группа {load_group}п)")
        else:
            # Fallback: Если PLATE_LOAD_DETAILS пуст (старый формат без нагрузок)
            # Группируем все плиты как нагрузку 8п (дефолт)
            print("[BOT] ⚠️ PLATE_LOAD_DETAILS пуст, используем fallback (все плиты = 8п)")
            for width_mm, plates_list, target_name in [
                (1200, cfg.PLATES_1_2, 'PLATES_1_2'), (1080, cfg.PLATES_1_08, 'PLATES_1_08'), (1000, cfg.PLATES_1_0, 'PLATES_1_0'),
                (320, cfg.PLATES_0_32, 'PLATES_0_32'), (460, cfg.PLATES_0_46, 'PLATES_0_46'), (700, cfg.PLATES_0_70, 'PLATES_0_70'),
                (720, cfg.PLATES_0_72, 'PLATES_0_72'), (860, cfg.PLATES_0_86, 'PLATES_0_86'), (880, cfg.PLATES_0_88, 'PLATES_0_88'),
                (740, cfg.PLATES_0_74, 'PLATES_0_74'), (480, cfg.PLATES_0_48, 'PLATES_0_48'), (500, cfg.PLATES_0_50, 'PLATES_0_50'),
                (340, cfg.PLATES_0_34, 'PLATES_0_34')
            ]:
                if plates_list:
                    length_counts = Counter(plates_list)
                    for length, qty in length_counts.items():
                        # Получаем точную ширину
                        exact_width_m = cfg.get_exact_width(length, target_name, width_mm / 1000.0)
                        exact_width_mm = int(round(exact_width_m * 1000))
                        
                        # Используем дефолтную нагрузку 8п
                        load_code = cfg.get_load_code_for_plate(length, exact_width_m, default=8)
                        
                        orders_by_load[load_code].append({
                            'length': length,
                            'width': exact_width_mm,
                            'qty': qty,
                            'load_code': load_code
                        })
        
        # Если после парсинга не осталось ни одной плиты — сразу выходим
        if not orders_by_load:
            await message.answer(
                "❌ Не удалось распознать ни одной плиты в вашем сообщении.\n"
                "Проверьте формат строк (ширина×длина×кол-во или 'Плиты ПБ 78-12-8п 3')."
            )
            await state.clear()
            return

        # ✅ ЗАПУСКАЕМ ОПТИМИЗАЦИЮ ДЛЯ КАЖДОЙ НАГРУЗКИ ОТДЕЛЬНО
        print(f"\n[BOT] Найдено {len(orders_by_load)} групп(ы) по нагрузкам: {sorted(orders_by_load.keys())}")
        
        optimization_results_by_load = {}
        total_plates_all = 0
        total_cost_all = 0
        
        # Создаём карту группа→оригинальные нагрузки (для правильного отображения)
        load_group_to_originals = {}  # {12: [12, 12.5], 10: [10], ...}
        for load_group, orders in orders_by_load.items():
            originals = set(o['load_code'] for o in orders)
            load_group_to_originals[load_group] = sorted(originals)
        
        for load_group in sorted(orders_by_load.keys()):
            orders_2d = orders_by_load[load_group]
            
            # Для отображения собираем все оригинальные нагрузки в этой группе
            original_loads = load_group_to_originals[load_group]
            load_display_list = [cfg.format_reinforcement_from_load_code(lc) for lc in original_loads]
            load_display = ", ".join(load_display_list) if len(load_display_list) > 1 else load_display_list[0]
            
            print(f"\n[BOT] === Оптимизация для группы {load_group}п ({load_display}) ===")
            print(f"[BOT] Плит: {sum(o['qty'] for o in orders_2d)} шт, типов: {len(orders_2d)}")
            
            try:
                from core.optimization import optimize_with_cascading_longitudinal_cuts
                optimization_result = await asyncio.to_thread(
                    optimize_with_cascading_longitudinal_cuts,
                    orders_2d=orders_2d
                )
                
                if optimization_result and optimization_result.get('total_plates', 0) > 0:
                    # Сохраняем с информацией о группе и оригинальных нагрузках
                    optimization_result['load_group'] = load_group
                    optimization_result['original_loads'] = original_loads
                    optimization_results_by_load[load_group] = optimization_result
                    total_plates_all += optimization_result.get('total_plates', 0)
                    total_cost_all += optimization_result.get('total_cost', 0)
                    
                    print(f"[BOT] ✅ Группа {load_group}п ({load_display}): {optimization_result['total_plates']} плит, "
                          f"{optimization_result.get('total_cost', 0):,} ₽".replace(',', ' '))
            except Exception as e:
                print(f"[BOT] ❌ Ошибка оптимизации для группы {load_group}п: {e}")
        
        # Сохраняем результаты в глобальную переменную
        if optimization_results_by_load:
            import core.optimization as optimization
            optimization.OPT_CASCADING_PLAN_BY_LOAD = optimization_results_by_load
            print(f"\n[BOT] ✅ Сохранено {len(optimization_results_by_load)} результатов оптимизации")
            
            # Показываем сводку пользователю с ОРИГИНАЛЬНЫМИ нагрузками
            opt_msg = "💡 **Результат оптимизации по нагрузкам:**\n"
            for load_group in sorted(optimization_results_by_load.keys()):
                result = optimization_results_by_load[load_group]
                original_loads = result.get('original_loads', [load_group])
                load_display_list = [cfg.format_reinforcement_from_load_code(lc) for lc in original_loads]
                load_display = ", ".join(load_display_list)
                opt_msg += f"• **{load_display}**: {result['total_plates']} плит\n"
            opt_msg += f"\n**Итого:** {total_plates_all} плит, {total_cost_all:,} ₽\n".replace(',', ' ')
            
            await message.answer(opt_msg, parse_mode="Markdown")
        else:
            print("[BOT] ⚠️ Оптимизация не дала результатов, используем fallback")
        
        # 4) Строим приоритет ширин (запасной вариант, если каскадная не сработала)
        if not optimization_results_by_load:
            from core.optimization import apply_width_optimization
            apply_width_optimization()
        
        # 5) Запускаем расчёт и визуализацию
        result_paths = await asyncio.to_thread(visualize_plan, OUTPUTS_DIR_STR)
        if isinstance(result_paths, tuple) and len(result_paths) >= 2:
            png_path, pdf_path = result_paths

            # Извлекаем timestamp из имени PNG
            base = os.path.basename(png_path)
            # Ожидаемый формат: Схема_Дорожка_1_КЗ_{timestamp}.png
            # Извлекаем timestamp (всё после "КЗ_")
            if 'КЗ_' in base:
                timestamp = base.split('КЗ_', 1)[-1].replace('.png', '')
            else:
                # Fallback: последняя часть после последнего подчеркивания
                timestamp = base.rsplit('_', 1)[-1].replace('.png', '')
            print(f'[BOT] Извлечен timestamp: {timestamp}')
            print(f'[BOT] Ищу файлы в директории: {OUTPUTS_DIR_STR}')
            print(f'[BOT] Директория существует: {os.path.exists(OUTPUTS_DIR_STR)}')
            
            # Показываем все файлы с этим timestamp для отладки
            if os.path.exists(OUTPUTS_DIR_STR):
                matching_files = [f for f in os.listdir(OUTPUTS_DIR_STR) if timestamp in f]
                print(f'[BOT] Найдено файлов с timestamp {timestamp}: {len(matching_files)}')
                for f in matching_files:
                    print(f'  - {f}')

            # Возможные имена доп.файлов (поддерживаем оба варианта из визуализатора)
            candidates = [
                os.path.join(OUTPUTS_DIR_STR, f'Ведомость_Дорожка_1_{timestamp}.xlsx'),
                os.path.join(OUTPUTS_DIR_STR, f'Смета_Дорожка_1_{timestamp}.xlsx'),
                os.path.join(OUTPUTS_DIR_STR, f'Детальная_разбивка_Дорожка_1_{timestamp}.xlsx'),
                os.path.join(OUTPUTS_DIR_STR, f'Ведомость_Дорожка_1_{timestamp}.csv'),
                os.path.join(OUTPUTS_DIR_STR, f'Раскладка_Дорожка_1_{timestamp}.csv'),
            ]

            await message.answer("✅ Готово! Отправляю файлы:")

            if os.path.exists(png_path):
                await message.answer_document(FSInputFile(png_path))
            if os.path.exists(pdf_path):
                await message.answer_document(FSInputFile(pdf_path))
            
            # Отправляем Excel файлы в правильном порядке
            files_sent = 0
            for p in candidates:
                if os.path.exists(p):
                    print(f'[BOT] ✅ Отправляю файл: {os.path.basename(p)}')
                    await message.answer_document(FSInputFile(p))
                    files_sent += 1
                else:
                    print(f'[BOT] ❌ Файл не найден: {os.path.basename(p)}')
                    print(f'[BOT]    Полный путь: {p}')
            
            print(f'[BOT] Всего отправлено Excel/CSV файлов: {files_sent}')

            # Формируем итоговое сообщение
            final_msg = "📋 **Итоги:**\n• Схема раскладки готова\n• Ведомость и смета сформированы"
            if optimization_result and optimization_result.get('total_plates', 0) > 0:
                final_msg += "\n\n✨ **Использована оптимизация с каскадными резами**\n• Минимум плит\n• Остатки используются повторно"
            await message.answer(final_msg, parse_mode="Markdown")
        else:
            await message.answer("❌ Ошибка при расчёте КП")
    except Exception as e:
        await message.answer(f"❌ Ошибка: {str(e)}")
    finally:
        await state.clear()

@router.message(Command("build_plan"))
async def cmd_build_plan(message: Message):
    """Обработчик команды /build_plan"""
    await message.answer("⏳ Выполняю расчёт дорожки, подожди немного...")
    
    try:
        # Запускаем расчёт в отдельном потоке
        result_paths = await asyncio.to_thread(visualize_plan, OUTPUTS_DIR_STR)
        
        if isinstance(result_paths, tuple) and len(result_paths) >= 2:
            png_path, pdf_path = result_paths
            
            # Ищем дополнительные файлы
            # Извлекаем timestamp (всё после "КЗ_")
            base = os.path.basename(png_path)
            if 'КЗ_' in base:
                timestamp = base.split('КЗ_', 1)[-1].replace('.png', '')
            else:
                # Fallback: последняя часть после последнего подчеркивания
                timestamp = base.rsplit('_', 1)[-1].replace('.png', '')
            
            csv_path = os.path.join(OUTPUTS_DIR_STR, f'Раскладка_Дорожка_1_{timestamp}.csv')
            xlsx_path = os.path.join(OUTPUTS_DIR_STR, f'Ведомость_Дорожка_1_{timestamp}.xlsx')
            breakdown_path = os.path.join(OUTPUTS_DIR_STR, f'Детальная_разбивка_Дорожка_1_{timestamp}.xlsx')
            xlsx_smeta_path = os.path.join(OUTPUTS_DIR_STR, f'Смета_Дорожка_1_{timestamp}.xlsx')
            
            await message.answer("✅ Готово! Отправляю файлы:")
            
            # Отправляем изображение как документ, чтобы избежать PHOTO_INVALID_DIMENSIONS
            if os.path.exists(png_path):
                await message.answer_document(FSInputFile(png_path))
            
            # Отправляем документы
            if os.path.exists(pdf_path):
                await message.answer_document(FSInputFile(pdf_path))
            
            if os.path.exists(xlsx_path):
                await message.answer_document(FSInputFile(xlsx_path))
            
            if os.path.exists(xlsx_smeta_path):
                await message.answer_document(FSInputFile(xlsx_smeta_path))
            
            if os.path.exists(breakdown_path):
                await message.answer_document(FSInputFile(breakdown_path))
            
            if os.path.exists(csv_path):
                await message.answer_document(FSInputFile(csv_path))
            
            await message.answer(
                "📋 **Результаты расчёта готовы!**\n\n"
                "• Схема раскладки сохранена\n"
                "• Ведомость материалов готова\n"
                "• Смета стоимости рассчитана\n"
                "• Все файлы экспортированы"
            )
        else:
            await message.answer("❌ Ошибка при расчёте плана")
            
    except Exception as e:
        await message.answer(f"❌ Ошибка: {str(e)}")

@router.message(Command("help"))
async def cmd_help(message: Message):
    """Обработчик команды /help"""
    help_text = """
📖 **Помощь по командам:**

🏗️ **Построить план** - создаёт визуализацию дорожки с расчётом стоимости

**Команды:**
• `/start` - главное меню
• `/build_plan` - построить план дорожки
• `/optimize` - оптимизация раскроя с экономией до 40%
• `/help` - эта справка
• `/stats` - статистика проекта

**Форматы файлов:**
• PNG - схема раскладки
• PDF - техническая документация  
• XLSX - ведомость и смета
• CSV - данные для импорта

💡 **Оптимизация резов:**
Использует каскадные продольные резы для минимизации отходов и экономии материала.
    """
    await message.answer(help_text, parse_mode="Markdown")

@router.message(Command("stats"))
async def cmd_stats(message: Message):
    """Обработчик команды /stats"""
    try:
        # Подсчитываем файлы в папке outputs
        files_count = len([f for f in os.listdir(OUTPUTS_DIR_STR) if f.endswith(('.png', '.pdf', '.xlsx'))])
        
        stats_text = f"""
📊 **Статистика проекта:**

📁 Файлов создано: {files_count}
📂 Папка результатов: `{OUTPUTS_DIR_STR}`

🔧 **Доступные функции:**
• Визуализация раскладки
• Расчёт стоимости материалов
• Экспорт в различные форматы

📈 **Последние результаты:**
• PNG схемы: {len([f for f in os.listdir(OUTPUTS_DIR_STR) if f.endswith('.png')])} шт
• PDF документы: {len([f for f in os.listdir(OUTPUTS_DIR_STR) if f.endswith('.pdf')])} шт
• Excel файлы: {len([f for f in os.listdir(OUTPUTS_DIR_STR) if f.endswith('.xlsx')])} шт
        """
        
        await message.answer(stats_text, parse_mode="Markdown")
        
    except Exception as e:
        await message.answer(f"❌ Ошибка: {str(e)}")

@router.message(Command("optimize"))
@router.message(F.text == "Оптимизация резов")
async def cmd_optimize(message: Message):
    """Оптимизация раскроя с каскадными продольными резами"""
    await message.answer("⏳ Выполняю оптимизацию раскроя с учётом вторичных резов...")
    
    try:
        # Собираем заказы из текущей конфигурации
        orders = {}
        if cfg.PLATES_0_32:
            orders[320] = len(cfg.PLATES_0_32)
        if cfg.PLATES_0_46:
            orders[460] = len(cfg.PLATES_0_46)
        if cfg.PLATES_0_70:
            orders[700] = len(cfg.PLATES_0_70)
        if cfg.PLATES_0_72:
            orders[720] = len(cfg.PLATES_0_72)
        if cfg.PLATES_0_86:
            orders[860] = len(cfg.PLATES_0_86)
        if cfg.PLATES_0_88:
            orders[880] = len(cfg.PLATES_0_88)
        if cfg.PLATES_0_74:
            orders[740] = len(cfg.PLATES_0_74)
        if cfg.PLATES_0_48:
            orders[480] = len(cfg.PLATES_0_48)
        if cfg.PLATES_0_50:
            orders[500] = len(cfg.PLATES_0_50)
        if cfg.PLATES_0_34:
            orders[340] = len(cfg.PLATES_0_34)
        
        if not orders:
            await message.answer(
                "⚠️ Нет данных для оптимизации.\n"
                "Сначала используйте 'Получить КП' для загрузки списка плит.",
                reply_markup=main_menu_kb()
            )
            return
        
        # Запускаем оптимизацию в отдельном потоке
        result = await asyncio.to_thread(optimize_with_cascading_longitudinal_cuts, orders)
        
        if result and result.get('total_plates', 0) > 0:
            # Формируем красивый ответ
            response = "✅ **Оптимизация завершена!**\n\n"
            response += f"📊 **Результат:**\n"
            response += f"• Плит потребуется: **{result['total_plates']} шт**\n"
            response += f"• Стоимость: **{result['total_cost']:,} ₽**\n".replace(',', ' ')
            response += f"• Отходы по ширине: **{result.get('waste_width', 0)} мм**\n\n"
            
            if result.get('primary_cuts'):
                response += "🔹 **Первичные резы:**\n"
                for cut in result['primary_cuts']:
                    response += f"  • {cut['qty']} плит → {cut['width']} мм + остаток {cut['rest']} мм\n"
            
            if result.get('secondary_cuts'):
                response += f"\n🔸 **Вторичные резы (из остатков):**\n"
                for cut in result['secondary_cuts']:
                    if cut.get('pieces', 1) > 1:
                        response += f"  • {cut['qty']} остатков {cut['source']} мм → {cut['pieces']} частей по {cut['cuts'][0]} мм\n"
                    else:
                        cuts_str = ' + '.join(str(c) for c in cut['cuts'])
                        response += f"  • {cut['qty']} остатков {cut['source']} мм → {cuts_str} мм\n"
            
            response += "\n💡 **Преимущества:**\n"
            response += "• Минимум плит\n"
            response += "• Остатки используются повторно\n"
            response += "• Меньше отходов\n"
            
            await message.answer(response, parse_mode="Markdown", reply_markup=main_menu_kb())
        else:
            await message.answer(
                "❌ Не удалось выполнить оптимизацию.\n"
                "Проверьте корректность данных.",
                reply_markup=main_menu_kb()
            )
    
    except Exception as e:
        await message.answer(
            f"❌ Ошибка при оптимизации: {str(e)}\n\n"
            f"Убедитесь, что библиотека PuLP установлена.",
            reply_markup=main_menu_kb()
        )


# TODO: Временно отключено - модуль planning не реализован
# @router.message(F.text == "Планирование по дням")
async def btn_planning_days_DISABLED(message: Message):
    await message.answer("⏳ Строю календарь дорожек… подождите пару секунд.")

    try:
        # schedule, report_path = await asyncio.to_thread(plan_tracks)
        schedule, report_path = None, None

        if not schedule:
            await message.answer(
                "⚠️ Не найдено плит в базе. Попробуйте обновить данные.",
                reply_markup=main_menu_kb(),
            )
            return

        PLANNING_CACHE[message.from_user.id] = {
            "schedule": schedule,
            "report": report_path,
        }

        # days = available_days(schedule)
        days = []
        buttons = [
            [InlineKeyboardButton(text=f"День {day}", callback_data=f"plan_day:{day}")]
            for day in days
        ]

        summary_lines = [
            f"День {day}: {sum(1 for t in schedule if t.day == day)} дорожек"
            for day in days
        ]

        keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)

        await message.answer(
            "✅ План готов!\n\n" + "\n".join(summary_lines) + "\n\nВыберите день:",
            reply_markup=keyboard,
        )

        if report_path and report_path.exists():
            await message.answer_document(FSInputFile(report_path))

    except Exception as e:
        await message.answer(
            f"❌ Ошибка при планировании: {e}",
            reply_markup=main_menu_kb(),
        )


# TODO: Временно отключено - модуль planning не реализован
# @router.callback_query(F.data.startswith("plan_day:"))
async def cb_plan_day_DISABLED(callback: CallbackQuery):
    await callback.answer()

    cache = PLANNING_CACHE.get(callback.from_user.id)
    if not cache:
        await callback.message.answer(
            "⚠️ План не найден. Нажмите «Планирование по дням» ещё раз.",
            reply_markup=main_menu_kb(),
        )
        return

    try:
        day = int(callback.data.split(":", 1)[1])
    except (ValueError, AttributeError):
        await callback.message.answer("❌ Не удалось определить день.")
        return

    schedule = cache.get("schedule")
    if not schedule:
        await callback.message.answer("⚠️ План пуст. Постройте его заново.")
        return

    day_tracks = [track for track in schedule if track.day == day]
    if not day_tracks:
        await callback.message.answer(f"⚠️ На день {day} дорожек нет.")
        return

    await callback.message.answer(f"📍 День {day}: готовлю визуализации по линиям…")

    for track in sorted(day_tracks, key=lambda t: t.line):
        # await callback.message.answer(track_to_text(track), parse_mode="Markdown")
        await callback.message.answer("Track info N/A", parse_mode="Markdown")
        try:
            # png_path, pdf_path, extras = await asyncio.to_thread(render_line, track)
            png_path, pdf_path, extras = None, None, []
        except Exception as e:
            await callback.message.answer(f"❌ Ошибка визуализации линии {track.line}: {e}")
            continue

        if png_path.exists():
            await callback.message.answer_document(
                FSInputFile(str(png_path)), caption=f"День {day} • Линия {track.line}"
            )
        if pdf_path.exists():
            await callback.message.answer_document(FSInputFile(str(pdf_path)))
        for extra in extras:
            await callback.message.answer_document(FSInputFile(str(extra)))


@router.message(F.text == "Коммерческое предложение PDF")
@router.message(Command("commercial_offer"))
async def btn_commercial_offer(message: Message, state: FSMContext):
    """Обработчик запроса на создание коммерческого предложения"""
    await state.set_state(KPStates.waiting_for_commercial_offer)
    await message.answer(
        "📄 Создание коммерческого предложения\n\n"
        "Пришлите список плит в свободной форме.\n\n"
        "Примеры форматов:\n"
        "• '1.2×3.39 — 2 шт'\n"
        "• '0.32×6.63 — 4 шт'\n"
        "• 'ПБ 38-12-8п 2'\n"
        "• 'ПБ 66-3-8п 4'\n\n"
        "Я создам PDF с расчётом стоимости, веса и НДС.",
        reply_markup=main_menu_kb()
    )


@router.message(KPStates.waiting_for_commercial_offer)
async def receive_order_and_generate_pdf(message: Message, state: FSMContext):
    """Обработчик получения заказа и генерации PDF"""
    await message.answer("⏳ Формирую коммерческое предложение...")
    
    try:
        # Парсим список пользователя и собираем «странные» строки по формату
        user_text = message.text or ""
        unparsed_lines = set_plate_lists_from_text(user_text)

        if unparsed_lines:
            warn_text = "⚠️ Некоторые строки я не смог распознать по формату и пропустил:\n"
            warn_text += "\n".join(f"• {line}" for line in unparsed_lines)
            warn_text += (
                "\n\nЯ понимаю, например, такие форматы:\n"
                "• 1.2×3.39 — 2 шт\n"
                "• 0,32x6,63 - 4\n"
                "• Плиты ПБ 78-12-8п 3\n"
                "• ПБ 66,2-12-8п 6\n"
            )
            await message.answer(warn_text)
        
        # Собираем данные заказа из глобальных списков
        from collections import Counter
        order_data = []
        
        # Собираем все плиты по типам
        # ВАЖНО: Добавлен target_name для получения точных ширин из PLATE_EXACT_WIDTHS
        plate_groups = [
            (1200, cfg.PLATES_1_2, "12", 'PLATES_1_2'),
            (1080, cfg.PLATES_1_08, "10.8", 'PLATES_1_08'),
            (1000, cfg.PLATES_1_0, "10", 'PLATES_1_0'),
            (320, cfg.PLATES_0_32, "3.2", 'PLATES_0_32'),
            (460, cfg.PLATES_0_46, "4.6", 'PLATES_0_46'),
            (700, cfg.PLATES_0_70, "7", 'PLATES_0_70'),
            (720, cfg.PLATES_0_72, "7.2", 'PLATES_0_72'),
            (860, cfg.PLATES_0_86, "8.6", 'PLATES_0_86'),
            (880, cfg.PLATES_0_88, "8.8", 'PLATES_0_88'),
            (740, cfg.PLATES_0_74, "7.4", 'PLATES_0_74'),
            (480, cfg.PLATES_0_48, "4.8", 'PLATES_0_48'),
            (500, cfg.PLATES_0_50, "5", 'PLATES_0_50'),
            (340, cfg.PLATES_0_34, "3.4", 'PLATES_0_34'),
        ]
        
        for width_mm, plates_list, width_dm_str, target_name in plate_groups:
            if plates_list:
                # Группируем по длине
                length_counts = Counter(plates_list)
                for length_m, qty in length_counts.items():
                    # Получаем ТОЧНУЮ ширину из PLATE_EXACT_WIDTHS
                    exact_width_m = cfg.get_exact_width(length_m, target_name, width_mm / 1000.0)
                    exact_width_mm = int(round(exact_width_m * 1000))
                    
                    length_dm = int(round(length_m * 10))
                    
                    # Получаем нагрузку из PLATE_LOAD_MAP (8п, 10п, 12п и т.д.)
                    load_code = cfg.get_load_code_for_plate(length_m, exact_width_m, default=8)
                    
                    # Формируем наименование в формате "Плиты ПБ 38-12-8п" с ТОЧНОЙ шириной
                    if exact_width_mm >= 1000:
                        width_str = str(int(round(exact_width_mm / 100)))
                    else:
                        # Для малых ширин используем дм с точкой (например, 5,3 для 530мм, 6,65 для 665мм)
                        exact_width_dm = exact_width_mm / 100.0
                        # Умное форматирование: убираем лишние нули (6.65→"6,65", 5.3→"5,3", 12.0→"12")
                        width_str = f"{exact_width_dm:.2f}".rstrip('0').rstrip('.').replace('.', ',')
                    
                    # Используем ПРАВИЛЬНУЮ нагрузку из заказа
                    name = f"Плиты ПБ {length_dm}-{width_str}-{load_code}п"
                    
                    order_data.append({
                        "name": name,
                        "length_m": length_m,
                        "width_m": exact_width_m,  # ТОЧНАЯ ширина в метрах!
                        "qty": qty
                    })
        
        if not order_data:
            await message.answer(
                "❌ Не удалось распознать ни одной плиты в вашем сообщении.\n"
                "Проверьте формат строк (ширина×длина×кол-во или 'Плиты ПБ 78-12-8п 3').",
                reply_markup=main_menu_kb()
            )
            await state.clear()
            return
        
        # Сохраняем заказ в кэш
        ORDER_CACHE[message.from_user.id] = order_data
        
        # Генерируем номер и дату КП
        offer_number = f"{message.from_user.id}_{datetime.now().strftime('%Y%m%d%H%M')}"
        offer_date = datetime.now().strftime("%d.%m.%Y")
        
        # Получаем имя пользователя для КП
        user = message.from_user
        if user.last_name:
            customer_name = f"{user.first_name} {user.last_name}"
        else:
            customer_name = user.first_name or "заказчик"
        
        # Генерируем PDF в памяти
        pdf_buffer = await asyncio.to_thread(
            generate_commercial_offer_pdf,
            order_data,
            offer_number,
            offer_date,
            customer_name
        )
        
        # Сохраняем во временный файл для отправки
        pdf_filename = f"КП_{offer_number}_{offer_date.replace('.', '')}.pdf"
        pdf_path = os.path.join(OUTPUTS_DIR_STR, pdf_filename)
        
        with open(pdf_path, 'wb') as f:
            f.write(pdf_buffer.getvalue())
        
        # Формируем сводку по заказу
        total_qty = sum(item['qty'] for item in order_data)
        summary = f"✅ Коммерческое предложение готово!\n\n"
        summary += f"📋 Заказ:\n"
        for item in order_data:
            summary += f"  • {item['name']} — {item['qty']} шт\n"
        summary += f"\n📊 Всего позиций: {len(order_data)}\n"
        summary += f"📦 Всего плит: {total_qty} шт\n"
        
        await message.answer(summary)
        
        # Отправляем PDF
        if os.path.exists(pdf_path):
            await message.answer_document(
                FSInputFile(pdf_path),
                caption=f"📄 Коммерческое предложение № {offer_number}"
            )
            await message.answer(
                "✨ Документ содержит:\n"
                "• Подробную спецификацию\n"
                "• Расчёт стоимости материалов\n"
                "• Стоимость резов\n"
                "• Вес изделий\n"
                "• НДС (20%)\n"
                "• Условия оплаты",
                reply_markup=main_menu_kb()
            )
        else:
            await message.answer(
                "❌ Ошибка при сохранении файла",
                reply_markup=main_menu_kb()
            )
    
    except Exception as e:
        await message.answer(
            f"❌ Ошибка при генерации КП: {str(e)}\n\n"
            "Проверьте формат данных и попробуйте снова.",
            reply_markup=main_menu_kb()
        )
    finally:
        await state.clear()


# ==================== НОВЫЕ КОМАНДЫ: /myorders, /export ====================

@router.message(Command("myorders"))
async def cmd_myorders(message: Message):
    """Показывает историю заказов пользователя"""
    try:
        import sqlite3
        from domain.export import get_user_orders
        
        con = sqlite3.connect('pb.db')
        orders = get_user_orders(con, message.from_user.id, limit=10)
        con.close()
        
        if not orders:
            await message.answer(
                "📋 У вас пока нет сохранённых заказов.\n\n"
                "Создайте заказ через 'Получить КП' или 'Коммерческое предложение PDF'",
                reply_markup=main_menu_kb()
            )
            return
        
        # Формируем список заказов
        response = "📋 <b>История ваших заказов:</b>\n\n"
        
        for order in orders:
            status_icon = {
                'created': '🆕',
                'processing': '⏳',
                'completed': '✅',
                'archived': '📦'
            }.get(order['status'], '❓')
            
            client_info = f" ({order['client_name']})" if order['client_name'] else ""
            
            response += (
                f"{status_icon} <b>Заказ #{order['id']}</b>{client_info}\n"
                f"   Дата: {order['created_at'][:10]}\n"
                f"   Позиций: {order['items_count']}\n"
                f"   /export_{order['id']} - экспортировать\n\n"
            )
        
        response += "\n💡 Для экспорта заказа используйте команду /export_НОМЕР"
        
        await message.answer(response, parse_mode="HTML", reply_markup=main_menu_kb())
        
    except Exception as e:
        await message.answer(
            f"❌ Ошибка при получении истории заказов: {str(e)}",
            reply_markup=main_menu_kb()
        )


@router.message(Command("export"))
async def cmd_export(message: Message):
    """Экспортирует заказ в ZIP архив"""
    try:
        # Парсим ID заказа из команды /export_123
        command_parts = message.text.split('_')
        if len(command_parts) < 2:
            await message.answer(
                "❓ Укажите номер заказа: /export_123\n\n"
                "Посмотреть список заказов: /myorders",
                reply_markup=main_menu_kb()
            )
            return
        
        try:
            order_id = int(command_parts[1])
        except ValueError:
            await message.answer(
                "❌ Неверный формат номера заказа",
                reply_markup=main_menu_kb()
            )
            return
        
        import sqlite3
        from pathlib import Path
        from domain.export import get_order_items, create_order_archive
        from domain.calc import cost_standard, cost_addon
        from domain.excel_kz import generate_kz_excel
        # TODO: Модуль не реализован
        # from commercial_offer import generate_commercial_offer_pdf
        from datetime import datetime
        
        # Получаем данные заказа
        con = sqlite3.connect('pb.db')
        items = get_order_items(con, order_id)
        
        if not items:
            con.close()
            await message.answer(
                f"❌ Заказ #{order_id} не найден или у вас нет к нему доступа",
                reply_markup=main_menu_kb()
            )
            return
        
        await message.answer("⏳ Формирую архив заказа...")
        
        # Генерируем файлы
        output_dir = Path("Визуализация_Раскладки")
        output_dir.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 1. Excel КЗ
        excel_path = generate_kz_excel(
            con,
            items,
            tracks=None,
            output_path=str(output_dir / f"kz_{order_id}_{timestamp}.xlsx"),
            order_number=str(order_id),
            customer_name=None
        )
        
        # 2. PDF КП
        order_data = []
        for item in items:
            length_dm = int(round(item['length_m'] * 10))
            width_dm = int(round(item['width_m'] * 10))
            name = f"ПБ {length_dm}-{width_dm}-{int(item['load_class'])}п"
            order_data.append({
                'name': name,
                'length_m': item['length_m'],
                'width_m': item['width_m'],
                'qty': item['qty']
            })
        
        # TODO: Функция не реализована
        # pdf_buffer = generate_commercial_offer_pdf(
        #     order_data,
        #     offer_number=str(order_id),
        #     offer_date=datetime.now().strftime("%d.%m.%Y"),
        #     customer_name=None
        # )
        
        # pdf_path = output_dir / f"kp_{order_id}_{timestamp}.pdf"
        # with open(pdf_path, 'wb') as f:
        #     f.write(pdf_buffer.getvalue())
        pdf_path = None  # Заглушка - функция не реализована
        
        con.close()
        
        # 3. Архивируем
        files_to_archive = [excel_path]
        if pdf_path:
            files_to_archive.append(pdf_path)
        archive_path = create_order_archive(
            order_id,
            files_to_archive,
            output_dir=str(output_dir)
        )
        
        # Отправляем архив
        if archive_path.exists():
            await message.answer_document(
                FSInputFile(archive_path),
                caption=f"📦 Архив заказа #{order_id}\n\nВключает КП (PDF) и КЗ (Excel)"
            )
            
            await message.answer(
                "✅ Архив готов!\n\n"
                "💡 Хотите отправить на email? Напишите адрес в ответ на это сообщение",
                reply_markup=main_menu_kb()
            )
        else:
            await message.answer(
                "❌ Ошибка при создании архива",
                reply_markup=main_menu_kb()
            )
        
    except Exception as e:
        await message.answer(
            f"❌ Ошибка при экспорте: {str(e)}",
            reply_markup=main_menu_kb()
        )

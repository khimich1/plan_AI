"""Обработчики информации о плитах ПБ в работе"""
import os
from datetime import datetime
from pathlib import Path

from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, FSInputFile

# Импорты из проекта
import sys
BOT_DIR = Path(__file__).parent.parent
PROJECT_ROOT = BOT_DIR.parent
sys.path.insert(0, str(PROJECT_ROOT))

from core import kp_db
from ..keyboards import main_menu_kb, pb_info_kb
from ..bot_config import OUTPUTS_DIR_STR

router = Router()


# Маппинг колонок для плит в производстве (kp_plates)
PRODUCTION_COLUMNS = {
    'id': '№ записи',
    'kp_id': '№ КП',
    'position_number': '№ позиции',
    'plate_name': 'Наименование плиты',
    'length_m': 'Длина (м)',
    'width_m': 'Ширина (м)',
    'load_class': 'Класс нагрузки',
    'qty': 'Количество',
    'unit_weight': 'Вес единицы (кг)',
    'total_weight': 'Общий вес (кг)',
    'discounted_price': 'Цена со скидкой',
    'customer_name': 'Заказчик',
    'execution_terms': 'Срок выполнения'
}

# Маппинг колонок для выполненных плит (completed_plates)
COMPLETED_COLUMNS = {
    'id': '№ записи',
    'kp_id': '№ КП',
    'plate_name': 'Наименование плиты',
    'length_m': 'Длина (м)',
    'width_m': 'Ширина (м)',
    'load_class': 'Класс нагрузки',
    'qty': 'Количество',
    'completed_date': 'Дата выполнения',
    'production_day': 'День производства',
    'customer_name': 'Заказчик',
    'execution_terms': 'Срок выполнения'
}


def generate_excel_from_data(data: list, columns_map: dict, filename: str) -> str:
    """
    Генерирует Excel файл из данных.
    
    Args:
        data: список словарей с данными
        columns_map: маппинг английских названий колонок на русские
        filename: имя файла (без расширения)
    
    Returns:
        Путь к созданному файлу
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise ImportError("Для работы требуется библиотека openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Плиты"
    
    # Стили
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Определяем колонки на основе маппинга
    columns = list(columns_map.keys())
    headers = list(columns_map.values())
    
    # Записываем заголовки
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_alignment
    
    # Записываем данные
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_alignment
    
    # Автоширина колонок
    for col_idx, col_name in enumerate(columns, 1):
        max_length = len(headers[col_idx - 1])
        for row in ws.iter_rows(min_row=2, max_row=len(data) + 1, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
    
    # Сохраняем файл
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(OUTPUTS_DIR_STR, f"{filename}_{timestamp}.xlsx")
    
    # Создаём папку если её нет
    os.makedirs(OUTPUTS_DIR_STR, exist_ok=True)
    
    wb.save(filepath)
    return filepath


@router.message(F.text == "Информация о ПБ в работе")
async def btn_pb_info(message: Message):
    """Обработчик кнопки 'Информация о ПБ в работе'"""
    await message.answer(
        "📊 Информация о плитах ПБ\n\n"
        "Выберите тип отчёта:",
        reply_markup=pb_info_kb()
    )


@router.callback_query(F.data == "plates_in_production")
async def export_plates_in_production(callback: CallbackQuery):
    """Экспорт плит в производстве в Excel"""
    await callback.message.answer("⏳ Загружаю данные о плитах в производстве...")
    
    try:
        # Получаем данные из БД
        db_path = PROJECT_ROOT / "plita.db"
        plates = kp_db.get_all_plates_in_production(str(db_path))
        
        if not plates:
            await callback.message.answer(
                "📭 Нет плит в производстве.\n\n"
                "Все заказы выполнены или нет КП в работе.",
                reply_markup=main_menu_kb()
            )
            await callback.answer()
            return
        
        # Генерируем Excel
        filepath = generate_excel_from_data(
            data=plates,
            columns_map=PRODUCTION_COLUMNS,
            filename="Плиты_в_производстве"
        )
        
        # Подсчитываем статистику
        total_qty = sum(p.get('qty', 0) for p in plates)
        unique_kp = len(set(p.get('kp_id') for p in plates))
        
        # Отправляем файл
        await callback.message.answer_document(
            FSInputFile(filepath),
            caption=(
                f"🏭 Плиты в производстве\n\n"
                f"📊 Статистика:\n"
                f"• Всего записей: {len(plates)}\n"
                f"• Общее кол-во плит: {total_qty}\n"
                f"• КП в работе: {unique_kp}"
            )
        )
        
        await callback.message.answer(
            "✅ Файл готов!",
            reply_markup=main_menu_kb()
        )
        
    except Exception as e:
        await callback.message.answer(
            f"❌ Ошибка при экспорте: {str(e)}",
            reply_markup=main_menu_kb()
        )
    
    await callback.answer()


@router.callback_query(F.data == "completed_plates_export")
async def export_completed_plates(callback: CallbackQuery):
    """Экспорт выполненных плит в Excel"""
    await callback.message.answer("⏳ Загружаю данные о выполненных плитах...")
    
    try:
        # Получаем данные из БД
        db_path = PROJECT_ROOT / "plita.db"
        plates = kp_db.get_all_completed_plates(str(db_path))
        
        if not plates:
            await callback.message.answer(
                "📭 Нет выполненных плит.\n\n"
                "Пока не было завершено ни одного дня производства.",
                reply_markup=main_menu_kb()
            )
            await callback.answer()
            return
        
        # Генерируем Excel
        filepath = generate_excel_from_data(
            data=plates,
            columns_map=COMPLETED_COLUMNS,
            filename="Выполненные_плиты"
        )
        
        # Подсчитываем статистику
        total_qty = sum(p.get('qty', 0) for p in plates)
        unique_kp = len(set(p.get('kp_id') for p in plates))
        unique_days = len(set(p.get('production_day') for p in plates if p.get('production_day')))
        
        # Отправляем файл
        await callback.message.answer_document(
            FSInputFile(filepath),
            caption=(
                f"✅ Выполненные плиты\n\n"
                f"📊 Статистика:\n"
                f"• Всего записей: {len(plates)}\n"
                f"• Общее кол-во плит: {total_qty}\n"
                f"• КП затронуто: {unique_kp}\n"
                f"• Дней производства: {unique_days}"
            )
        )
        
        await callback.message.answer(
            "✅ Файл готов!",
            reply_markup=main_menu_kb()
        )
        
    except Exception as e:
        await callback.message.answer(
            f"❌ Ошибка при экспорте: {str(e)}",
            reply_markup=main_menu_kb()
        )
    
    await callback.answer()


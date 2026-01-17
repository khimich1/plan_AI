"""Обработчики сравнения КП завода и нашей сметы"""
import asyncio
import os
import re
import logging
from dataclasses import dataclass
from typing import Any

logger = logging.getLogger(__name__)

from aiogram import Router, F
from aiogram.types import Message, FSInputFile
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext

try:
    import pandas as pd
except Exception:
    pd = None

from ..states import CompareStates

router = Router()


@dataclass
class PlateRow:
    """Одна строка из Excel с данными по плите."""

    name: str
    qty: float
    weight: float
    price: float
    total: float

    @property
    def price_per_kg(self) -> float | None:
        if self.weight and self.weight > 0:
            return self.price / self.weight
        return None


def _to_float(value) -> float:
    """Переводит '9 130,80' → 9130.80."""
    try:
        s = str(value).replace(" ", "").replace("\u00a0", "").replace(",", ".")
        return float(s)
    except Exception:
        return 0.0


def _to_qty(value) -> float:
    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except Exception:
        return 0.0


def _normalize_name(name: str) -> str:
    """
    Нормализуем наименование плиты:
    - убираем лишние пробелы;
    - приводим к одному регистру;
    - оставляем только текст без лишних пробелов.
    """
    return " ".join(str(name).strip().lower().split())


def _make_plate_key(name: str):
    """
    Делает «умный» ключ для сравнения плит.

    Логика:
    - из строки вида 'Плиты ПБ 66,2-12-8п' вытаскиваем два первых числа: 66,2 и 12;
    - это длина и ширина в дециметрах;
    - переводим их в float, округляем до целых дм и берём пару (length_dm, width_dm);
      так '66,2-12-8п' и '66-12-8п' дадут один ключ (66, 12);
    - если распарсить не удалось — используем нормализованный текст как раньше.
    """
    s = str(name)

    # Ищем шаблон "ЧИСЛО-ЧИСЛО" с поддержкой запятой/точки
    m = re.search(r'(\d[\d,\.]*)\s*-\s*(\d[\d,\.]*)', s)
    if m:
        length_dm_str = m.group(1).replace(" ", "").replace(",", ".")
        width_dm_str = m.group(2).replace(" ", "").replace(",", ".")
        try:
            # Округляем длину и ширину в дециметрах до целых
            length_dm = int(round(float(length_dm_str)))
            width_dm = int(round(float(width_dm_str)))
            return (length_dm, width_dm)
        except ValueError:
            # если вдруг не смогли преобразовать — упадём в текстовый ключ
            pass

    # Фолбэк: сравниваем по тексту, как раньше
    return _normalize_name(name)


def _find_column(df, *candidates: str) -> str:
    """
    Находит колонку в DataFrame по нескольким вариантам имени
    (по вхождению подстроки без учёта регистра и пробелов).
    """
    columns = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        cand_norm = cand.strip().lower()
        for key, real_name in columns.items():
            if cand_norm == key or cand_norm in key:
                return real_name
    raise KeyError(f"Не найдена колонка среди вариантов: {candidates}")


def _aggregate_row(result: dict, row: PlateRow) -> None:
    """
    Добавляет строку в словарь, суммируя данные,
    если такая плита уже встречалась (даже если она записана по‑разному).

    Ключ строим по (длина, ширина) в дециметрах, а не по сырому тексту:
    так 'Плиты ПБ 66,2-12-8п' и 'Плиты ПБ 66-12-8п' попадут в одну позицию.
    """
    key = _make_plate_key(row.name)
    prev = result.get(key)

    if prev is not None:
        qty = prev.qty + row.qty
        total = prev.total + row.total
        # Вес и цена считаем одинаковыми для одинаковой плиты,
        # поэтому просто берём первое ненулевое значение.
        weight = prev.weight or row.weight
        price = prev.price or row.price
        result[key] = PlateRow(
            name=prev.name,  # храним оригинальное имя первой строки
            qty=qty,
            weight=weight,
            price=price,
            total=total,
        )
    else:
        result[key] = row


def _prepare_kp_sheet(sheet_df):
    """
    Пытается найти строку с заголовками ('Товары (работы, услуги)', 'Количество', 'Цена', 'Сумма')
    если она не в первой строке, и переставить её как header DataFrame.
    """
    # Если заголовки уже в columns, просто возвращаем как есть
    try:
        _find_column(
            sheet_df,
            "товары (работы, услуги)",
            "наименование",
            "товар",
            "товары",
            "работы",
            "услуги",
        )
        return sheet_df
    except KeyError:
        pass

    header_candidates = (
        "товары (работы, услуги)",
        "наименование",
        "товар",
        "товары",
        "работы",
        "услуги",
    )
    max_rows = min(40, len(sheet_df))
    for i in range(max_rows):
        row = sheet_df.iloc[i]
        row_lower = [str(v).strip().lower() for v in row]
        if any(
            any(cand in cell for cand in header_candidates)
            for cell in row_lower
        ):
            # Делаем эту строку заголовком
            new_df = sheet_df.iloc[i + 1 :].copy()
            new_df.columns = row
            return new_df
    return sheet_df


def load_kp_excel(path: str) -> dict:
    """
    Загружает РЕАЛЬНОЕ КП завода.
    Здесь может не быть столбца веса, а наименование может быть
    в колонке «Товары (работы, услуги)».
    """
    if pd is None:
        raise RuntimeError("pandas не установлена, сравнение КП недоступно.")

    all_sheets = pd.read_excel(path, sheet_name=None)
    df = None
    name_col = qty_col = price_col = total_col = None

    for sheet_df in all_sheets.values():
        sheet_df = _prepare_kp_sheet(sheet_df)
        try:
            name_col = _find_column(
                sheet_df,
                "товары (работы, услуги)",
                "наименование",
                "наимен",
                "товар",
                "товары",
                "работы",
                "услуги",
            )
            qty_col = _find_column(sheet_df, "кол-во", "количество")
            price_col = _find_column(sheet_df, "цена", "стоимость")
            total_col = _find_column(sheet_df, "сумма", "итого")
        except KeyError:
            continue
        df = sheet_df
        break

    if df is None:
        raise RuntimeError("Не удалось найти нужные столбцы в файле КП.")

    result: dict = {}

    for _, row in df.iterrows():
        name = str(row.get(name_col, "")).strip()
        if not name:
            continue
        if "итог" in name.lower():
            # Строки с итогами пропускаем
            continue

        qty = _to_qty(row.get(qty_col, 0))
        price = _to_float(row.get(price_col, 0))
        total = _to_float(row.get(total_col, 0))

        # Нас интересуют только плиты ПБ
        if "плиты пб" not in _normalize_name(name):
            continue

        plate_row = PlateRow(
            name=name,
            qty=qty,
            weight=0.0,  # в КП веса нет
            price=price,
            total=total,
        )
        _aggregate_row(result, plate_row)

    return result


def load_smeta_excel(path: str) -> dict:
    """
    Загружает НАШУ СМЕТУ (Смета_Дорожка_1_...xlsx).
    Здесь обязательно есть колонка веса.
    """
    if pd is None:
        raise RuntimeError("pandas не установлена, сравнение КП недоступно.")

    all_sheets = pd.read_excel(path, sheet_name=None)
    df = None
    name_col = qty_col = weight_col = price_col = total_col = None

    for sheet_df in all_sheets.values():
        try:
            name_col = _find_column(sheet_df, "наименование", "наимен")
            qty_col = _find_column(sheet_df, "кол-во", "количество")
            weight_col = _find_column(sheet_df, "вес", "вес(кг)", "масса")
            price_col = _find_column(sheet_df, "цена")
            total_col = _find_column(sheet_df, "сумма")
        except KeyError:
            continue
        df = sheet_df
        break

    if df is None:
        raise RuntimeError("Не удалось найти нужные столбцы в файле сметы.")

    result: dict = {}

    for _, row in df.iterrows():
        name = str(row.get(name_col, "")).strip()
        if not name:
            continue
        if "итого" in name.lower():
            continue

        qty = _to_qty(row.get(qty_col, 0))
        weight = _to_float(row.get(weight_col, 0))
        price = _to_float(row.get(price_col, 0))
        total = _to_float(row.get(total_col, 0))

        plate_row = PlateRow(
            name=name,
            qty=qty,
            weight=weight,
            price=price,
            total=total,
        )
        _aggregate_row(result, plate_row)

    return result


def _explain_price_breakdown(name: str, our_price: float, smeta_path: str) -> str:
    """
    Пытается найти детальную разбивку цены для плиты из файла детальной разбивки рядом со сметой.
    Возвращает строку с объяснением компонентов цены.
    """
    try:
        # Ищем файл детальной разбивки (может быть в том же Excel-файле на другом листе)
        all_sheets = pd.read_excel(smeta_path, sheet_name=None)
        breakdown_df = None
        
        for sheet_name, df in all_sheets.items():
            if 'детал' in sheet_name.lower() or 'разбивка' in sheet_name.lower() or 'компонент' in sheet_name.lower():
                breakdown_df = df
                break
        
        if breakdown_df is None:
            return "  💡 Детальная разбивка не найдена в файле сметы"
        
        # Ищем строку с нашим наименованием
        for _, row in breakdown_df.iterrows():
            row_name = str(row.get('Наименование', '')).strip()
            # Сравниваем по ключу (length, width)
            if _normalize_name(row_name) == _normalize_name(name) or _make_plate_key(row_name) == _make_plate_key(name):
                # Извлекаем компоненты
                base = _to_float(row.get('Базовая цена', 0))
                long_cuts = _to_float(row.get('Продольные резы', 0))
                trans_cuts = _to_float(row.get('Поперечные резы', 0))
                rest = _to_float(row.get('Остатки', 0))
                waste = _to_float(row.get('Отходы', 0))
                
                parts = []
                if base > 0:
                    parts.append(f"базовая {base:,.2f} ₽")
                if long_cuts > 0:
                    parts.append(f"прод.резы {long_cuts:,.2f} ₽")
                if trans_cuts > 0:
                    parts.append(f"попер.резы {trans_cuts:,.2f} ₽")
                if rest > 0:
                    parts.append(f"остатки {rest:,.2f} ₽")
                if waste > 0:
                    parts.append(f"отходы {waste:,.2f} ₽")
                
                if parts:
                    formula = " + ".join(parts)
                    return f"  💡 Разбивка: {formula} = {our_price:,.2f} ₽"
                else:
                    return f"  💡 Разбивка: {our_price:,.2f} ₽ (детали не найдены в таблице)"
        
        return "  💡 Позиция не найдена в детальной разбивке"
    
    except Exception as e:
        return f"  💡 Ошибка чтения разбивки: {e}"


def create_comparison_excel(kp_path: str, smeta_path: str) -> str:
    """
    Создаёт Excel файл с двумя таблицами (смета и КП) рядом + пояснения о ценообразовании.
    
    Возвращает путь к созданному файлу.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    if pd is None:
        raise RuntimeError("pandas не установлена, сравнение КП недоступно.")
    
    # Загружаем данные из КП и сметы
    kp = load_kp_excel(kp_path)
    smeta = load_smeta_excel(smeta_path)
    
    # Ключи, которые есть в обоих файлах
    common_names = sorted(set(kp.keys()) & set(smeta.keys()), key=str)
    
    if not common_names:
        raise RuntimeError("Не нашёл общих позиций по наименованию плит в двух файлах.")
    
    # Создаём новый Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сравнение"
    
    # ==================== СТИЛИ ====================
    # Заголовок
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Подзаголовки
    subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    subheader_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Названия плит
    plate_name_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    plate_name_font = Font(bold=True, size=10)
    
    # Разница (положительная - зелёный, отрицательная - красный)
    positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Рамки
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ==================== ГЛАВНЫЙ ЗАГОЛОВОК ====================
    ws.merge_cells('A1:N1')
    cell = ws['A1']
    cell.value = "📊 СРАВНЕНИЕ СМЕТЫ И КП ЗАВОДА"
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    
    # ==================== ЗАГОЛОВКИ ТАБЛИЦ ====================
    row = 3
    
    # Заголовок "НАША СМЕТА"
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws[f'A{row}']
    cell.value = "📋 НАША СМЕТА"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    
    # Заголовок "КП ЗАВОДА"
    ws.merge_cells(f'H{row}:N{row}')
    cell = ws[f'H{row}']
    cell.value = "🏭 КП ЗАВОДА"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    
    ws.row_dimensions[row].height = 20
    row += 1
    
    # ==================== СТОЛБЦЫ ТАБЛИЦ ====================
    # Столбцы для сметы
    smeta_cols = ['№', 'Наименование', 'Кол-во', 'Вес, кг', 'Цена, ₽', 'Сумма, ₽', 'Примечание']
    for col_idx, col_name in enumerate(smeta_cols, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = col_name
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = subheader_align
        cell.border = thin_border
    
    # Столбцы для КП
    kp_cols = ['№', 'Наименование', 'Кол-во', 'Цена, ₽', 'Сумма, ₽', 'Разница цены', 'Разница суммы']
    for col_idx, col_name in enumerate(kp_cols, start=8):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = col_name
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = subheader_align
        cell.border = thin_border
    
    ws.row_dimensions[row].height = 30
    row += 1
    
    # ==================== ДАННЫЕ ====================
    # Пытаемся найти файл детальной разбивки
    breakdown_data = {}
    try:
        # Ищем файл детальной разбивки рядом со сметой
        smeta_dir = os.path.dirname(smeta_path)
        smeta_name = os.path.basename(smeta_path)
        
        # Извлекаем timestamp из имени сметы (например, "Смета_Дорожка_1_20250128_143022.xlsx")
        timestamp_match = re.search(r'_(\d{8}_\d{6})\.xlsx$', smeta_name)
        if timestamp_match:
            timestamp = timestamp_match.group(1)
            breakdown_path = os.path.join(smeta_dir, f'Детальная_разбивка_Дорожка_1_{timestamp}.xlsx')
            
            if os.path.exists(breakdown_path):
                # Читаем файл детальной разбивки
                breakdown_df = pd.read_excel(breakdown_path)
                current_plate = None
                
                for _, br_row in breakdown_df.iterrows():
                    comp = str(br_row.get('Компонент', '')).strip()
                    calc = str(br_row.get('Расчёт', '')).strip()
                    summa = str(br_row.get('Сумма', '')).strip()
                    
                    # Если это заголовок плиты (наименование)
                    if comp and not calc and not summa and 'Плиты ПБ' in comp:
                        current_plate = _make_plate_key(comp)
                        breakdown_data[current_plate] = []
                    elif current_plate and comp:
                        # Добавляем компонент к текущей плите
                        breakdown_data[current_plate].append({
                            'component': comp,
                            'formula': calc,
                            'sum': summa
                        })
    except Exception as e:
        logger.exception(f"Не удалось загрузить детальную разбивку: {e}")
    
    # Итоги
    total_smeta = 0.0
    total_kp = 0.0
    
    idx = 1
    for key in common_names:
        s = smeta[key]  # строка из нашей сметы
        k = kp[key]     # строка из КП завода
        
        name = s.name or k.name
        
        # ===== СМЕТА (столбцы A-G) =====
        # Номер
        cell = ws.cell(row=row, column=1, value=idx)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
        # Наименование
        cell = ws.cell(row=row, column=2, value=name)
        cell.font = plate_name_font
        cell.fill = plate_name_fill
        cell.border = thin_border
        
        # Количество
        cell = ws.cell(row=row, column=3, value=s.qty)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        cell.number_format = '0'
        
        # Вес
        cell = ws.cell(row=row, column=4, value=s.weight)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        cell.number_format = '0.00'
        
        # Цена
        cell = ws.cell(row=row, column=5, value=s.price)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        cell.number_format = '#,##0.00'
        
        # Сумма
        cell = ws.cell(row=row, column=6, value=s.total)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        cell.number_format = '#,##0.00'
        
        # Примечание (цена за кг)
        if s.price_per_kg:
            note = f"{s.price_per_kg:.2f} ₽/кг"
        else:
            note = ""
        cell = ws.cell(row=row, column=7, value=note)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
        # ===== КП (столбцы H-N) =====
        # Номер
        cell = ws.cell(row=row, column=8, value=idx)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
        # Наименование
        cell = ws.cell(row=row, column=9, value=name)
        cell.font = plate_name_font
        cell.fill = plate_name_fill
        cell.border = thin_border
        
        # Количество
        cell = ws.cell(row=row, column=10, value=k.qty)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        cell.number_format = '0'
        
        # Цена
        cell = ws.cell(row=row, column=11, value=k.price)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        cell.number_format = '#,##0.00'
        
        # Сумма
        cell = ws.cell(row=row, column=12, value=k.total)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        cell.number_format = '#,##0.00'
        
        # Разница цены (наша - их)
        price_diff = s.price - k.price
        cell = ws.cell(row=row, column=13, value=price_diff)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        cell.number_format = '#,##0.00'
        if price_diff > 0:
            cell.fill = negative_fill  # Мы дороже - красный
        elif price_diff < 0:
            cell.fill = positive_fill  # Мы дешевле - зелёный
        
        # Разница суммы (наша - их)
        total_diff = s.total - k.total
        cell = ws.cell(row=row, column=14, value=total_diff)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        cell.number_format = '#,##0.00'
        if total_diff > 0:
            cell.fill = negative_fill  # Мы дороже - красный
        elif total_diff < 0:
            cell.fill = positive_fill  # Мы дешевле - зелёный
        
        total_smeta += s.total
        total_kp += k.total
        
        row += 1
        
        # ===== ДЕТАЛЬНАЯ РАЗБИВКА ДЛЯ ЭТОЙ ПЛИТЫ =====
        if key in breakdown_data and breakdown_data[key]:
            # Заголовок разбивки
            ws.merge_cells(f'A{row}:N{row}')
            cell = ws[f'A{row}']
            cell.value = "💡 Детальная разбивка цены из нашей сметы:"
            cell.font = Font(bold=True, italic=True, size=9, color="404040")
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 18
            row += 1
            
            # Строки разбивки
            for comp_data in breakdown_data[key]:
                # Компонент
                cell = ws.cell(row=row, column=2, value=comp_data['component'])
                cell.font = Font(size=9, italic=True)
                cell.alignment = Alignment(horizontal="left")
                
                # Расчёт
                ws.merge_cells(f'C{row}:F{row}')
                cell = ws.cell(row=row, column=3, value=comp_data['formula'])
                cell.font = Font(size=9, italic=True, color="404040")
                cell.alignment = Alignment(horizontal="left")
                
                # Сумма
                cell = ws.cell(row=row, column=7, value=comp_data['sum'])
                cell.font = Font(size=9, italic=True)
                cell.alignment = Alignment(horizontal="right")
                
                ws.row_dimensions[row].height = 15
                row += 1
            
            # Пустая строка после разбивки
            row += 1
        
        idx += 1
    
    # ==================== ИТОГО ====================
    row += 1
    
    # ИТОГО для сметы
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "ИТОГО (наша смета):"
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="right")
    cell.border = thin_border
    
    cell = ws.cell(row=row, column=6, value=total_smeta)
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="right")
    cell.border = thin_border
    cell.number_format = '#,##0.00'
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # ИТОГО для КП
    ws.merge_cells(f'H{row}:K{row}')
    cell = ws[f'H{row}']
    cell.value = "ИТОГО (КП завода):"
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="right")
    cell.border = thin_border
    
    cell = ws.cell(row=row, column=12, value=total_kp)
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="right")
    cell.border = thin_border
    cell.number_format = '#,##0.00'
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Общая разница
    ws.merge_cells(f'M{row}:N{row}')
    cell = ws[f'M{row}']
    diff_total = total_smeta - total_kp
    cell.value = diff_total
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="right")
    cell.border = thin_border
    cell.number_format = '#,##0.00'
    if diff_total > 0:
        cell.fill = negative_fill  # Мы дороже
    elif diff_total < 0:
        cell.fill = positive_fill  # Мы дешевле
    
    ws.row_dimensions[row].height = 25
    
    # ==================== ВЫВОДЫ ====================
    row += 2
    ws.merge_cells(f'A{row}:N{row}')
    cell = ws[f'A{row}']
    if diff_total > 0:
        conclusion = f"⚠️ Наша смета на {abs(diff_total):,.2f} ₽ ДОРОЖЕ, чем КП завода"
    elif diff_total < 0:
        conclusion = f"✅ Наша смета на {abs(diff_total):,.2f} ₽ ДЕШЕВЛЕ, чем КП завода (экономия!)"
    else:
        conclusion = "✅ Цены СОВПАДАЮТ!"
    cell.value = conclusion
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="70AD47" if diff_total <= 0 else "E74C3C", 
                            end_color="70AD47" if diff_total <= 0 else "E74C3C", 
                            fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30
    
    # ==================== ШИРИНА СТОЛБЦОВ ====================
    ws.column_dimensions['A'].width = 5   # №
    ws.column_dimensions['B'].width = 30  # Наименование (смета)
    ws.column_dimensions['C'].width = 8   # Кол-во
    ws.column_dimensions['D'].width = 10  # Вес
    ws.column_dimensions['E'].width = 12  # Цена
    ws.column_dimensions['F'].width = 14  # Сумма
    ws.column_dimensions['G'].width = 12  # Примечание
    
    ws.column_dimensions['H'].width = 5   # №
    ws.column_dimensions['I'].width = 30  # Наименование (КП)
    ws.column_dimensions['J'].width = 8   # Кол-во
    ws.column_dimensions['K'].width = 12  # Цена
    ws.column_dimensions['L'].width = 14  # Сумма
    ws.column_dimensions['M'].width = 14  # Разница цены
    ws.column_dimensions['N'].width = 14  # Разница суммы
    
    # Сохраняем файл
    user_id = os.path.basename(kp_path).split('_')[0]
    output_path = os.path.join("tmp", f"{user_id}_comparison.xlsx")
    wb.save(output_path)
    
    return output_path


@router.message(Command("compare"))
@router.message(F.text == "Сравнение результатов")
async def start_comparison(message: Message, state: FSMContext):
    """Старт диалога сравнения КП и сметы."""
    if pd is None:
        await message.answer(
            "⚠️ Модуль pandas не установлен, сравнение Excel-файлов недоступно.\n"
            "Установите зависимости из requirements.txt (pandas, openpyxl)."
        )
        return

    await state.set_state(CompareStates.waiting_kp)
    await message.answer(
        "📄 Отправьте файл **реального КП завода** в формате Excel (.xlsx)\n\n"
        "Файл должен содержать столбцы:\n"
        "• Наименование (или 'Товары (работы, услуги)')\n"
        "• Кол-во (или 'Количество')\n"
        "• Цена\n"
        "• Сумма\n\n"
        "💡 Столбец 'Вес' не обязателен",
        parse_mode="Markdown"
    )


@router.message(CompareStates.waiting_kp, F.document)
async def receive_kp(message: Message, state: FSMContext):
    """Принимаем файл КП и сохраняем его."""
    user_id = message.from_user.id
    os.makedirs("tmp", exist_ok=True)
    kp_path = os.path.join("tmp", f"{user_id}_kp.xlsx")

    await message.bot.download(message.document, destination=kp_path)
    await state.update_data(kp_path=kp_path)
    await state.set_state(CompareStates.waiting_smeta)

    await message.answer(
        "✅ КП завода получил!\n\n"
        "📄 Теперь отправьте файл **нашей сметы** (Excel .xlsx)\n\n"
        "Это должен быть файл 'Смета_Дорожка_...' или 'Детальная_разбивка_...',\n"
        "который создаётся при расчёте КП через бота.",
        parse_mode="Markdown"
    )


@router.message(CompareStates.waiting_kp)
async def receive_kp_wrong(message: Message, state: FSMContext):  # noqa: ARG001
    await message.answer("Нужно отправить именно файл КП (Excel-документ). Попробуйте ещё раз.")


@router.message(CompareStates.waiting_smeta, F.document)
async def receive_smeta(message: Message, state: FSMContext):
    """Принимаем файл нашей сметы и запускаем сравнение."""
    user_id = message.from_user.id
    os.makedirs("tmp", exist_ok=True)
    smeta_path = os.path.join("tmp", f"{user_id}_smeta.xlsx")

    await message.bot.download(message.document, destination=smeta_path)

    data = await state.get_data()
    kp_path = data.get("kp_path")

    if not kp_path or not os.path.exists(kp_path):
        await message.answer("Не нашёл сохранённый файл КП. Давай начнём сначала: /compare")
        await state.clear()
        return

    await message.answer("📊 Сравниваю файлы и создаю Excel отчёт, подождите...")

    try:
        # Создаём Excel файл со сравнением
        comparison_path = await asyncio.to_thread(create_comparison_excel, kp_path, smeta_path)
        
        # Отправляем Excel файл
        if os.path.exists(comparison_path):
            await message.answer(
                "✅ Сравнение готово!\n\n"
                "📋 В файле вы найдёте:\n"
                "• Две таблицы рядом (наша смета и КП завода)\n"
                "• Сравнение цен по каждой позиции\n"
                "• Детальную разбивку ценообразования\n"
                "• Цветовую индикацию разницы (зелёный=экономия, красный=дороже)\n"
                "• Итоговые суммы и выводы"
            )
            await message.answer_document(
                FSInputFile(comparison_path),
                caption="📊 Сравнительная таблица КП и сметы"
            )
        else:
            await message.answer("❌ Не удалось создать файл сравнения")
            
    except Exception as e:  # pylint: disable=broad-except
        # Отключаем разбор разметки, чтобы спецсимволы из текста ошибки не ломали сообщение
        logger.exception(f"Не удалось сравнить файлы: {e}")
        await message.answer(
            "❌ Не удалось сравнить файлы.\n"
            "Подробности в logs/bot.log.",
            parse_mode=None
        )
    finally:
        await state.clear()


@router.message(CompareStates.waiting_smeta)
async def receive_smeta_wrong(message: Message, state: FSMContext):  # noqa: ARG001
    await message.answer("Нужно отправить именно файл нашей сметы (Excel-документ). Попробуйте ещё раз.")


#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Тесты превью XLSX списка плит после замены широких."""

import sys
import tempfile
import unittest
from pathlib import Path

TESTS_DIR = Path(__file__).parent
PROJECT_ROOT = TESTS_DIR.parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    from openpyxl import load_workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from core.plates_preview_xlsx import (
    build_plates_reconciliation_preview_xlsx,
    name_qty_pairs_for_contributions,
    preview_row_triples_for_contributions,
    qty_for_contribution_key,
)
from core.config_and_data import LineContributionKey, set_plate_lists_from_text


@unittest.skipUnless(HAS_OPENPYXL, "openpyxl не установлен")
class TestPlatesPreviewXlsx(unittest.TestCase):
    def test_two_contribution_keys_share_wide_qty(self) -> None:
        text = "ПБ 59-15-8п 2"
        _, contribs, _line_loads = set_plate_lists_from_text(text)
        self.assertEqual(len(contribs), 1)
        keys = contribs[0]
        self.assertEqual(len(keys), 2)

        import core.config_and_data as cfg

        details = dict(cfg.PLATE_LOAD_DETAILS)
        p1 = name_qty_pairs_for_contributions(keys, details)
        self.assertEqual(len(p1), 2)
        self.assertEqual(p1[0][1], 2)
        self.assertEqual(p1[1][1], 2)

    def test_preview_xlsx_column_a_and_two_pairs(self) -> None:
        text = "ПБ 59-15-8п 2"
        user_lines = ["ПБ 59-15-8п 2 шт"]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "preview.xlsx"
            build_plates_reconciliation_preview_xlsx(
                path,
                plates_text=text,
                initial_user_plate_lines=user_lines,
            )
            ws = load_workbook(path).active
            self.assertEqual(ws.cell(row=2, column=1).value, "ПБ 59-15-8п 2 шт")
            self.assertEqual(ws.cell(row=3, column=1).value, "ПБ 59-15-8п 2 шт")
            b2 = ws.cell(row=2, column=2).value
            c2 = ws.cell(row=2, column=3).value
            b3 = ws.cell(row=3, column=2).value
            c3 = ws.cell(row=3, column=3).value
            self.assertTrue(b2)
            self.assertTrue(b3)
            self.assertNotEqual(str(b2).strip(), str(b3).strip())
            self.assertEqual(c2, "2")
            self.assertEqual(c3, "2")
            # Обе позиции раскола должны присутствовать рядом (порядок не фиксируем).
            names = f"{b2}\n{b3}"
            self.assertIn("12-8", names)
            self.assertIn("0.3", names)
            self.assertEqual(ws.cell(row=2, column=4).value, b2)
            self.assertEqual(ws.cell(row=2, column=5).value, c2)
            self.assertEqual(ws.cell(row=3, column=4).value, b3)
            self.assertEqual(ws.cell(row=3, column=5).value, c3)

    def test_preview_duplicate_user_lines_blank_kp_columns(self) -> None:
        """Две одинаковые строки ввода: второй раз D–E пустые; C — построчное кол-во."""
        text = "ПБ 78-12-8п 3\nПБ 78-12-8п 3"
        user_lines = ["ПБ 78-12-8п 3", "ПБ 78-12-8п 3"]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "preview.xlsx"
            build_plates_reconciliation_preview_xlsx(
                path,
                plates_text=text,
                initial_user_plate_lines=user_lines,
            )
            ws = load_workbook(path).active
            self.assertEqual(ws.cell(row=2, column=3).value, "3")
            self.assertEqual(ws.cell(row=3, column=3).value, "3")
            self.assertTrue(ws.cell(row=2, column=4).value)
            self.assertEqual(ws.cell(row=2, column=5).value, "6")
            self.assertEqual(ws.cell(row=3, column=4).value, None)
            self.assertEqual(ws.cell(row=3, column=5).value, None)

    def test_preview_three_duplicate_user_lines_kp_global_sum(self) -> None:
        """Три одинаковые строки ввода: E — сумма по заказу на первой строке; D–E пустые на 2–3."""
        text = "ПБ 78-12-8п 3\nПБ 78-12-8п 3\nПБ 78-12-8п 3"
        user_lines = ["ПБ 78-12-8п 3", "ПБ 78-12-8п 3", "ПБ 78-12-8п 3"]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "preview.xlsx"
            build_plates_reconciliation_preview_xlsx(
                path,
                plates_text=text,
                initial_user_plate_lines=user_lines,
            )
            ws = load_workbook(path).active
            self.assertEqual(ws.cell(row=2, column=5).value, "9")
            self.assertTrue(ws.cell(row=2, column=4).value)
            self.assertEqual(ws.cell(row=3, column=4).value, None)
            self.assertEqual(ws.cell(row=3, column=5).value, None)
            self.assertEqual(ws.cell(row=4, column=4).value, None)
            self.assertEqual(ws.cell(row=4, column=5).value, None)

    def test_preview_grouped_single_kp_header_for_20_6_14(self) -> None:
        text = "ПБ 69-12-8п 20\nПБ 69-12-8п 6\nПБ 69-12-8п 14"
        user_lines = ["ПБ 69-12-8п 20", "ПБ 69-12-8п 6", "ПБ 69-12-8п 14"]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "preview.xlsx"
            build_plates_reconciliation_preview_xlsx(
                path,
                plates_text=text,
                initial_user_plate_lines=user_lines,
            )
            ws = load_workbook(path).active

            self.assertEqual(ws.cell(row=2, column=1).value, "ПБ 69-12-8п 20")
            self.assertEqual(ws.cell(row=3, column=1).value, "ПБ 69-12-8п 6")
            self.assertEqual(ws.cell(row=4, column=1).value, "ПБ 69-12-8п 14")

            self.assertEqual(ws.cell(row=2, column=3).value, "20")
            self.assertEqual(ws.cell(row=3, column=3).value, "6")
            self.assertEqual(ws.cell(row=4, column=3).value, "14")

            self.assertEqual(ws.cell(row=2, column=4).value, "Плиты ПБ 69-12-8п")
            self.assertEqual(ws.cell(row=2, column=5).value, "40")
            self.assertEqual(ws.cell(row=3, column=4).value, None)
            self.assertEqual(ws.cell(row=3, column=5).value, None)
            self.assertEqual(ws.cell(row=4, column=4).value, None)
            self.assertEqual(ws.cell(row=4, column=5).value, None)

    def test_preview_unparsed_line_empty_kp_columns_no_crash(self) -> None:
        """Строка без ключей вклада (не распознана): D–E пустые; ранее падало на d_val, e_val = \"\"."""
        text = "xxx"
        user_lines = ["xxx"]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "preview.xlsx"
            build_plates_reconciliation_preview_xlsx(
                path,
                plates_text=text,
                initial_user_plate_lines=user_lines,
            )
            ws = load_workbook(path).active
            self.assertEqual(ws.cell(row=2, column=1).value, "xxx")
            self.assertEqual(ws.cell(row=2, column=4).value, None)
            self.assertEqual(ws.cell(row=2, column=5).value, None)


class TestPreviewRowTriplesLimits(unittest.TestCase):
    def test_max_pairs_one_returns_single_row(self) -> None:
        _, contribs, _ = set_plate_lists_from_text("ПБ 59-15-8п 2")
        keys = contribs[0]
        import core.config_and_data as cfg

        details = dict(cfg.PLATE_LOAD_DETAILS)
        triples = preview_row_triples_for_contributions(
            keys, details, details, max_pairs=1
        )
        self.assertEqual(len(triples), 1)

    def test_three_distinct_keys_all_triples_without_limit(self) -> None:
        k1: LineContributionKey = (6.0, 1.2, 8, "60")
        k2: LineContributionKey = (6.0, 0.3, 8, "60")
        k3: LineContributionKey = (6.0, 1.0, 8, "60")
        det = {k1: 1, k2: 1, k3: 1}
        triples = preview_row_triples_for_contributions([k1, k2, k3], det, det, max_pairs=None)
        self.assertEqual(len(triples), 3)
        names = [t[0] for t in triples]
        self.assertEqual(len(set(names)), 3)
        self.assertIn("0.3", names[0])
        self.assertIn("10-8", names[1])  # ширина 1.0 м в марке (10 дм)
        self.assertIn("12-8", names[2])


class TestQtyForContribution(unittest.TestCase):
    def test_qty_for_split_width_uses_15_record(self) -> None:
        set_plate_lists_from_text("ПБ 59-15-8п 3")
        import core.config_and_data as cfg

        details = dict(cfg.PLATE_LOAD_DETAILS)
        k: LineContributionKey = (5.9, 1.2, 8.0, "59")
        self.assertEqual(qty_for_contribution_key(k, details), 3)


if __name__ == "__main__":
    unittest.main()

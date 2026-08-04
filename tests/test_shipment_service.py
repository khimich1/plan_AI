"""SHIP-200..205 + SHIP-500: ShipmentService CRUD/propose/confirm/complete/cancel, SGP-guard, XLSX."""

from __future__ import annotations

import json
import sqlite3
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.domain.enums import KpStatus, PlateStatus, PlateTransitionReason
from app.schemas.logistics import ShipmentItemInput, ShipmentOrderPatch
from app.services.archive_service import ArchiveService
from app.services.sgp_service import SgpError, SgpService
from app.services.shipment_service import ShipmentError, ShipmentService
from tests.helpers import kp_db_fixtures as fx

PLATE = "ПБ 60-12-8п"
UNIT_WEIGHT_KG = 2040.0  # 2.83333333 * 60 * 12
DATE = "2026-07-31"


@pytest.fixture
def db(tmp_path) -> str:
    return fx.make_iso_db(tmp_path)


@pytest.fixture
def svc(db: str) -> ShipmentService:
    return ShipmentService(db_path=db)


def _seed_completed(
    db_path: str,
    kp_id: int,
    qty: int,
    *,
    plate: str = PLATE,
    completed_date: str = "27.07.2026",
    day: int = 1,
) -> int:
    with sqlite3.connect(db_path) as conn:
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO completed_plates (
                kp_id, plate_name, length_m, width_m, load_class,
                qty, completed_date, production_day, plan_id
            ) VALUES (?, ?, 6.0, 1.2, 800, ?, ?, ?, 'plan-1')
            """,
            (kp_id, plate, qty, completed_date, day),
        )
        conn.commit()
        return int(cur.lastrowid)


def _create(svc: ShipmentService, kp_ids: list[int]) -> int:
    return svc.create(
        shipment_date=DATE, delivery_type="delivery", kp_ids=kp_ids, actor="tester"
    ).id


def _set_ya(svc: ShipmentService, shipment_id: int, kp_ids: list[int], ya: str) -> None:
    svc.patch(
        shipment_id,
        fields={},
        orders=[ShipmentOrderPatch(kp_id=kp_id, ya_order_no=ya) for kp_id in kp_ids],
    )


def _scalar(db_path: str, sql: str, params: tuple = ()):
    with sqlite3.connect(db_path) as conn:
        return conn.execute(sql, params).fetchone()[0]


# ---------------------------------------------------------------------------
# CRUD (SHIP-200)
# ---------------------------------------------------------------------------


def test_create_requires_kp(svc: ShipmentService) -> None:
    with pytest.raises(ShipmentError) as exc:
        svc.create(shipment_date=DATE, delivery_type="delivery", kp_ids=[])
    assert exc.value.code == "shipment_no_orders"


def test_create_validates_date(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    with pytest.raises(ShipmentError) as exc:
        svc.create(shipment_date="31.07.2026", delivery_type="delivery", kp_ids=[1])
    assert exc.value.code == "shipment_invalid_date"


def test_create_unknown_kp_404ish(svc: ShipmentService) -> None:
    with pytest.raises(ShipmentError) as exc:
        svc.create(shipment_date=DATE, delivery_type="delivery", kp_ids=[777])
    assert exc.value.code == "shipment_kp_not_found"


def test_create_and_get_card(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1, customer_name="КлиентА")
    fx.seed_kp_offer(db, 2, customer_name="КлиентБ")
    cp_id = _seed_completed(db, 1, 5)

    card = svc.create(
        shipment_date=DATE, delivery_type="delivery", kp_ids=[1, 2], actor="tester"
    )
    assert card.status == "in_work"
    assert [order.kp_id for order in card.orders] == [1, 2]
    assert card.orders[0].customer_name == "КлиентА"
    assert card.orders[0].ya_order_no is None
    assert card.items == []

    fetched = svc.get(card.id)
    assert fetched.available_by_kp[0].kp_id == 1
    assert fetched.available_by_kp[0].plates[0].completed_plate_id == cp_id
    assert fetched.available_by_kp[0].plates[0].available_qty == 5
    assert fetched.available_by_kp[1].kp_id == 2
    assert fetched.available_by_kp[1].plates == []

    with pytest.raises(ShipmentError) as exc:
        svc.get(4242)
    assert exc.value.code == "shipment_not_found"


def test_ya_order_prefill_from_latest_shipment(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    first_id = _create(svc, [1])
    _set_ya(svc, first_id, [1], "ЯР-101")

    second = svc.create(
        shipment_date="2026-08-01", delivery_type="pickup", kp_ids=[1]
    )
    assert second.orders[0].ya_order_no == "ЯР-101"


def test_patch_fields_and_orders(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    fx.seed_kp_offer(db, 2)
    shipment_id = _create(svc, [1])

    card = svc.patch(
        shipment_id,
        fields={"driver_name": "Пётр", "vehicle_class": "t20", "attention": 1},
        orders=[ShipmentOrderPatch(kp_id=2, ya_order_no="ЯР-202")],
    )
    assert card.driver_name == "Пётр"
    assert card.vehicle_class == "t20"
    assert card.attention is True
    assert [order.kp_id for order in card.orders] == [2]
    assert card.orders[0].ya_order_no == "ЯР-202"


def test_patch_orders_full_replacement_add_remove_edit(svc: ShipmentService, db: str) -> None:
    """Один PATCH: добавление КП + удаление КП + правка ЯР — транзакционно."""
    fx.seed_kp_offer(db, 1)
    fx.seed_kp_offer(db, 2)
    fx.seed_kp_offer(db, 3)
    shipment_id = _create(svc, [1, 2])
    _set_ya(svc, shipment_id, [1, 2], "ЯР-old")

    card = svc.patch(
        shipment_id,
        fields={},
        orders=[
            ShipmentOrderPatch(kp_id=2, ya_order_no="ЯР-new"),  # правка ЯР
            ShipmentOrderPatch(kp_id=3, ya_order_no="ЯР-3"),  # добавление
        ],  # КП 1 удалён
    )
    assert [(order.kp_id, order.ya_order_no) for order in card.orders] == [
        (2, "ЯР-new"),
        (3, "ЯР-3"),
    ]
    # PATCH возвращает полную карточку — идентичную GET.
    fetched = svc.get(shipment_id)
    assert card.model_dump() == fetched.model_dump()

    with sqlite3.connect(db) as conn:
        rows = conn.execute(
            "SELECT kp_id, ya_order_no FROM shipment_orders WHERE shipment_id = ? ORDER BY id",
            (shipment_id,),
        ).fetchall()
    assert rows == [(2, "ЯР-new"), (3, "ЯР-3")]


def test_patch_orders_unknown_kp_rolls_back(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    shipment_id = _create(svc, [1])
    with pytest.raises(ShipmentError) as exc:
        svc.patch(
            shipment_id,
            fields={},
            orders=[ShipmentOrderPatch(kp_id=999, ya_order_no="ЯР-x")],
        )
    assert exc.value.code == "shipment_kp_not_found"
    # Старый состав заказов не пострадал.
    assert [order.kp_id for order in svc.get(shipment_id).orders] == [1]


def test_patch_unknown_carrier_rejected(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    shipment_id = _create(svc, [1])
    with pytest.raises(ShipmentError) as exc:
        svc.patch(shipment_id, fields={"carrier_id": 999})
    assert exc.value.code == "shipment_carrier_not_found"


def test_list_shipments_filters(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    fx.seed_kp_offer(db, 2)
    first = _create(svc, [1])
    second = svc.create(
        shipment_date="2026-08-01", delivery_type="pickup", kp_ids=[2]
    ).id
    svc.patch(first, fields={"attention": 1})

    assert svc.list_shipments().count == 2
    assert svc.list_shipments(kp_id=1).items[0].id == first
    assert svc.list_shipments(delivery_type="pickup").items[0].id == second
    assert svc.list_shipments(attention=True).items[0].id == first
    assert svc.list_shipments(date_from="2026-08-01").items[0].id == second
    assert svc.list_shipments(status="done").count == 0
    assert svc.list_shipments(no_upd=True).count == 2


# ---------------------------------------------------------------------------
# Propose (SHIP-201)
# ---------------------------------------------------------------------------


def test_propose_fifo_and_vehicle_limit(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    first_cp = _seed_completed(db, 1, 5, completed_date="25.07.2026", day=1)
    second_cp = _seed_completed(db, 1, 8, completed_date="27.07.2026", day=2)
    shipment_id = _create(svc, [1])

    response = svc.propose(shipment_id, vehicle_class="t20")
    # 20000 // 2040 = 9 плит: первая строка целиком (5), вторая — 4 шт.
    assert response.vehicle_class == "t20"
    assert response.vehicle_class_limits_kg["t20"] == 19800
    assert response.total_weight_kg == pytest.approx(9 * UNIT_WEIGHT_KG)
    assert response.overload is False
    assert [(item.completed_plate_id, item.qty) for item in response.items] == [
        (first_cp, 5),
        (second_cp, 4),
    ]
    assert [(item.completed_plate_id, item.qty) for item in response.not_fit] == [
        (second_cp, 4)
    ]


def test_propose_without_vehicle_class_uses_t20_v2(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    _seed_completed(db, 1, 13)
    shipment_id = _create(svc, [1])

    response = svc.propose(shipment_id)
    assert response.vehicle_class == "t20"
    assert response.total_weight_kg == pytest.approx(9 * UNIT_WEIGHT_KG)
    assert sum(item.qty for item in response.items) == 9
    assert sum(item.qty for item in response.not_fit) == 4
    assert response.overload is False


def test_propose_skips_reserved_qty(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    cp_id = _seed_completed(db, 1, 5)
    first_id = _create(svc, [1])
    svc.put_items(
        first_id,
        [ShipmentItemInput(item_type="plate", completed_plate_id=cp_id, qty=3)],
    )
    second_id = _create(svc, [1])

    response = svc.propose(second_id)
    assert [(item.completed_plate_id, item.qty) for item in response.items] == [(cp_id, 2)]


def test_propose_saves_snapshot(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    _seed_completed(db, 1, 2)
    shipment_id = _create(svc, [1])
    svc.propose(shipment_id, vehicle_class="t30plus")
    snapshot_raw = _scalar(
        db, "SELECT propose_snapshot FROM shipments WHERE id = ?", (shipment_id,)
    )
    snapshot = json.loads(snapshot_raw)
    assert snapshot["vehicle_class"] == "t30plus"
    assert len(snapshot["items"]) == 1
    assert snapshot["layout"] is None


def test_propose_saves_layout_snapshot(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    _seed_completed(db, 1, 2)
    shipment_id = _create(svc, [1])

    response = svc.propose(shipment_id, vehicle_class="t20")
    layout = response.layout
    assert layout is not None
    assert layout.body_length_m == pytest.approx(13.2)
    assert layout.body_used_m == pytest.approx(6.0)
    assert len(layout.stacks) == 1
    assert layout.stacks[0].marking_length_m == pytest.approx(6.0)
    assert [u.plate_name for u in layout.stacks[0].tiers[0].units] == [PLATE, PLATE]
    assert [(s.step, s.description) for s in layout.loading_steps] == [
        (1, f"{PLATE} ×2")
    ]

    snapshot_raw = _scalar(
        db, "SELECT propose_snapshot FROM shipments WHERE id = ?", (shipment_id,)
    )
    snapshot = json.loads(snapshot_raw)
    assert snapshot["layout"] is not None
    assert snapshot["layout"]["stacks"][0]["tiers"][0]["units"][0]["plate_name"] == PLATE
    assert snapshot["layout"]["loading_steps"][0]["description"] == f"{PLATE} ×2"


def test_propose_unknown_vehicle_class(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    shipment_id = _create(svc, [1])
    with pytest.raises(ShipmentError) as exc:
        svc.propose(shipment_id, vehicle_class="ufo")
    assert exc.value.code == "shipment_invalid_vehicle_class"


# ---------------------------------------------------------------------------
# Confirm (SHIP-202)
# ---------------------------------------------------------------------------


def test_put_items_weights_and_overrides(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    cp_id = _seed_completed(db, 1, 5)
    with sqlite3.connect(db) as conn:
        conn.execute(
            "INSERT INTO pile_catalog (mark, length_m, section_mm, volume_m3, weight_kg, pcs_per_20t) "
            "VALUES ('С60.30', 6.0, 300, 0.216, 1060.0, 18)"
        )
        conn.commit()
    shipment_id = _create(svc, [1])

    card = svc.put_items(
        shipment_id,
        [
            ShipmentItemInput(item_type="plate", completed_plate_id=cp_id, qty=2),
            ShipmentItemInput(item_type="free", mark="С60.30", qty=2),
            ShipmentItemInput(item_type="free", mark="НЕИЗВЕСТНО", qty=1, weight_kg=500.0),
            ShipmentItemInput(item_type="free", mark="НЕИЗВЕСТНО2", qty=1, kp_id=1),
            ShipmentItemInput(
                item_type="plate", completed_plate_id=cp_id, qty=1, weight_kg=999.0
            ),
        ],
    )
    plate_item = card.items[0]
    assert plate_item.unit_weight_kg == pytest.approx(UNIT_WEIGHT_KG)
    assert plate_item.weight_kg == pytest.approx(2 * UNIT_WEIGHT_KG)
    assert card.items[1].weight_kg == pytest.approx(2 * 1060.0)
    assert card.items[2].weight_kg == pytest.approx(500.0)  # ручная правка сохранена
    assert card.items[3].weight_kg is None
    assert card.items[3].kp_id == 1
    assert card.items[4].weight_kg == pytest.approx(999.0)  # override важнее формулы


def test_put_items_no_availability(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    cp_id = _seed_completed(db, 1, 5)
    shipment_id = _create(svc, [1])

    with pytest.raises(ShipmentError) as exc:
        svc.put_items(
            shipment_id,
            [ShipmentItemInput(item_type="plate", completed_plate_id=cp_id, qty=6)],
        )
    assert exc.value.code == "shipment_no_availability"

    # Две строки на одну плиту суммируются.
    with pytest.raises(ShipmentError) as exc:
        svc.put_items(
            shipment_id,
            [
                ShipmentItemInput(item_type="plate", completed_plate_id=cp_id, qty=3),
                ShipmentItemInput(item_type="plate", completed_plate_id=cp_id, qty=3),
            ],
        )
    assert exc.value.code == "shipment_no_availability"
    # Откат: состав пуст.
    assert svc.get(shipment_id).items == []


def test_put_items_full_replacement(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    cp_id = _seed_completed(db, 1, 5)
    shipment_id = _create(svc, [1])
    svc.put_items(
        shipment_id,
        [ShipmentItemInput(item_type="plate", completed_plate_id=cp_id, qty=3)],
    )
    card = svc.put_items(
        shipment_id,
        [ShipmentItemInput(item_type="plate", completed_plate_id=cp_id, qty=5)],
    )
    # Замена проходит: резерв старого состава (3) не мешает новому (5).
    assert [(item.qty) for item in card.items] == [5]


def test_put_items_plate_must_be_linked(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    with sqlite3.connect(db) as conn:
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO completed_plates (
                kp_id, plate_name, length_m, width_m, load_class, qty, completed_date
            ) VALUES (NULL, ?, 6.0, 1.2, 800, 2, '27.07.2026')
            """,
            (PLATE,),
        )
        conn.commit()
        unlinked_cp = int(cur.lastrowid)
    shipment_id = _create(svc, [1])
    with pytest.raises(ShipmentError) as exc:
        svc.put_items(
            shipment_id,
            [ShipmentItemInput(item_type="plate", completed_plate_id=unlinked_cp, qty=1)],
        )
    assert exc.value.code == "shipment_plate_unlinked"


# ---------------------------------------------------------------------------
# Complete (SHIP-203)
# ---------------------------------------------------------------------------


def _prepared_shipment(
    svc: ShipmentService, db: str, kp_id: int, cp_qty: int, ship_qty: int
) -> tuple[int, int]:
    fx.seed_kp_offer(db, kp_id)
    with sqlite3.connect(db) as conn:
        conn.execute(
            "UPDATE kp_meta SET ordered_qty = ? WHERE kp_id = ?", (cp_qty, kp_id)
        )
        conn.commit()
    cp_id = _seed_completed(db, kp_id, cp_qty)
    shipment_id = _create(svc, [kp_id])
    svc.put_items(
        shipment_id,
        [ShipmentItemInput(item_type="plate", completed_plate_id=cp_id, qty=ship_qty)],
    )
    return shipment_id, cp_id


def test_complete_partial_deduct_and_audit(svc: ShipmentService, db: str) -> None:
    shipment_id, cp_id = _prepared_shipment(svc, db, 1, 5, 3)
    _set_ya(svc, shipment_id, [1], "ЯР-1")

    response = svc.complete(shipment_id, actor="logist")
    assert response.status == "done"

    assert _scalar(db, "SELECT qty FROM completed_plates WHERE id = ?", (cp_id,)) == 2
    audit = _scalar(
        db,
        "SELECT COUNT(*) FROM plate_status_log WHERE reason = ? AND shipment_id = ?",
        (PlateTransitionReason.SGP_SHIP.value, shipment_id),
    )
    assert audit == 1
    row = _scalar(
        db,
        "SELECT from_status || '>' || to_status || '>' || qty FROM plate_status_log "
        "WHERE shipment_id = ?",
        (shipment_id,),
    )
    assert row == f"{PlateStatus.ON_SGP.value}>{PlateStatus.SHIPPED.value}>3"
    assert _scalar(db, "SELECT actor FROM plate_status_log WHERE shipment_id = ?", (shipment_id,)) == "logist"
    card = svc.get(shipment_id)
    assert card.status == "done"
    assert card.completed_at is not None
    # Частичная отгрузка → КП НЕ выполнено (остаток на складе → «На СГП»).
    assert _scalar(db, "SELECT status FROM kp_meta WHERE kp_id = 1") == KpStatus.ON_SGP.value


def test_complete_full_sets_kp_done(svc: ShipmentService, db: str) -> None:
    shipment_id, cp_id = _prepared_shipment(svc, db, 1, 5, 5)
    _set_ya(svc, shipment_id, [1], "ЯР-1")
    svc.complete(shipment_id)

    assert _scalar(db, "SELECT status FROM kp_meta WHERE kp_id = 1") == KpStatus.DONE.value
    # Строка склада обнулена, но не удалена (FK + карточка рейса).
    assert _scalar(db, "SELECT qty FROM completed_plates WHERE id = ?", (cp_id,)) == 0
    assert svc.get(shipment_id).items[0].plate_name == PLATE


def test_complete_missing_ya_order_atomic(svc: ShipmentService, db: str) -> None:
    shipment_id, cp_id = _prepared_shipment(svc, db, 1, 5, 3)

    with pytest.raises(ShipmentError) as exc:
        svc.complete(shipment_id)
    assert exc.value.code == "shipment_missing_ya_order"

    # 422 → БД неизменна.
    assert _scalar(db, "SELECT qty FROM completed_plates WHERE id = ?", (cp_id,)) == 5
    assert _scalar(db, "SELECT status FROM shipments WHERE id = ?", (shipment_id,)) == "in_work"
    assert _scalar(db, "SELECT COUNT(*) FROM plate_status_log") == 0


def test_complete_no_availability_atomic(svc: ShipmentService, db: str) -> None:
    shipment_id, cp_id = _prepared_shipment(svc, db, 1, 5, 3)
    _set_ya(svc, shipment_id, [1], "ЯР-1")
    with sqlite3.connect(db) as conn:
        conn.execute("UPDATE completed_plates SET qty = 1 WHERE id = ?", (cp_id,))
        conn.commit()

    with pytest.raises(ShipmentError) as exc:
        svc.complete(shipment_id)
    assert exc.value.code == "shipment_no_availability"
    assert _scalar(db, "SELECT status FROM shipments WHERE id = ?", (shipment_id,)) == "in_work"
    assert _scalar(db, "SELECT COUNT(*) FROM plate_status_log") == 0


def test_complete_empty_items_rejected(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    shipment_id = _create(svc, [1])
    _set_ya(svc, shipment_id, [1], "ЯР-1")
    with pytest.raises(ShipmentError) as exc:
        svc.complete(shipment_id)
    assert exc.value.code == "shipment_no_items"


def test_complete_done_shipment_rejected(svc: ShipmentService, db: str) -> None:
    shipment_id, _cp_id = _prepared_shipment(svc, db, 1, 5, 5)
    _set_ya(svc, shipment_id, [1], "ЯР-1")
    svc.complete(shipment_id)

    with pytest.raises(ShipmentError) as exc:
        svc.complete(shipment_id)
    assert exc.value.code == "shipment_not_in_work"
    with pytest.raises(ShipmentError) as exc:
        svc.patch(shipment_id, fields={"driver_name": "X"})
    assert exc.value.code == "shipment_not_in_work"


def test_complete_multi_kp_done_boundary(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    fx.seed_kp_offer(db, 2)
    with sqlite3.connect(db) as conn:
        conn.executemany(
            "UPDATE kp_meta SET ordered_qty = ? WHERE kp_id = ?",
            [(5, 1), (4, 2)],
        )
        conn.commit()
    cp1 = _seed_completed(db, 1, 5)
    cp2 = _seed_completed(db, 2, 4)
    shipment_id = _create(svc, [1, 2])
    svc.put_items(
        shipment_id,
        [
            ShipmentItemInput(item_type="plate", completed_plate_id=cp1, qty=5),
            ShipmentItemInput(item_type="plate", completed_plate_id=cp2, qty=3),
        ],
    )
    _set_ya(svc, shipment_id, [1, 2], "ЯР-77")
    svc.complete(shipment_id)

    # КП 1 отгружено полностью → выполнено; КП 2 частично → «На СГП» (остаток на складе).
    assert _scalar(db, "SELECT status FROM kp_meta WHERE kp_id = 1") == KpStatus.DONE.value
    assert _scalar(db, "SELECT status FROM kp_meta WHERE kp_id = 2") == KpStatus.ON_SGP.value
    assert _scalar(
        db, "SELECT COUNT(*) FROM plate_status_log WHERE reason = ?",
        (PlateTransitionReason.SGP_SHIP.value,),
    ) == 2


def test_complete_event_flag_on_off(svc: ShipmentService, db: str, tmp_path) -> None:
    shipment_id, _cp_id = _prepared_shipment(svc, db, 1, 5, 5)
    _set_ya(svc, shipment_id, [1], "ЯР-9")
    export_dir = tmp_path / "events"

    svc.complete(shipment_id, events_enabled=False, export_dir=str(export_dir))
    assert not export_dir.exists()

    shipment2_id, _cp2 = _prepared_shipment(svc, db, 3, 2, 2)
    _set_ya(svc, shipment2_id, [3], "ЯР-10")
    svc.complete(shipment2_id, events_enabled=True, export_dir=str(export_dir))

    files = list(export_dir.glob("shipment_completed_*.json"))
    assert len(files) == 1
    payload = json.loads(files[0].read_text(encoding="utf-8"))
    assert payload["event"] == "shipment_completed"
    assert payload["version"] == 1
    assert payload["shipment_id"] == shipment2_id
    assert payload["orders"] == [{"kp_id": 3, "ya_order_no": "ЯР-10", "uid_kp": None}]
    assert payload["items"][0]["type"] == "plate"
    assert payload["items"][0]["plate_name"] == PLATE
    assert payload["items"][0]["qty"] == 2
    assert payload["total_weight_kg"] == pytest.approx(2 * UNIT_WEIGHT_KG)


# ---------------------------------------------------------------------------
# Cancel (SHIP-204)
# ---------------------------------------------------------------------------


def test_cancel_restores_availability_no_audit(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    cp_id = _seed_completed(db, 1, 5)
    shipment_id = _create(svc, [1])
    svc.put_items(
        shipment_id,
        [ShipmentItemInput(item_type="plate", completed_plate_id=cp_id, qty=3)],
    )
    assert svc.get(shipment_id).available_by_kp[0].plates[0].available_qty == 2

    response = svc.cancel(shipment_id)
    assert response.ok is True

    with pytest.raises(ShipmentError) as exc:
        svc.get(shipment_id)
    assert exc.value.code == "shipment_not_found"
    assert _scalar(db, "SELECT COUNT(*) FROM shipment_items") == 0
    assert _scalar(db, "SELECT COUNT(*) FROM shipment_orders") == 0
    assert _scalar(db, "SELECT COUNT(*) FROM plate_status_log") == 0

    other_id = _create(svc, [1])
    assert svc.get(other_id).available_by_kp[0].plates[0].available_qty == 5


def test_cancel_done_rejected(svc: ShipmentService, db: str) -> None:
    shipment_id, _cp_id = _prepared_shipment(svc, db, 1, 5, 5)
    _set_ya(svc, shipment_id, [1], "ЯР-1")
    svc.complete(shipment_id)
    with pytest.raises(ShipmentError) as exc:
        svc.cancel(shipment_id)
    assert exc.value.code == "shipment_not_in_work"


# ---------------------------------------------------------------------------
# SGP guard (SHIP-205)
# ---------------------------------------------------------------------------


def test_sgp_unlink_blocked_on_allocated_qty(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    cp_id = _seed_completed(db, 1, 5)
    shipment_id = _create(svc, [1])
    svc.put_items(
        shipment_id,
        [ShipmentItemInput(item_type="plate", completed_plate_id=cp_id, qty=3)],
    )

    sgp = SgpService(db_path=db)
    with pytest.raises(SgpError) as exc:
        sgp.unlink(cp_id, 3, actor="test")
    assert exc.value.code == "sgp_row_allocated"
    assert f"#{shipment_id}" in str(exc.value)

    # Свободная часть (2 из 5) по-прежнему отвязывается как раньше.
    sgp.unlink(cp_id, 2, actor="test")
    assert _scalar(db, "SELECT COALESCE(SUM(qty),0) FROM kp_plates WHERE kp_id = 1") == 2
    assert _scalar(db, "SELECT qty FROM completed_plates WHERE id = ?", (cp_id,)) == 3


def test_sgp_relink_blocked_on_allocated_qty(svc: ShipmentService, db: str) -> None:
    """relink работает только с отвязанными строками — резерв на них приходит
    из legacy-данных (shipment_items на уже отвязанную плиту)."""
    fx.seed_kp_offer(db, 1)
    fx.seed_plate(
        db, kp_id=1, plate_name=PLATE, length_m=6.0, width_m=1.2,
        load_class=800, qty=5, status="в производстве",
    )
    with sqlite3.connect(db) as conn:
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO completed_plates (
                kp_id, plate_name, length_m, width_m, load_class, qty, completed_date
            ) VALUES (NULL, ?, 6.0, 1.2, 800, 5, '27.07.2026')
            """,
            (PLATE,),
        )
        cp_id = int(cur.lastrowid)
        conn.commit()
    shipment_id = _create(svc, [1])
    with sqlite3.connect(db) as conn:
        conn.execute(
            "INSERT INTO shipment_items (shipment_id, item_type, completed_plate_id, qty) "
            "VALUES (?, 'plate', ?, 3)",
            (shipment_id, cp_id),
        )
        conn.commit()

    sgp = SgpService(db_path=db)
    with pytest.raises(SgpError) as exc:
        sgp.relink(cp_id, target_kp_id=1, qty=3, actor="test")
    assert exc.value.code == "sgp_row_allocated"

    # Свободная часть (2 из 5) перепривязывается как раньше.
    sgp.relink(cp_id, target_kp_id=1, qty=2, actor="test")
    assert _scalar(
        db, "SELECT COALESCE(SUM(qty),0) FROM completed_plates WHERE kp_id = 1"
    ) == 2


# ---------------------------------------------------------------------------
# Лист отгрузки XLSX (SHIP-500)
# ---------------------------------------------------------------------------


def test_export_shipment_sheet_xlsx(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1, customer_name="КлиентА")
    cp_id = _seed_completed(db, 1, 5)
    shipment_id = _create(svc, [1])
    _set_ya(svc, shipment_id, [1], "ЯР-5")
    svc.patch(shipment_id, fields={"driver_name": "Пётр", "vehicle_text": "MAN А123ВС77"})
    svc.put_items(
        shipment_id,
        [
            ShipmentItemInput(
                item_type="free", mark="С60.30", qty=1, weight_kg=1060.0, sort_order=1
            ),
            ShipmentItemInput(
                item_type="plate",
                completed_plate_id=cp_id,
                qty=2,
                sort_order=0,
                note="хрупкое",
            ),
        ],
    )

    content = svc.export_shipment_sheet_xlsx(shipment_id)
    wb = load_workbook(BytesIO(content))
    ws = wb.active

    assert "Лист отгрузки" in ws["A1"].value
    header_cells = [ws.cell(row=r, column=1).value for r in range(2, 8)]
    assert any("Заказы" in str(value) and "ЯР-5" in str(value) for value in header_cells)
    assert any("Заказчик" in str(value) and "КлиентА" in str(value) for value in header_cells)
    assert any("Водитель" in str(value) and "Пётр" in str(value) for value in header_cells)

    table_header_row = next(
        row[0].row for row in ws.iter_rows(min_col=1, max_col=1) if row[0].value == "№"
    )
    assert [cell.value for cell in ws[table_header_row]] == [
        "№", "Марка / плита", "Размеры, м", "Кол-во, шт", "Вес, кг", "Примечание",
    ]
    # sort_order: плита (0) первая, свая (1) вторая.
    row1 = [cell.value for cell in ws[table_header_row + 1]]
    row2 = [cell.value for cell in ws[table_header_row + 2]]
    assert row1[1] == PLATE and row1[3] == 2
    assert row1[4] == pytest.approx(2 * UNIT_WEIGHT_KG)
    assert row1[5] == "хрупкое"
    assert row2[1] == "С60.30" and row2[3] == 1
    totals = [cell.value for cell in ws[table_header_row + 3]]
    assert totals[1] == "Итого"
    assert totals[3] == 3
    assert totals[4] == pytest.approx(2 * UNIT_WEIGHT_KG + 1060.0)


# ---------------------------------------------------------------------------
# Реестр: фильтры date_to / carrier / «без УПД» (SC-1, дополнение)
# ---------------------------------------------------------------------------


def test_list_shipments_date_to_filter(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    july_id = _create(svc, [1])
    august_id = svc.create(
        shipment_date="2026-08-15", delivery_type="delivery", kp_ids=[1]
    ).id

    assert [item.id for item in svc.list_shipments(date_to="2026-08-01").items] == [july_id]
    ranged = svc.list_shipments(date_from="2026-07-01", date_to="2026-08-31")
    assert {item.id for item in ranged.items} == {july_id, august_id}


def _seed_carrier(db_path: str, name: str = 'ООО «Альфа»', normalized: str = "альфа") -> int:
    with sqlite3.connect(db_path) as conn:
        cur = conn.execute(
            "INSERT INTO carriers (name, name_normalized) VALUES (?, ?)",
            (name, normalized),
        )
        conn.commit()
        return int(cur.lastrowid)


def test_list_shipments_carrier_filter(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    carrier_id = _seed_carrier(db)
    with_carrier = _create(svc, [1])
    svc.patch(with_carrier, fields={"carrier_id": carrier_id})
    without_carrier = _create(svc, [1])

    result = svc.list_shipments(carrier_id=carrier_id)
    assert [item.id for item in result.items] == [with_carrier]
    assert result.items[0].carrier_name == 'ООО «Альфа»'
    assert without_carrier not in [item.id for item in result.items]


def test_list_shipments_no_upd_excludes_filled_upd(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    with_upd = _create(svc, [1])
    svc.patch(with_upd, fields={"upd_no": "101"})
    without_upd = _create(svc, [1])

    assert [item.id for item in svc.list_shipments(no_upd=True).items] == [without_upd]


# ---------------------------------------------------------------------------
# Propose: мульти-заказный рейс (SC-2/SC-6)
# ---------------------------------------------------------------------------


def test_propose_multi_kp_fifo_per_order_and_foreign_plates_excluded(
    svc: ShipmentService, db: str
) -> None:
    fx.seed_kp_offer(db, 1)
    fx.seed_kp_offer(db, 2)
    fx.seed_kp_offer(db, 3)
    cp1_old = _seed_completed(db, 1, 2, completed_date="25.07.2026", day=1)
    cp1_new = _seed_completed(db, 1, 3, completed_date="27.07.2026", day=2)
    cp2 = _seed_completed(db, 2, 4, completed_date="26.07.2026", day=1)
    _seed_completed(db, 3, 9)  # КП вне рейса — предлагаться не должно
    with sqlite3.connect(db) as conn:
        conn.execute(
            """
            INSERT INTO completed_plates (
                kp_id, plate_name, length_m, width_m, load_class, qty, completed_date
            ) VALUES (NULL, ?, 6.0, 1.2, 800, 7, '27.07.2026')
            """,
            (PLATE,),
        )
        conn.commit()
    shipment_id = _create(svc, [1, 2])

    response = svc.propose(shipment_id)

    # Порядок: КП в порядке добавления к рейсу, внутри — FIFO по completed_date.
    assert [(item.kp_id, item.completed_plate_id, item.qty) for item in response.items] == [
        (1, cp1_old, 2),
        (1, cp1_new, 3),
        (2, cp2, 4),
    ]
    assert response.not_fit == []
    assert response.total_weight_kg == pytest.approx(9 * UNIT_WEIGHT_KG)


# ---------------------------------------------------------------------------
# Complete: мульти-заказный рейс, ЯР обязателен у каждого КП (SC-6)
# ---------------------------------------------------------------------------


def test_complete_multi_kp_missing_ya_on_one_order_atomic(
    svc: ShipmentService, db: str
) -> None:
    fx.seed_kp_offer(db, 1)
    fx.seed_kp_offer(db, 2)
    cp1 = _seed_completed(db, 1, 5)
    cp2 = _seed_completed(db, 2, 4)
    shipment_id = _create(svc, [1, 2])
    svc.put_items(
        shipment_id,
        [
            ShipmentItemInput(item_type="plate", completed_plate_id=cp1, qty=2),
            ShipmentItemInput(item_type="plate", completed_plate_id=cp2, qty=2),
        ],
    )
    svc.patch(
        shipment_id,
        fields={},
        orders=[
            ShipmentOrderPatch(kp_id=1, ya_order_no="ЯР-1"),
            ShipmentOrderPatch(kp_id=2, ya_order_no=None),
        ],
    )

    with pytest.raises(ShipmentError) as exc:
        svc.complete(shipment_id)
    assert exc.value.code == "shipment_missing_ya_order"

    # Откат: ни одна плита не списана, audit пуст, рейс остался в работе.
    assert _scalar(db, "SELECT qty FROM completed_plates WHERE id = ?", (cp1,)) == 5
    assert _scalar(db, "SELECT qty FROM completed_plates WHERE id = ?", (cp2,)) == 4
    assert _scalar(db, "SELECT COUNT(*) FROM plate_status_log") == 0
    assert _scalar(db, "SELECT status FROM shipments WHERE id = ?", (shipment_id,)) == "in_work"


# ---------------------------------------------------------------------------
# Complete: границы критерия KpStatus.DONE (SC-5, Never «DONE при неполном qty»)
# ---------------------------------------------------------------------------


def test_complete_full_but_plates_in_production_blocks_kp_done(
    svc: ShipmentService, db: str
) -> None:
    """Отгружено всё заказное M, но в kp_plates осталась потребность (в производстве)
    → «выполнено» не ставится, даже когда shipped >= ordered_qty."""
    shipment_id, cp_id = _prepared_shipment(svc, db, 1, 5, 5)
    fx.seed_plate(
        db, kp_id=1, plate_name=PLATE, length_m=6.0, width_m=1.2,
        qty=2, status="в производстве",
    )
    _set_ya(svc, shipment_id, [1], "ЯР-1")

    svc.complete(shipment_id)

    assert _scalar(db, "SELECT status FROM kp_meta WHERE kp_id = 1") == KpStatus.IN_WORK.value
    # Списание при этом прошло штатно.
    assert _scalar(db, "SELECT qty FROM completed_plates WHERE id = ?", (cp_id,)) == 0


def test_complete_without_ordered_qty_never_marks_done(svc: ShipmentService, db: str) -> None:
    """Deviation 4: ordered_qty <= 0 (M неизвестно) → автоматический DONE запрещён."""
    shipment_id, _cp_id = _prepared_shipment(svc, db, 1, 5, 5)
    # M «не зафиксировано» (guard срабатывает на ordered_qty <= 0).
    with sqlite3.connect(db) as conn:
        conn.execute("UPDATE kp_meta SET ordered_qty = 0 WHERE kp_id = 1")
        conn.commit()
    _set_ya(svc, shipment_id, [1], "ЯР-1")

    svc.complete(shipment_id)

    assert _scalar(db, "SELECT status FROM kp_meta WHERE kp_id = 1") != KpStatus.DONE.value


# ---------------------------------------------------------------------------
# Архив КП: бейдж «отгружено X/M» (SHIP-301, SC-4)
# ---------------------------------------------------------------------------


class _ArchiveRepoStub:
    """Минимальный репозиторий архива поверх реальной БД (для shipped_progress)."""

    def __init__(self, db_path: str, raw: dict) -> None:
        self.db_path = db_path
        self._raw = raw

    def list_by_section(self, section: str, *, product_type: str = "all", **filters) -> list:
        return [self._raw]

    def get_completion_percentage(self, kp_id: int) -> dict:
        return {"percentage": 0.0}


def _archive_shipped_badge(db_path: str, tmp_path, kp_id: int) -> dict | None:
    status = _scalar(db_path, "SELECT status FROM kp_meta WHERE kp_id = ?", (kp_id,))
    raw = {
        "kp_id": kp_id,
        "creation_date": "2026-01-01",
        "customer_name": "КлиентА",
        "status": status,
    }
    service = ArchiveService(
        repository=_ArchiveRepoStub(db_path, raw), outputs_dir=tmp_path
    )
    items = service.list_offers("in_production", user={"id": 1, "role": "admin"})
    return items[0].shipped_progress


def test_archive_shipped_progress_partial_shipment(svc: ShipmentService, db: str, tmp_path) -> None:
    """Отгружено 3 из 5 → архив показывает «отгружено 3/5», КП не выполнен."""
    shipment_id, _cp_id = _prepared_shipment(svc, db, 1, 5, 3)
    _set_ya(svc, shipment_id, [1], "ЯР-1")
    svc.complete(shipment_id)

    assert _archive_shipped_badge(db, tmp_path, 1) == {"x": 3, "m": 5}


def test_archive_shipped_progress_accumulates_across_done_shipments(
    svc: ShipmentService, db: str, tmp_path
) -> None:
    shipment_id, cp_id = _prepared_shipment(svc, db, 1, 5, 3)
    _set_ya(svc, shipment_id, [1], "ЯР-1")
    svc.complete(shipment_id)

    second_id = _create(svc, [1])
    svc.put_items(
        second_id,
        [ShipmentItemInput(item_type="plate", completed_plate_id=cp_id, qty=2)],
    )
    _set_ya(svc, second_id, [1], "ЯР-2")
    svc.complete(second_id)

    assert _archive_shipped_badge(db, tmp_path, 1) == {"x": 5, "m": 5}
    assert _scalar(db, "SELECT status FROM kp_meta WHERE kp_id = 1") == KpStatus.DONE.value


# ---------------------------------------------------------------------------
# Переиспользование транспорта (REUSE-100)
# ---------------------------------------------------------------------------


def test_reuse_transport_copies_whitelist_only(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1, customer_name="КлиентА")
    fx.seed_kp_offer(db, 2, customer_name="КлиентБ")
    carrier_id = _seed_carrier(db)
    cp_id = _seed_completed(db, 1, 3)
    source_id = _create(svc, [1])
    _set_ya(svc, source_id, [1], "ЯР-SRC")
    svc.patch(
        source_id,
        fields={
            "carrier_id": carrier_id,
            "driver_name": "Пётр",
            "vehicle_text": "MAN А123ВС77",
            "vehicle_class": "t20",
            "proxy_no": "дов-9",
            "upd_no": "УПД-77",
            "freight_request_no": "З-100",
            "planned_cost": 45000.0,
            "attention": 1,
            "attention_comment": "кран",
        },
    )
    svc.put_items(
        source_id,
        [ShipmentItemInput(item_type="plate", completed_plate_id=cp_id, qty=2)],
    )

    card = svc.reuse_transport(
        source_id,
        shipment_date="2026-08-02",
        delivery_type="pickup",
        kp_ids=[2],
        actor="tester",
    )

    assert card.id != source_id
    assert card.status == "in_work"
    assert card.shipment_date == "2026-08-02"
    assert card.delivery_type == "pickup"
    assert card.carrier_id == carrier_id
    assert card.driver_name == "Пётр"
    assert card.vehicle_text == "MAN А123ВС77"
    assert card.vehicle_class == "t20"
    assert card.proxy_no == "дов-9"
    assert card.upd_no is None
    assert card.freight_request_no is None
    assert card.planned_cost is None
    assert card.attention is False
    assert card.attention_comment is None
    assert card.items == []
    assert [order.kp_id for order in card.orders] == [2]
    assert all(order.ya_order_no != "ЯР-SRC" for order in card.orders)


def test_reuse_transport_works_for_done_source(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    fx.seed_kp_offer(db, 2)
    source_id = _create(svc, [1])
    svc.patch(source_id, fields={"driver_name": "Иван", "vehicle_class": "t30"})
    with sqlite3.connect(db) as conn:
        conn.execute("UPDATE shipments SET status = 'done' WHERE id = ?", (source_id,))
        conn.commit()

    card = svc.reuse_transport(
        source_id,
        shipment_date="2026-08-03",
        delivery_type="delivery",
        kp_ids=[2],
    )
    assert card.status == "in_work"
    assert card.driver_name == "Иван"
    assert card.vehicle_class == "t30"


def test_reuse_transport_empty_source_transport(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    fx.seed_kp_offer(db, 2)
    source_id = _create(svc, [1])

    card = svc.reuse_transport(
        source_id,
        shipment_date="2026-08-02",
        delivery_type="delivery",
        kp_ids=[2],
    )
    assert card.status == "in_work"
    assert card.carrier_id is None
    assert card.driver_name is None
    assert card.vehicle_text is None
    assert card.vehicle_class is None
    assert card.proxy_no is None


def test_reuse_transport_source_not_found(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    with pytest.raises(ShipmentError) as exc:
        svc.reuse_transport(
            9999,
            shipment_date=DATE,
            delivery_type="delivery",
            kp_ids=[1],
        )
    assert exc.value.code == "shipment_not_found"


def test_reuse_transport_requires_kp(svc: ShipmentService, db: str) -> None:
    fx.seed_kp_offer(db, 1)
    source_id = _create(svc, [1])
    with pytest.raises(ShipmentError) as exc:
        svc.reuse_transport(
            source_id,
            shipment_date=DATE,
            delivery_type="delivery",
            kp_ids=[],
        )
    assert exc.value.code == "shipment_no_orders"


# ---------------------------------------------------------------------------
# Событие shipment_completed: полный контракт папки обмена (SC-13)
# ---------------------------------------------------------------------------


def test_event_payload_matches_exchange_contract(svc: ShipmentService, db: str) -> None:
    """Plate + free строки, перевозчик, водитель, УПД — по контракту-черновику 1С."""
    fx.seed_kp_offer(db, 1)
    carrier_id = _seed_carrier(db, name="ООО ТрансЛогистик", normalized="транслогистик")
    with sqlite3.connect(db) as conn:
        conn.execute("UPDATE kp_meta SET ordered_qty = 5 WHERE kp_id = 1")
        conn.commit()
    cp_id = _seed_completed(db, 1, 5)
    shipment_id = _create(svc, [1])
    svc.patch(
        shipment_id,
        fields={
            "carrier_id": carrier_id,
            "driver_name": "Иванов И.И.",
            "vehicle_text": "Volvo FH / а123бв77",
            "upd_no": "1234",
        },
    )
    svc.put_items(
        shipment_id,
        [
            ShipmentItemInput(item_type="plate", completed_plate_id=cp_id, qty=3),
            ShipmentItemInput(item_type="free", mark="С60.30", qty=14, weight_kg=19320.0),
        ],
    )
    _set_ya(svc, shipment_id, [1], "ЯР-0001467")
    svc.complete(shipment_id, events_enabled=False)

    payload = svc.build_shipment_event_payload(shipment_id)

    assert payload["event"] == "shipment_completed"
    assert payload["version"] == 1
    assert payload["shipment_id"] == shipment_id
    assert payload["shipment_date"] == DATE
    assert payload["completed_at"] is not None
    assert payload["delivery_type"] == "delivery"
    assert payload["orders"] == [{"kp_id": 1, "ya_order_no": "ЯР-0001467", "uid_kp": None}]
    plate_item, free_item = payload["items"]
    assert plate_item["type"] == "plate"
    assert plate_item["plate_name"] == PLATE
    assert plate_item["length_m"] == 6.0
    assert plate_item["width_m"] == 1.2
    assert plate_item["nomenclature_id"] is None
    assert plate_item["qty"] == 3
    assert plate_item["weight_kg"] == pytest.approx(3 * UNIT_WEIGHT_KG)
    assert free_item == {"type": "free", "mark": "С60.30", "qty": 14, "weight_kg": 19320.0}
    assert payload["carrier"] == {"name": "ООО ТрансЛогистик"}
    assert payload["driver_name"] == "Иванов И.И."
    assert payload["vehicle_text"] == "Volvo FH / а123бв77"
    assert payload["upd_no"] == "1234"
    assert payload["total_weight_kg"] == pytest.approx(3 * UNIT_WEIGHT_KG + 19320.0)


def test_search_pile_catalog_sqlite_error_message_sanitized(
    svc: ShipmentService, monkeypatch: pytest.MonkeyPatch
) -> None:
    class BoomCursor:
        def execute(self, *args, **kwargs):
            raise sqlite3.Error("no such table: secret_schema_xyz")

    class BoomConn:
        row_factory = None

        def cursor(self):
            return BoomCursor()

        def close(self):
            return None

    monkeypatch.setattr(
        "app.services.shipment_service._connect", lambda db_path: BoomConn()
    )
    with pytest.raises(ShipmentError) as exc:
        svc.search_pile_catalog("С60")
    assert exc.value.code == "pile_catalog_read_failed"
    assert str(exc.value) == "Ошибка чтения каталога свай"
    assert "secret_schema_xyz" not in str(exc.value)
    assert "no such table" not in str(exc.value)

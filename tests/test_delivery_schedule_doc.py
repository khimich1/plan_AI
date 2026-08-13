"""T8: документ графика поставки — XLSX/PDF build + generate_document без перезаписи."""

from __future__ import annotations

from pathlib import Path

import pytest

pytest.importorskip("openpyxl")
pytest.importorskip("reportlab")

from openpyxl import load_workbook

from app.domain.enums import KpStatus
from app.schemas.delivery_schedule import BatchIn, BatchItemIn, DeliverySchedulePut
from app.services.delivery_schedule_service import DeliveryScheduleService
from core import kp_db_schema
from core.delivery_schedule_pdf import build_document as build_pdf_document
from core.delivery_schedule_xlsx import (
    DOC_HEADERS,
    build_document as build_xlsx_document,
)
from core.kp_db_common import _connect

ADMIN = {"id": 1, "role": "admin"}

_SAMPLE_SCHEDULE = {
    "kp_id": 42,
    "invoice_number": "СЧ-101",
    "contract_number": "Д-5",
    "customer_name": "ООО Тест",
    "batches": [
        {
            "name": "1 этаж",
            "deliver_from": "2026-09-01",
            "deliver_to": "2026-09-10",
            "produce_by": "2026-08-25",
            "items": [
                {"plate_id": 100, "plate_name": "ПБ 60-12-8п", "qty": 3},
            ],
        }
    ],
}


def _fresh_db(tmp_path: Path, name: str = "plita.db") -> str:
    db_path = str(tmp_path / name)
    kp_db_schema._schema_ready.clear()
    kp_db_schema.ensure_schema(db_path)
    return db_path


def _seed_kp(
    db_path: str,
    *,
    kp_id: int = 1,
    plate_qty: int = 10,
    plate_name: str = "ПБ 60-12-8п",
) -> int:
    with _connect(db_path) as conn:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO KP_offers (kp_id, creation_date, customer_name) "
            "VALUES (?, '2026-08-07', 'ООО Тест')",
            (kp_id,),
        )
        cur.execute(
            "INSERT INTO kp_meta (kp_id, status, owner_user_id, product_type) "
            "VALUES (?, ?, ?, 'plates')",
            (kp_id, KpStatus.IN_WORK.value, 1),
        )
        cur.execute(
            """
            INSERT INTO kp_plates (
                kp_id, position_number, plate_name, qty,
                length_m, width_m, load_class
            )
            VALUES (?, 1, ?, ?, NULL, NULL, NULL)
            """,
            (kp_id, plate_name, plate_qty),
        )
        plate_id = int(cur.lastrowid)
        conn.commit()
    return plate_id


def _payload(plate_id: int, *, qty: int = 3) -> DeliverySchedulePut:
    return DeliverySchedulePut(
        invoice_number="СЧ-101",
        contract_number="Д-5",
        batches=[
            BatchIn(
                name="1 этаж",
                deliver_from="2026-09-01",
                deliver_to="2026-09-10",
                produce_by="2026-08-25",
                items=[BatchItemIn(plate_id=plate_id, qty=qty)],
                sort_order=0,
            )
        ],
    )


def test_build_document_xlsx_writes_header_and_rows(tmp_path: Path) -> None:
    path = tmp_path / "schedule.xlsx"
    result = build_xlsx_document(_SAMPLE_SCHEDULE, path)

    assert result == path
    assert path.is_file()
    assert path.stat().st_size > 0

    wb = load_workbook(path)
    ws = wb.active
    assert ws.title == "График поставки"
    assert ws.cell(row=1, column=1).value == "График поставки"
    assert ws.cell(row=2, column=1).value == "Договор: Д-5"
    assert ws.cell(row=3, column=1).value == "Счёт: СЧ-101"
    assert ws.cell(row=4, column=1).value == "Заказчик: ООО Тест"

    # пустая строка, затем заголовки таблицы
    header_row = 6
    for col, header in enumerate(DOC_HEADERS, start=1):
        assert ws.cell(row=header_row, column=col).value == header

    data_row = header_row + 1
    assert ws.cell(row=data_row, column=1).value == 1
    assert ws.cell(row=data_row, column=2).value == "1 этаж"
    assert ws.cell(row=data_row, column=3).value == "01.09.2026"
    assert ws.cell(row=data_row, column=4).value == "10.09.2026"
    assert ws.cell(row=data_row, column=5).value == "25.08.2026"
    assert ws.cell(row=data_row, column=6).value == "ПБ 60-12-8п"
    assert ws.cell(row=data_row, column=7).value == 3


def test_build_document_pdf_creates_nonempty_pdf(tmp_path: Path) -> None:
    path = tmp_path / "schedule.pdf"
    result = build_pdf_document(_SAMPLE_SCHEDULE, path)

    assert result == path
    assert path.is_file()
    content = path.read_bytes()
    assert len(content) > 0
    assert content.startswith(b"%PDF")


def test_generate_document_twice_writes_distinct_paths(tmp_path: Path) -> None:
    db_path = _fresh_db(tmp_path)
    plate_id = _seed_kp(db_path, kp_id=1)
    outputs = tmp_path / "outputs"
    service = DeliveryScheduleService(db_path=db_path, outputs_dir=outputs)
    service.replace(1, _payload(plate_id, qty=3), ADMIN)

    first = service.generate_document(1, "xlsx", ADMIN)
    second = service.generate_document(1, "xlsx", ADMIN)

    assert first != second
    assert first.is_file()
    assert second.is_file()
    assert first.exists()
    assert second.exists()
    assert first.read_bytes()  # первый файл не затёрт
    assert first.name.startswith("График_КП1_ред_")
    assert second.name.startswith("График_КП1_ред_")
    assert first.suffix == ".xlsx"
    assert second.suffix == ".xlsx"

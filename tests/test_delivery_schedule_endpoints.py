"""T6/T7b: HTTP endpoints графика поставки (GET/PUT + /template, /import)."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock

import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient

from app.core.settings import get_settings
from app.dependencies.services import get_delivery_schedule_service
from app.main import create_app
from app.schemas.delivery_schedule import (
    BatchDraftItemOut,
    BatchDraftOut,
    BatchItemOut,
    BatchOut,
    DeliveryScheduleView,
    ImportDraftResponse,
    UnmatchedRowOut,
)
from app.security.offer_access import FORBIDDEN_OFFER_DETAIL
from app.security.session import create_session_token
from app.services.delivery_schedule_service import (
    DeliveryScheduleNotFoundError,
    DeliveryScheduleValidationError,
)
from tests.helpers.auth_fixtures import patch_auth_users
from tests.helpers.csrf import CsrfAwareTestClient

API = "/api/v1/commercial/archive/{kp_id}/delivery-schedule"
TEMPLATE_API = API + "/template"
IMPORT_API = API + "/import"
DOCUMENT_API = API + "/document"
XLSX_MEDIA = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
PDF_MEDIA = "application/pdf"

USERS = [
    {
        "id": 1,
        "username": "admin",
        "role": "admin",
        "manager_id": None,
        "is_active": 1,
        "created_at": "2026-01-01 00:00:00",
        "session_version": 0,
    },
    {
        "id": 2,
        "username": "manager_a",
        "role": "manager",
        "manager_id": None,
        "is_active": 1,
        "created_at": "2026-01-01 00:00:00",
        "session_version": 0,
    },
    {
        "id": 3,
        "username": "customer_user",
        "role": "customer",
        "manager_id": None,
        "is_active": 1,
        "created_at": "2026-01-01 00:00:00",
        "session_version": 0,
    },
    {
        "id": 4,
        "username": "manager_b",
        "role": "manager",
        "manager_id": None,
        "is_active": 1,
        "created_at": "2026-01-01 00:00:00",
        "session_version": 0,
    },
]


@pytest.fixture()
def fake_service() -> MagicMock:
    return MagicMock()


@pytest.fixture()
def client(
    monkeypatch: pytest.MonkeyPatch,
    fake_service: MagicMock,
) -> TestClient:
    monkeypatch.setenv("APP_SECRET_KEY", "test-secret-key-for-pytest-must-be-32-chars-min")
    get_settings.cache_clear()
    patch_auth_users(monkeypatch, USERS)
    app = create_app()
    app.dependency_overrides[get_delivery_schedule_service] = lambda: fake_service
    return CsrfAwareTestClient(app)


def _session_cookie(user_id: int, role: str, username: str) -> dict[str, str]:
    return {
        "app_session": create_session_token(
            {"id": user_id, "username": username, "role": role},
            ttl_seconds=300,
        )
    }


def _admin_cookie() -> dict[str, str]:
    return _session_cookie(1, "admin", "admin")


def _manager_cookie() -> dict[str, str]:
    return _session_cookie(2, "manager", "manager_a")


def _manager_b_cookie() -> dict[str, str]:
    return _session_cookie(4, "manager", "manager_b")


def _customer_cookie() -> dict[str, str]:
    return _session_cookie(3, "customer", "customer_user")


def _fake_view(
    *,
    kp_id: int = 42,
    schedule_id: int = 7,
    plate_id: int = 100,
    qty: int = 3,
) -> DeliveryScheduleView:
    return DeliveryScheduleView(
        id=schedule_id,
        kp_id=kp_id,
        invoice_number="СЧ-101",
        contract_number="Д-5",
        status="draft",
        batches=[
            BatchOut(
                id=1,
                name="1 этаж",
                deliver_from="2026-09-01",
                deliver_to="2026-09-10",
                produce_by="2026-08-25",
                items=[
                    BatchItemOut(
                        plate_id=plate_id,
                        qty=qty,
                        plate_name="ПБ 60-12-8п",
                    )
                ],
                sort_order=0,
            )
        ],
        updated_at="2026-08-07T12:00:00",
    )


def _put_payload(*, plate_id: int = 100, qty: int = 3) -> dict:
    return {
        "invoice_number": "СЧ-101",
        "contract_number": "Д-5",
        "batches": [
            {
                "name": "1 этаж",
                "deliver_from": "2026-09-01",
                "deliver_to": "2026-09-10",
                "produce_by": "2026-08-25",
                "items": [{"plate_id": plate_id, "qty": qty}],
                "sort_order": 0,
            }
        ],
    }


def _filled_template_bytes(tmp_path: Path, rows: list[tuple]) -> bytes:
    """Пустой шаблон + строки данных через openpyxl; возвращает bytes файла."""
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    from core.delivery_schedule_xlsx import build_template

    path = tmp_path / "filled_schedule.xlsx"
    build_template(path)
    wb = load_workbook(path)
    ws = wb.active
    for r_idx, values in enumerate(rows, start=2):
        for c_idx, value in enumerate(values, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(path)
    return path.read_bytes()


def _import_draft_from_bytes(
    kp_id: int, file_bytes: bytes, *, user: dict | None = None
) -> ImportDraftResponse:
    """Разбор XLSX как в сервисе (без БД) — для side_effect MagicMock."""
    from core.delivery_schedule_xlsx import parse_template

    _ = kp_id
    _ = user
    drafts, unmatched = parse_template(
        file_bytes,
        [{"id": 10, "plate_name": "ПБ 60-12-8"}],
    )
    return ImportDraftResponse(
        batches=[
            BatchDraftOut(
                name=draft.name,
                deliver_from=draft.deliver_from,
                deliver_to=draft.deliver_to,
                produce_by=draft.produce_by,
                items=[
                    BatchDraftItemOut(
                        plate_id=item.plate_id,
                        plate_name=item.plate_name,
                        qty=item.qty,
                    )
                    for item in draft.items
                ],
            )
            for draft in drafts
        ],
        unmatched_rows=[
            UnmatchedRowOut(
                row_number=row.row_number,
                reason=row.reason,
                raw=row.raw,
            )
            for row in unmatched
        ],
    )


def test_get_requires_auth(client: TestClient) -> None:
    response = client.get(API.format(kp_id=42))
    assert response.status_code in (401, 403)


def test_get_customer_forbidden(client: TestClient) -> None:
    response = client.get(API.format(kp_id=42), cookies=_customer_cookie())
    assert response.status_code == 403
    assert response.json()["detail"] == "Forbidden"


def test_put_customer_forbidden(client: TestClient, fake_service: MagicMock) -> None:
    response = client.put(
        API.format(kp_id=42),
        json=_put_payload(),
        cookies=_customer_cookie(),
    )
    assert response.status_code == 403
    fake_service.replace.assert_not_called()


@pytest.mark.parametrize(
    "cookies,user",
    [
        (_admin_cookie(), USERS[0]),
        (_manager_cookie(), USERS[1]),
    ],
    ids=["admin", "manager"],
)
def test_put_creates_and_get_returns(
    client: TestClient,
    fake_service: MagicMock,
    cookies: dict[str, str],
    user: dict,
) -> None:
    view = _fake_view(kp_id=42, qty=3)
    fake_service.replace.return_value = view
    fake_service.get.return_value = view

    put_response = client.put(
        API.format(kp_id=42),
        json=_put_payload(qty=3),
        cookies=cookies,
    )
    assert put_response.status_code == 200
    put_payload = put_response.json()
    assert put_payload["kp_id"] == 42
    assert put_payload["invoice_number"] == "СЧ-101"
    assert len(put_payload["batches"]) == 1
    assert put_payload["batches"][0]["items"][0]["qty"] == 3
    fake_service.replace.assert_called_once()
    call_args = fake_service.replace.call_args
    assert call_args.args[0] == 42
    assert call_args.args[2]["id"] == user["id"]
    assert call_args.args[2]["role"] == user["role"]

    get_response = client.get(API.format(kp_id=42), cookies=cookies)
    assert get_response.status_code == 200
    get_payload = get_response.json()
    assert get_payload["kp_id"] == 42
    assert get_payload["batches"][0]["name"] == "1 этаж"
    assert get_payload["batches"][0]["items"][0]["qty"] == 3
    fake_service.get.assert_called_once()
    get_call = fake_service.get.call_args
    assert get_call.args[0] == 42
    assert get_call.kwargs["user"]["id"] == user["id"]
    assert get_call.kwargs["user"]["role"] == user["role"]


def test_get_not_found(
    client: TestClient,
    fake_service: MagicMock,
) -> None:
    fake_service.get.side_effect = DeliveryScheduleNotFoundError(
        "График поставки для КП №999 не найден"
    )

    response = client.get(API.format(kp_id=999), cookies=_admin_cookie())

    assert response.status_code == 404
    fake_service.get.assert_called_once()
    get_call = fake_service.get.call_args
    assert get_call.args[0] == 999
    assert get_call.kwargs["user"]["id"] == 1


def test_get_foreign_manager_returns_403(
    client: TestClient,
    fake_service: MagicMock,
) -> None:
    fake_service.get.side_effect = HTTPException(
        status_code=403, detail=FORBIDDEN_OFFER_DETAIL
    )

    response = client.get(API.format(kp_id=42), cookies=_manager_b_cookie())

    assert response.status_code == 403
    assert response.json()["detail"] == FORBIDDEN_OFFER_DETAIL
    fake_service.get.assert_called_once()
    assert fake_service.get.call_args.kwargs["user"]["id"] == 4


def test_import_foreign_manager_returns_403(
    client: TestClient,
    fake_service: MagicMock,
) -> None:
    fake_service.import_draft.side_effect = HTTPException(
        status_code=403, detail=FORBIDDEN_OFFER_DETAIL
    )

    response = client.post(
        IMPORT_API.format(kp_id=42),
        files={"file": ("schedule.xlsx", b"PK\x03\x04", XLSX_MEDIA)},
        cookies=_manager_b_cookie(),
    )

    assert response.status_code == 403
    assert response.json()["detail"] == FORBIDDEN_OFFER_DETAIL
    fake_service.import_draft.assert_called_once()
    assert fake_service.import_draft.call_args.kwargs["user"]["id"] == 4


def test_get_own_kp_without_schedule_returns_404(
    client: TestClient,
    fake_service: MagicMock,
) -> None:
    fake_service.get.side_effect = DeliveryScheduleNotFoundError(
        "График поставки для КП №42 не найден"
    )

    response = client.get(API.format(kp_id=42), cookies=_manager_cookie())

    assert response.status_code == 404
    fake_service.get.assert_called_once()
    assert fake_service.get.call_args.kwargs["user"]["id"] == 2


def test_put_qty_exceeded_returns_422(
    client: TestClient,
    fake_service: MagicMock,
) -> None:
    fake_service.replace.side_effect = DeliveryScheduleValidationError(
        "Сумма qty по plate_id=100 превышает qty позиции КП"
    )

    response = client.put(
        API.format(kp_id=42),
        json=_put_payload(qty=99),
        cookies=_admin_cookie(),
    )

    assert response.status_code == 422
    fake_service.replace.assert_called_once()


def test_openapi_contains_delivery_schedule_path(client: TestClient) -> None:
    response = client.get("/openapi.json")
    assert response.status_code == 200
    paths = response.json().get("paths", {})
    assert any("delivery-schedule" in path for path in paths)
    target = API.format(kp_id="{kp_id}")
    assert target in paths
    assert "get" in paths[target]
    assert "put" in paths[target]


def test_get_template_as_admin(
    client: TestClient,
    fake_service: MagicMock,
    tmp_path: Path,
) -> None:
    pytest.importorskip("openpyxl")
    from core.delivery_schedule_xlsx import build_template

    path = tmp_path / "delivery_schedule_template.xlsx"
    build_template(path)
    template_bytes = path.read_bytes()
    fake_service.build_template_bytes.return_value = template_bytes

    response = client.get(
        TEMPLATE_API.format(kp_id=42),
        cookies=_admin_cookie(),
    )

    assert response.status_code == 200
    assert XLSX_MEDIA in response.headers["content-type"]
    assert len(response.content) > 0
    assert response.content == template_bytes
    fake_service.build_template_bytes.assert_called_once()
    call = fake_service.build_template_bytes.call_args
    assert call.args[0] == 42
    assert call.kwargs["user"]["id"] == 1


def test_post_import_valid_xlsx_returns_draft_without_db_schedule(
    client: TestClient,
    fake_service: MagicMock,
    tmp_path: Path,
) -> None:
    xlsx_bytes = _filled_template_bytes(
        tmp_path,
        [
            ("1 этаж", "01.09.2026", "10.09.2026", "25.08.2026", "ПБ 60-12-8", 3),
        ],
    )
    fake_service.import_draft.side_effect = _import_draft_from_bytes
    fake_service.get.side_effect = DeliveryScheduleNotFoundError(
        "График поставки для КП №42 не найден"
    )

    import_response = client.post(
        IMPORT_API.format(kp_id=42),
        files={"file": ("schedule.xlsx", xlsx_bytes, XLSX_MEDIA)},
        cookies=_admin_cookie(),
    )

    assert import_response.status_code == 200
    payload = import_response.json()
    assert payload["batches"] or payload["unmatched_rows"]
    if payload["batches"]:
        assert payload["batches"][0]["name"] == "1 этаж"
        assert payload["batches"][0]["items"][0]["plate_id"] == 10
        assert payload["batches"][0]["items"][0]["qty"] == 3
        assert payload["unmatched_rows"] == []
    fake_service.import_draft.assert_called_once()
    call_args = fake_service.import_draft.call_args
    assert call_args.args[0] == 42
    assert call_args.args[1] == xlsx_bytes
    assert call_args.kwargs["user"]["id"] == 1
    fake_service.replace.assert_not_called()

    get_response = client.get(API.format(kp_id=42), cookies=_admin_cookie())
    assert get_response.status_code == 404
    fake_service.get.assert_called_once()
    assert fake_service.get.call_args.args[0] == 42
    assert fake_service.get.call_args.kwargs["user"]["id"] == 1


@pytest.mark.parametrize(
    "method,url_template",
    [
        ("get", TEMPLATE_API),
        ("post", IMPORT_API),
    ],
    ids=["template", "import"],
)
def test_template_import_customer_forbidden(
    client: TestClient,
    fake_service: MagicMock,
    method: str,
    url_template: str,
) -> None:
    url = url_template.format(kp_id=42)
    if method == "get":
        response = client.get(url, cookies=_customer_cookie())
    else:
        response = client.post(
            url,
            files={"file": ("schedule.xlsx", b"PK\x03\x04", XLSX_MEDIA)},
            cookies=_customer_cookie(),
        )

    assert response.status_code == 403
    assert response.json()["detail"] == "Forbidden"
    fake_service.build_template_bytes.assert_not_called()
    fake_service.import_draft.assert_not_called()


def test_get_template_kp_not_found(
    client: TestClient,
    fake_service: MagicMock,
) -> None:
    fake_service.build_template_bytes.side_effect = DeliveryScheduleNotFoundError(
        "КП №999 не найдено"
    )

    response = client.get(
        TEMPLATE_API.format(kp_id=999),
        cookies=_admin_cookie(),
    )

    assert response.status_code == 404
    fake_service.build_template_bytes.assert_called_once()
    assert fake_service.build_template_bytes.call_args.args[0] == 999


def test_get_document_xlsx_returns_200(
    client: TestClient,
    fake_service: MagicMock,
    tmp_path: Path,
) -> None:
    target = tmp_path / "График_КП42_ред_2026-08-07.xlsx"
    target.write_bytes(b"PK\x03\x04xlsx-doc")
    fake_service.generate_document.return_value = target

    response = client.get(
        DOCUMENT_API.format(kp_id=42) + "?fmt=xlsx",
        cookies=_admin_cookie(),
    )

    assert response.status_code == 200
    assert XLSX_MEDIA in response.headers["content-type"]
    assert response.content == b"PK\x03\x04xlsx-doc"
    fake_service.generate_document.assert_called_once()
    call_args = fake_service.generate_document.call_args
    assert call_args.args[0] == 42
    assert call_args.args[1] == "xlsx"


def test_get_document_pdf_returns_200(
    client: TestClient,
    fake_service: MagicMock,
    tmp_path: Path,
) -> None:
    target = tmp_path / "График_КП42_ред_2026-08-07.pdf"
    target.write_bytes(b"%PDF-1.4 mock")
    fake_service.generate_document.return_value = target

    response = client.get(
        DOCUMENT_API.format(kp_id=42) + "?fmt=pdf",
        cookies=_admin_cookie(),
    )

    assert response.status_code == 200
    assert PDF_MEDIA in response.headers["content-type"]
    assert response.content.startswith(b"%PDF")
    fake_service.generate_document.assert_called_once()
    call_args = fake_service.generate_document.call_args
    assert call_args.args[0] == 42
    assert call_args.args[1] == "pdf"

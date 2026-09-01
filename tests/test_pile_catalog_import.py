"""SHIP-100 / PT-001: парсинг прайса свай + upsert каталога (идемпотентность)."""

from __future__ import annotations

import sqlite3
from pathlib import Path

import pytest
from openpyxl import Workbook

from core.pile_catalog import (
    PileCatalogEntry,
    parse_bridge_pile_geometry,
    parse_pile_catalog_from_xlsx,
    parse_pile_mark,
    resolve_catalog_for_mark,
    upsert_pile_catalog,
)
from tests.helpers import kp_db_fixtures as fx

QUIRK_MARK = "С137,5.40"
REAL_XLSX = Path(__file__).resolve().parents[1] / "банк знаний" / "сваи вес и объем.xlsx"


def _build_price_xlsx(tmp_path, rows: list[tuple]) -> str:
    """Лист «Вес и объем»: заголовок в строке 2, данные со строки 3."""
    path = str(tmp_path / "price.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Вес и объем"
    ws.append(["Прайс цельных свай"])
    ws.append(["Марка", "Объем, м3", "Вес, кг", "Шт. на а/м 20тн"])
    for row in rows:
        ws.append(list(row))
    wb.save(path)
    return path


def _standard_rows(count: int = 43) -> list[tuple]:
    rows = [(f"С{30 + i}.30", 0.1, 1000 + i, 10) for i in range(count)]
    rows.append((QUIRK_MARK, 0.55, 2750, "-"))
    return rows


@pytest.mark.parametrize(
    ("mark", "expected_length", "expected_section"),
    [
        ("С60.30", 6.0, 300),
        (QUIRK_MARK, 13.75, 400),
        ("С120.35", 12.0, 350),
        ("мусор", None, None),
        ("С60", None, None),
    ],
)
def test_parse_pile_mark(mark, expected_length, expected_section) -> None:
    assert parse_pile_mark(mark) == (expected_length, expected_section)


def test_parse_full_sheet_44_rows_and_quirks(tmp_path) -> None:
    xlsx = _build_price_xlsx(
        tmp_path,
        _standard_rows()
        + [
            (QUIRK_MARK, 0.55, 2750, "-"),  # дубликат марки — пропускается
            ("С99.30", None, None, 5),  # без веса — пропускается
            (None, 0.1, 100, 5),  # пустая марка — пропускается
        ],
    )
    entries = parse_pile_catalog_from_xlsx(xlsx)
    assert len(entries) == 44
    quirk = next(entry for entry in entries if entry.mark == QUIRK_MARK)
    assert quirk.length_m == pytest.approx(13.75)
    assert quirk.section_mm == 400
    assert quirk.weight_kg == pytest.approx(2750.0)
    assert quirk.pcs_per_20t is None  # "-" → NULL


def test_parse_missing_sheet_raises(tmp_path) -> None:
    xlsx = _build_price_xlsx(tmp_path, _standard_rows(1))
    with pytest.raises(ValueError, match="не найден"):
        parse_pile_catalog_from_xlsx(xlsx, sheet="Нет такого")


def test_upsert_idempotent(tmp_path) -> None:
    db_path = fx.make_iso_db(tmp_path)
    entries = [
        PileCatalogEntry(
            mark="С60.30",
            length_m=6.0,
            section_mm=300,
            volume_m3=0.216,
            weight_kg=1060.0,
            pcs_per_20t=18,
        ),
        PileCatalogEntry(
            mark=QUIRK_MARK,
            length_m=13.75,
            section_mm=400,
            volume_m3=0.55,
            weight_kg=2750.0,
            pcs_per_20t=None,
        ),
    ]
    inserted, updated = upsert_pile_catalog(db_path, entries)
    assert (inserted, updated) == (2, 0)
    inserted, updated = upsert_pile_catalog(db_path, entries)
    assert (inserted, updated) == (0, 2)

    changed = [
        PileCatalogEntry(
            mark="С60.30",
            length_m=6.0,
            section_mm=300,
            volume_m3=0.216,
            weight_kg=1111.0,
            pcs_per_20t=17,
        )
    ]
    inserted, updated = upsert_pile_catalog(db_path, changed)
    assert (inserted, updated) == (0, 1)
    with sqlite3.connect(db_path) as conn:
        row = conn.execute(
            "SELECT weight_kg, pcs_per_20t FROM pile_catalog WHERE mark = ?",
            ("С60.30",),
        ).fetchone()
        total = conn.execute("SELECT COUNT(*) FROM pile_catalog").fetchone()[0]
    assert row == (1111.0, 17)
    assert total == 2


def test_end_to_end_parse_then_upsert(tmp_path) -> None:
    db_path = fx.make_iso_db(tmp_path)
    xlsx = _build_price_xlsx(tmp_path, _standard_rows())
    entries = parse_pile_catalog_from_xlsx(xlsx)
    inserted, updated = upsert_pile_catalog(db_path, entries)
    assert (inserted, updated) == (44, 0)
    inserted, updated = upsert_pile_catalog(db_path, entries)
    assert (inserted, updated) == (0, 44)
    with sqlite3.connect(db_path) as conn:
        total = conn.execute("SELECT COUNT(*) FROM pile_catalog").fetchone()[0]
    assert total == 44


def _build_list1_xlsx(tmp_path, rows: list[tuple], *, extra_title: bool = True) -> str:
    """Лист «Лист1»: опциональная строка-заголовок блока, затем шапка с «автомобильный 20тн»."""
    path = str(tmp_path / "list1.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"
    if extra_title:
        ws.append([None, None, None, "Количество штук на транспорте", None, None])
    ws.append(
        [
            "марка сваи",
            "объем м3",
            "вес, кг",
            "автомобильный г/п 20тн",
            "ж/д МПС г/п 66тн",
            "ж/д трансконтейнер г/п 68тн.",
        ]
    )
    for row in rows:
        ws.append(list(row))
    wb.save(path)
    return path


def test_parse_list1_sheet_fallback_and_dash_pcs(tmp_path) -> None:
    xlsx = _build_list1_xlsx(
        tmp_path,
        [
            ("С140.40", 2.26, 5650, 3, "-", 12),
            ("С160.40", 2.58, 6450, "-", "-", 10),
            ("С160.30", 1.45, 3625, "", "-", 18),
        ],
    )
    entries = parse_pile_catalog_from_xlsx(xlsx)
    by_mark = {e.mark: e for e in entries}
    assert set(by_mark) == {"С140.40", "С160.40", "С160.30"}
    assert by_mark["С140.40"].weight_kg == pytest.approx(5650.0)
    assert by_mark["С140.40"].pcs_per_20t == 3
    assert by_mark["С160.40"].pcs_per_20t is None
    assert by_mark["С160.30"].pcs_per_20t is None


def test_parse_detects_header_and_auto_20t_column(tmp_path) -> None:
    """Шапка не во 2-й строке; колонка авто 20 т не четвёртая."""
    path = str(tmp_path / "shifted.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"
    ws.append(["служебная"])
    ws.append(["ещё одна"])
    ws.append(
        [
            "марка сваи",
            "ж/д МПС г/п 66тн",
            "объем м3",
            "вес, кг",
            "автомобильный г/п 20тн",
        ]
    )
    ws.append(["С140.40", "-", 2.26, 5650, 3])
    ws.append(["С160.40", "-", 2.58, 6450, "-"])
    wb.save(path)

    entries = parse_pile_catalog_from_xlsx(path)
    by_mark = {e.mark: e for e in entries}
    assert by_mark["С140.40"].weight_kg == pytest.approx(5650.0)
    assert by_mark["С140.40"].pcs_per_20t == 3
    assert by_mark["С140.40"].volume_m3 == pytest.approx(2.26)
    assert by_mark["С160.40"].pcs_per_20t is None


def test_parse_real_xlsx_44_marks_if_present() -> None:
    if not REAL_XLSX.is_file():
        return
    entries = parse_pile_catalog_from_xlsx(str(REAL_XLSX), sheet="Лист1")
    assert len(entries) == 44
    by_mark = {e.mark: e for e in entries}
    assert by_mark["С140.40"].weight_kg == pytest.approx(5650.0)
    assert by_mark["С140.40"].pcs_per_20t == 3
    for mark in ("С160.30", "С160.35", "С160.40"):
        assert by_mark[mark].pcs_per_20t is None
    fallback = parse_pile_catalog_from_xlsx(str(REAL_XLSX))
    assert {e.mark for e in fallback} == {e.mark for e in entries}


@pytest.mark.parametrize(
    ("mark", "expected"),
    [
        ("C14-40T4", (14.0, 400)),
        ("С14-40Т4", (14.0, 400)),
        ("C9-35T6", (9.0, 350)),
        ("C18-40T8", (18.0, 400)),
        ("С140.40", (None, None)),
        ("мусор", (None, None)),
    ],
)
def test_parse_bridge_pile_geometry(mark, expected) -> None:
    assert parse_bridge_pile_geometry(mark) == expected


def test_resolve_catalog_exact_and_geometry(tmp_path) -> None:
    xlsx = _build_list1_xlsx(
        tmp_path,
        [
            ("С140.40", 2.26, 5650, 3, "-", 12),
            ("С90.35", 1.12, 2800, 7, 23, "-"),
        ],
    )
    entries = parse_pile_catalog_from_xlsx(xlsx)
    row_cyr = resolve_catalog_for_mark("С140.40", entries)
    row_lat = resolve_catalog_for_mark("C14-40T4", entries)
    row_cyr_bridge = resolve_catalog_for_mark("С14-40Т4", entries)
    assert row_cyr is not None and row_lat is not None and row_cyr_bridge is not None
    assert row_cyr.mark == row_lat.mark == row_cyr_bridge.mark == "С140.40"
    assert row_lat.weight_kg == pytest.approx(5650.0)
    assert row_lat.pcs_per_20t == 3
    assert resolve_catalog_for_mark("C9-35T6", entries).mark == "С90.35"
    assert resolve_catalog_for_mark("C18-40T8", entries) is None
    assert resolve_catalog_for_mark("нет-такой", entries) is None


def test_resolve_catalog_prefers_pcs_when_geometry_ties(tmp_path) -> None:
    entries = [
        PileCatalogEntry("С140.40-A", 14.0, 400, 2.26, 5650.0, None),
        PileCatalogEntry("С140.40", 14.0, 400, 2.26, 5650.0, 3),
    ]
    resolved = resolve_catalog_for_mark("C14-40T4", entries)
    assert resolved is not None
    assert resolved.pcs_per_20t == 3
    assert resolved.mark == "С140.40"

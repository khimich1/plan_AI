"""Stage 2.8: GSM usage report — acceptance (May 2026, Tugella О848ХР44) + API."""

from __future__ import annotations

import io
import json
import shutil
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from app.core.settings import get_settings
from app.main import create_app
from app.repositories.gsm_repository import GsmRepository
from app.services.gsm_report_service import GsmReportService
from core import kp_db_schema
from core.gsm.balance import burn_for_km
from tests.helpers.auth_fixtures import patch_auth_users
from tests.helpers.csrf import CsrfAwareTestClient
from tests.helpers.production_api_fixtures import VALID_APP_SECRET_KEY, session_cookie

PREFIX = "/api/v1/gsm"
REPORT = f"{PREFIX}/report/usage"

PROJECT_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = PROJECT_ROOT / "core" / "gsm" / "templates" / "gsm_usage_report.xlsx"
FLOAT_TOL = 0.02

# Paper reference: блок «май 2026», Geely Tugella О848ХР44
MAY_ROWS: list[dict[str, Any]] = [
    # day, driver, km, odo_start, fuel_start, received, destination
    {
        "day": date(2026, 5, 4),
        "driver": "Кулигин Никита Валерьевич",
        "km": 560,
        "odo": 71514,
        "fuel_start": 16.25,
        "received": 54.87,
        "dest": "Вологда",
    },
    {
        "day": date(2026, 5, 7),
        "driver": "Скрябин Алексей Александрович",
        "km": 190,
        "odo": 72074,
        "fuel_start": 18.48,
        "received": 25.33,
        "dest": "Ярославль",
    },
    {
        "day": date(2026, 5, 12),
        "driver": "Полякова Любовь Фёдоровна",
        "km": 190,
        "odo": 72264,
        "fuel_start": 25.95,
        "received": 30.0,
        "dest": "Ярославль",
    },
    {
        "day": date(2026, 5, 14),
        "driver": "Скрябин Алексей Александрович",
        "km": 560,
        "odo": 72454,
        "fuel_start": 38.09,
        "received": 40.0,
        "dest": "Вологда",
    },
    {
        "day": date(2026, 5, 19),
        "driver": "Скрябин Алексей Александрович",
        "km": 400,
        "odo": 73014,
        "fuel_start": 25.45,
        "received": 28.0,
        "dest": "Переславль-Залесский",
    },
    {
        "day": date(2026, 5, 20),
        "driver": "Лоншакова Наталья Евгеньевна",
        "km": 190,
        "odo": 73414,
        "fuel_start": 15.85,
        "received": 29.0,
        "dest": "Ярославль",
    },
    {
        "day": date(2026, 5, 22),
        "driver": "Кулигин Никита Валерьевич",
        "km": 350,
        "odo": 73604,
        "fuel_start": 26.99,
        "received": 25.34,
        "dest": "Рыбинск",
    },
    {
        "day": date(2026, 5, 26),
        "driver": "Скрябин Алексей Александрович",
        "km": 450,
        "odo": 73954,
        "fuel_start": 19.43,
        "received": 30.14,
        "dest": "Владимир",
    },
    {
        "day": date(2026, 5, 28),
        "driver": "Кулигин Никита Валерьевич",
        "km": 530,
        "odo": 74404,
        "fuel_start": 7.27,
        "received": 51.47,
        "dest": "Сергиев Посад",
    },
    {
        "day": date(2026, 5, 29),
        "driver": "Полякова Любовь Фёдоровна",
        "km": 190,
        "odo": 74934,
        "fuel_start": 8.92,
        "received": 21.45,
        "dest": "Ярославль",
    },
]

NORM = 9.4
EXPECTED_FUEL_START = 16.25
EXPECTED_BURN = 339.34
EXPECTED_RECEIVED = 335.60
EXPECTED_FUEL_END = 12.51
EXPECTED_ODO_START = 71514
EXPECTED_ODO_END = 75124
EXPECTED_DATES = [4, 7, 12, 14, 19, 20, 22, 26, 28, 29]

TEST_USERS = [
    {
        "id": 1,
        "username": "admin",
        "role": "admin",
        "manager_id": None,
        "is_active": 1,
        "session_version": 0,
        "created_at": "2026-01-01 00:00:00",
    },
    {
        "id": 5,
        "username": "accountant_user",
        "role": "accountant",
        "manager_id": None,
        "is_active": 1,
        "session_version": 0,
        "created_at": "2026-01-01 00:00:00",
    },
    {
        "id": 3,
        "username": "manager_a",
        "role": "manager",
        "manager_id": None,
        "is_active": 1,
        "session_version": 0,
        "created_at": "2026-01-01 00:00:00",
    },
]


def _fresh_db(tmp_path: Path, name: str = "gsm_usage.db") -> str:
    db_path = str(tmp_path / name)
    kp_db_schema._schema_ready.clear()
    kp_db_schema.ensure_schema(db_path)
    return db_path


@pytest.fixture()
def db_path(tmp_path: Path) -> str:
    return _fresh_db(tmp_path)


@pytest.fixture()
def repo(db_path: str) -> GsmRepository:
    return GsmRepository(db_path=db_path)


@pytest.fixture()
def api_client(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> CsrfAwareTestClient:
    db = tmp_path / "plita.db"
    monkeypatch.setenv("APP_SECRET_KEY", VALID_APP_SECRET_KEY)
    monkeypatch.setenv("PLITA_DB_PATH", str(db))
    monkeypatch.setenv("PB_DB_PATH", str(db))
    get_settings.cache_clear()
    kp_db_schema._schema_ready.clear()
    kp_db_schema.ensure_schema(str(db))
    patch_auth_users(monkeypatch, TEST_USERS)
    return CsrfAwareTestClient(create_app())


def _auth(username: str = "accountant_user") -> dict[str, str]:
    by_user = {
        "admin": (1, "admin"),
        "accountant_user": (5, "accountant"),
        "manager_a": (3, "manager"),
    }
    user_id, role = by_user[username]
    return session_cookie(user_id, role, username)


def _driver_id(repo: GsmRepository, full_name: str) -> int:
    for d in repo.list_drivers():
        if d["full_name"] == full_name:
            return int(d["id"])
    return repo.create_driver(
        full_name=full_name,
        license_number=f"44 00 {abs(hash(full_name)) % 1000000:06d}",
        license_issued_at="2015-01-01",
    )


def seed_may_848(repo: GsmRepository) -> dict[str, Any]:
    """Minimal confirmed WBs + fuel txs matching paper May 2026 block for 848."""
    assert TEMPLATE.exists(), "gsm_usage_report.xlsx template missing"

    driver_ids: dict[str, int] = {}
    for row in MAY_ROWS:
        name = row["driver"]
        if name not in driver_ids:
            driver_ids[name] = _driver_id(repo, name)

    vehicle_id = repo.create_vehicle(
        name="Geely Tugella 848",
        plate_number="О 848 ХР 44",
        tank_volume_liters=55.0,
        norm_summer=9.4,
        norm_winter=10.3,
        primary_driver_id=driver_ids[MAY_ROWS[0]["driver"]],
    )
    card_id = repo.create_card(
        card_number="0000000000000266",
        vehicle_id=vehicle_id,
        assigned_at="2025-01-01",
    )
    batch_id = repo.create_import_batch(
        filename="may848_fixture.xls",
        uploaded_at="2026-06-01T12:00:00",
        uploaded_by="test",
        period_from="2026-05-01",
        period_to="2026-05-31",
    )
    # Summer mode for May (switch to summer before period)
    repo.set_setting(
        "season_switches",
        json.dumps([{"date": "2026-04-01", "mode": "summer"}], ensure_ascii=False),
    )

    for row in MAY_ROWS:
        burn = burn_for_km(int(row["km"]), NORM)
        fuel_end = round(float(row["fuel_start"]) + float(row["received"]) - burn, 2)
        odo_end = int(row["odo"]) + int(row["km"])
        route = [
            {"from": "Завод", "to": row["dest"], "km": int(row["km"]) // 2},
            {"from": row["dest"], "to": "Завод", "km": int(row["km"]) - int(row["km"]) // 2},
        ]
        repo.upsert_waybill(
            vehicle_id=vehicle_id,
            date=row["day"],
            driver_id=driver_ids[row["driver"]],
            status="confirmed",
            source="manual",
            odometer_start=int(row["odo"]),
            odometer_end=odo_end,
            fuel_start=float(row["fuel_start"]),
            fuel_issued=float(row["received"]),
            fuel_end=fuel_end,
            route_json=json.dumps(route, ensure_ascii=False),
        )
        # One fuel tx on the same day → attaches to the PL row
        ts = datetime(row["day"].year, row["day"].month, row["day"].day, 12, 0, 0)
        repo.insert_transaction(
            card_id=card_id,
            ts=ts.isoformat(sep="T"),
            service_type="fuel",
            fuel_grade="АИ-95",
            qty_liters=float(row["received"]),
            amount=float(row["received"]) * 50.0,
            raw_address="АЗС",
            batch_id=batch_id,
        )

    return {"vehicle_id": vehicle_id, "card_id": card_id, "driver_ids": driver_ids}


def fake_convert(src: Path, fmt: str, outdir: Path, workdir: Path, timeout: int) -> Path:
    out = outdir / f"{src.stem}.{fmt}"
    if fmt == "xls":
        # Keep xlsx bytes so openpyxl can still open the stub for assertions
        out.write_bytes(src.read_bytes())
    else:
        shutil.copy2(src, out)
    return out


def _seed_simple_vehicle(
    repo: GsmRepository,
    *,
    name: str,
    plate: str,
    license_number: str,
    days: list[date],
    status: str = "draft",
    warnings_by_day: dict[date, str] | None = None,
) -> dict[str, Any]:
    """Minimal PL rows for kit/export tests (fuel_end stays > 0)."""
    driver_id = repo.create_driver(
        full_name=f"Водитель {plate}",
        license_number=license_number,
        license_issued_at="2015-01-01",
    )
    vehicle_id = repo.create_vehicle(
        name=name,
        plate_number=plate,
        tank_volume_liters=55.0,
        norm_summer=9.4,
        norm_winter=10.3,
        primary_driver_id=driver_id,
    )
    warnings_by_day = warnings_by_day or {}
    for i, day in enumerate(days):
        km = 100
        odo = 10_000 + i * km
        repo.upsert_waybill(
            vehicle_id=vehicle_id,
            date=day,
            driver_id=driver_id,
            status=status,
            source="manual",
            odometer_start=odo,
            odometer_end=odo + km,
            fuel_start=20.0,
            fuel_issued=10.0,
            fuel_end=20.5,
            route_json=json.dumps(
                [{"from": "Завод", "to": "Клиент", "km": km}],
                ensure_ascii=False,
            ),
            warnings_json=warnings_by_day.get(day),
        )
    repo.set_setting(
        "season_switches",
        json.dumps([{"date": "2026-04-01", "mode": "summer"}], ensure_ascii=False),
    )
    return {"vehicle_id": vehicle_id, "driver_id": driver_id, "days": list(days)}


def _post_usage(
    api_client: CsrfAwareTestClient,
    *,
    period_from: str,
    period_to: str,
    vehicle_ids: list[int] | None,
):
    with patch(
        "app.services.gsm_export_service.convert_with_soffice",
        side_effect=fake_convert,
    ), patch(
        "app.services.gsm_report_service.convert_with_soffice",
        side_effect=fake_convert,
    ):
        return api_client.post(
            REPORT,
            json={
                "period_from": period_from,
                "period_to": period_to,
                "vehicle_ids": vehicle_ids,
            },
            cookies=_auth(),
        )


def _summary_row_count(zip_bytes: bytes, *, plate_hint: str) -> int:
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    reports = [
        n
        for n in zf.namelist()
        if n.startswith("Отчет по использованию ГСМ") and plate_hint in n
    ]
    assert len(reports) == 1, reports
    wb = load_workbook(io.BytesIO(zf.read(reports[0])), data_only=False)
    ws = wb.active
    count = 0
    row_i = 18
    while True:
        val = ws.cell(row_i, 1).value
        if val is None or str(val).upper().startswith("ИТОГО"):
            break
        count += 1
        row_i += 1
    return count


def _waybill_statuses(
    repo: GsmRepository, vehicle_id: int, period_from: date, period_to: date
) -> list[str]:
    return [
        str(wb["status"])
        for wb in repo.list_waybills(
            vehicle_id=vehicle_id,
            period_from=period_from,
            period_to=period_to,
        )
    ]


# =============================================================================
# Acceptance: May 848 numbers
# =============================================================================


def test_template_exists() -> None:
    assert TEMPLATE.exists()


def test_may_848_usage_report_numbers(repo: GsmRepository, tmp_path: Path) -> None:
    seeded = seed_may_848(repo)
    service = GsmReportService(repo=repo, template_path=TEMPLATE)

    xlsx_path = tmp_path / "report.xlsx"
    service.build_vehicle_report_xlsx(
        vehicle_id=seeded["vehicle_id"],
        period_from=date(2026, 5, 1),
        period_to=date(2026, 5, 31),
        out_path=xlsx_path,
    )
    wb = load_workbook(xlsx_path, data_only=False)
    ws = wb.active

    # First data row after header block (row 18 in Образец layout)
    first = 18
    assert ws.cell(first, 6).value == pytest.approx(EXPECTED_FUEL_START, abs=FLOAT_TOL)
    assert ws.cell(first, 7).value == EXPECTED_ODO_START

    dates_notes = []
    burns = []
    facts = []
    received = []
    for i in range(10):
        row = first + i
        burns.append(float(ws.cell(row, 11).value))
        facts.append(float(ws.cell(row, 12).value))
        received.append(float(ws.cell(row, 13).value))
        note = str(ws.cell(row, 15).value or "")
        dates_notes.append(note)

    assert len(dates_notes) == 10
    for day_num, note in zip(EXPECTED_DATES, dates_notes, strict=True):
        assert note.startswith(f"{day_num:02d} мая") or note.startswith(f"{day_num} мая")

    assert sum(burns) == pytest.approx(EXPECTED_BURN, abs=FLOAT_TOL)
    assert sum(facts) == pytest.approx(EXPECTED_BURN, abs=FLOAT_TOL)
    assert sum(received) == pytest.approx(EXPECTED_RECEIVED, abs=FLOAT_TOL)
    for b, f in zip(burns, facts, strict=True):
        assert b == pytest.approx(f, abs=FLOAT_TOL)

    last = first + 9
    assert ws.cell(last, 8).value == EXPECTED_ODO_END
    assert ws.cell(last, 14).value == pytest.approx(EXPECTED_FUEL_END, abs=FLOAT_TOL)

    # ИТОГО row
    total_row = first + 10
    assert str(ws.cell(total_row, 1).value).upper().startswith("ИТОГО")
    assert ws.cell(total_row, 11).value == pytest.approx(EXPECTED_BURN, abs=FLOAT_TOL)
    assert ws.cell(total_row, 12).value == pytest.approx(EXPECTED_BURN, abs=FLOAT_TOL)
    assert ws.cell(total_row, 13).value == pytest.approx(EXPECTED_RECEIVED, abs=FLOAT_TOL)
    assert ws.cell(total_row, 14).value == pytest.approx(EXPECTED_FUEL_END, abs=FLOAT_TOL)
    assert ws.cell(total_row, 6).value == pytest.approx(EXPECTED_FUEL_START, abs=FLOAT_TOL)

    # Approval date = period_to
    assert "31" in str(ws["O5"].value)
    assert "мая" in str(ws["O5"].value).lower()
    assert "2026" in str(ws["O5"].value)


# =============================================================================
# API
# =============================================================================


def test_usage_report_requires_accounting(api_client: CsrfAwareTestClient) -> None:
    settings = get_settings()
    repo = GsmRepository(db_path=settings.plita_db_path)
    seeded = seed_may_848(repo)

    r = api_client.post(
        REPORT,
        json={
            "period_from": "2026-05-01",
            "period_to": "2026-05-31",
            "vehicle_ids": [seeded["vehicle_id"]],
        },
        cookies=_auth("manager_a"),
    )
    assert r.status_code == 403


def test_usage_report_invalid_period(api_client: CsrfAwareTestClient) -> None:
    settings = get_settings()
    repo = GsmRepository(db_path=settings.plita_db_path)
    seeded = seed_may_848(repo)

    r = api_client.post(
        REPORT,
        json={
            "period_from": "2026-05-31",
            "period_to": "2026-05-01",
            "vehicle_ids": [seeded["vehicle_id"]],
        },
        cookies=_auth(),
    )
    assert r.status_code == 400, r.text
    detail = r.json().get("detail") or r.json()
    code = detail.get("code") if isinstance(detail, dict) else None
    assert code == "gsm_report_invalid_period"


def test_usage_report_zip_contains_report_and_waybills(
    api_client: CsrfAwareTestClient,
) -> None:
    settings = get_settings()
    repo = GsmRepository(db_path=settings.plita_db_path)
    seeded = seed_may_848(repo)

    with patch(
        "app.services.gsm_export_service.convert_with_soffice",
        side_effect=fake_convert,
    ), patch(
        "app.services.gsm_report_service.convert_with_soffice",
        side_effect=fake_convert,
    ):
        r = api_client.post(
            REPORT,
            json={
                "period_from": "2026-05-01",
                "period_to": "2026-05-31",
                "vehicle_ids": [seeded["vehicle_id"]],
            },
            cookies=_auth(),
        )
    assert r.status_code == 200, r.text
    assert "application/zip" in r.headers.get("content-type", "")
    cd = r.headers.get("content-disposition", "")
    assert "gsm_usage_report_2026-05-01_2026-05-31.zip" in cd

    names = zipfile.ZipFile(io.BytesIO(r.content)).namelist()
    report_names = [n for n in names if n.startswith("Отчет по использованию ГСМ")]
    waybill_names = [n for n in names if n.startswith("ПЛ ")]
    assert len(report_names) == 1
    assert "848" in report_names[0]
    assert len(waybill_names) == 10


def test_usage_report_no_data(api_client: CsrfAwareTestClient) -> None:
    settings = get_settings()
    repo = GsmRepository(db_path=settings.plita_db_path)
    driver_id = repo.create_driver(
        full_name="Тестов Тест Тестович",
        license_number="00 00 000000",
    )
    vehicle_id = repo.create_vehicle(
        name="Empty Car",
        plate_number="А 000 АА 00",
        tank_volume_liters=50.0,
        norm_summer=9.0,
        norm_winter=10.0,
        primary_driver_id=driver_id,
    )

    with patch(
        "app.services.gsm_report_service.convert_with_soffice",
        side_effect=fake_convert,
    ):
        r = api_client.post(
            REPORT,
            json={
                "period_from": "2026-05-01",
                "period_to": "2026-05-31",
                "vehicle_ids": [vehicle_id],
            },
            cookies=_auth(),
        )
    assert r.status_code in (404, 422), r.text
    detail = r.json().get("detail") or r.json()
    code = detail.get("code") if isinstance(detail, dict) else None
    assert code == "gsm_report_no_data"


def test_usage_report_draft_kit_marks_exported(api_client: CsrfAwareTestClient) -> None:
    settings = get_settings()
    repo = GsmRepository(db_path=settings.plita_db_path)
    days = [date(2026, 7, 6), date(2026, 7, 8), date(2026, 7, 10)]
    seeded = _seed_simple_vehicle(
        repo,
        name="Palisade Kit",
        plate="О 111 КТ 44",
        license_number="44 11 111111",
        days=days,
        status="draft",
    )
    period_from = date(2026, 7, 1)
    period_to = date(2026, 7, 31)

    r = _post_usage(
        api_client,
        period_from=period_from.isoformat(),
        period_to=period_to.isoformat(),
        vehicle_ids=[seeded["vehicle_id"]],
    )
    assert r.status_code == 200, r.text
    names = zipfile.ZipFile(io.BytesIO(r.content)).namelist()
    waybill_names = [n for n in names if n.startswith("ПЛ ")]
    assert _summary_row_count(r.content, plate_hint="111") == len(days)
    assert len(waybill_names) == len(days)
    assert _waybill_statuses(repo, seeded["vehicle_id"], period_from, period_to) == [
        "exported"
    ] * len(days)


def test_usage_report_skips_red_vehicle_keeps_clean_neighbor(
    api_client: CsrfAwareTestClient,
) -> None:
    settings = get_settings()
    repo = GsmRepository(db_path=settings.plita_db_path)
    red_days = [date(2026, 7, 3), date(2026, 7, 4)]
    clean_days = [date(2026, 7, 10), date(2026, 7, 11)]
    red = _seed_simple_vehicle(
        repo,
        name="Monjaro Red",
        plate="О 222 КР 44",
        license_number="44 22 222222",
        days=red_days,
        status="draft",
        warnings_by_day={
            date(2026, 7, 3): json.dumps(
                [{"code": "manual_intervention", "detail": "бак не сходится"}],
                ensure_ascii=False,
            ),
        },
    )
    clean = _seed_simple_vehicle(
        repo,
        name="Palisade Clean",
        plate="О 333 ЧС 44",
        license_number="44 33 333333",
        days=clean_days,
        status="draft",
        warnings_by_day={
            date(2026, 7, 10): json.dumps(["hook_above_threshold"]),
        },
    )
    period_from = date(2026, 7, 1)
    period_to = date(2026, 7, 31)

    r = _post_usage(
        api_client,
        period_from=period_from.isoformat(),
        period_to=period_to.isoformat(),
        vehicle_ids=[red["vehicle_id"], clean["vehicle_id"]],
    )
    assert r.status_code == 200, r.text
    names = zipfile.ZipFile(io.BytesIO(r.content)).namelist()
    reports = [n for n in names if n.startswith("Отчет по использованию ГСМ")]
    waybills = [n for n in names if n.startswith("ПЛ ")]
    assert len(reports) == 1
    assert "333" in reports[0]
    assert "222" not in reports[0]
    assert _summary_row_count(r.content, plate_hint="333") == len(clean_days)
    assert sorted(waybills) == ["ПЛ 10.07.26.xls", "ПЛ 11.07.26.xls"]
    assert _waybill_statuses(repo, red["vehicle_id"], period_from, period_to) == [
        "draft",
        "draft",
    ]
    assert _waybill_statuses(repo, clean["vehicle_id"], period_from, period_to) == [
        "exported",
        "exported",
    ]


def test_usage_report_no_data_when_all_red(api_client: CsrfAwareTestClient) -> None:
    settings = get_settings()
    repo = GsmRepository(db_path=settings.plita_db_path)
    seeded = _seed_simple_vehicle(
        repo,
        name="All Red",
        plate="О 444 КР 44",
        license_number="44 44 444444",
        days=[date(2026, 7, 15)],
        status="confirmed",
        warnings_by_day={
            date(2026, 7, 15): json.dumps(["manual_intervention"]),
        },
    )
    period_from = date(2026, 7, 1)
    period_to = date(2026, 7, 31)

    r = _post_usage(
        api_client,
        period_from=period_from.isoformat(),
        period_to=period_to.isoformat(),
        vehicle_ids=[seeded["vehicle_id"]],
    )
    assert r.status_code == 404, r.text
    detail = r.json().get("detail") or r.json()
    code = detail.get("code") if isinstance(detail, dict) else None
    assert code == "gsm_report_no_data"
    assert _waybill_statuses(repo, seeded["vehicle_id"], period_from, period_to) == [
        "confirmed"
    ]


def test_usage_report_includes_yellow_only_vehicle(
    api_client: CsrfAwareTestClient,
) -> None:
    settings = get_settings()
    repo = GsmRepository(db_path=settings.plita_db_path)
    days = [date(2026, 7, 12), date(2026, 7, 13)]
    seeded = _seed_simple_vehicle(
        repo,
        name="Yellow Only",
        plate="О 555 ЖЛ 44",
        license_number="44 55 555555",
        days=days,
        status="draft",
        warnings_by_day={
            date(2026, 7, 12): json.dumps(["hook_above_threshold"]),
            date(2026, 7, 13): json.dumps(
                [{"code": "weekend_anchor", "detail": "выходной"}],
                ensure_ascii=False,
            ),
        },
    )
    period_from = date(2026, 7, 1)
    period_to = date(2026, 7, 31)

    r = _post_usage(
        api_client,
        period_from=period_from.isoformat(),
        period_to=period_to.isoformat(),
        vehicle_ids=[seeded["vehicle_id"]],
    )
    assert r.status_code == 200, r.text
    names = zipfile.ZipFile(io.BytesIO(r.content)).namelist()
    reports = [n for n in names if n.startswith("Отчет по использованию ГСМ")]
    waybills = [n for n in names if n.startswith("ПЛ ")]
    assert len(reports) == 1
    assert "555" in reports[0]
    assert _summary_row_count(r.content, plate_hint="555") == len(days)
    assert sorted(waybills) == ["ПЛ 12.07.26.xls", "ПЛ 13.07.26.xls"]
    assert _waybill_statuses(repo, seeded["vehicle_id"], period_from, period_to) == [
        "exported",
        "exported",
    ]


def test_usage_report_null_vehicle_ids_exports_all_active(
    api_client: CsrfAwareTestClient,
) -> None:
    settings = get_settings()
    repo = GsmRepository(db_path=settings.plita_db_path)
    first = _seed_simple_vehicle(
        repo,
        name="Active First",
        plate="О 666 АА 44",
        license_number="44 66 666666",
        days=[date(2026, 7, 6), date(2026, 7, 8)],
        status="draft",
    )
    second = _seed_simple_vehicle(
        repo,
        name="Active Second",
        plate="О 777 ББ 44",
        license_number="44 77 777777",
        days=[date(2026, 7, 20), date(2026, 7, 22)],
        status="draft",
    )
    period_from = date(2026, 7, 1)
    period_to = date(2026, 7, 31)

    r = _post_usage(
        api_client,
        period_from=period_from.isoformat(),
        period_to=period_to.isoformat(),
        vehicle_ids=None,
    )
    assert r.status_code == 200, r.text
    names = zipfile.ZipFile(io.BytesIO(r.content)).namelist()
    reports = [n for n in names if n.startswith("Отчет по использованию ГСМ")]
    waybills = [n for n in names if n.startswith("ПЛ ")]
    assert len(reports) == 2
    joined = " ".join(reports)
    assert "666" in joined
    assert "777" in joined
    assert _summary_row_count(r.content, plate_hint="666") == 2
    assert _summary_row_count(r.content, plate_hint="777") == 2
    assert sorted(waybills) == [
        "ПЛ 06.07.26.xls",
        "ПЛ 08.07.26.xls",
        "ПЛ 20.07.26.xls",
        "ПЛ 22.07.26.xls",
    ]
    assert _waybill_statuses(repo, first["vehicle_id"], period_from, period_to) == [
        "exported",
        "exported",
    ]
    assert _waybill_statuses(repo, second["vehicle_id"], period_from, period_to) == [
        "exported",
        "exported",
    ]


def test_usage_report_explicit_vehicle_ids_excludes_sibling(
    api_client: CsrfAwareTestClient,
) -> None:
    settings = get_settings()
    repo = GsmRepository(db_path=settings.plita_db_path)
    selected = _seed_simple_vehicle(
        repo,
        name="Selected",
        plate="О 888 ВВ 44",
        license_number="44 88 888888",
        days=[date(2026, 7, 6), date(2026, 7, 8)],
        status="draft",
    )
    sibling = _seed_simple_vehicle(
        repo,
        name="Sibling",
        plate="О 999 ГГ 44",
        license_number="44 99 999999",
        days=[date(2026, 7, 20), date(2026, 7, 22)],
        status="draft",
    )
    period_from = date(2026, 7, 1)
    period_to = date(2026, 7, 31)

    r = _post_usage(
        api_client,
        period_from=period_from.isoformat(),
        period_to=period_to.isoformat(),
        vehicle_ids=[selected["vehicle_id"]],
    )
    assert r.status_code == 200, r.text
    names = zipfile.ZipFile(io.BytesIO(r.content)).namelist()
    reports = [n for n in names if n.startswith("Отчет по использованию ГСМ")]
    waybills = [n for n in names if n.startswith("ПЛ ")]
    assert len(reports) == 1
    assert "888" in reports[0]
    assert "999" not in reports[0]
    assert _summary_row_count(r.content, plate_hint="888") == 2
    assert sorted(waybills) == ["ПЛ 06.07.26.xls", "ПЛ 08.07.26.xls"]
    assert _waybill_statuses(repo, selected["vehicle_id"], period_from, period_to) == [
        "exported",
        "exported",
    ]
    assert _waybill_statuses(repo, sibling["vehicle_id"], period_from, period_to) == [
        "draft",
        "draft",
    ]


def test_usage_report_already_exported_stays_in_kit(
    api_client: CsrfAwareTestClient,
) -> None:
    settings = get_settings()
    repo = GsmRepository(db_path=settings.plita_db_path)
    days = [date(2026, 7, 18), date(2026, 7, 19)]
    seeded = _seed_simple_vehicle(
        repo,
        name="Already Exported",
        plate="О 101 ЭК 44",
        license_number="44 10 101010",
        days=days,
        status="exported",
    )
    period_from = date(2026, 7, 1)
    period_to = date(2026, 7, 31)

    r = _post_usage(
        api_client,
        period_from=period_from.isoformat(),
        period_to=period_to.isoformat(),
        vehicle_ids=[seeded["vehicle_id"]],
    )
    assert r.status_code == 200, r.text
    assert _summary_row_count(r.content, plate_hint="101") == len(days)
    names = zipfile.ZipFile(io.BytesIO(r.content)).namelist()
    waybills = [n for n in names if n.startswith("ПЛ ")]
    assert sorted(waybills) == ["ПЛ 18.07.26.xls", "ПЛ 19.07.26.xls"]
    assert _waybill_statuses(repo, seeded["vehicle_id"], period_from, period_to) == [
        "exported",
        "exported",
    ]

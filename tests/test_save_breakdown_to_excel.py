"""Breakdown XLSX export must match reference layout (readable labels + formulas)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from core.commercial_offer import save_breakdown_to_excel


def _sample_tables() -> list[dict]:
    return [
        {
            "name": "Плиты ПБ 23-6,65-8п",
            "rows": [
                ["Базовая цена (0,67м)", "7 945,00 × (0,67 / 1.2)", "4 402,85 руб"],
                ["Продольный рез", "460 × 2,3 × 4 / 4", "1 058,00 руб"],
                ["Отходы (535 + 535 + 535 + 535мм)", "(535 + 535 + 535 + 535 / 1200) × 7 945,00 / 4", "3 542,15 руб"],
                ["ИТОГО за 1 плиту", "", "9 003,00 руб"],
                ["Округлено", "", "9 003,00 руб"],
                ["За 4 плит", "9 003,00 × 4", "36 012,00 руб"],
            ],
        },
        {
            "name": "Плиты ПБ 28-7,2-8п",
            "rows": [
                ["Базовая цена (0,72м)", "7 945,00 × (0,72 / 1.2)", "4 767,00 руб"],
                ["Поперечный рез", "1200 × 1", "1 200,00 руб"],
                ["Остаток после поперечного реза (0,20м)", "8 603,00 × 0,0665", "572,10 руб"],
                ["ИТОГО за 1 плиту", "", "6 539,10 руб"],
                ["Округлено", "", "6 539,10 руб"],
                ["За 1 плит", "6 539,10 × 1", "6 539,10 руб"],
            ],
        },
    ]


def test_save_breakdown_headers_and_block_structure(tmp_path: Path) -> None:
    out = tmp_path / "breakdown.xlsx"
    assert save_breakdown_to_excel(_sample_tables(), str(out)) is True

    ws = load_workbook(out).active
    assert [ws.cell(1, c).value for c in range(1, 4)] == ["Компонент", "Расчёт", "Сумма"]

    assert ws.cell(2, 1).value == "Плиты ПБ 23-6,65-8п"
    assert ws.cell(3, 1).value == "Базовая цена (0,67м)"
    assert ws.cell(3, 2).value == "7 945,00 × (0,67 / 1.2)"
    assert ws.cell(3, 3).value == "4 402,85 руб"
    assert ws.cell(8, 1).value == "За 4 плит"
    assert ws.cell(8, 2).value == "9 003,00 × 4"

    # Empty separator between product blocks
    assert ws.cell(9, 1).value in (None, "")
    assert ws.cell(9, 2).value in (None, "")
    assert ws.cell(9, 3).value in (None, "")
    assert ws.cell(10, 1).value == "Плиты ПБ 28-7,2-8п"


def test_save_breakdown_preserves_full_labels_not_truncated(tmp_path: Path) -> None:
    out = tmp_path / "breakdown.xlsx"
    assert save_breakdown_to_excel(_sample_tables(), str(out)) is True
    ws = load_workbook(out).active

    labels = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value]
    assert "Базовая цена (0,67м)" in labels
    assert "Продольный рез" in labels
    assert "Поперечный рез" in labels
    assert "Остаток после поперечного реза (0,20м)" in labels
    assert "ИТОГО за 1 плиту" in labels
    assert "Округлено" in labels
    # Must not look like the cramped UI truncation
    assert "Базовая" not in labels  # bare truncated token without "цена"
    assert all(not str(label).startswith("Попереч") or label == "Поперечный рез" for label in labels)
    assert all(not str(label).startswith("Продоль") or label == "Продольный рез" for label in labels)


def test_save_breakdown_currency_suffix_and_readable_columns(tmp_path: Path) -> None:
    out = tmp_path / "breakdown.xlsx"
    assert save_breakdown_to_excel(_sample_tables(), str(out)) is True
    ws = load_workbook(out).active

    sums = [
        ws.cell(r, 3).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, 3).value not in (None, "")
    ]
    assert sums
    assert all(str(v).endswith(" руб") for v in sums)
    assert all("," in str(v) for v in sums)

    # Narrow default columns visually truncate; require readable widths.
    assert float(ws.column_dimensions["A"].width or 0) >= 40
    assert float(ws.column_dimensions["B"].width or 0) >= 35
    assert float(ws.column_dimensions["C"].width or 0) >= 16

"""Unit-тесты XLSX-шаблона графика поставки (core/delivery_schedule_xlsx).

build_template / parse_template: заголовки, round-trip, unmatched, группировка.
"""

from __future__ import annotations

from pathlib import Path

import pytest

pytest.importorskip("openpyxl")

from openpyxl import load_workbook

from core.delivery_schedule_xlsx import (
    HEADERS,
    REASON_BAD_DATE,
    REASON_BAD_QTY,
    REASON_CONFLICTING_BATCH_DATES,
    REASON_UNKNOWN_MARK,
    SECTION_IN_SCHEDULE,
    SECTION_REMAINDER,
    BatchDraft,
    BatchDraftItem,
    build_template,
    parse_template,
)

_KP_PLATES = [
    {"id": 10, "plate_name": "ПБ 60-12-8"},
    {"id": 20, "plate_name": "ПБ 72-15-8"},
]


def _write_data_rows(path: Path, rows: list[tuple]) -> bytes:
    """Дописывает строки данных в шаблон (начиная со строки 2) и возвращает bytes."""
    wb = load_workbook(path)
    ws = wb.active
    for r_idx, values in enumerate(rows, start=2):
        for c_idx, value in enumerate(values, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(path)
    return path.read_bytes()


def test_build_template_writes_headers(tmp_path: Path) -> None:
    path = tmp_path / "template.xlsx"
    result = build_template(path)

    assert result == path
    assert path.is_file()

    wb = load_workbook(path)
    ws = wb.active
    assert ws.title == "График поставки"
    for col, header in enumerate(HEADERS, start=1):
        assert ws.cell(row=1, column=col).value == header
    assert ws.cell(row=2, column=1).value is None


def test_build_template_prefills_kp_plates(tmp_path: Path) -> None:
    path = tmp_path / "prefilled.xlsx"
    plates = [
        {"id": 10, "plate_name": "ПБ 69-12-8п", "qty": 40},
        {"id": 20, "plate_name": "ПБ 45-10,8-8п", "qty": 4},
    ]

    build_template(path, plates=plates)

    wb = load_workbook(path)
    ws = wb.active
    assert ws.cell(row=2, column=1).value == SECTION_REMAINDER
    assert [ws.cell(row=3, column=c).value for c in range(1, 7)] == [
        None,
        None,
        None,
        None,
        "ПБ 69-12-8п",
        40,
    ]
    assert [ws.cell(row=4, column=c).value for c in range(1, 7)] == [
        None,
        None,
        None,
        None,
        "ПБ 45-10,8-8п",
        4,
    ]
    assert ws.cell(row=5, column=5).value is None


def test_round_trip_parse_batches_iso_dates_and_items(tmp_path: Path) -> None:
    path = tmp_path / "filled.xlsx"
    build_template(path)
    data = _write_data_rows(
        path,
        [
            ("Партия А", "01.04.2026", "10.04.2026", "25.03.2026", "ПБ 60-12-8", 5),
            ("Партия Б", "15.04.2026", "20.04.2026", "05.04.2026", "ПБ 72-15-8", 3),
        ],
    )

    batches, unmatched = parse_template(data, _KP_PLATES)

    assert unmatched == []
    assert len(batches) == 2

    assert batches[0] == BatchDraft(
        name="Партия А",
        deliver_from="2026-04-01",
        deliver_to="2026-04-10",
        produce_by="2026-03-25",
        items=[BatchDraftItem(plate_id=10, plate_name="ПБ 60-12-8", qty=5)],
    )
    assert batches[1] == BatchDraft(
        name="Партия Б",
        deliver_from="2026-04-15",
        deliver_to="2026-04-20",
        produce_by="2026-04-05",
        items=[BatchDraftItem(plate_id=20, plate_name="ПБ 72-15-8", qty=3)],
    )


def test_unknown_mark_goes_to_unmatched(tmp_path: Path) -> None:
    path = tmp_path / "unknown.xlsx"
    build_template(path)
    data = _write_data_rows(
        path,
        [
            ("П1", "01.04.2026", "10.04.2026", "25.03.2026", "НЕТ ТАКОЙ МАРКИ", 2),
        ],
    )

    batches, unmatched = parse_template(data, _KP_PLATES)

    assert batches == []
    assert len(unmatched) == 1
    assert unmatched[0].row_number == 2
    assert unmatched[0].reason == REASON_UNKNOWN_MARK
    assert unmatched[0].raw is not None
    assert unmatched[0].raw["Марка"] == "НЕТ ТАКОЙ МАРКИ"


def test_bad_date_goes_to_unmatched(tmp_path: Path) -> None:
    path = tmp_path / "bad_date.xlsx"
    build_template(path)
    data = _write_data_rows(
        path,
        [
            ("П1", "2026-04-01", "10.04.2026", "25.03.2026", "ПБ 60-12-8", 2),
        ],
    )

    batches, unmatched = parse_template(data, _KP_PLATES)

    assert batches == []
    assert len(unmatched) == 1
    assert unmatched[0].row_number == 2
    assert unmatched[0].reason == REASON_BAD_DATE


@pytest.mark.parametrize(
    "qty",
    [0, -3, "abc", 1.5, "2.5"],
    ids=["zero", "negative", "non_numeric", "float", "float_str"],
)
def test_bad_qty_goes_to_unmatched(tmp_path: Path, qty: object) -> None:
    path = tmp_path / "bad_qty.xlsx"
    build_template(path)
    data = _write_data_rows(
        path,
        [
            ("П1", "01.04.2026", "10.04.2026", "25.03.2026", "ПБ 60-12-8", qty),
        ],
    )

    batches, unmatched = parse_template(data, _KP_PLATES)

    assert batches == []
    assert len(unmatched) == 1
    assert unmatched[0].row_number == 2
    assert unmatched[0].reason == REASON_BAD_QTY


def test_same_batch_name_groups_into_one_draft(tmp_path: Path) -> None:
    path = tmp_path / "group.xlsx"
    build_template(path)
    data = _write_data_rows(
        path,
        [
            ("Партия X", "01.04.2026", "10.04.2026", "25.03.2026", "ПБ 60-12-8", 4),
            ("Партия X", "01.04.2026", "10.04.2026", "25.03.2026", "ПБ 72-15-8", 7),
        ],
    )

    batches, unmatched = parse_template(data, _KP_PLATES)

    assert unmatched == []
    assert len(batches) == 1
    assert batches[0].name == "Партия X"
    assert batches[0].deliver_from == "2026-04-01"
    assert batches[0].deliver_to == "2026-04-10"
    assert batches[0].produce_by == "2026-03-25"
    assert batches[0].items == [
        BatchDraftItem(plate_id=10, plate_name="ПБ 60-12-8", qty=4),
        BatchDraftItem(plate_id=20, plate_name="ПБ 72-15-8", qty=7),
    ]


def test_conflicting_batch_dates_go_to_unmatched(tmp_path: Path) -> None:
    path = tmp_path / "conflict.xlsx"
    build_template(path)
    data = _write_data_rows(
        path,
        [
            ("Партия X", "01.04.2026", "10.04.2026", "25.03.2026", "ПБ 60-12-8", 4),
            ("Партия X", "15.04.2026", "20.04.2026", "05.04.2026", "ПБ 72-15-8", 7),
        ],
    )

    batches, unmatched = parse_template(data, _KP_PLATES)

    assert len(batches) == 1
    assert batches[0].name == "Партия X"
    assert batches[0].deliver_from == "2026-04-01"
    assert batches[0].deliver_to == "2026-04-10"
    assert batches[0].produce_by == "2026-03-25"
    assert batches[0].items == [
        BatchDraftItem(plate_id=10, plate_name="ПБ 60-12-8", qty=4),
    ]
    assert len(unmatched) == 1
    assert unmatched[0].row_number == 3
    assert unmatched[0].reason == REASON_CONFLICTING_BATCH_DATES
    assert unmatched[0].raw is not None
    assert unmatched[0].raw["Марка"] == "ПБ 72-15-8"


def test_round_trip_prefilled_template_after_manager_fills_dates(
    tmp_path: Path,
) -> None:
    """Сценарий: шаблон с плитами КП → менеджер пишет партии/даты → parse."""
    path = tmp_path / "prefilled_filled.xlsx"
    plates = [
        {"id": 10, "plate_name": "ПБ 60-12-8", "qty": 5},
        {"id": 20, "plate_name": "ПБ 72-15-8", "qty": 3},
    ]
    build_template(path, plates=plates)

    wb = load_workbook(path)
    ws = wb.active
    ws.cell(row=3, column=1, value="Партия 1")
    ws.cell(row=3, column=2, value="01.04.2026")
    ws.cell(row=3, column=3, value="10.04.2026")
    ws.cell(row=3, column=4, value="25.03.2026")
    ws.cell(row=4, column=1, value="Партия 2")
    ws.cell(row=4, column=2, value="15.04.2026")
    ws.cell(row=4, column=3, value="20.04.2026")
    ws.cell(row=4, column=4, value="05.04.2026")
    wb.save(path)

    batches, unmatched = parse_template(path.read_bytes(), plates)

    assert unmatched == []
    assert batches == [
        BatchDraft(
            name="Партия 1",
            deliver_from="2026-04-01",
            deliver_to="2026-04-10",
            produce_by="2026-03-25",
            items=[BatchDraftItem(plate_id=10, plate_name="ПБ 60-12-8", qty=5)],
        ),
        BatchDraft(
            name="Партия 2",
            deliver_from="2026-04-15",
            deliver_to="2026-04-20",
            produce_by="2026-04-05",
            items=[BatchDraftItem(plate_id=20, plate_name="ПБ 72-15-8", qty=3)],
        ),
    ]


def test_unfilled_prefilled_rows_are_skipped(tmp_path: Path) -> None:
    """Незаполненные строки шаблона (марка+кол-во, без партии/дат) не идут в unmatched."""
    path = tmp_path / "partial.xlsx"
    plates = [
        {"id": 10, "plate_name": "ПБ 60-12-8", "qty": 5},
        {"id": 20, "plate_name": "ПБ 72-15-8", "qty": 3},
    ]
    build_template(path, plates=plates)

    wb = load_workbook(path)
    ws = wb.active
    ws.cell(row=3, column=1, value="Партия 1")
    ws.cell(row=3, column=2, value="01.04.2026")
    ws.cell(row=3, column=3, value="10.04.2026")
    ws.cell(row=3, column=4, value="25.03.2026")
    wb.save(path)

    batches, unmatched = parse_template(path.read_bytes(), plates)

    assert unmatched == []
    assert len(batches) == 1
    assert batches[0].name == "Партия 1"
    assert batches[0].items == [
        BatchDraftItem(plate_id=10, plate_name="ПБ 60-12-8", qty=5)
    ]


def test_parse_template_skips_section_titles(tmp_path: Path) -> None:
    path = tmp_path / "banded.xlsx"
    build_template(path)
    data = _write_data_rows(
        path,
        [
            (SECTION_IN_SCHEDULE, None, None, None, None, None),
            ("Партия 1", "01.04.2026", "10.04.2026", "25.03.2026", "ПБ 60-12-8", 10),
            (SECTION_REMAINDER, None, None, None, None, None),
            (None, None, None, None, "ПБ 60-12-8", 30),
        ],
    )

    batches, unmatched = parse_template(data, _KP_PLATES)

    assert unmatched == []
    assert batches == [
        BatchDraft(
            name="Партия 1",
            deliver_from="2026-04-01",
            deliver_to="2026-04-10",
            produce_by="2026-03-25",
            items=[BatchDraftItem(plate_id=10, plate_name="ПБ 60-12-8", qty=10)],
        )
    ]


def test_parse_template_skips_section_title_in_any_column(tmp_path: Path) -> None:
    path = tmp_path / "title_in_e.xlsx"
    build_template(path)
    data = _write_data_rows(
        path,
        [
            ("П1", "01.04.2026", "10.04.2026", "25.03.2026", SECTION_REMAINDER, 2),
            ("П1", "01.04.2026", "10.04.2026", "25.03.2026", "ПБ 60-12-8", 2),
        ],
    )

    batches, unmatched = parse_template(data, _KP_PLATES)

    assert unmatched == []
    assert len(batches) == 1
    assert batches[0].items[0].qty == 2


def test_build_template_remainder_only_writes_section(tmp_path: Path) -> None:
    path = tmp_path / "remainder.xlsx"
    plates = [
        {"id": 10, "plate_name": "ПБ 69-12-8п", "qty": 40},
        {"id": 20, "plate_name": "ПБ 45-10,8-8п", "qty": 4},
    ]

    build_template(path, plates=plates)

    wb = load_workbook(path)
    ws = wb.active
    assert ws.cell(row=2, column=1).value == SECTION_REMAINDER
    assert [ws.cell(row=3, column=c).value for c in range(1, 7)] == [
        None,
        None,
        None,
        None,
        "ПБ 69-12-8п",
        40,
    ]
    assert [ws.cell(row=4, column=c).value for c in range(1, 7)] == [
        None,
        None,
        None,
        None,
        "ПБ 45-10,8-8п",
        4,
    ]


def test_build_template_two_band_with_batches(tmp_path: Path) -> None:
    path = tmp_path / "two_band.xlsx"
    plates = [{"id": 10, "plate_name": "ПБ 60-12-8", "qty": 40}]
    batches = [
        {
            "name": "Партия 1",
            "deliver_from": "2026-04-01",
            "deliver_to": "2026-04-10",
            "produce_by": "2026-03-25",
            "items": [{"plate_id": 10, "plate_name": "ПБ 60-12-8", "qty": 10}],
        }
    ]

    build_template(path, plates=plates, batches=batches)

    wb = load_workbook(path)
    ws = wb.active
    assert ws.cell(row=2, column=1).value == SECTION_IN_SCHEDULE
    assert [ws.cell(row=3, column=c).value for c in range(1, 7)] == [
        "Партия 1",
        "01.04.2026",
        "10.04.2026",
        "25.03.2026",
        "ПБ 60-12-8",
        10,
    ]
    assert ws.cell(row=4, column=1).value == SECTION_REMAINDER
    assert [ws.cell(row=5, column=c).value for c in range(1, 7)] == [
        None,
        None,
        None,
        None,
        "ПБ 60-12-8",
        30,
    ]


def test_round_trip_two_band_file_keeps_only_scheduled_rows(tmp_path: Path) -> None:
    path = tmp_path / "round_trip.xlsx"
    plates = [
        {"id": 10, "plate_name": "ПБ 60-12-8", "qty": 40},
        {"id": 20, "plate_name": "ПБ 72-15-8", "qty": 3},
    ]
    batches = [
        {
            "name": "Партия 1",
            "deliver_from": "2026-04-01",
            "deliver_to": "2026-04-10",
            "produce_by": "2026-03-25",
            "items": [{"plate_id": 10, "qty": 12}],
        }
    ]
    build_template(path, plates=plates, batches=batches)

    parsed, unmatched = parse_template(path.read_bytes(), plates)

    assert unmatched == []
    assert parsed == [
        BatchDraft(
            name="Партия 1",
            deliver_from="2026-04-01",
            deliver_to="2026-04-10",
            produce_by="2026-03-25",
            items=[BatchDraftItem(plate_id=10, plate_name="ПБ 60-12-8", qty=12)],
        )
    ]


def test_old_file_without_sections_still_parses(tmp_path: Path) -> None:
    path = tmp_path / "legacy.xlsx"
    build_template(path)
    data = _write_data_rows(
        path,
        [
            ("Партия А", "01.04.2026", "10.04.2026", "25.03.2026", "ПБ 60-12-8", 5),
        ],
    )

    batches, unmatched = parse_template(data, _KP_PLATES)

    assert unmatched == []
    assert batches[0].name == "Партия А"
    assert batches[0].items[0].qty == 5

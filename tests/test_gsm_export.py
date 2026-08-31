"""Task T14: Waybill blank export (zip) — TDD end-to-end.

Acceptance:
- Zip: one file per waybill day named «ПЛ DD.MM.YY.xls»
- Formulas recalculated; values match DB ±0.01; reverse side has legs with АЗС
- Norm in BS41 matches vehicle season for that date
- soffice isolated profile + timeout; failure → 500 clear error
- POST /gsm/waybills/export under REQUIRE_ACCOUNTING
"""

from __future__ import annotations

import io
import json
import shutil
import zipfile
from datetime import date
from pathlib import Path
from typing import Any
from unittest.mock import patch

import pytest
import xlrd
from openpyxl import load_workbook

from app.core.settings import get_settings
from app.main import create_app
from app.repositories.gsm_repository import GsmRepository
from core import kp_db_schema
from core.gsm.balance import burn_for_km
from core.gsm.blank import (
    BlankDriver,
    BlankLeg,
    BlankWaybill,
    fill_workbook,
    format_bs41_formula,
    waybill_export_filename,
)
from tests.helpers.auth_fixtures import patch_auth_users
from tests.helpers.csrf import CsrfAwareTestClient
from tests.helpers.production_api_fixtures import VALID_APP_SECRET_KEY, session_cookie

PREFIX = "/api/v1/gsm"
EXPORT = f"{PREFIX}/waybills/export"

PROJECT_ROOT = Path(__file__).resolve().parents[1]
BLANK_TEMPLATE = PROJECT_ROOT / "core" / "gsm" / "templates" / "waybill_blank.xlsx"
FIXTURE_TEMPLATE = PROJECT_ROOT / "tests" / "fixtures" / "gsm" / "waybill_blank.xlsx"

HAS_SOFFICE = shutil.which("soffice") is not None
FLOAT_TOL = 0.01

TEST_USERS = [
    {
        "id": 1,
        "username": "admin",
        "role": "admin",
        "manager_id": None,
        "is_active": 1,
        "session_version": 0,
        "created_at": "2026-01-01 00:00:00",
    },
    {
        "id": 5,
        "username": "accountant_user",
        "role": "accountant",
        "manager_id": None,
        "is_active": 1,
        "session_version": 0,
        "created_at": "2026-01-01 00:00:00",
    },
    {
        "id": 3,
        "username": "manager_a",
        "role": "manager",
        "manager_id": None,
        "is_active": 1,
        "session_version": 0,
        "created_at": "2026-01-01 00:00:00",
    },
]

STATION_ADDR = "г.Кострома, АЗС Татнефть, ул.Индустриальная, 5"


# =============================================================================
# Unit: blank mapping (no soffice)
# =============================================================================


def test_waybill_export_filename_format() -> None:
    assert waybill_export_filename(date(2025, 4, 3)) == "ПЛ 03.04.25.xls"
    assert waybill_export_filename(date(2025, 12, 9)) == "ПЛ 09.12.25.xls"


def test_format_bs41_patches_norm() -> None:
    assert format_bs41_formula(9.5) == "='стр.2'!C19*9.5/100"
    assert format_bs41_formula(10.3) == "='стр.2'!C19*10.3/100"
    assert format_bs41_formula(14.5) == "='стр.2'!C19*14.5/100"


def test_fill_workbook_writes_date_driver_legs_and_norm(tmp_path: Path) -> None:
    src = FIXTURE_TEMPLATE if FIXTURE_TEMPLATE.exists() else BLANK_TEMPLATE
    assert src.exists(), "blank template missing — run Phase 0 or copy fixture"
    dst = tmp_path / "filled.xlsx"
    shutil.copy2(src, dst)

    data = BlankWaybill(
        day=date(2025, 4, 7),
        vehicle_mark="Легковой универсал Geely Tugella",
        plate_number="О 848 ХР 44",
        driver=BlankDriver(
            full_name="Кулигин Никита Валерьевич",
            license_number="44 21 846315",
            license_issued_at="30.07.2015",
            personnel_number="143",
            snils="103-862-039-30",
        ),
        odometer_start=30107,
        fuel_start=40.0,
        fuel_issued=35.5,
        norm_l_per_100=9.4,
        legs=(
            BlankLeg(
                from_addr="Завод",
                to_addr=STATION_ADDR,
                km=95,
                dep_time="07:10",
                arr_time="09:40",
            ),
            BlankLeg(
                from_addr=STATION_ADDR,
                to_addr="Завод",
                km=95,
                dep_time="12:05",
                arr_time="14:30",
            ),
        ),
    )
    wb = load_workbook(dst, data_only=False)
    fill_workbook(wb, data)
    wb.save(dst)

    wb2 = load_workbook(dst, data_only=False)
    ws1 = wb2["стр.1"]
    ws2 = wb2["стр.2"]

    assert ws1["W3"].value == 7
    assert ws1["AB3"].value == "апреля"
    assert ws1["AL3"].value == 2025
    assert ws1["R12"].value == "Кулигин Никита Валерьевич"
    assert "44 21 846315" in str(ws1["X14"].value)
    assert ws1["BZ20"].value == 30107
    assert ws1["BS34"].value == pytest.approx(40.0)
    assert ws1["BS38"].value == pytest.approx(35.5)
    assert ws1["BS41"].value == "='стр.2'!C19*9.4/100"
    assert ws1["AN11"].value == "О 848 ХР 44"

    assert ws2["C5"].value == "Завод"
    assert ws2["D5"].value == STATION_ADDR
    assert ws2["I5"].value == 95
    assert ws2["C6"].value == STATION_ADDR
    assert ws2["D6"].value == "Завод"
    assert "АЗС" in str(ws2["D5"].value)


# =============================================================================
# Fixtures / seed
# =============================================================================


def _fresh_db(tmp_path: Path, name: str = "gsm_export.db") -> str:
    db_path = str(tmp_path / name)
    kp_db_schema._schema_ready.clear()
    kp_db_schema.ensure_schema(db_path)
    return db_path


@pytest.fixture()
def db_path(tmp_path: Path) -> str:
    return _fresh_db(tmp_path)


@pytest.fixture()
def repo(db_path: str) -> GsmRepository:
    return GsmRepository(db_path=db_path)


@pytest.fixture()
def api_client(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> CsrfAwareTestClient:
    db = tmp_path / "plita.db"
    monkeypatch.setenv("APP_SECRET_KEY", VALID_APP_SECRET_KEY)
    monkeypatch.setenv("PLITA_DB_PATH", str(db))
    monkeypatch.setenv("PB_DB_PATH", str(db))
    get_settings.cache_clear()
    kp_db_schema._schema_ready.clear()
    kp_db_schema.ensure_schema(str(db))
    patch_auth_users(monkeypatch, TEST_USERS)
    return CsrfAwareTestClient(create_app())


def _auth(username: str = "accountant_user") -> dict[str, str]:
    by_user = {
        "admin": (1, "admin"),
        "accountant_user": (5, "accountant"),
        "manager_a": (3, "manager"),
    }
    user_id, role = by_user[username]
    return session_cookie(user_id, role, username)


def _seed_export_bundle(
    repo: GsmRepository,
    *,
    day: date = date(2025, 4, 7),
    norm_summer: float = 9.4,
    norm_winter: float = 10.3,
    fuel_start: float = 40.0,
    fuel_issued: float = 35.0,
    odometer_start: int = 30000,
    km: int = 190,
    with_azs: bool = True,
) -> dict[str, Any]:
    driver_id = repo.create_driver(
        full_name="Кулигин Никита Валерьевич",
        license_number="44 21 846315",
        license_issued_at="2015-07-30",
        personnel_number="143",
        snils="103-862-039-30",
    )
    vehicle_id = repo.create_vehicle(
        name="Geely Tugella 848",
        plate_number="О 848 ХР 44",
        tank_volume_liters=55.0,
        norm_summer=norm_summer,
        norm_winter=norm_winter,
        primary_driver_id=driver_id,
    )
    station_id = repo.create_station(address=STATION_ADDR, brand="TATNEFT")
    burn = burn_for_km(km, norm_summer if day.month < 11 else norm_winter)
    fuel_end = fuel_start + fuel_issued - burn
    legs: list[dict[str, Any]]
    if with_azs:
        half = km // 2
        legs = [
            {
                "from": "Завод",
                "to": STATION_ADDR,
                "km": half,
                "station_id": station_id,
                "dep_time": "07:10",
                "arr_time": "09:40",
            },
            {
                "from": STATION_ADDR,
                "to": "Завод",
                "km": km - half,
                "dep_time": "12:05",
                "arr_time": "14:30",
            },
        ]
    else:
        legs = [{"from": "Завод", "to": "Клиент", "km": km}]
    route_json = json.dumps(legs, ensure_ascii=False)
    waybill_id = repo.upsert_waybill(
        vehicle_id=vehicle_id,
        date=day,
        driver_id=driver_id,
        status="draft",
        source="manual",
        odometer_start=odometer_start,
        odometer_end=odometer_start + km,
        fuel_start=fuel_start,
        fuel_issued=fuel_issued,
        fuel_end=fuel_end,
        route_json=route_json,
    )
    repo.set_setting(
        "season_switches",
        json.dumps([{"date": "2025-11-01", "mode": "winter"}]),
    )
    return {
        "vehicle_id": vehicle_id,
        "driver_id": driver_id,
        "station_id": station_id,
        "waybill_id": waybill_id,
        "day": day,
        "fuel_start": fuel_start,
        "fuel_issued": fuel_issued,
        "fuel_end": fuel_end,
        "odometer_start": odometer_start,
        "odometer_end": odometer_start + km,
        "km": km,
        "norm_summer": norm_summer,
        "norm_winter": norm_winter,
    }


# =============================================================================
# Service / API
# =============================================================================


def test_export_requires_accounting(api_client: CsrfAwareTestClient) -> None:
    settings = get_settings()
    repo = GsmRepository(db_path=settings.plita_db_path)
    seeded = _seed_export_bundle(repo)

    r = api_client.post(
        EXPORT,
        json={
            "vehicle_ids": [seeded["vehicle_id"]],
            "from": "2025-04-01",
            "to": "2025-04-30",
        },
        cookies=_auth("manager_a"),
    )
    assert r.status_code == 403


@pytest.mark.skipif(not HAS_SOFFICE, reason="soffice not installed")
def test_export_zip_roundtrip_values_and_azs(
    api_client: CsrfAwareTestClient,
    tmp_path: Path,
) -> None:
    settings = get_settings()
    repo = GsmRepository(db_path=settings.plita_db_path)
    seeded = _seed_export_bundle(repo, day=date(2025, 4, 7), with_azs=True)

    r = api_client.post(
        EXPORT,
        json={
            "vehicle_ids": [seeded["vehicle_id"]],
            "from": "2025-04-01",
            "to": "2025-04-30",
        },
        cookies=_auth(),
    )
    assert r.status_code == 200, r.text
    assert "application/zip" in r.headers.get("content-type", "")
    assert "attachment" in r.headers.get("content-disposition", "").lower()

    zf = zipfile.ZipFile(io.BytesIO(r.content))
    names = zf.namelist()
    expected_name = "ПЛ 07.04.25.xls"
    assert expected_name in names
    assert len(names) == 1

    xls_bytes = zf.read(expected_name)
    xls_path = tmp_path / expected_name
    xls_path.write_bytes(xls_bytes)

    book = xlrd.open_workbook(str(xls_path))
    sh1 = book.sheet_by_index(0)
    sh2 = book.sheet_by_index(1)

    # BS34 / BS38 / BS39 / BS41 / BZ20 / BY45 — 0-based from import map
    fuel_start = float(sh1.cell_value(33, 70))
    fuel_issued = float(sh1.cell_value(37, 70))
    fuel_end = float(sh1.cell_value(38, 70))
    burn = float(sh1.cell_value(40, 70))  # BS41
    odo_start = float(sh1.cell_value(19, 77))
    odo_end = float(sh1.cell_value(44, 76))

    assert fuel_start == pytest.approx(seeded["fuel_start"], abs=FLOAT_TOL)
    assert fuel_issued == pytest.approx(seeded["fuel_issued"], abs=FLOAT_TOL)
    assert fuel_end == pytest.approx(seeded["fuel_end"], abs=FLOAT_TOL)
    expected_burn = burn_for_km(seeded["km"], seeded["norm_summer"])
    assert burn == pytest.approx(expected_burn, abs=FLOAT_TOL)
    assert odo_start == pytest.approx(seeded["odometer_start"], abs=FLOAT_TOL)
    assert odo_end == pytest.approx(seeded["odometer_end"], abs=FLOAT_TOL)

    # Reverse side: АЗС in legs
    reverse_text = " ".join(
        str(sh2.cell_value(r, c))
        for r in range(min(sh2.nrows, 20))
        for c in range(min(sh2.ncols, 10))
    )
    assert "АЗС" in reverse_text
    assert STATION_ADDR.split(",")[0] in reverse_text or "Татнефть" in reverse_text


@pytest.mark.skipif(not HAS_SOFFICE, reason="soffice not installed")
def test_export_winter_norm_in_bs41(
    api_client: CsrfAwareTestClient,
    tmp_path: Path,
) -> None:
    settings = get_settings()
    repo = GsmRepository(db_path=settings.plita_db_path)
    winter_day = date(2025, 12, 5)
    seeded = _seed_export_bundle(
        repo,
        day=winter_day,
        norm_summer=9.4,
        norm_winter=10.3,
        km=200,
        with_azs=True,
    )
    # Recompute fuel_end with winter norm (seed helper used month<11 heuristic)
    burn = burn_for_km(200, 10.3)
    fuel_end = seeded["fuel_start"] + seeded["fuel_issued"] - burn
    repo.upsert_waybill(
        vehicle_id=seeded["vehicle_id"],
        date=winter_day,
        driver_id=seeded["driver_id"],
        status="draft",
        source="manual",
        odometer_start=seeded["odometer_start"],
        odometer_end=seeded["odometer_start"] + 200,
        fuel_start=seeded["fuel_start"],
        fuel_issued=seeded["fuel_issued"],
        fuel_end=fuel_end,
        route_json=repo.get_waybill(seeded["vehicle_id"], winter_day)["route_json"],
    )

    r = api_client.post(
        EXPORT,
        json={
            "vehicle_ids": [seeded["vehicle_id"]],
            "from": "2025-12-01",
            "to": "2025-12-31",
        },
        cookies=_auth(),
    )
    assert r.status_code == 200, r.text
    zf = zipfile.ZipFile(io.BytesIO(r.content))
    name = "ПЛ 05.12.25.xls"
    assert name in zf.namelist()
    path = tmp_path / name
    path.write_bytes(zf.read(name))

    # Check formula was patched: open via converting back is heavy;
    # values after recalc use winter norm.
    book = xlrd.open_workbook(str(path))
    sh1 = book.sheet_by_index(0)
    burn_cell = float(sh1.cell_value(40, 70))
    assert burn_cell == pytest.approx(burn_for_km(200, 10.3), abs=FLOAT_TOL)
    fuel_end_cell = float(sh1.cell_value(38, 70))
    assert fuel_end_cell == pytest.approx(fuel_end, abs=FLOAT_TOL)


def test_export_soffice_failure_returns_500(
    api_client: CsrfAwareTestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    settings = get_settings()
    repo = GsmRepository(db_path=settings.plita_db_path)
    seeded = _seed_export_bundle(repo)

    def _boom(*_a: Any, **_k: Any) -> None:
        raise RuntimeError("soffice failed: boom")

    monkeypatch.setattr(
        "app.services.gsm_export_service.run_soffice",
        _boom,
    )

    r = api_client.post(
        EXPORT,
        json={
            "vehicle_ids": [seeded["vehicle_id"]],
            "from": "2025-04-01",
            "to": "2025-04-30",
        },
        cookies=_auth(),
    )
    assert r.status_code == 500, r.text
    body = r.json()
    detail = body.get("detail") or body
    text = json.dumps(detail, ensure_ascii=False) if not isinstance(detail, str) else detail
    assert "soffice" in text.lower() or "экспорт" in text.lower() or "export" in text.lower()


def test_export_multi_day_zip_names(api_client: CsrfAwareTestClient) -> None:
    """Without soffice: mock convert to produce stub .xls files; check zip names."""
    settings = get_settings()
    repo = GsmRepository(db_path=settings.plita_db_path)
    d1 = _seed_export_bundle(repo, day=date(2025, 4, 7))
    # second day same vehicle
    repo.upsert_waybill(
        vehicle_id=d1["vehicle_id"],
        date=date(2025, 4, 8),
        driver_id=d1["driver_id"],
        status="draft",
        source="manual",
        odometer_start=d1["odometer_end"],
        odometer_end=d1["odometer_end"] + 100,
        fuel_start=d1["fuel_end"],
        fuel_issued=0.0,
        fuel_end=d1["fuel_end"] - burn_for_km(100, d1["norm_summer"]),
        route_json=json.dumps(
            [
                {"from": "Завод", "to": STATION_ADDR, "km": 50, "station_id": d1["station_id"]},
                {"from": STATION_ADDR, "to": "Завод", "km": 50},
            ],
            ensure_ascii=False,
        ),
    )

    def fake_convert(src: Path, fmt: str, outdir: Path, workdir: Path, timeout: int) -> Path:
        out = outdir / f"{src.stem}.{fmt}"
        if fmt == "xls":
            # Minimal BIFF-ish file xlrd can open is hard; write bytes + skip xlrd.
            # Service only needs the file to exist for zipping.
            out.write_bytes(b"stub-xls")
        else:
            shutil.copy2(src, out)
        return out

    with patch("app.services.gsm_export_service.convert_with_soffice", side_effect=fake_convert):
        r = api_client.post(
            EXPORT,
            json={
                "vehicle_ids": [d1["vehicle_id"]],
                "from": "2025-04-01",
                "to": "2025-04-30",
            },
            cookies=_auth(),
        )
    assert r.status_code == 200, r.text
    names = sorted(zipfile.ZipFile(io.BytesIO(r.content)).namelist())
    assert names == ["ПЛ 07.04.25.xls", "ПЛ 08.04.25.xls"]


def _fake_xls_convert(src: Path, fmt: str, outdir: Path, workdir: Path, timeout: int) -> Path:
    out = outdir / f"{src.stem}.{fmt}"
    out.write_bytes(b"stub-xls")
    return out


def _seed_named_export_vehicle(
    repo: GsmRepository,
    *,
    name: str,
    plate: str,
    license_number: str,
    days: list[date],
    status: str = "draft",
) -> dict[str, Any]:
    driver_id = repo.create_driver(
        full_name=f"Водитель {plate}",
        license_number=license_number,
        license_issued_at="2015-01-01",
    )
    vehicle_id = repo.create_vehicle(
        name=name,
        plate_number=plate,
        tank_volume_liters=55.0,
        norm_summer=9.4,
        norm_winter=10.3,
        primary_driver_id=driver_id,
    )
    for i, day in enumerate(days):
        km = 100
        odo = 10_000 + i * km
        repo.upsert_waybill(
            vehicle_id=vehicle_id,
            date=day,
            driver_id=driver_id,
            status=status,
            source="manual",
            odometer_start=odo,
            odometer_end=odo + km,
            fuel_start=20.0,
            fuel_issued=10.0,
            fuel_end=20.5,
            route_json=json.dumps([{"from": "Завод", "to": "Клиент", "km": km}]),
        )
    repo.set_setting(
        "season_switches",
        json.dumps([{"date": "2026-04-01", "mode": "summer"}], ensure_ascii=False),
    )
    return {"vehicle_id": vehicle_id, "driver_id": driver_id}


def _waybill_statuses(
    repo: GsmRepository, vehicle_id: int, period_from: date, period_to: date
) -> list[str]:
    return [
        str(wb["status"])
        for wb in repo.list_waybills(
            vehicle_id=vehicle_id,
            period_from=period_from,
            period_to=period_to,
        )
    ]


def test_export_august_monjaro_july_tail_does_not_flip(
    api_client: CsrfAwareTestClient,
) -> None:
    settings = get_settings()
    repo = GsmRepository(db_path=settings.plita_db_path)
    monjaro = _seed_named_export_vehicle(
        repo,
        name="Monjaro Tail",
        plate="О 301 МН 44",
        license_number="44 30 130111",
        days=[date(2026, 7, 10), date(2026, 8, 5)],
    )
    august_from = date(2026, 8, 1)
    august_to = date(2026, 8, 31)

    with patch(
        "app.services.gsm_export_service.convert_with_soffice",
        side_effect=_fake_xls_convert,
    ):
        r = api_client.post(
            EXPORT,
            json={
                "vehicle_ids": [monjaro["vehicle_id"]],
                "from": august_from.isoformat(),
                "to": august_to.isoformat(),
            },
            cookies=_auth(),
        )
    assert r.status_code in (400, 404, 422), r.text
    detail = r.json().get("detail") or r.json()
    code = detail.get("code") if isinstance(detail, dict) else None
    assert code == "gsm_kit_tail"
    assert _waybill_statuses(repo, monjaro["vehicle_id"], august_from, august_to) == [
        "draft"
    ]


def test_export_august_mix_exports_clean_skips_tail(
    api_client: CsrfAwareTestClient,
) -> None:
    settings = get_settings()
    repo = GsmRepository(db_path=settings.plita_db_path)
    monjaro = _seed_named_export_vehicle(
        repo,
        name="Monjaro Tail",
        plate="О 302 МН 44",
        license_number="44 30 130222",
        days=[date(2026, 7, 10), date(2026, 8, 5)],
    )
    palisade = _seed_named_export_vehicle(
        repo,
        name="Palisade Clean",
        plate="О 303 ПЛ 44",
        license_number="44 30 130333",
        days=[date(2026, 8, 10)],
    )
    august_from = date(2026, 8, 1)
    august_to = date(2026, 8, 31)

    with patch(
        "app.services.gsm_export_service.convert_with_soffice",
        side_effect=_fake_xls_convert,
    ):
        r = api_client.post(
            EXPORT,
            json={
                "vehicle_ids": [monjaro["vehicle_id"], palisade["vehicle_id"]],
                "from": august_from.isoformat(),
                "to": august_to.isoformat(),
            },
            cookies=_auth(),
        )
    assert r.status_code == 200, r.text
    names = sorted(zipfile.ZipFile(io.BytesIO(r.content)).namelist())
    assert names == ["ПЛ 10.08.26.xls"]
    assert _waybill_statuses(repo, palisade["vehicle_id"], august_from, august_to) == [
        "exported"
    ]
    assert _waybill_statuses(repo, monjaro["vehicle_id"], august_from, august_to) == [
        "draft"
    ]

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""XLSX сверки ввода заказа плит: как прислали → распознано → как в КП."""

from __future__ import annotations

import logging
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    Workbook = None  # type: ignore

from core.config_and_data import LineContributionKey, load_code_for_price_match, make_plate_name

logger = logging.getLogger(__name__)

_WARN_FILL = PatternFill(start_color="FFF4B084", end_color="FFF4B084", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")


def split_plate_text_lines(text: str) -> list[str]:
    """Такая же разбивка строк, как в ``set_plate_lists_from_text`` после нормализации текста."""
    t = (text or "").replace("\u00d7", "x").replace("×", "x")
    t = t.replace("\u00a0", " ")
    return [re.sub(r"\s+", " ", ln).strip() for ln in re.split(r"[\n;]+", t) if ln.strip()]


def _unparsed_line_bases(unparsed_lines: list[str]) -> set[str]:
    bases: set[str] = set()
    for u in unparsed_lines:
        s = (u or "").strip()
        if not s:
            continue
        if " (пропущено:" in s:
            bases.add(s.split(" (пропущено:", 1)[0].strip())
        bases.add(s)
    return bases


def is_line_unparsed(norm_line: str, unparsed_lines: list[str]) -> bool:
    nl = (norm_line or "").strip()
    if not nl:
        return False
    bases = _unparsed_line_bases(unparsed_lines)
    return nl in bases


def _order_item_load_code(item: dict[str, Any]) -> Optional[float]:
    lc = item.get("load_code")
    if lc is not None:
        try:
            return float(lc)
        except (TypeError, ValueError):
            pass
    load_class = item.get("load_class")
    if load_class is not None:
        try:
            return float(load_class) / 100.0
        except (TypeError, ValueError):
            pass
    return None


def _matches_key(item: dict[str, Any], key: LineContributionKey) -> bool:
    lm_k, wm_k, lc_k, ldr_k = key
    try:
        lm = float(item.get("length_m", 0))
        wm = float(item.get("width_m", 0))
    except (TypeError, ValueError):
        return False
    if abs(lm - lm_k) >= 0.01 or abs(wm - wm_k) >= 0.01:
        return False
    ldr_item = (item.get("length_dm_raw") or "").strip()
    ldr_key = (ldr_k or "").strip()
    if ldr_key and ldr_item and ldr_key != ldr_item:
        if ldr_key.replace(",", ".") != ldr_item.replace(",", "."):
            return False
    od_lc = _order_item_load_code(item)
    if lc_k is not None and od_lc is not None:
        if load_code_for_price_match(od_lc) != load_code_for_price_match(lc_k):
            return False
    return True


def format_kp_for_keys(order_data: list[dict[str, Any]], keys: list[LineContributionKey]) -> str:
    """Собирает текст для столбца «Как в КП» по ключам вклада строки."""
    if not keys:
        return ""
    totals: dict[str, int] = defaultdict(int)
    for key in dict.fromkeys(keys):
        for item in order_data:
            if not _matches_key(item, key):
                continue
            name = str(item.get("name") or "").strip()
            qty = int(item.get("qty") or 0)
            if name and qty > 0:
                totals[name] += qty
    if totals:
        lines = [f"{name} — {qty} шт" for name, qty in sorted(totals.items())]
        return "\n".join(lines)
    # Fallback: канонические имена по ключам
    fallback: list[str] = []
    seen_k: set[LineContributionKey] = set()
    for key in dict.fromkeys(keys):
        if key in seen_k:
            continue
        seen_k.add(key)
        lm, wm, lc, ldr = key
        name = make_plate_name(
            lm,
            wm,
            load_code=lc if lc is not None else None,  # float допустим (12.5п)
            length_dm_raw=ldr or None,
        )
        fallback.append(name)
    return "\n".join(fallback) if fallback else ""


_PB_WIDE_RE = re.compile(
    r"(?i)\b(?P<prefix>п[бк])\s*"
    r"(?P<length>[\d,.]+)\s*-\s*"
    r"(?P<width>[\d,.]+)\s*-\s*"
    r"(?P<load>[\d,.]+)\s*п?\s*"
    r"(?P<qty>\d+)?\s*(?:шт\.?|штук)?\s*$"
)


def wide_replacement_hint(norm_line: str) -> str:
    """Подсказка замены для плит шире 12 дм (в т.ч. 15 дм в марке ПБ)."""
    line = (norm_line or "").strip()
    if not line:
        return ""
    from core.plate_text_normalizer import get_wide_plate_lines

    wide = get_wide_plate_lines(line)
    if not wide:
        return ""

    m = _PB_WIDE_RE.search(line)
    if m:
        prefix = m.group("prefix").upper()
        length_part = m.group("length").replace(".", ",")
        load_part = m.group("load").replace(".", ",")
        qty = (m.group("qty") or "").strip()
        qty_suffix = f" {qty}" if qty else ""
        return (
            f"Вариант замены: {prefix} {length_part}-12-{load_part}п{qty_suffix}\n"
            f"+ {prefix} {length_part}-3,0-{load_part}п{qty_suffix}"
        )
    return (
        "Ширина > 12 дм: в КП для ленты 1.5 м используются две позиции — 1.2 м и 0.3 м "
        "(см. столбец «Как в КП»)."
    )


def build_reconciliation_xlsx(
    path: str | Path,
    *,
    plates_text: str,
    raw_plate_lines: list[str],
    unparsed_lines: list[str],
    line_contributions: list[list[LineContributionKey]],
    order_data: list[dict[str, Any]],
    is_photo: bool,
) -> None:
    """Создаёт файл сверки. При ``is_photo`` колонка «как прислал» не заполняется."""
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl не установлен — сверка XLSX недоступна")

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    norm_lines = split_plate_text_lines(plates_text)
    n = max(len(norm_lines), len(line_contributions))
    if len(norm_lines) != len(line_contributions):
        logger.warning(
            "[reconciliation] len(norm_lines)=%s != len(line_contributions)=%s",
            len(norm_lines),
            len(line_contributions),
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Сверка"

    if is_photo:
        headers = ["Распознано", "Как в КП", "Подсказка (ширина >12 дм)"]
    else:
        headers = ["Как прислал пользователь", "Распознано", "Как в КП", "Подсказка (ширина >12 дм)"]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP

    for i in range(n):
        r = i + 2
        norm = norm_lines[i] if i < len(norm_lines) else ""
        contrib = line_contributions[i] if i < len(line_contributions) else []
        kp_text = format_kp_for_keys(order_data, contrib)
        hint = wide_replacement_hint(norm)
        bad = is_line_unparsed(norm, unparsed_lines)

        if is_photo:
            c1, c2, c3 = 1, 2, 3
            ws.cell(row=r, column=c1, value=norm).alignment = _WRAP
            ws.cell(row=r, column=c2, value=kp_text).alignment = _WRAP
            ws.cell(row=r, column=c3, value=hint).alignment = _WRAP
            fill_row = (c1, c2, c3)
        else:
            raw = raw_plate_lines[i] if i < len(raw_plate_lines) else ""
            ws.cell(row=r, column=1, value=raw).alignment = _WRAP
            ws.cell(row=r, column=2, value=norm).alignment = _WRAP
            ws.cell(row=r, column=3, value=kp_text).alignment = _WRAP
            ws.cell(row=r, column=4, value=hint).alignment = _WRAP
            fill_row = (1, 2, 3, 4)

        if bad:
            for c in fill_row:
                ws.cell(row=r, column=c).fill = _WARN_FILL

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 36 if col <= 2 else 40

    wb.save(str(path))

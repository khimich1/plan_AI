#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Модуль генерации коммерческого предложения в формате XLSX
Создаёт документ по образцу КП с расчётными формулами
"""

import io
import logging
import os
from typing import Any, List, Dict, Optional

try:
    import pandas as pd
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.drawing.image import Image as XLImage
    HAS_EXCEL_LIBS = True
except ImportError:
    HAS_EXCEL_LIBS = False
    XLImage = None
    logging.getLogger(__name__).warning(
        "pandas или openpyxl не установлены; генерация XLSX будет недоступна"
    )

# Единый расчёт веса строки КП (plate_weights → approximate)
try:
    from .cargo_delivery_pricing import (
        cargo_delivery_trips_count,
        delivery_service_charge_rub,
        total_order_cargo_weight_kg,
    )
    from .kp_plate_weight import resolve_kp_line_weight_kg
except ImportError:
    from cargo_delivery_pricing import (
        cargo_delivery_trips_count,
        delivery_service_charge_rub,
        total_order_cargo_weight_kg,
    )
    from kp_plate_weight import resolve_kp_line_weight_kg


# ==================== КОНСТАНТЫ ====================

# Вычисляем абсолютный путь к корню проекта (папка выше core/)
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Реквизиты компании
COMPANY_NAME = "ООО «Комбинат ЖБК»"
COMPANY_ADDRESS = "150020, г Ярославль г, проезд Домостроителей, дом 1, строение 3"
COMPANY_PHONE = "8 (4852) 595 000"
COMPANY_EMAIL = "info@zhbk.ru"

# Путь к базе данных с ценами (абсолютный)
DB_PATH = os.path.join(PROJECT_ROOT, "pb.db")

logger = logging.getLogger(__name__)

try:
    from core.project_paths import resolve_commercial_offer_logo_path
except ImportError:
    from project_paths import resolve_commercial_offer_logo_path

# Коэффициент скидки в процентах (0 = без скидки, 5 = скидка 5%, и т.д.)
DISCOUNT_PERCENT = 0


# ==================== ФУНКЦИИ ====================

def get_plate_price(length_m: float, width_m: float, load_class: int = 800) -> float:
    """
    Получает цену плиты из базы данных по длине, ширине и классу нагрузки
    
    Args:
        length_m: длина плиты в метрах
        width_m: ширина плиты в метрах  
        load_class: класс нагрузки (по умолчанию 800 кг/м²)
    
    Returns:
        Цена плиты в рублях
    
    Raises:
        PriceNotFoundError: если цена отсутствует в прайс-листе
    """
    from core.commercial_pricing import lookup_plate_price

    return lookup_plate_price(length_m, width_m, load_class, db_path=DB_PATH)


from core.commercial_line_format import format_line_name  # noqa: E402
from core.commercial_offer_layout import (  # noqa: E402
    commercial_offer_table_headers,
    is_unified_commercial_document,
    line_product_type,
    product_type_label,
)
from core.commercial_pricing import (  # noqa: E402
    VAT_RATE,
    calculate_total_cost as _calculate_total_cost,
    format_phone,
    is_bridge_pile_order,
    is_fbs_order,
    is_march_order,
    is_pile_order,
    is_step_order,
    kp_delivery_export_lines,
    lookup_bridge_pile_price,
    lookup_fbs_price,
    lookup_march_price,
    lookup_pile_price,
    lookup_step_price,
)


def calculate_total_cost(
    order_data: List[Dict],
    discount_percent: float = 0,
    logistics_cost: float = 0,
    *,
    pile_logistics_cost: float = 0,
    pile_trip_overrides: dict | None = None,
    pile_catalog_db_path: str | None = None,
) -> Dict:
    return _calculate_total_cost(
        order_data,
        discount_percent,
        logistics_cost,
        db_path=DB_PATH,
        pile_logistics_cost=pile_logistics_cost,
        pile_trip_overrides=pile_trip_overrides,
        pile_catalog_db_path=pile_catalog_db_path,
    )


def _resolve_line_unit_price(item: Dict, *, order_mode: str) -> float:
    """Resolve unit price for one line (unified uses per-line type)."""
    if "unit_price" in item and item["unit_price"] is not None:
        return float(item["unit_price"])

    pt = line_product_type(item) if order_mode == "unified" else order_mode
    mark = str(item.get("mark") or item.get("name") or "").strip()
    grade = str(item.get("concrete_grade") or "B25").strip()

    if pt in ("piles", "pile"):
        return lookup_pile_price(mark, grade, db_path=DB_PATH)
    if pt in ("bridge_piles", "bridge_pile"):
        return lookup_bridge_pile_price(mark, grade, db_path=DB_PATH)
    if pt == "fbs":
        return lookup_fbs_price(mark, grade, db_path=DB_PATH)
    if pt in ("marches", "march"):
        return lookup_march_price(mark, grade, db_path=DB_PATH)
    if pt in ("steps", "step"):
        return lookup_step_price(mark, db_path=DB_PATH)

    name = item.get("name", "Плиты ПБ")
    length_m = item.get("length_m", 0)
    width_m = item.get("width_m", 0)
    load_class = item.get("load_class")
    if load_class is None:
        try:
            from config_and_data import parse_load_code_from_name

            load_code = parse_load_code_from_name(name, default=8)
            load_class = max(1, load_code) * 100
        except ImportError:
            load_class = 800
    return get_plate_price(length_m, width_m, load_class)


def generate_commercial_offer_xlsx(
    order_data: List[Dict],
    offer_number: str,
    offer_date: str,
    customer_name: Optional[str] = None,
    manager_name: Optional[str] = None,
    manager_phone: Optional[str] = None,
    manager_email: Optional[str] = None,
    discount_percent: float = 0,
    delivery_conditions: Optional[str] = None,
    payment_conditions: Optional[str] = None,
    kp_db_id: Optional[int] = None,
    logistics_cost: float = 0.0,
    append_batches: Optional[List[Dict]] = None,
    pile_logistics_cost: float = 0.0,
    pile_trip_overrides: Optional[Dict] = None,
    pile_catalog_db_path: Optional[str] = None,
) -> io.BytesIO:
    """
    Генерирует коммерческое предложение в формате XLSX с расчётными формулами
    
    Args:
        order_data: список позиций заказа с полями name, length_m, width_m, qty
        offer_number: номер КП
        offer_date: дата КП
        customer_name: имя заказчика (будет в строке 16)
        manager_name: имя менеджера (выводится в подписи)
        manager_phone: телефон менеджера (выводится под именем)
        manager_email: email менеджера (выводится под телефоном)
        discount_percent: процент скидки (0-100, по умолчанию 0)
        delivery_conditions: условия поставки (строка 28, если указано)
        payment_conditions: условия оплаты (строка 29, если указано)
        kp_db_id: номер КП из базы данных (если КП сохранен в БД)
        logistics_cost: стоимость одного рейса (без НДС по плитам; итог по формуле рейсов)
        append_batches: metadata заходов (для unified при multi-append одного типа)
    
    Returns:
        BytesIO буфер с XLSX файлом
    """
    if not HAS_EXCEL_LIBS:
        raise ImportError("Для генерации XLSX необходимы pandas и openpyxl")
    
    buffer = io.BytesIO()
    
    # Определяем, есть ли логотип (плашка; будет вставлен выше заголовка)
    logo_path = resolve_commercial_offer_logo_path()
    has_logo = logo_path is not None
    
    # Формируем заголовок документа
    # Если есть логотип, добавляем пустые строки для него
    header_data = []
    
    if has_logo:
        # Резервируем место под логотип (7 строк как в образце)
        header_data.extend([[""], [""], [""], [""], [""], [""], [""]])
    
    # Формируем строку с номером КП и датой
    if kp_db_id:
        date_line = f"КП № {kp_db_id} от {offer_date}"
    else:
        date_line = f"от {offer_date}"
    
    # Формируем строку с контактами (МЕНЕДЖЕРА, а не компании!)
    if manager_phone and manager_email:
        contacts_line = f"Тел.: {format_phone(manager_phone)}, email: {manager_email}"
    elif manager_phone:
        contacts_line = f"Тел.: {format_phone(manager_phone)}"
    elif manager_email:
        contacts_line = f"Email: {manager_email}"
    else:
        # Fallback: контакты компании, если нет контактов менеджера
        contacts_line = f"Тел.: {COMPANY_PHONE}, Email: {COMPANY_EMAIL}"
    
    header_data.extend([
        [f"{COMPANY_ADDRESS}"],
        [contacts_line],  # 🆕 КОНТАКТЫ МЕНЕДЖЕРА (не компании!)
        [""],
        [date_line],
        ["Срок действия коммерческого предложения 3 дня"],
        [""],
        ["Коммерческое предложение для"],
        [customer_name or "Заказчик не указан"],
        [""],
    ])
    
    # Создаем DataFrame для заголовка
    df_header = pd.DataFrame(header_data)
    
    # Формируем таблицу товаров
    table_data = []
    unified = is_unified_commercial_document(order_data, append_batches=append_batches)
    pile_order = is_pile_order(order_data)
    bridge_pile_order = is_bridge_pile_order(order_data)
    fbs_order = is_fbs_order(order_data)
    march_order = is_march_order(order_data)
    step_order = is_step_order(order_data)
    table_headers = commercial_offer_table_headers(order_data, append_batches=append_batches)
    col_count = len(table_headers)

    if unified:
        order_mode = "unified"
    elif pile_order:
        order_mode = "piles"
    elif bridge_pile_order:
        order_mode = "bridge_piles"
    elif fbs_order:
        order_mode = "fbs"
    elif march_order:
        order_mode = "marches"
    elif step_order:
        order_mode = "steps"
    else:
        order_mode = "plates"

    total_weight = 0.0

    for idx, item in enumerate(order_data, start=1):
        qty = item.get('qty', 0)
        unit_price = _resolve_line_unit_price(item, order_mode=order_mode)
        discounted_price = unit_price * (1 - discount_percent / 100)

        if unified:
            pt = line_product_type(item)
            if pt == "plates":
                _, line_kg = resolve_kp_line_weight_kg(item)
                total_weight += line_kg
            table_data.append({
                '№': idx,
                'Тип': product_type_label(pt),
                'Наименование': format_line_name(item),
                'Кол-во': qty,
                'Цена': discounted_price,
                'Сумма': discounted_price * qty,
            })
            continue

        if pile_order or bridge_pile_order or fbs_order or march_order:
            table_data.append({
                '№': idx,
                'Наименование': str(item.get('mark') or item.get('name') or ''),
                'Класс бетона': str(item.get('concrete_grade') or 'B25'),
                'Кол-во': qty,
                'Цена': discounted_price,
                'Сумма': discounted_price * qty,
            })
            continue

        if step_order:
            table_data.append({
                '№': idx,
                'Наименование': str(item.get('mark') or item.get('name') or ''),
                'Кол-во': qty,
                'Цена': discounted_price,
                'Сумма': discounted_price * qty,
            })
            continue

        name = item.get('name', 'Плиты ПБ')
        _, total_item_weight = resolve_kp_line_weight_kg(item)
        total_weight += total_item_weight
        table_data.append({
            '№': idx,
            'Наименование': name,
            'Кол-во': qty,
            'Ед.': 'шт',
            'Вес(кг)': total_item_weight,
            'Цена': discounted_price,
            'Сумма': discounted_price * qty
        })
    
    trip_cost = max(0.0, float(logistics_cost or 0.0))
    pile_trip = max(0.0, float(pile_logistics_cost or 0.0))
    totals = calculate_total_cost(
        order_data,
        discount_percent,
        logistics_cost=trip_cost,
        pile_logistics_cost=pile_trip,
        pile_trip_overrides=pile_trip_overrides,
        pile_catalog_db_path=pile_catalog_db_path,
    )
    if unified:
        plates_kg = total_order_cargo_weight_kg(order_data, product_types={"plates"})
        total_weight = plates_kg
    delivery_export = kp_delivery_export_lines(
        totals, plate_trip_cost=trip_cost, pile_trip_cost=pile_trip
    )
    for line in delivery_export:
        row: dict[str, Any] = {
            "№": len(table_data) + 1,
            "Наименование": line["label"],
            "Кол-во": line["trips"],
            "Цена": line["unit_price"],
            "Сумма": line["amount"],
        }
        if unified:
            row["Тип"] = ""
        elif not pile_order and not bridge_pile_order and not fbs_order and not march_order and not step_order:
            row["Ед."] = "рейс"
            row["Вес(кг)"] = 0.0
        table_data.append(row)

    delivery_row_count = len(delivery_export)

    df_table = pd.DataFrame(table_data, columns=table_headers)

    # Создаем строку итогов
    total_items = len(table_data)
    
    # Записываем в Excel
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Записываем заголовок (без заголовков столбцов)
        df_header.to_excel(writer, sheet_name='КП', index=False, header=False, startrow=0)
        
        # Записываем таблицу товаров (с заголовками)
        start_row = len(header_data)
        df_table.to_excel(writer, sheet_name='КП', index=False, startrow=start_row)
        
        # Получаем workbook и worksheet для форматирования
        workbook = writer.book
        worksheet = writer.sheets['КП']
        
        # === ВСТАВКА ЛОГОТИПА ===
        if has_logo and XLImage is not None:
            try:
                # Создаем объект изображения
                logo_img = XLImage(str(logo_path))
                
                # Масштабируем логотип (оригинал 1456x207)
                # Делаем ширину больше для лучшей читаемости
                logo_img.width = 1000  # пикселей (было 600)
                logo_img.height = int(logo_img.width * (207 / 1456.0))  # сохраняем пропорции
                
                # Объединяем ячейки для логотипа (7 строк x 7 колонок, как в образце)
                worksheet.merge_cells('A1:G7')
                
                # Вставляем в ячейку A1
                worksheet.add_image(logo_img, 'A1')
                
                # Увеличиваем высоту первых 7 строк для логотипа
                total_logo_height = logo_img.height
                for row_idx in range(1, 8):
                    worksheet.row_dimensions[row_idx].height = total_logo_height / 7
                    
            except Exception as e:
                logger.warning("Не удалось вставить логотип в XLSX: %s", e)
        
        # === ФОРМАТИРОВАНИЕ ===
        
        # Шрифты и стили
        header_font = Font(name='Tahoma', size=11, bold=True)
        title_font = Font(name='Tahoma', size=18, bold=True)
        customer_font = Font(name='Tahoma', size=16, bold=True)
        table_header_font = Font(name='Tahoma', size=11, bold=True)
        table_font = Font(name='Tahoma', size=10)
        summary_font = Font(name='Tahoma', size=11, bold=True)
        
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Границы для таблицы
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заливка для заголовка таблицы
        header_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        
        # Определяем смещение строк (если есть логотип, добавляем 7 строк)
        logo_offset = 7 if has_logo else 0
        
        # Форматируем заголовок документа (с учетом смещения для логотипа)
        company_row = logo_offset + 1
        address_row = logo_offset + 2
        phone_row = logo_offset + 3
        number_row = logo_offset + 5
        validity_row = logo_offset + 6
        title_row = logo_offset + 8
        customer_row = logo_offset + 9
        
        worksheet[f'A{company_row}'].font = header_font
        worksheet[f'A{address_row}'].font = header_font
        worksheet[f'A{phone_row}'].font = header_font
        worksheet[f'A{number_row}'].font = header_font
        worksheet[f'A{validity_row}'].font = header_font
        worksheet[f'A{title_row}'].font = title_font
        worksheet[f'A{title_row}'].alignment = center_align
        worksheet[f'A{customer_row}'].font = customer_font
        worksheet[f'A{customer_row}'].alignment = center_align
        
        # Объединяем ячейки для заголовка
        worksheet.merge_cells(f'A{company_row}:G{company_row}')
        worksheet.merge_cells(f'A{address_row}:G{address_row}')
        worksheet.merge_cells(f'A{phone_row}:G{phone_row}')
        worksheet.merge_cells(f'A{number_row}:G{number_row}')
        worksheet.merge_cells(f'A{validity_row}:G{validity_row}')
        worksheet.merge_cells(f'A{title_row}:G{title_row}')
        worksheet.merge_cells(f'A{customer_row}:G{customer_row}')
        
        # Форматируем заголовок таблицы
        table_header_row = start_row + 1
        for col_idx in range(1, col_count + 1):
            cell = worksheet.cell(row=table_header_row, column=col_idx)
            cell.font = table_header_font
            cell.alignment = center_align
            cell.fill = header_fill
            cell.border = thin_border
        
        # Форматируем строки таблицы
        table_rows_count = len(table_data)
        qty_col = table_headers.index('Кол-во') + 1
        price_col = table_headers.index('Цена') + 1
        sum_col = table_headers.index('Сумма') + 1
        qty_letter = chr(ord('A') + qty_col - 1)
        price_letter = chr(ord('A') + price_col - 1)
        sum_letter = chr(ord('A') + sum_col - 1)

        for row_idx in range(table_header_row + 1, table_header_row + 1 + table_rows_count):
            for col_idx in range(1, col_count + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.font = table_font
                header_name = table_headers[col_idx - 1]
                if header_name in ("№", "Тип", "Кол-во", "Ед.", "Класс бетона"):
                    cell.alignment = center_align
                elif header_name in ("Цена", "Сумма", "Вес(кг)"):
                    cell.alignment = right_align
                    cell.number_format = '#,##0.00'
                else:
                    cell.alignment = left_align

            sum_cell = worksheet.cell(row=row_idx, column=sum_col)
            sum_cell.value = f"={qty_letter}{row_idx}*{price_letter}{row_idx}"
            sum_cell.alignment = right_align
            sum_cell.border = thin_border
            sum_cell.font = table_font
            sum_cell.number_format = '#,##0.00'
        
        # Добавляем итоговые строки
        summary_row = table_header_row + table_rows_count + 2
        
        # Итого по таблице: сумма позиций (цены с НДС) + услуга доставки при наличии.
        subtotal_row = summary_row
        pre_sum_letter = chr(ord('A') + sum_col - 2)
        worksheet.merge_cells(f'A{subtotal_row}:{pre_sum_letter}{subtotal_row}')
        worksheet[f'A{subtotal_row}'] = f"Всего наименований {total_items}, общим весом {total_weight:,.3f} кг."
        worksheet[f'A{subtotal_row}'].font = summary_font
        worksheet[f'A{subtotal_row}'].alignment = left_align

        first_data_row = table_header_row + 1
        last_data_row = table_header_row + table_rows_count
        subtotal_cell = worksheet[f'{sum_letter}{subtotal_row}']
        subtotal_cell.value = f"=SUM({sum_letter}{first_data_row}:{sum_letter}{last_data_row})"
        subtotal_cell.font = summary_font
        subtotal_cell.number_format = '#,##0.00'
        subtotal_cell.alignment = right_align

        # НДС 22% только от суммы позиций (без строки «Услуга по доставке грузов»).
        last_plate_row = last_data_row - delivery_row_count if delivery_row_count else last_data_row
        vat_row = summary_row + 1
        worksheet.merge_cells(f'A{vat_row}:{pre_sum_letter}{vat_row}')
        worksheet[f'A{vat_row}'] = "в том числе НДС (22%)"
        worksheet[f'A{vat_row}'].font = summary_font
        worksheet[f'A{vat_row}'].alignment = left_align
        worksheet[f'{sum_letter}{vat_row}'] = (
            f"=SUM({sum_letter}{first_data_row}:{sum_letter}{last_plate_row})*{VAT_RATE}"
        )
        worksheet[f'{sum_letter}{vat_row}'].font = summary_font
        worksheet[f'{sum_letter}{vat_row}'].number_format = '#,##0.00'
        worksheet[f'{sum_letter}{vat_row}'].alignment = right_align
        
        # Условия
        conditions_row = vat_row + 2
        
        # Условия поставки (строка 28 с учётом смещения)
        if delivery_conditions:
            worksheet[f'A{conditions_row}'] = f"1. Условия поставки: {delivery_conditions}"
        else:
            worksheet[f'A{conditions_row}'] = "1. Условия поставки:"
        worksheet[f'A{conditions_row}'].font = table_font
        
        conditions_row += 1
        
        # Условия оплаты (строка 29 с учётом смещения)
        if payment_conditions:
            worksheet[f'A{conditions_row}'] = f"2. Условия оплаты: {payment_conditions}"
        else:
            worksheet[f'A{conditions_row}'] = "2. Условия оплаты: Предварительная оплата в размере 100%"
        worksheet[f'A{conditions_row}'].font = table_font
        
        # Подпись
        signature_row = conditions_row + 2
        worksheet[f'A{signature_row}'] = "С уважением,"
        worksheet[f'A{signature_row}'].font = summary_font
        
        signature_row += 1
        # Используем только имя менеджера (телефон и email теперь в шапке)
        worksheet[f'A{signature_row}'] = manager_name or "Менеджер"
        worksheet[f'A{signature_row}'].font = table_font

        # Настройка ширины столбцов
        worksheet.column_dimensions['A'].width = 5
        if unified:
            worksheet.column_dimensions['B'].width = 14
            worksheet.column_dimensions['C'].width = 45
            worksheet.column_dimensions['D'].width = 10
            worksheet.column_dimensions['E'].width = 15
            worksheet.column_dimensions['F'].width = 18
        else:
            worksheet.column_dimensions['B'].width = 45
            if pile_order or bridge_pile_order or fbs_order or march_order:
                worksheet.column_dimensions['C'].width = 14
                worksheet.column_dimensions['D'].width = 8
                worksheet.column_dimensions['E'].width = 15
                worksheet.column_dimensions['F'].width = 18
            elif step_order:
                worksheet.column_dimensions['C'].width = 8
                worksheet.column_dimensions['D'].width = 15
                worksheet.column_dimensions['E'].width = 18
            else:
                worksheet.column_dimensions['C'].width = 10
                worksheet.column_dimensions['D'].width = 8
                worksheet.column_dimensions['E'].width = 15
                worksheet.column_dimensions['F'].width = 18
                worksheet.column_dimensions['G'].width = 18
        
        # Высота строк
        worksheet.row_dimensions[table_header_row].height = 25
    
    buffer.seek(0)
    return buffer


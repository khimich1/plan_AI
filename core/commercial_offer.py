#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Модуль генерации коммерческого предложения в формате PDF и XLSX
Создаёт документ по образцу КП № 1133 от 16.10.2025
"""

import io
import os
import sqlite3
from datetime import datetime
from typing import List, Dict, Optional
from xml.sax.saxutils import escape
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.drawing.image import Image as XLImage
    HAS_EXCEL_LIBS = True
except ImportError:
    HAS_EXCEL_LIBS = False
    XLImage = None
    print("ВНИМАНИЕ: pandas или openpyxl не установлены. Генерация XLSX будет недоступна.")

# Импортируем функцию расчёта веса
try:
    from config_and_data import approximate_weight_kg
except ImportError:
    # Если не удалось импортировать, используем локальную функцию
    def approximate_weight_kg(length_m: float, width_m: float, thickness_m: float = 0.22) -> float:
        """Примерный расчёт веса плиты в килограммах"""
        volume = length_m * width_m * thickness_m
        return round(volume * 2400, 2)


# ==================== КОНСТАНТЫ ====================

# Реквизиты компании (согласно примеру КП)
COMPANY_NAME = "ООО «Комбинат ЖБК»"
COMPANY_ADDRESS = "150020, г Ярославль г, проезд Домостроителей, дом 1, строение 3"
COMPANY_PHONE = "8 (4852) 595 000"
COMPANY_EMAIL = "info@zhbk.ru"
COMPANY_INN = "4705123456"
COMPANY_KPP = "470501001"

# Банковские реквизиты
BANK_NAME = "ПАО Сбербанк"
BANK_BIK = "044030653"
BANK_ACCOUNT = "40702810123456789012"
BANK_CORR_ACCOUNT = "30101810500000000653"

# Путь к базе данных с ценами
DB_PATH = "pb.db"

# Путь к логотипу
LOGO_PATH = "банк знаний/ЖБЛСТАРТ.png"


# ==================== РЕГИСТРАЦИЯ ШРИФТОВ ====================

def register_fonts():
    """
    Регистрирует русские шрифты для ReportLab и старается выбрать Tahoma,
    чтобы повторить фирменное оформление КП.
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    font_paths = []
    local_font_dirs = [
        os.path.join(script_dir, 'fonts'),
        os.path.join(script_dir, 'банк знаний'),
        os.path.join(script_dir, 'bank'),
    ]
    
    for local_dir in local_font_dirs:
        if os.path.exists(local_dir):
            font_paths.append(local_dir)
    
    linux_font_paths = [
        '/usr/share/fonts/truetype/dejavu/',
        '/usr/share/fonts/truetype/liberation/',
        '/usr/share/fonts/truetype/msttcorefonts/',
        '/usr/share/fonts/truetype/freefont/',
        '/usr/share/fonts/TTF/',
        '/usr/share/fonts/',
        '~/.fonts/',
    ]
    
    windows_fonts = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts')
    if os.path.exists(windows_fonts):
        font_paths.append(windows_fonts)
    
    for path in linux_font_paths:
        expanded = os.path.expanduser(path)
        if os.path.exists(expanded):
            font_paths.append(expanded)
    
    # Убираем дубликаты, сохраняя порядок
    seen_paths = set()
    unique_font_paths = []
    for path in font_paths:
        if path not in seen_paths:
            unique_font_paths.append(path)
            seen_paths.add(path)
    
    fonts_to_register = [
        ('Tahoma', ('Tahoma.ttf', 'tahoma.ttf')),
        ('Tahoma-Bold', ('tahomabd.ttf', 'Tahoma-Bold.ttf', 'Tahoma Bold.ttf', 'tahoma-bold.ttf')),
        ('DejaVuSans', ('DejaVuSans.ttf',)),
        ('DejaVuSans-Bold', ('DejaVuSans-Bold.ttf',)),
        ('LiberationSans', ('LiberationSans-Regular.ttf', 'LiberationSans.ttf')),
        ('LiberationSans-Bold', ('LiberationSans-Bold.ttf',)),
        ('Arial', ('Arial.ttf', 'arial.ttf')),
        ('Arial-Bold', ('Arial-Bold.ttf', 'arialbd.ttf')),
        ('TimesNewRoman', ('TimesNewRoman.ttf', 'times.ttf')),
        ('TimesNewRoman-Bold', ('TimesNewRoman-Bold.ttf', 'timesbd.ttf')),
    ]
    
    registered_fonts = set()
    
    for font_name, candidates in fonts_to_register:
        for font_dir in unique_font_paths:
            for candidate in candidates:
                font_path = os.path.join(font_dir, candidate)
                if os.path.exists(font_path):
                    try:
                        pdfmetrics.registerFont(TTFont(font_name, font_path))
                        registered_fonts.add(font_name)
                        break
                    except Exception:
                        continue
            if font_name in registered_fonts:
                break
    
    preferred_pairs = [
        ('Tahoma', 'Tahoma-Bold'),
        ('DejaVuSans', 'DejaVuSans-Bold'),
        ('LiberationSans', 'LiberationSans-Bold'),
        ('Arial', 'Arial-Bold'),
        ('TimesNewRoman', 'TimesNewRoman-Bold'),
    ]
    
    for normal, bold in preferred_pairs:
        if normal in registered_fonts:
            globals()['FONT_NORMAL'] = normal
            globals()['FONT_BOLD'] = bold if bold in registered_fonts else normal
            break
    else:
        if registered_fonts:
            fallback = next(iter(registered_fonts))
            globals()['FONT_NORMAL'] = fallback
            globals()['FONT_BOLD'] = fallback
        else:
            globals()['FONT_NORMAL'] = 'Helvetica'
            globals()['FONT_BOLD'] = 'Helvetica-Bold'
            print("ВНИМАНИЕ: Не найдены шрифты с поддержкой кириллицы! Русский текст может отображаться некорректно.")
    
    return bool(registered_fonts)


# Регистрируем шрифты при импорте модуля
HAS_CYRILLIC_FONTS = register_fonts()
FONT_NORMAL = globals().get('FONT_NORMAL', 'Helvetica')
FONT_BOLD = globals().get('FONT_BOLD', 'Helvetica-Bold')


# ==================== ФУНКЦИИ ====================

def get_plate_price(length_m: float, width_m: float, load_class: int = 800) -> float:
    """
    Получает цену плиты из базы данных по длине, ширине и классу нагрузки
    
    Args:
        length_m: длина плиты в метрах
        width_m: ширина плиты в метрах  
        load_class: класс нагрузки (по умолчанию 800 кг/м²)
    
    Returns:
        Цена плиты в рублях
    """
    try:
        # Преобразуем в дециметры для поиска в базе
        length_dm = int(round(length_m * 10))
        
        # Определяем код нагрузки (8 = 800 кг/м², 10 = 1000 кг/м²)
        load_code = load_class // 100
        
        con = sqlite3.connect(DB_PATH)
        cur = con.cursor()
        
        # Ищем цену в таблице prices
        result = cur.execute(
            "SELECT price FROM prices WHERE length_dm = ? AND load_code = ?",
            (length_dm, load_code)
        ).fetchone()
        
        con.close()
        
        if result:
            return float(result[0])
        else:
            # Если нет точной цены, используем базовую формулу
            # Примерная цена: 4000 руб/м² * площадь плиты
            area_m2 = length_m * width_m
            return round(area_m2 * 4000, 2)
            
    except Exception as e:
        print(f"Ошибка получения цены: {e}")
        # Возвращаем примерную цену
        area_m2 = length_m * width_m
        return round(area_m2 * 4000, 2)


def calculate_total_cost(order_data: List[Dict]) -> Dict:
    """
    Рассчитывает общую стоимость заказа
    
    Args:
        order_data: список позиций заказа с полями name, length_m, width_m, qty, unit_price (опционально)
    
    Returns:
        Словарь с итоговыми суммами
    """
    total_qty = 0
    total_cost = 0.0
    
    for item in order_data:
        qty = item.get('qty', 0)
        
        # 🔥 ПРИОРИТЕТ: Если цена уже рассчитана (с учётом резов/отходов), используем её!
        if 'unit_price' in item and item['unit_price'] is not None:
            unit_price = item['unit_price']
        else:
            # Fallback: старая логика (только базовая цена из БД)
            length_m = item.get('length_m', 0)
            width_m = item.get('width_m', 0)
            load_class = item.get('load_class', 800)
            unit_price = get_plate_price(length_m, width_m, load_class)
        
        # Считаем сумму по позиции
        item_cost = unit_price * qty
        
        total_qty += qty
        total_cost += item_cost
    
    # НДС 20%
    vat_amount = round(total_cost * 0.20, 2)
    total_with_vat = round(total_cost + vat_amount, 2)
    
    return {
        'total_qty': total_qty,
        'subtotal': round(total_cost, 2),
        'vat_amount': vat_amount,
        'total_with_vat': total_with_vat
    }


def generate_commercial_offer_pdf(
    order_data: List[Dict],
    offer_number: str,
    offer_date: str,
    customer_name: Optional[str] = None
) -> io.BytesIO:
    """
    Генерирует коммерческое предложение в формате PDF, повторяя фирменное
    оформление КП № 1133 от 16.10.2025.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=15 * mm
    )
    
    content_width = doc.width
    styles = getSampleStyleSheet()
    
    style_header_info = ParagraphStyle(
        'HeaderInfo',
        parent=styles['Normal'],
        fontName=FONT_BOLD,
        fontSize=11,
        leading=14,
        spaceAfter=2 * mm
    )
    
    style_doc_number = ParagraphStyle(
        'DocNumber',
        parent=styles['Normal'],
        fontName=FONT_BOLD,
        fontSize=11,
        leading=14,
        spaceAfter=4 * mm
    )
    
    style_title = ParagraphStyle(
        'OfferTitle',
        parent=styles['Normal'],
        fontName=FONT_BOLD,
        fontSize=18,
        leading=22,
        alignment=1,
        spaceAfter=1 * mm
    )
    
    style_customer = ParagraphStyle(
        'OfferCustomer',
        parent=styles['Normal'],
        fontName=FONT_BOLD,
        fontSize=16,
        leading=20,
        alignment=1,
        spaceAfter=6 * mm
    )
    
    style_table_text = ParagraphStyle(
        'TableText',
        parent=styles['Normal'],
        fontName=FONT_NORMAL,
        fontSize=10,
        leading=12
    )
    
    style_summary = ParagraphStyle(
        'Summary',
        parent=styles['Normal'],
        fontName=FONT_BOLD,
        fontSize=11,
        leading=15,
        spaceAfter=6 * mm
    )
    
    style_conditions = ParagraphStyle(
        'Conditions',
        parent=styles['Normal'],
        fontName=FONT_NORMAL,
        fontSize=11,
        leading=14,
        spaceAfter=2 * mm
    )
    
    style_signature_label = ParagraphStyle(
        'SignatureLabel',
        parent=styles['Normal'],
        fontName=FONT_BOLD,
        fontSize=12,
        leading=16
    )
    
    style_signature = ParagraphStyle(
        'Signature',
        parent=styles['Normal'],
        fontName=FONT_NORMAL,
        fontSize=12,
        leading=16
    )
    
    style_note = ParagraphStyle(
        'Note',
        parent=styles['Normal'],
        fontName=FONT_NORMAL,
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#333333'),
        spaceBefore=2 * mm
    )
    
    story = []
    
    if os.path.exists(LOGO_PATH):
        try:
            logo_width = content_width
            logo_height = logo_width * (207 / 1456.0)
            story.append(Image(LOGO_PATH, width=logo_width, height=logo_height))
            story.append(Spacer(1, 4 * mm))
        except Exception as exc:
            print(f"Ошибка загрузки логотипа: {exc}")
    
    story.append(Paragraph(
        f"{COMPANY_ADDRESS}<br/>Тел.: {COMPANY_PHONE}",
        style_header_info
    ))
    
    story.append(Paragraph(
        f"№ {offer_number} от {offer_date}",
        style_doc_number
    ))
    
    story.append(Paragraph(
        "Коммерческое предложение для",
        style_title
    ))
    
    customer_display = escape(customer_name.strip()) if customer_name else "Заказчик не указан"
    story.append(Paragraph(customer_display, style_customer))
    
    story.append(Spacer(1, 6 * mm))
    
    table_data = [['№', 'Наименование', 'Кол-во', 'Ед.', 'Вес(кг)', 'Цена', 'Сумма']]
    
    totals = calculate_total_cost(order_data)
    total_weight = 0.0
    
    for idx, item in enumerate(order_data, start=1):
        name_raw = item.get('name', 'Плиты ПБ')
        name = escape(name_raw)
        qty = item.get('qty', 0)
        length_m = item.get('length_m', 0)
        width_m = item.get('width_m', 0)
        load_class = item.get('load_class')

        # 🔥 ПРИОРИТЕТ: Если цена уже рассчитана (с учётом резов/отходов), используем её!
        if 'unit_price' in item and item['unit_price'] is not None:
            unit_price = item['unit_price']
        else:
            # Fallback: старая логика (только базовая цена из БД)
            # Если класс нагрузки явно не передан, пробуем вытащить его из имени плиты.
            # Формат имени: "Плиты ПБ 71-12-10п", "ПБ 69-12-12,5п" и т.п.
            if load_class is None:
                try:
                    from config_and_data import parse_load_code_from_name
                except ImportError:
                    load_class = 800
                else:
                    load_code = parse_load_code_from_name(name_raw, default=8)
                    load_class = max(1, load_code) * 100  # 8 -> 800, 10 -> 1000 и т.п.
            
            unit_price = get_plate_price(length_m, width_m, load_class)
        
        item_sum = unit_price * qty
        
        # Вес: если уже передан в item, используем его, иначе рассчитываем
        if 'weight' in item and item['weight'] is not None:
            total_item_weight = item['weight']
        else:
            unit_weight = approximate_weight_kg(length_m, width_m)
            total_item_weight = unit_weight * qty
        
        total_weight += total_item_weight
        
        weight_str = f"{total_item_weight:,.2f}".replace(',', 'X').replace('.', ',').replace('X', ' ')
        price_str = f"{unit_price:,.2f}".replace(',', 'X').replace('.', ',').replace('X', ' ')
        sum_str = f"{item_sum:,.2f}".replace(',', 'X').replace('.', ',').replace('X', ' ')
        
        table_data.append([
            str(idx),
            Paragraph(name, style_table_text),
            str(qty),
            'шт',
            weight_str,
            price_str,
            sum_str
        ])
    
    no_width = 10 * mm
    qty_width = 14 * mm
    unit_width = 11 * mm
    weight_width = 20 * mm
    price_width = 25 * mm
    sum_width = 25 * mm
    
    fixed_total = no_width + qty_width + unit_width + weight_width + price_width + sum_width
    name_width = content_width - fixed_total
    
    if name_width <= 0:
        ratios = (0.05, 0.45, 0.08, 0.06, 0.18, 0.09, 0.09)
        col_widths = [content_width * ratio for ratio in ratios]
    else:
        col_widths = [
            no_width,
            name_width,
            qty_width,
            unit_width,
            weight_width,
            price_width,
            sum_width
        ]
    
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        
        ('FONTNAME', (0, 1), (-1, -1), FONT_NORMAL),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (2, 1), (3, -1), 'CENTER'),
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        
        ('BOX', (0, 0), (-1, -1), 1.2, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.6, colors.black),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 6 * mm))
    
    total_items = len(order_data)
    weight_summary = f"{total_weight:,.3f}".replace(',', 'X').replace('.', ',').replace('X', ' ')
    subtotal_str = f"{totals['subtotal']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', ' ')
    vat_str = f"{totals['vat_amount']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', ' ')
    
    summary_text = (
        f"Всего наименований {total_items}, общим весом {weight_summary} кг., "
        f"на сумму {subtotal_str} руб., в том числе НДС {vat_str} руб."
    )
    story.append(Paragraph(summary_text, style_summary))
    
    story.append(Paragraph("1.Условия поставки:", style_conditions))
    story.append(Paragraph("2.Условия оплаты: Предварительная оплата в размере 100%", style_conditions))
    
    story.append(Spacer(1, 6 * mm))
    
    story.append(Paragraph("С уважением,", style_signature_label))
    story.append(Paragraph("Шишов Александр Васильевич", style_signature))
    story.append(Paragraph("8 920 640 55 85", style_signature))
    
    story.append(Paragraph(
        'Доборные плиты ПБ отгружаются только при наличии в наименовании "+доб", '
        "для получения доборов, просим сообщить Вашему менеджеру до начала изготовления.<br/>"
        "Все доборы по умолчанию отправляем на утилизацию.",
        style_note
    ))
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_commercial_offer_xlsx(
    order_data: List[Dict],
    offer_number: str,
    offer_date: str,
    customer_name: Optional[str] = None
) -> io.BytesIO:
    """
    Генерирует коммерческое предложение в формате XLSX с расчётными формулами
    
    Args:
        order_data: список позиций заказа с полями name, length_m, width_m, qty
        offer_number: номер КП
        offer_date: дата КП
        customer_name: имя заказчика
    
    Returns:
        BytesIO буфер с XLSX файлом
    """
    if not HAS_EXCEL_LIBS:
        raise ImportError("Для генерации XLSX необходимы pandas и openpyxl")
    
    buffer = io.BytesIO()
    
    # Определяем, есть ли логотип (будет вставлен выше заголовка)
    has_logo = os.path.exists(LOGO_PATH)
    
    # Формируем заголовок документа
    # Если есть логотип, добавляем пустые строки для него
    header_data = []
    
    if has_logo:
        # Резервируем место под логотип (7 строк как в образце)
        header_data.extend([[""], [""], [""], [""], [""], [""], [""]])
    
    header_data.extend([
        [COMPANY_NAME],
        [f"{COMPANY_ADDRESS}"],
        [f"Тел.: {COMPANY_PHONE}, Email: {COMPANY_EMAIL}"],
        [""],
        [f"№ {offer_number} от {offer_date}"],
        [""],
        ["Коммерческое предложение для"],
        [customer_name or "Заказчик не указан"],
        [""],
    ])
    
    # Создаем DataFrame для заголовка
    df_header = pd.DataFrame(header_data)
    
    # Формируем таблицу товаров
    table_data = []
    table_headers = ['№', 'Наименование', 'Кол-во', 'Ед.', 'Вес(кг)', 'Цена', 'Сумма']
    
    total_weight = 0.0
    
    for idx, item in enumerate(order_data, start=1):
        name = item.get('name', 'Плиты ПБ')
        qty = item.get('qty', 0)
        length_m = item.get('length_m', 0)
        width_m = item.get('width_m', 0)
        load_class = item.get('load_class')
        
        # 🔥 ПРИОРИТЕТ: Если цена уже рассчитана (с учётом резов/отходов), используем её!
        if 'unit_price' in item and item['unit_price'] is not None:
            unit_price = item['unit_price']
        else:
            # Fallback: старая логика (только базовая цена из БД)
            # Определяем класс нагрузки из имени, если не передан
            if load_class is None:
                try:
                    from config_and_data import parse_load_code_from_name
                    load_code = parse_load_code_from_name(name, default=8)
                    load_class = max(1, load_code) * 100
                except ImportError:
                    load_class = 800
            
            unit_price = get_plate_price(length_m, width_m, load_class)
        
        # Вес: если уже передан в item, используем его, иначе рассчитываем
        if 'weight' in item and item['weight'] is not None:
            unit_weight = item['weight'] / qty  # Общий вес делим на количество
        else:
            unit_weight = approximate_weight_kg(length_m, width_m)
        
        total_item_weight = unit_weight * qty
        total_weight += total_item_weight
        
        table_data.append({
            '№': idx,
            'Наименование': name,
            'Кол-во': qty,
            'Ед.': 'шт',
            'Вес(кг)': total_item_weight,
            'Цена': unit_price,
            'Сумма': unit_price * qty  # Сначала вставим значение, потом заменим на формулу
        })
    
    df_table = pd.DataFrame(table_data)
    
    # Рассчитываем итоги
    totals = calculate_total_cost(order_data)
    
    # Создаем строку итогов
    total_items = len(order_data)
    
    # Записываем в Excel
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Записываем заголовок (без заголовков столбцов)
        df_header.to_excel(writer, sheet_name='КП', index=False, header=False, startrow=0)
        
        # Записываем таблицу товаров (с заголовками)
        start_row = len(header_data)
        df_table.to_excel(writer, sheet_name='КП', index=False, startrow=start_row)
        
        # Получаем workbook и worksheet для форматирования
        workbook = writer.book
        worksheet = writer.sheets['КП']
        
        # === ВСТАВКА ЛОГОТИПА ===
        if has_logo and XLImage is not None:
            try:
                # Создаем объект изображения
                logo_img = XLImage(LOGO_PATH)
                
                # Масштабируем логотип (оригинал 1456x207)
                # Делаем ширину примерно на всю ширину таблицы (7 колонок)
                logo_img.width = 600  # пикселей
                logo_img.height = int(logo_img.width * (207 / 1456.0))  # сохраняем пропорции
                
                # Объединяем ячейки для логотипа (7 строк x 7 колонок, как в образце)
                worksheet.merge_cells('A1:G7')
                
                # Вставляем в ячейку A1
                worksheet.add_image(logo_img, 'A1')
                
                # Увеличиваем высоту первых 7 строк для логотипа
                total_logo_height = logo_img.height
                for row_idx in range(1, 8):
                    worksheet.row_dimensions[row_idx].height = total_logo_height / 7
                    
            except Exception as e:
                print(f"[XLSX] Не удалось вставить логотип: {e}")
        
        # === ФОРМАТИРОВАНИЕ ===
        
        # Шрифты и стили
        header_font = Font(name='Tahoma', size=11, bold=True)
        title_font = Font(name='Tahoma', size=18, bold=True)
        customer_font = Font(name='Tahoma', size=16, bold=True)
        table_header_font = Font(name='Tahoma', size=11, bold=True)
        table_font = Font(name='Tahoma', size=10)
        summary_font = Font(name='Tahoma', size=11, bold=True)
        
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Границы для таблицы
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заливка для заголовка таблицы
        header_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        
        # Определяем смещение строк (если есть логотип, добавляем 7 строк)
        logo_offset = 7 if has_logo else 0
        
        # Форматируем заголовок документа (с учетом смещения для логотипа)
        company_row = logo_offset + 1
        address_row = logo_offset + 2
        phone_row = logo_offset + 3
        number_row = logo_offset + 5
        title_row = logo_offset + 7
        customer_row = logo_offset + 8
        
        worksheet[f'A{company_row}'].font = header_font
        worksheet[f'A{address_row}'].font = header_font
        worksheet[f'A{phone_row}'].font = header_font
        worksheet[f'A{number_row}'].font = header_font
        worksheet[f'A{title_row}'].font = title_font
        worksheet[f'A{title_row}'].alignment = center_align
        worksheet[f'A{customer_row}'].font = customer_font
        worksheet[f'A{customer_row}'].alignment = center_align
        
        # Объединяем ячейки для заголовка
        worksheet.merge_cells(f'A{company_row}:G{company_row}')
        worksheet.merge_cells(f'A{address_row}:G{address_row}')
        worksheet.merge_cells(f'A{phone_row}:G{phone_row}')
        worksheet.merge_cells(f'A{number_row}:G{number_row}')
        worksheet.merge_cells(f'A{title_row}:G{title_row}')
        worksheet.merge_cells(f'A{customer_row}:G{customer_row}')
        
        # Форматируем заголовок таблицы
        table_header_row = start_row + 1
        for col_idx in range(1, 8):  # 7 столбцов
            cell = worksheet.cell(row=table_header_row, column=col_idx)
            cell.font = table_header_font
            cell.alignment = center_align
            cell.fill = header_fill
            cell.border = thin_border
        
        # Форматируем строки таблицы
        for row_idx in range(table_header_row + 1, table_header_row + 1 + len(order_data)):
            # №
            worksheet.cell(row=row_idx, column=1).alignment = center_align
            worksheet.cell(row=row_idx, column=1).border = thin_border
            worksheet.cell(row=row_idx, column=1).font = table_font
            
            # Наименование
            worksheet.cell(row=row_idx, column=2).alignment = left_align
            worksheet.cell(row=row_idx, column=2).border = thin_border
            worksheet.cell(row=row_idx, column=2).font = table_font
            
            # Кол-во
            worksheet.cell(row=row_idx, column=3).alignment = center_align
            worksheet.cell(row=row_idx, column=3).border = thin_border
            worksheet.cell(row=row_idx, column=3).font = table_font
            
            # Ед.
            worksheet.cell(row=row_idx, column=4).alignment = center_align
            worksheet.cell(row=row_idx, column=4).border = thin_border
            worksheet.cell(row=row_idx, column=4).font = table_font
            
            # Вес
            worksheet.cell(row=row_idx, column=5).alignment = right_align
            worksheet.cell(row=row_idx, column=5).border = thin_border
            worksheet.cell(row=row_idx, column=5).font = table_font
            worksheet.cell(row=row_idx, column=5).number_format = '#,##0.00'
            
            # Цена
            worksheet.cell(row=row_idx, column=6).alignment = right_align
            worksheet.cell(row=row_idx, column=6).border = thin_border
            worksheet.cell(row=row_idx, column=6).font = table_font
            worksheet.cell(row=row_idx, column=6).number_format = '#,##0.00'
            
            # Сумма - добавляем ФОРМУЛУ!
            sum_cell = worksheet.cell(row=row_idx, column=7)
            sum_cell.value = f"=C{row_idx}*F{row_idx}"  # Кол-во * Цена
            sum_cell.alignment = right_align
            sum_cell.border = thin_border
            sum_cell.font = table_font
            sum_cell.number_format = '#,##0.00'
        
        # Добавляем итоговые строки
        summary_row = table_header_row + len(order_data) + 2
        
        # Итоговая сумма без НДС
        subtotal_row = summary_row
        worksheet.merge_cells(f'A{subtotal_row}:F{subtotal_row}')
        worksheet[f'A{subtotal_row}'] = f"Всего наименований {total_items}, общим весом {total_weight:,.3f} кг."
        worksheet[f'A{subtotal_row}'].font = summary_font
        worksheet[f'A{subtotal_row}'].alignment = left_align
        
        # Формула для подсчёта суммы
        first_data_row = table_header_row + 1
        last_data_row = table_header_row + len(order_data)
        worksheet[f'G{subtotal_row}'] = f"=SUM(G{first_data_row}:G{last_data_row})"
        worksheet[f'G{subtotal_row}'].font = summary_font
        worksheet[f'G{subtotal_row}'].number_format = '#,##0.00'
        worksheet[f'G{subtotal_row}'].alignment = right_align
        
        # НДС 20%
        vat_row = summary_row + 1
        worksheet.merge_cells(f'A{vat_row}:F{vat_row}')
        worksheet[f'A{vat_row}'] = "в том числе НДС (20%)"
        worksheet[f'A{vat_row}'].font = summary_font
        worksheet[f'A{vat_row}'].alignment = left_align
        worksheet[f'G{vat_row}'] = f"=G{subtotal_row}*0.2"  # Формула для НДС
        worksheet[f'G{vat_row}'].font = summary_font
        worksheet[f'G{vat_row}'].number_format = '#,##0.00'
        worksheet[f'G{vat_row}'].alignment = right_align
        
        # Условия
        conditions_row = vat_row + 2
        worksheet[f'A{conditions_row}'] = "1. Условия поставки:"
        worksheet[f'A{conditions_row}'].font = table_font
        
        conditions_row += 1
        worksheet[f'A{conditions_row}'] = "2. Условия оплаты: Предварительная оплата в размере 100%"
        worksheet[f'A{conditions_row}'].font = table_font
        
        # Подпись
        signature_row = conditions_row + 2
        worksheet[f'A{signature_row}'] = "С уважением,"
        worksheet[f'A{signature_row}'].font = summary_font
        
        signature_row += 1
        worksheet[f'A{signature_row}'] = "Шишов Александр Васильевич"
        worksheet[f'A{signature_row}'].font = table_font
        
        signature_row += 1
        worksheet[f'A{signature_row}'] = "8 920 640 55 85"
        worksheet[f'A{signature_row}'].font = table_font
        
        # Примечание
        note_row = signature_row + 2
        worksheet[f'A{note_row}'] = ('Доборные плиты ПБ отгружаются только при наличии в наименовании "+доб", '
                                      'для получения доборов, просим сообщить Вашему менеджеру до начала изготовления. '
                                      'Все доборы по умолчанию отправляем на утилизацию.')
        worksheet[f'A{note_row}'].font = Font(name='Tahoma', size=9)
        worksheet[f'A{note_row}'].alignment = left_align
        worksheet.merge_cells(f'A{note_row}:G{note_row}')
        
        # Настройка ширины столбцов
        worksheet.column_dimensions['A'].width = 5
        worksheet.column_dimensions['B'].width = 45
        worksheet.column_dimensions['C'].width = 10
        worksheet.column_dimensions['D'].width = 8
        worksheet.column_dimensions['E'].width = 15
        worksheet.column_dimensions['F'].width = 18
        worksheet.column_dimensions['G'].width = 18
        
        # Высота строк
        worksheet.row_dimensions[table_header_row].height = 25
    
    buffer.seek(0)
    return buffer


# ==================== ТЕСТИРОВАНИЕ ====================

if __name__ == "__main__":
    # Тестовый заказ
    test_order = [
        {"name": "ПБ 56-6-8п", "length_m": 5.6, "width_m": 0.6, "qty": 1, "load_class": 800},
        {"name": "ПБ 78-0.3-8п", "length_m": 7.8, "width_m": 0.3, "qty": 3, "load_class": 800},
        {"name": "ПБ 68-6-8п", "length_m": 6.8, "width_m": 0.6, "qty": 2, "load_class": 800},
        {"name": "ПБ 56-9-8п", "length_m": 5.6, "width_m": 0.9, "qty": 1, "load_class": 800},
        {"name": "ПБ 78-9-8п", "length_m": 7.8, "width_m": 0.9, "qty": 3, "load_class": 800},
        {"name": "ПБ 52-11-8п", "length_m": 5.2, "width_m": 1.1, "qty": 1, "load_class": 800},
    ]
    
    # Генерируем PDF
    pdf_buffer = generate_commercial_offer_pdf(
        order_data=test_order,
        offer_number="1133",
        offer_date=datetime.now().strftime("%d.%m.%Y"),
        customer_name="ООО «Тестовая компания»"
    )
    
    # Сохраняем в файл
    output_path = "test_commercial_offer.pdf"
    with open(output_path, 'wb') as f:
        f.write(pdf_buffer.getvalue())
    
    print(f"[OK] Test KP created: {output_path}")


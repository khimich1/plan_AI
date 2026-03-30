#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Модуль для создания Excel-файла с диаграммой Ганта производства.

Показывает план производства по КП:
- Строки: номера КП с информацией о заказчике и дедлайне
- Столбцы: дни недели (верхняя строка), затем даты производства; после последнего
  дня с планом добавляются пустые столбцы для будущего горизонта
- Ячейки: закрашены в дни, когда производятся плиты из КП

Цветовая кодировка:
- Зелёный: КП завершается до дедлайна
- Жёлтый: КП завершается в день дедлайна  
- Красный: КП завершается после дедлайна (опаздываем!)
"""
from __future__ import annotations

import os
from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from core.work_calendar import (
    is_working_day,
    load_extra_workdays,
    load_holidays,
    nth_working_day,
)

_WEEKDAYS_RU: tuple[str, ...] = ("Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс")


def _weekday_ru(dt: datetime) -> str:
    """Короткое русское обозначение дня недели (понедельник = 0)."""
    return _WEEKDAYS_RU[dt.weekday()]


def create_gantt_excel(
    all_tracks_list: list,
    tracks_count: int,
    plate_lookup_exact: dict,
    plate_lookup_by_length: dict,
    output_dir: str,
    start_date: datetime | None = None,
    extra_calendar_days: int = 14,
) -> str:
    """
    Создаёт Excel-файл с диаграммой Ганта для планирования производства.

    Сетка: строка 1 — дни недели над датами; строка 2 — заголовки и даты;
    строка 3 — загруженность дорожек; далее строки КП и итоги.

    Args:
        all_tracks_list: список всех дорожек с плитами
        tracks_count: количество дорожек в день
        plate_lookup_exact: словарь для поиска информации о плитах по (length, width)
        plate_lookup_by_length: словарь для поиска информации о плитах по длине
        output_dir: директория для сохранения файла
        start_date: дата начала производства (по умолчанию - сегодня)
        extra_calendar_days: сколько дней добавить после последнего дня с планом
            (пустые столбцы для будущего планирования)

    Returns:
        Путь к созданному Excel-файлу
    """
    if start_date is None:
        start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    holidays = load_holidays()
    extra_workdays = load_extra_workdays()

    def _working_dt(day_number: int) -> datetime:
        work_date = nth_working_day(start_date.date(), day_number, holidays, extra_workdays)
        return datetime.combine(work_date, datetime.min.time())
    
    # === ШАГ 1: Собираем информацию о КП из дорожек ===
    kp_production_info = {}  # {kp_id: {'start_day', 'end_day', 'customer', 'deadline', 'plate_count'}}
    
    # Создаём копию lookup для подсчёта (чтобы не модифицировать оригинал)
    import copy
    lookup_copy = copy.deepcopy(plate_lookup_exact)
    lookup_by_length_copy = copy.deepcopy(plate_lookup_by_length)
    
    def get_plate_info(length, width):
        """Получить информацию о плите из lookup"""
        key = (round(length, 2), width)
        entries = lookup_copy.get(key, [])
        
        for entry in entries:
            if entry.get('qty_remaining', 0) > 0:
                entry['qty_remaining'] -= 1
                return entry.copy()
        
        # Fallback по длине
        length_key = round(length, 2)
        entries = lookup_by_length_copy.get(length_key, [])
        for entry in entries:
            if entry.get('qty_remaining', 0) > 0:
                entry['qty_remaining'] -= 1
                return entry.copy()
        
        return None
    
    # Проходим по всем дорожкам
    for track_idx, track in enumerate(all_tracks_list):
        # НОВАЯ ЛОГИКА: Берём день из самой дорожки (если есть)
        # Если нет — используем старую формулу для совместимости со старыми планами
        if isinstance(track, dict) and 'production_day' in track:
            production_day = track['production_day']
        else:
            # Старая формула (для обратной совместимости)
            production_day = (track_idx // tracks_count) + 1
        
        for item in track.get('items', []):
            if item is None:
                continue
            
            length = item.get('length', 0)
            if not length:
                continue
            
            # НОВАЯ ЛОГИКА: Приоритет — kp_id из самого элемента
            kp_id = item.get('kp_id')
            customer = item.get('customer')
            kp_date = item.get('kp_date')
            
            # Если kp_id есть в элементе, используем его напрямую (точная информация!)
            if kp_id:
                if kp_id not in kp_production_info:
                    kp_production_info[kp_id] = {
                        'start_day': production_day,
                        'end_day': production_day,
                        'customer': customer or 'неизвестно',
                        'deadline': kp_date or 'неизвестно',
                        'plate_count': 1,
                        'plates_by_day': defaultdict(int)
                    }
                else:
                    # Обновляем день начала (минимальный) и конца (максимальный)
                    kp_production_info[kp_id]['start_day'] = min(
                        kp_production_info[kp_id]['start_day'], 
                        production_day
                    )
                    kp_production_info[kp_id]['end_day'] = max(
                        kp_production_info[kp_id]['end_day'], 
                        production_day
                    )
                    kp_production_info[kp_id]['plate_count'] += 1
                
                # Считаем плиты по дням
                kp_production_info[kp_id]['plates_by_day'][production_day] += 1
            else:
                # Fallback: пытаемся найти по размерам (старая логика)
                # Определяем ширину
                mode = item.get('mode', 'solid')
                if mode == 'transverse' and item.get('width'):
                    width = round(item['width'] * 1000)  # round для корректного округления float
                elif mode == 'split' and item.get('main_w'):
                    width = round(item['main_w'] * 1000)  # round для корректного округления
                else:
                    width = 1200
                
                # Получаем информацию о плите через lookup
                plate_info = get_plate_info(length, width)
                
                if plate_info:
                    kp_id = plate_info.get('kp_id')
                    if kp_id:
                        if kp_id not in kp_production_info:
                            kp_production_info[kp_id] = {
                                'start_day': production_day,
                                'end_day': production_day,
                                'customer': plate_info.get('customer', 'неизвестно'),
                                'deadline': plate_info.get('kp_date', 'неизвестно'),
                                'plate_count': 1,
                                'plates_by_day': defaultdict(int)
                            }
                        else:
                            # Обновляем день начала (минимальный) и конца (максимальный)
                            kp_production_info[kp_id]['start_day'] = min(
                                kp_production_info[kp_id]['start_day'], 
                                production_day
                            )
                            kp_production_info[kp_id]['end_day'] = max(
                                kp_production_info[kp_id]['end_day'], 
                                production_day
                            )
                            kp_production_info[kp_id]['plate_count'] += 1
                        
                        # Считаем плиты по дням
                        kp_production_info[kp_id]['plates_by_day'][production_day] += 1
            
            # Обрабатываем вторичные резы
            secondary_cuts = item.get('secondary_cuts', []) or []
            for sec_cut in secondary_cuts:
                sec_width_m = sec_cut.get('width', 0)
                if sec_width_m <= 0:
                    continue
                
                sec_width = round(sec_width_m * 1000)  # round для корректного округления float
                sec_length = sec_cut.get('target_length') or length
                
                # НОВАЯ ЛОГИКА: Сначала проверяем kp_id в элементе
                # (для вторичных резов — это kp_id родительской плиты)
                sec_kp_id = item.get('kp_id')  # Берем из основного элемента!
                sec_customer = item.get('customer')
                sec_kp_date = item.get('kp_date')
                
                if sec_kp_id:
                    # Используем точную информацию из элемента
                    if sec_kp_id not in kp_production_info:
                        kp_production_info[sec_kp_id] = {
                            'start_day': production_day,
                            'end_day': production_day,
                            'customer': sec_customer or 'неизвестно',
                            'deadline': sec_kp_date or 'неизвестно',
                            'plate_count': 1,
                            'plates_by_day': defaultdict(int)
                        }
                    else:
                        kp_production_info[sec_kp_id]['start_day'] = min(
                            kp_production_info[sec_kp_id]['start_day'], 
                            production_day
                        )
                        kp_production_info[sec_kp_id]['end_day'] = max(
                            kp_production_info[sec_kp_id]['end_day'], 
                            production_day
                        )
                        kp_production_info[sec_kp_id]['plate_count'] += 1
                    
                    kp_production_info[sec_kp_id]['plates_by_day'][production_day] += 1
                else:
                    # Fallback: пытаемся найти по размерам
                    sec_plate_info = get_plate_info(sec_length, sec_width)
                    
                    if sec_plate_info:
                        sec_kp_id = sec_plate_info.get('kp_id')
                        if sec_kp_id:
                            if sec_kp_id not in kp_production_info:
                                kp_production_info[sec_kp_id] = {
                                    'start_day': production_day,
                                    'end_day': production_day,
                                    'customer': sec_plate_info.get('customer', 'неизвестно'),
                                    'deadline': sec_plate_info.get('kp_date', 'неизвестно'),
                                    'plate_count': 1,
                                    'plates_by_day': defaultdict(int)
                                }
                            else:
                                kp_production_info[sec_kp_id]['start_day'] = min(
                                    kp_production_info[sec_kp_id]['start_day'], 
                                    production_day
                                )
                                kp_production_info[sec_kp_id]['end_day'] = max(
                                    kp_production_info[sec_kp_id]['end_day'], 
                                    production_day
                                )
                                kp_production_info[sec_kp_id]['plate_count'] += 1
                            
                            kp_production_info[sec_kp_id]['plates_by_day'][production_day] += 1
    
    # === РЕЗЕРВНЫЙ ВАРИАНТ: Собираем плиты без КП ===
    # Если не нашли КП для некоторых плит, собираем их в отдельную группу
    unknown_kp_count = 0
    for track_idx, track in enumerate(all_tracks_list):
        # Определяем день производства
        if isinstance(track, dict) and 'production_day' in track:
            production_day = track['production_day']
        else:
            production_day = (track_idx // tracks_count) + 1
        
        # Подсчитываем плиты в дорожке
        plates_in_track = len([item for item in track.get('items', []) if item is not None])
        unknown_kp_count += plates_in_track
    
    # Вычитаем найденные плиты
    found_plates = sum(info['plate_count'] for info in kp_production_info.values())
    unknown_kp_count -= found_plates
    
    # Если есть плиты без КП, добавляем их как отдельную группу
    if unknown_kp_count > 0:
        print(f"[GANTT] ⚠️ Найдено {unknown_kp_count} плит без информации о КП")
        # Добавляем резервную группу
        kp_production_info['НЕИЗВЕСТНО'] = {
            'start_day': 1,
            'end_day': max((info['end_day'] for info in kp_production_info.values()), default=1),
            'customer': 'Информация отсутствует',
            'deadline': 'неизвестно',
            'plate_count': unknown_kp_count,
            'plates_by_day': defaultdict(int)
        }
    
    if not kp_production_info:
        print("[GANTT] ⚠️ Нет данных о плитах для создания диаграммы")
        # Создаём минимальную диаграмму
        kp_production_info['ПУСТО'] = {
            'start_day': 1,
            'end_day': 1,
            'customer': 'План пуст',
            'deadline': 'неизвестно',
            'plate_count': 0,
            'plates_by_day': defaultdict(int)
        }
    
    # === ШАГ 2: Определяем диапазон дней ===
    plan_days = max(info['end_day'] for info in kp_production_info.values())
    total_days = plan_days + max(0, extra_calendar_days)
    
    # === ШАГ 3: Создаём Excel ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Диаграмма Ганта"
    
    # Стили
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Цвета для статусов
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Успеваем
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Впритык
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Опаздываем
    
    # === ШАГ 3.5: Подсчитываем дорожки по дням ===
    tracks_by_day = defaultdict(int)  # {day_number: count_of_tracks}
    
    for track_idx, track in enumerate(all_tracks_list):
        # Определяем день производства для дорожки
        if isinstance(track, dict) and 'production_day' in track:
            production_day = track['production_day']
        else:
            production_day = (track_idx // tracks_count) + 1
        
        tracks_by_day[production_day] += 1
    
    # === ШАГ 4: Заголовки ===
    headers = ["КП", "Заказчик", "Дедлайн", "Плит"]
    
    # Добавляем даты
    date_columns = []
    for day in range(1, total_days + 1):
        date = _working_dt(day)
        date_str = date.strftime("%d.%m")
        headers.append(date_str)
        date_columns.append(date)

    subheader_gray = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    weekend_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # === Строка 1: дни недели над датами ===
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    cell_a1 = ws.cell(row=1, column=1, value="")
    cell_a1.font = Font(bold=True, size=10)
    cell_a1.fill = subheader_gray
    cell_a1.alignment = center_align
    cell_a1.border = thin_border

    for day in range(1, total_days + 1):
        col = 4 + day
        dt = _working_dt(day)
        cell = ws.cell(row=1, column=col, value=_weekday_ru(dt))
        cell.font = header_font
        if not is_working_day(dt.date(), holidays, extra_workdays):
            cell.fill = weekend_fill
        else:
            cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # === Строка 2: заголовки (КП, …, даты) ===
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        if col > 4 and not is_working_day(date_columns[col - 5].date(), holidays, extra_workdays):
            cell.fill = weekend_fill
        else:
            cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # === ШАГ 4.5: Строка 3 — загруженность дорожек ===
    # Добавляем информацию о дорожках: "3/5" = 3 дорожки занято из 5 максимум
    MAX_TRACKS = 5  # Максимальное количество дорожек в день

    # Первые 4 колонки объединяем (КП, Заказчик, Дедлайн, Плит)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
    cell = ws.cell(row=3, column=1, value="Загруженность дорожек")
    cell.font = Font(bold=True, size=10, italic=True)
    cell.fill = subheader_gray
    cell.alignment = center_align
    cell.border = thin_border

    # Для каждого дня показываем загруженность
    for day in range(1, total_days + 1):
        col = 4 + day
        occupied_tracks = tracks_by_day.get(day, 0)
        tracks_info = f"{occupied_tracks}/{MAX_TRACKS}"

        cell = ws.cell(row=3, column=col, value=tracks_info)
        cell.font = Font(bold=False, size=9)
        cell.alignment = center_align
        cell.border = thin_border
        
        # Цветовая индикация загруженности
        if occupied_tracks == 0:
            # Пусто - белый
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        elif occupied_tracks < MAX_TRACKS:
            # Есть свободные места - светло-зелёный
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        elif occupied_tracks == MAX_TRACKS:
            # Полностью занято - жёлтый
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        else:
            # Перегружено (больше 5) - красный
            cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    # === ШАГ 5: Данные по КП ===
    # Сортируем КП по дедлайну
    sorted_kps = sorted(
        kp_production_info.items(),
        key=lambda x: _parse_date(x[1]['deadline'])
    )
    
    row = 4  # После строк: дни недели, заголовки, загруженность
    plates_by_day_total = defaultdict(int)  # Итого плит по дням
    
    for kp_id, info in sorted_kps:
        # Колонка КП
        ws.cell(row=row, column=1, value=kp_id).border = thin_border
        ws.cell(row=row, column=1).alignment = center_align
        
        # Колонка Заказчик
        customer = info['customer']
        if len(customer) > 20:
            customer = customer[:18] + ".."
        ws.cell(row=row, column=2, value=customer).border = thin_border
        
        # Колонка Дедлайн
        ws.cell(row=row, column=3, value=info['deadline']).border = thin_border
        ws.cell(row=row, column=3).alignment = center_align
        
        # Колонка Плит
        ws.cell(row=row, column=4, value=info['plate_count']).border = thin_border
        ws.cell(row=row, column=4).alignment = center_align
        
        # Определяем цвет на основе дедлайна
        deadline_date = _parse_date(info['deadline'])
        end_production_date = _working_dt(info['end_day'])
        
        if deadline_date:
            if end_production_date < deadline_date:
                fill_color = green_fill  # Успеваем с запасом
            elif end_production_date == deadline_date:
                fill_color = yellow_fill  # Впритык
            else:
                fill_color = red_fill  # Опаздываем!
        else:
            fill_color = green_fill  # Если дедлайн неизвестен - зелёный
        
        # Закрашиваем ячейки для дней производства
        start_day = info['start_day']
        end_day = info['end_day']
        
        for day in range(1, total_days + 1):
            col = 4 + day  # Столбец даты (после 4 информационных столбцов)
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center_align
            
            if start_day <= day <= end_day:
                cell.fill = fill_color
                # Показываем количество плит в этот день
                plates_in_day = info['plates_by_day'].get(day, 0)
                if plates_in_day > 0:
                    cell.value = plates_in_day
                    plates_by_day_total[day] += plates_in_day
        
        row += 1
    
    # === ШАГ 6: Строка итогов ===
    row += 1  # Пустая строка
    
    ws.cell(row=row, column=1, value="ИТОГО").font = header_font
    ws.cell(row=row, column=1).border = thin_border
    
    total_plates = sum(info['plate_count'] for info in kp_production_info.values())
    ws.cell(row=row, column=4, value=total_plates).font = header_font
    ws.cell(row=row, column=4).border = thin_border
    ws.cell(row=row, column=4).alignment = center_align
    
    for day in range(1, total_days + 1):
        col = 4 + day
        cell = ws.cell(row=row, column=col)
        cell.value = plates_by_day_total.get(day, 0)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align
        cell.fill = header_fill
    
    # === ШАГ 7: Настройка ширины столбцов ===
    ws.column_dimensions['A'].width = 6   # КП
    ws.column_dimensions['B'].width = 22  # Заказчик
    ws.column_dimensions['C'].width = 10  # Дедлайн
    ws.column_dimensions['D'].width = 6   # Плит
    
    # Столбцы с датами
    for col in range(5, 5 + total_days):
        ws.column_dimensions[get_column_letter(col)].width = 7
    
    # === ШАГ 8: Легенда ===
    legend_row = row + 3
    ws.cell(row=legend_row, column=1, value="Легенда:").font = Font(bold=True)
    
    ws.cell(row=legend_row + 1, column=1, value="")
    ws.cell(row=legend_row + 1, column=1).fill = green_fill
    ws.cell(row=legend_row + 1, column=2, value="Успеваем до дедлайна")
    
    ws.cell(row=legend_row + 2, column=1, value="")
    ws.cell(row=legend_row + 2, column=1).fill = yellow_fill
    ws.cell(row=legend_row + 2, column=2, value="Завершаем в день дедлайна")
    
    ws.cell(row=legend_row + 3, column=1, value="")
    ws.cell(row=legend_row + 3, column=1).fill = red_fill
    ws.cell(row=legend_row + 3, column=2, value="Опаздываем!")
    
    # === ШАГ 9: Сохраняем файл ===
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Диаграмма_Ганта_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    wb.save(filepath)
    print(f"[GANTT] Diagramma Ganta sohranena: {filepath}")
    
    return filepath


def _parse_date(date_str: str) -> datetime:
    """
    Парсит строку даты в datetime.
    Поддерживает форматы: "ДД.ММ.ГГГГ", "ГГГГ-ММ-ДД"
    
    Returns:
        datetime или None если не удалось распарсить
    """
    if not date_str or date_str == 'неизвестно':
        return None
    
    # Формат "ДД.ММ.ГГГГ"
    try:
        return datetime.strptime(date_str, '%d.%m.%Y')
    except ValueError:
        pass
    
    # Формат "ГГГГ-ММ-ДД"
    try:
        return datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        pass
    
    return None

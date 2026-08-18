"""Pure mapping: WaybillDay / export DTO → ОКУД 0345001 blank cells.

No I/O except operating on an already-loaded openpyxl workbook.
Cell coordinates verified against real ПЛ (Phase 0 + import_gsm_history).
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from typing import Any

from openpyxl.workbook.workbook import Workbook

SHEET_FRONT = "стр.1"
SHEET_BACK = "стр.2"

# openpyxl 1-based coordinates (стр.1)
CELL_DATE_DAY = "W3"
CELL_DATE_MONTH = "AB3"
CELL_DATE_YEAR = "AL3"
CELL_VALID_DAY = "V4"
CELL_VALID_MONTH = "AA4"
CELL_VALID_YEAR = "AK4"
CELL_VEHICLE_MARK = "AA10"
CELL_PLATE = "AN11"
CELL_DRIVER_FIO = "R12"
CELL_PERSONNEL = "BU12"
CELL_LICENSE = "X14"
CELL_SNILS_LINE = "F17"
CELL_DRIVER_SHORT_BU27 = "BU27"
CELL_DRIVER_SHORT_AH32 = "AH32"
CELL_DRIVER_SHORT_AH36 = "AH36"
CELL_DRIVER_SHORT_AG45 = "AG45"
CELL_ODO_START = "BZ20"  # input
CELL_FUEL_START = "BS34"  # input
CELL_FUEL_ISSUED = "BS38"  # input
CELL_BS41 = "BS41"  # formula — patch norm only
CELL_ADDR_PODACHI = "F28"

# стр.2 leg rows 5..7 (I/J merged for km)
LEG_ROW_START = 5
LEG_ROW_END = 7  # inclusive
MAX_LEGS = LEG_ROW_END - LEG_ROW_START + 1

_MONTHS_GENITIVE = (
    "",
    "января",
    "февраля",
    "марта",
    "апреля",
    "мая",
    "июня",
    "июля",
    "августа",
    "сентября",
    "октября",
    "ноября",
    "декабря",
)

_TIME_RE = re.compile(r"^(\d{1,2}):(\d{2})$")


@dataclass(frozen=True, slots=True)
class BlankDriver:
    full_name: str
    license_number: str
    license_issued_at: str | None = None
    personnel_number: str | None = None
    snils: str | None = None


@dataclass(frozen=True, slots=True)
class BlankLeg:
    from_addr: str
    to_addr: str
    km: int | float
    dep_time: str | None = None
    arr_time: str | None = None


@dataclass(frozen=True, slots=True)
class BlankWaybill:
    day: date
    vehicle_mark: str
    plate_number: str
    driver: BlankDriver
    odometer_start: int
    fuel_start: float
    fuel_issued: float
    norm_l_per_100: float
    legs: tuple[BlankLeg, ...]
    addr_podachi: str | None = None


def waybill_export_filename(day: date) -> str:
    """«ПЛ DD.MM.YY.xls» — zip entry name."""
    return f"ПЛ {day.strftime('%d.%m.%y')}.xls"


def format_bs41_formula(norm_l_per_100: float) -> str:
    """Patched расход-по-норме formula for machine/season."""
    norm = _format_norm(norm_l_per_100)
    return f"='стр.2'!C19*{norm}/100"


def short_driver_name(full_name: str) -> str:
    """«Фамилия И.О.» from «Фамилия Имя Отчество»."""
    parts = [p for p in (full_name or "").split() if p]
    if not parts:
        return ""
    if len(parts) == 1:
        return parts[0]
    initials = "".join(f"{p[0]}." for p in parts[1:] if p)
    return f"{parts[0]} {initials}".strip()


def format_license_line(license_number: str, issued_at: str | None) -> str:
    num = (license_number or "").strip()
    if not issued_at:
        return num
    issued = _normalize_issued_date(issued_at)
    if not issued:
        return num
    return f"{num}, выдано {issued}г."


def format_snils_line(full_name: str, snils: str | None) -> str:
    short = short_driver_name(full_name)
    if not snils:
        return short
    return f"{short}       СНИЛС {snils.strip()}"


def vehicle_mark_label(name: str) -> str:
    """Blank cell prefers «Легковой универсал …» prefix used in historical ПЛ."""
    raw = (name or "").strip()
    if not raw:
        return raw
    low = raw.lower()
    if low.startswith("легковой"):
        return raw
    return f"Легковой универсал {raw}"


def fill_workbook(wb: Workbook, data: BlankWaybill) -> None:
    """Write input cells; leave formula cells (except BS41 norm patch) intact."""
    ws1 = wb[SHEET_FRONT]
    ws2 = wb[SHEET_BACK]

    month_name = _MONTHS_GENITIVE[data.day.month]
    for day_cell, month_cell, year_cell in (
        (CELL_DATE_DAY, CELL_DATE_MONTH, CELL_DATE_YEAR),
        (CELL_VALID_DAY, CELL_VALID_MONTH, CELL_VALID_YEAR),
    ):
        ws1[day_cell] = data.day.day
        ws1[month_cell] = month_name
        ws1[year_cell] = data.day.year

    ws1[CELL_VEHICLE_MARK] = data.vehicle_mark
    ws1[CELL_PLATE] = data.plate_number
    ws1[CELL_DRIVER_FIO] = data.driver.full_name
    if data.driver.personnel_number:
        ws1[CELL_PERSONNEL] = str(data.driver.personnel_number)
    ws1[CELL_LICENSE] = format_license_line(
        data.driver.license_number, data.driver.license_issued_at
    )
    ws1[CELL_SNILS_LINE] = format_snils_line(data.driver.full_name, data.driver.snils)

    short = short_driver_name(data.driver.full_name)
    for coord in (
        CELL_DRIVER_SHORT_BU27,
        CELL_DRIVER_SHORT_AH32,
        CELL_DRIVER_SHORT_AH36,
        CELL_DRIVER_SHORT_AG45,
    ):
        ws1[coord] = short

    ws1[CELL_ODO_START] = int(data.odometer_start)
    ws1[CELL_FUEL_START] = float(data.fuel_start)
    ws1[CELL_FUEL_ISSUED] = float(data.fuel_issued)
    ws1[CELL_BS41] = format_bs41_formula(data.norm_l_per_100)

    if data.addr_podachi:
        ws1[CELL_ADDR_PODACHI] = data.addr_podachi

    _clear_and_write_legs(ws2, data.legs)


def legs_from_route_items(
    items: list[dict[str, Any]],
    *,
    stations_by_id: dict[int, str] | None = None,
) -> tuple[BlankLeg, ...]:
    """Build blank legs from route_json items; inject АЗС address when needed."""
    stations_by_id = stations_by_id or {}
    out: list[BlankLeg] = []
    for item in items:
        from_addr = str(
            item.get("from")
            or item.get("from_addr")
            or item.get("addr_from")
            or item.get("addr_a")
            or ""
        ).strip()
        to_addr = str(
            item.get("to")
            or item.get("to_addr")
            or item.get("addr_to")
            or item.get("addr_b")
            or ""
        ).strip()
        km = int(item.get("km") or 0)
        dep = item.get("dep_time") or item.get("time_dep")
        arr = item.get("arr_time") or item.get("time_ret")
        station_id = item.get("station_id")
        station_addr = ""
        if station_id is not None:
            try:
                station_addr = stations_by_id.get(int(station_id), "") or ""
            except (TypeError, ValueError):
                station_addr = ""

        if (
            station_addr
            and station_addr not in from_addr
            and station_addr not in to_addr
            and len(out) + 2 <= MAX_LEGS
        ):
            half = km // 2
            out.append(
                BlankLeg(
                    from_addr=from_addr or station_addr,
                    to_addr=station_addr,
                    km=half if half > 0 else km,
                    dep_time=str(dep) if dep else None,
                    arr_time=None,
                )
            )
            out.append(
                BlankLeg(
                    from_addr=station_addr,
                    to_addr=to_addr or station_addr,
                    km=km - half if half > 0 else 0,
                    dep_time=None,
                    arr_time=str(arr) if arr else None,
                )
            )
            continue

        if station_addr and station_addr not in from_addr and station_addr not in to_addr:
            # No room to split — put АЗС into destination so SC-4 holds.
            to_addr = station_addr if not to_addr else f"{to_addr} / {station_addr}"

        out.append(
            BlankLeg(
                from_addr=from_addr,
                to_addr=to_addr,
                km=km,
                dep_time=str(dep) if dep else None,
                arr_time=str(arr) if arr else None,
            )
        )
        if len(out) >= MAX_LEGS:
            break

    return tuple(out[:MAX_LEGS])


def _clear_and_write_legs(ws2: Any, legs: tuple[BlankLeg, ...]) -> None:
    for row in range(LEG_ROW_START, LEG_ROW_END + 1):
        ws2.cell(row=row, column=2).value = None  # B seq
        ws2.cell(row=row, column=3).value = None
        ws2.cell(row=row, column=4).value = None
        for col in range(5, 9):
            ws2.cell(row=row, column=col).value = None
        ws2.cell(row=row, column=9).value = None  # I (merged I:J)

    for idx, leg in enumerate(legs[:MAX_LEGS]):
        row = LEG_ROW_START + idx
        ws2.cell(row=row, column=2).value = str(idx + 1)
        ws2.cell(row=row, column=3).value = leg.from_addr
        ws2.cell(row=row, column=4).value = leg.to_addr
        dep_h, dep_m = _split_time(leg.dep_time)
        arr_h, arr_m = _split_time(leg.arr_time)
        if dep_h is not None:
            ws2.cell(row=row, column=5).value = dep_h
        if dep_m is not None:
            ws2.cell(row=row, column=6).value = dep_m
        if arr_h is not None:
            ws2.cell(row=row, column=7).value = arr_h
        if arr_m is not None:
            ws2.cell(row=row, column=8).value = arr_m
        ws2.cell(row=row, column=9).value = float(leg.km) if leg.km else None


def _split_time(value: str | None) -> tuple[str | int | None, str | int | None]:
    if not value:
        return None, None
    m = _TIME_RE.match(value.strip())
    if not m:
        return None, None
    return int(m.group(1)), m.group(2)


def _format_norm(norm: float) -> str:
    text = f"{float(norm):.10g}"
    return text


def _normalize_issued_date(raw: str) -> str:
    s = (raw or "").strip()
    if not s:
        return ""
    # ISO YYYY-MM-DD → DD.MM.YYYY
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        y, m, d = s.split("-")
        return f"{d}.{m}.{y}"
    s = s.rstrip(".")
    if s.lower().endswith("г"):
        s = s[:-1].rstrip(".")
    return s

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""XLSX-шаблон графика поставки: шаблон с плитами КП и парсер черновика партий.

Колонки: Партия | Поставка с | Поставка по | Произвести до | Марка | Кол-во.
Даты в файле — ДД.ММ.ГГГГ; в ``BatchDraft`` — ISO (YYYY-MM-DD).
Матчинг марки → позиция КП: точное совпадение ``plate_name``.
``build_template(..., plates=, batches=)`` пишет две полосы:
сверху уже в поставках, ниже остаток. Партии и даты в верхней полосе —
из ``batches``; в остатке они пустые.
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore

logger = logging.getLogger(__name__)

HEADERS = (
    "Партия",
    "Поставка с",
    "Поставка по",
    "Произвести до",
    "Марка",
    "Кол-во",
)

SECTION_IN_SCHEDULE = "Уже в поставках"
SECTION_REMAINDER = "Остаток"
SECTION_TITLES = frozenset({SECTION_IN_SCHEDULE, SECTION_REMAINDER})

# Колонки документа (MVP: строки партий/позиций, без этажей-колонок).
DOC_HEADERS = (
    "№",
    "Партия",
    "Поставка с",
    "Поставка по",
    "Произвести до",
    "Марка",
    "Кол-во",
)

_DATE_FMT = "%d.%m.%Y"
_HEADER_FONT = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
_SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

REASON_UNKNOWN_MARK = "unknown mark"
REASON_BAD_DATE = "bad date"
REASON_BAD_QTY = "bad qty"
REASON_CONFLICTING_BATCH_DATES = "conflicting batch dates"


@dataclass
class BatchDraftItem:
    """Позиция черновика партии (ссылка на ``kp_plates.id``)."""

    plate_id: int
    plate_name: str
    qty: int


@dataclass
class BatchDraft:
    """Черновик партии из шаблона; даты в ISO (YYYY-MM-DD)."""

    name: str
    deliver_from: str
    deliver_to: str
    produce_by: str
    items: list[BatchDraftItem] = field(default_factory=list)


@dataclass
class UnmatchedRow:
    """Строка шаблона, которую не удалось разобрать или сопоставить."""

    row_number: int
    reason: str
    raw: dict | None = None


def build_template(
    path: str | Path,
    plates: list[dict] | None = None,
    batches: list[Any] | None = None,
) -> Path:
    """Создаёт XLSX-шаблон графика: две полосы при наличии данных.

    Верх — строки ``batches`` (уже в поставках). Низ — остаток КП
    (qty позиции минус сумма qty черновика). Полосу и заголовок пишем
    только если в ней есть строки. Без ``plates`` и ``batches`` — только
    заголовки колонок.

    Returns:
        Path к записанному файлу.
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl не установлен — шаблон XLSX недоступен")

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "График поставки"

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP

    row_idx = 2
    for row in _iter_template_rows(plates or [], batches or []):
        if row["kind"] == "section":
            _write_section_row(ws, row_idx, row["title"])
        else:
            ws.cell(row=row_idx, column=1, value=row["batch_name"] or None)
            ws.cell(row=row_idx, column=2, value=row["deliver_from"] or None)
            ws.cell(row=row_idx, column=3, value=row["deliver_to"] or None)
            ws.cell(row=row_idx, column=4, value=row["produce_by"] or None)
            ws.cell(row=row_idx, column=5, value=row["plate_name"] or None)
            ws.cell(row=row_idx, column=6, value=row["qty"])
        row_idx += 1

    widths = (28, 14, 14, 14, 28, 10)
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(str(path))
    return path


def build_document(schedule_view_or_dict: Any, path: str | Path) -> Path:
    """Собирает XLSX-документ графика поставки (шапка + таблица партий).

    Макет MVP (R5): строки партий/позиций, не этажи-колонки ЯРПРОФИТ.
    ``schedule_view_or_dict`` — ``DeliveryScheduleView`` или dict; опционально
    ключ ``customer_name`` из КП для шапки.
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl не установлен — документ XLSX недоступен")

    data = schedule_as_dict(schedule_view_or_dict)
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "График поставки"

    header_lines = document_header_lines(data)
    row_idx = 1
    for line in header_lines:
        cell = ws.cell(row=row_idx, column=1, value=line)
        cell.font = _HEADER_FONT
        row_idx += 1
    row_idx += 1  # пустая строка перед таблицей

    for col, header in enumerate(DOC_HEADERS, start=1):
        cell = ws.cell(row=row_idx, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER

    for values in iter_document_table_rows(data):
        row_idx += 1
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = _WRAP if col in (2, 6) else _CENTER

    widths = (6, 28, 14, 14, 14, 28, 10)
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(str(path))
    return path


def document_header_lines(data: dict[str, Any]) -> list[str]:
    """Строки шапки документа: договор, счёт/КП, стороны."""
    kp_id = data.get("kp_id")
    contract = (data.get("contract_number") or "").strip() or None
    invoice = (data.get("invoice_number") or "").strip() or None
    customer = (data.get("customer_name") or "").strip() or None

    account_label = invoice if invoice else (
        f"КП №{kp_id}" if kp_id is not None else "КП"
    )

    lines = ["График поставки"]
    if contract:
        lines.append(f"Договор: {contract}")
    lines.append(f"Счёт: {account_label}")
    if customer:
        lines.append(f"Заказчик: {customer}")
    return lines


def iter_document_table_rows(data: dict[str, Any]) -> list[tuple[Any, ...]]:
    """Строки таблицы документа: №, партия, даты, марка, кол-во."""
    rows: list[tuple[Any, ...]] = []
    n = 0
    batches = data.get("batches") or []
    for batch in batches:
        if not isinstance(batch, dict):
            if hasattr(batch, "model_dump"):
                batch = batch.model_dump()
            else:
                continue
        name = str(batch.get("name") or "")
        d_from = _iso_to_display(batch.get("deliver_from"))
        d_to = _iso_to_display(batch.get("deliver_to"))
        produce = _iso_to_display(batch.get("produce_by"))
        items = batch.get("items") or []
        if not items:
            n += 1
            rows.append((n, name, d_from, d_to, produce, "", ""))
            continue
        for item in items:
            if not isinstance(item, dict):
                if hasattr(item, "model_dump"):
                    item = item.model_dump()
                else:
                    continue
            n += 1
            mark = str(item.get("plate_name") or "").strip()
            qty = item.get("qty")
            rows.append((n, name, d_from, d_to, produce, mark, qty))
    return rows


def schedule_as_dict(schedule_view_or_dict: Any) -> dict[str, Any]:
    """Нормализует view/dict графика к обычному dict."""
    if isinstance(schedule_view_or_dict, dict):
        return schedule_view_or_dict
    if hasattr(schedule_view_or_dict, "model_dump"):
        return schedule_view_or_dict.model_dump()
    raise TypeError(
        "ожидался DeliveryScheduleView или dict, "
        f"получено {type(schedule_view_or_dict)!r}"
    )


def _iso_to_display(value: Any) -> str:
    """ISO YYYY-MM-DD → ДД.ММ.ГГГГ; иначе строка как есть."""
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    try:
        return date.fromisoformat(text).strftime(_DATE_FMT)
    except ValueError:
        return text


def _iter_template_rows(
    plates: list[dict],
    batches: list[Any],
) -> list[dict[str, Any]]:
    """Ряды двухполосного шаблона: секции + позиции (даты ДД.ММ.ГГГГ)."""
    plates_by_id: dict[int, dict] = {}
    for plate in plates:
        plate_id = _plate_id(plate)
        if plate_id is not None:
            plates_by_id[plate_id] = plate

    allocated: dict[int, int] = {}
    scheduled: list[dict[str, Any]] = []
    for batch in batches:
        batch_map = _as_mapping(batch)
        items = batch_map.get("items") or []
        for item in items:
            item_map = _as_mapping(item)
            qty = _parse_qty(item_map.get("qty"))
            if qty is None:
                continue
            plate_id = _plate_id(item_map)
            if plate_id is None:
                continue
            plate = plates_by_id.get(plate_id)
            mark = str(item_map.get("plate_name") or "").strip()
            if not mark and plate is not None:
                mark = str(plate.get("plate_name") or "").strip()
            if not mark:
                continue
            allocated[plate_id] = allocated.get(plate_id, 0) + qty
            scheduled.append(
                {
                    "kind": "item",
                    "batch_name": str(batch_map.get("name") or ""),
                    "deliver_from": _iso_to_display(batch_map.get("deliver_from")),
                    "deliver_to": _iso_to_display(batch_map.get("deliver_to")),
                    "produce_by": _iso_to_display(batch_map.get("produce_by")),
                    "plate_name": mark,
                    "qty": qty,
                }
            )

    remainder: list[dict[str, Any]] = []
    for plate in plates:
        mark = str(plate.get("plate_name") or "").strip()
        if not mark:
            continue
        leftover = _kp_qty(plate) - allocated.get(_plate_id(plate), 0)
        if leftover <= 0:
            continue
        remainder.append(
            {
                "kind": "item",
                "batch_name": "",
                "deliver_from": "",
                "deliver_to": "",
                "produce_by": "",
                "plate_name": mark,
                "qty": leftover,
            }
        )

    rows: list[dict[str, Any]] = []
    if scheduled:
        rows.append({"kind": "section", "title": SECTION_IN_SCHEDULE})
        rows.extend(scheduled)
    if remainder:
        rows.append({"kind": "section", "title": SECTION_REMAINDER})
        rows.extend(remainder)
    return rows


def _write_section_row(ws: Any, row_idx: int, title: str) -> None:
    for col in range(1, 7):
        cell = ws.cell(row=row_idx, column=col, value=title if col == 1 else None)
        cell.fill = _SECTION_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)


def _as_mapping(obj: Any) -> dict[str, Any]:
    if isinstance(obj, dict):
        return obj
    if hasattr(obj, "model_dump"):
        return obj.model_dump()
    if hasattr(obj, "__dict__"):
        return dict(obj.__dict__)
    return {}


def _plate_id(obj: dict[str, Any] | None) -> int | None:
    if not obj:
        return None
    raw = obj.get("id") if obj.get("id") is not None else obj.get("plate_id")
    if raw is None or raw == "":
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


def _kp_qty(plate: dict[str, Any]) -> int:
    raw = plate.get("qty")
    if raw is None or raw == "":
        return 0
    try:
        return int(raw)
    except (TypeError, ValueError):
        parsed = _parse_qty(raw)
        return parsed if parsed is not None else 0


def _row_is_section_title(*values: Any) -> bool:
    for value in values:
        if value is None:
            continue
        if str(value).strip() in SECTION_TITLES:
            return True
    return False


def parse_template(
    data: bytes,
    kp_plates: list[dict],
) -> tuple[list[BatchDraft], list[UnmatchedRow]]:
    """Разбирает заполненный XLSX-шаблон в черновик партий.

    Матчинг марки — точное по ``plate_name``. Строки с одинаковым именем
    партии группируются в одну ``BatchDraft``. Несматченные строки
    возвращаются отдельно с причиной ``unknown mark`` / ``bad date`` /
    ``bad qty``.
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl не установлен — разбор XLSX недоступен")

    plates_by_name: dict[str, dict] = {}
    for plate in kp_plates:
        name = str(plate.get("plate_name") or "").strip()
        if name and name not in plates_by_name:
            plates_by_name[name] = plate

    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    try:
        ws = wb.active
        batches_order: list[str] = []
        batches_map: dict[str, BatchDraft] = {}
        unmatched: list[UnmatchedRow] = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=6, values_only=True), start=2):
            cells = list(row) + [None] * (6 - len(row))
            batch_name_raw, d_from_raw, d_to_raw, produce_raw, mark_raw, qty_raw = cells[:6]

            if _row_is_blank(batch_name_raw, d_from_raw, d_to_raw, produce_raw, mark_raw, qty_raw):
                continue
            if _row_is_section_title(
                batch_name_raw, d_from_raw, d_to_raw, produce_raw, mark_raw, qty_raw
            ):
                continue
            if _row_is_unfilled_template(batch_name_raw, d_from_raw, d_to_raw, produce_raw):
                continue

            raw = {
                "Партия": _cell_as_str(batch_name_raw),
                "Поставка с": _cell_as_str(d_from_raw),
                "Поставка по": _cell_as_str(d_to_raw),
                "Произвести до": _cell_as_str(produce_raw),
                "Марка": _cell_as_str(mark_raw),
                "Кол-во": _cell_as_str(qty_raw),
            }

            mark = str(mark_raw or "").strip()
            plate = plates_by_name.get(mark)
            if plate is None:
                unmatched.append(
                    UnmatchedRow(row_number=row_idx, reason=REASON_UNKNOWN_MARK, raw=raw)
                )
                continue

            iso_from = _parse_date_strict(d_from_raw)
            iso_to = _parse_date_strict(d_to_raw)
            iso_produce = _parse_date_strict(produce_raw)
            if iso_from is None or iso_to is None or iso_produce is None:
                unmatched.append(
                    UnmatchedRow(row_number=row_idx, reason=REASON_BAD_DATE, raw=raw)
                )
                continue

            qty = _parse_qty(qty_raw)
            if qty is None:
                unmatched.append(
                    UnmatchedRow(row_number=row_idx, reason=REASON_BAD_QTY, raw=raw)
                )
                continue

            batch_name = str(batch_name_raw or "").strip()
            plate_id = int(plate.get("id") if plate.get("id") is not None else plate["plate_id"])
            plate_name = str(plate.get("plate_name") or mark).strip()
            item = BatchDraftItem(plate_id=plate_id, plate_name=plate_name, qty=qty)

            existing = batches_map.get(batch_name)
            if existing is None:
                draft = BatchDraft(
                    name=batch_name,
                    deliver_from=iso_from,
                    deliver_to=iso_to,
                    produce_by=iso_produce,
                    items=[item],
                )
                batches_map[batch_name] = draft
                batches_order.append(batch_name)
            elif (
                existing.deliver_from != iso_from
                or existing.deliver_to != iso_to
                or existing.produce_by != iso_produce
            ):
                unmatched.append(
                    UnmatchedRow(
                        row_number=row_idx,
                        reason=REASON_CONFLICTING_BATCH_DATES,
                        raw=raw,
                    )
                )
            else:
                _merge_item(existing, item)

        return [batches_map[name] for name in batches_order], unmatched
    finally:
        wb.close()


def _row_is_blank(*values: Any) -> bool:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return False
    return True


def _row_is_unfilled_template(
    batch_name: Any,
    d_from: Any,
    d_to: Any,
    produce: Any,
) -> bool:
    """Строка шаблона: марка/кол-во есть, партия и даты ещё не заполнены."""
    return _row_is_blank(batch_name, d_from, d_to, produce)


def _cell_as_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime(_DATE_FMT)
    if isinstance(value, date):
        return value.strftime(_DATE_FMT)
    return str(value).strip()


def _parse_date_strict(value: Any) -> str | None:
    """Строгий разбор даты → ISO. Строки только ДД.ММ.ГГГГ; datetime/date — как есть."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.strptime(text, _DATE_FMT).date().isoformat()
    except ValueError:
        return None


def _parse_qty(value: Any) -> int | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        if isinstance(value, bool):
            return None
        if isinstance(value, int):
            qty = value
        elif isinstance(value, float):
            if not value.is_integer():
                return None
            qty = int(value)
        else:
            text = str(value).strip().replace(",", ".")
            as_float = float(text)
            if not as_float.is_integer():
                return None
            qty = int(as_float)
    except (TypeError, ValueError):
        return None
    if qty < 1:
        return None
    return qty


def _merge_item(batch: BatchDraft, item: BatchDraftItem) -> None:
    for existing in batch.items:
        if existing.plate_id == item.plate_id:
            existing.qty += item.qty
            return
    batch.items.append(item)

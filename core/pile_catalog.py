#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Мини-справочник свай ``pile_catalog``: парсинг прайса, upsert, резолв марок.

Источник — лист «Вес и объем» или «Лист1»: строка заголовка ищется по «марка»/«вес»,
колонка шт. — «автомобильный» / «20т» (иначе 4-я колонка). «-» и пусто → pcs NULL.
"""

from __future__ import annotations

import os
import re
import sqlite3
from dataclasses import dataclass
from typing import Optional, Sequence

from openpyxl import load_workbook

PILE_WEIGHT_SHEET = "Вес и объем"
PILE_FALLBACK_SHEET = "Лист1"
PILE_HEADER_ROW = 2  # legacy default if header cells are not found

_BRIDGE_GEOMETRY_RE = re.compile(
    r"^[СC]\s*(\d+(?:[.,]\d+)?)\s*-\s*(\d+)\s*[TТBВ]\s*\d+$",
    re.IGNORECASE | re.UNICODE,
)


@dataclass(frozen=True)
class PileCatalogEntry:
    mark: str
    length_m: Optional[float]
    section_mm: Optional[int]
    volume_m3: Optional[float]
    weight_kg: float
    pcs_per_20t: Optional[int]


def parse_pile_mark(mark: str) -> tuple[Optional[float], Optional[int]]:
    """«С137,5.40» → (13.75 м, 400 мм); «С60.30» → (6.0 м, 300 мм).

    Длина в марке — в дециметрах (допустима запятая как разделитель),
    сечение — в сантиметрах.
    """
    text = mark.strip().lstrip("СсCc").strip()
    length_part, sep, section_part = text.rpartition(".")
    if not sep:
        return None, None
    try:
        length_dm = float(length_part.replace(",", "."))
        section_cm = float(section_part.replace(",", "."))
    except ValueError:
        return None, None
    return length_dm / 10.0, int(round(section_cm * 10))


def parse_bridge_pile_geometry(mark: str) -> tuple[Optional[float], Optional[int]]:
    """C14-40T4 / С14-40Т4 → (14.0 м, 400 мм). Канон С140.40 сюда не подходит."""
    text = (mark or "").strip()
    text = re.sub(r"\s+", "", text)
    match = _BRIDGE_GEOMETRY_RE.match(text)
    if not match:
        return None, None
    length_m = float(match.group(1).replace(",", "."))
    section_cm = float(match.group(2).replace(",", "."))
    return length_m, int(round(section_cm * 10))


def normalize_pile_mark_key(mark: str) -> str:
    """Ключ override / точного совпадения: C↔С, T↔Т, B↔В, без пробелов."""
    text = str(mark or "").strip().upper()
    text = (
        text.replace("С", "C")
        .replace("В", "B")
        .replace("Т", "T")
    )
    return re.sub(r"\s+", "", text)


def _as_float(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text or text == "-":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _as_int(value: object) -> Optional[int]:
    number = _as_float(value)
    if number is None:
        return None
    return int(round(number))


def _cell_text(value: object) -> str:
    return str(value or "").strip().lower().replace("ё", "е")


def _find_header_row(rows: list[tuple]) -> int | None:
    for idx, row in enumerate(rows):
        texts = [_cell_text(cell) for cell in row]
        if any("марка" in t for t in texts) and any("вес" in t for t in texts):
            return idx
    return None


def _column_index(header: tuple, *needles: str) -> int | None:
    for idx, cell in enumerate(header):
        text = _cell_text(cell)
        if any(needle in text for needle in needles):
            return idx
    return None


def _resolve_sheet_name(sheetnames: Sequence[str], requested: str | None) -> str:
    names = list(sheetnames)
    if requested is not None:
        if requested in names:
            return requested
        raise ValueError(f"Лист «{requested}» не найден в файле")
    if PILE_WEIGHT_SHEET in names:
        return PILE_WEIGHT_SHEET
    if PILE_FALLBACK_SHEET in names:
        return PILE_FALLBACK_SHEET
    raise ValueError(
        f"Лист «{PILE_WEIGHT_SHEET}» или «{PILE_FALLBACK_SHEET}» не найден в файле"
    )


def parse_pile_catalog_from_xlsx(
    xlsx_path: str,
    *,
    sheet: str | None = None,
) -> list[PileCatalogEntry]:
    """Разобрать лист каталога в записи (марки без веса пропускаются)."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        try:
            sheet_name = _resolve_sheet_name(wb.sheetnames, sheet)
        except ValueError as exc:
            raise ValueError(f"{exc} {xlsx_path}") from exc
        ws = wb[sheet_name]
        raw_rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
        header_idx = _find_header_row(raw_rows)
        if header_idx is None:
            header_idx = PILE_HEADER_ROW - 1  # 1-based row 2
        if header_idx < 0 or header_idx >= len(raw_rows):
            return []
        header = raw_rows[header_idx]
        mark_col = _column_index(header, "марка")
        volume_col = _column_index(header, "объем")
        weight_col = _column_index(header, "вес")
        pcs_col = _column_index(header, "автомобильный", "20т")
        if mark_col is None:
            mark_col = 0
        if volume_col is None:
            volume_col = 1
        if weight_col is None:
            weight_col = 2
        if pcs_col is None:
            pcs_col = 3

        entries: list[PileCatalogEntry] = []
        seen: set[str] = set()
        for row in raw_rows[header_idx + 1 :]:
            if not row:
                continue
            mark_val = row[mark_col] if mark_col < len(row) else None
            mark = str(mark_val).strip() if mark_val is not None else ""
            if not mark or mark in seen:
                continue
            weight = _as_float(row[weight_col] if weight_col < len(row) else None)
            if weight is None or weight <= 0:
                continue
            length_m, section_mm = parse_pile_mark(mark)
            entries.append(
                PileCatalogEntry(
                    mark=mark,
                    length_m=length_m,
                    section_mm=section_mm,
                    volume_m3=_as_float(
                        row[volume_col] if volume_col < len(row) else None
                    ),
                    weight_kg=weight,
                    pcs_per_20t=_as_int(row[pcs_col] if pcs_col < len(row) else None),
                )
            )
            seen.add(mark)
        return entries
    finally:
        wb.close()


def resolve_catalog_for_mark(
    mark: str,
    entries: Sequence[PileCatalogEntry],
) -> Optional[PileCatalogEntry]:
    """Точный mark (C↔С), иначе геометрия length_m + section_mm.

    При нескольких строках с одной геометрией — любая с непустым pcs_per_20t,
    иначе любая с весом.
    """
    key = normalize_pile_mark_key(mark)
    if key:
        for entry in entries:
            if normalize_pile_mark_key(entry.mark) == key:
                return entry

    length_m, section_mm = parse_bridge_pile_geometry(mark)
    if length_m is None or section_mm is None:
        length_m, section_mm = parse_pile_mark(mark)
    if length_m is None or section_mm is None:
        return None

    matches = [
        entry
        for entry in entries
        if entry.length_m is not None
        and entry.section_mm is not None
        and abs(entry.length_m - length_m) < 1e-6
        and entry.section_mm == section_mm
    ]
    if not matches:
        return None
    with_pcs = [e for e in matches if e.pcs_per_20t and e.pcs_per_20t > 0]
    if with_pcs:
        return with_pcs[0]
    return matches[0]


def load_pile_catalog(db_path: str) -> list[PileCatalogEntry]:
    """Все строки ``pile_catalog`` (пустой список, если таблицы нет)."""
    try:
        path = os.fspath(db_path)
    except TypeError:
        return []
    if not path:
        return []
    try:
        conn = sqlite3.connect(path)
    except sqlite3.Error:
        return []
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT mark, length_m, section_mm, volume_m3, weight_kg, pcs_per_20t
            FROM pile_catalog
            """
        )
        rows = cur.fetchall()
    except sqlite3.Error:
        return []
    finally:
        conn.close()
    return [
        PileCatalogEntry(
            mark=row[0],
            length_m=row[1],
            section_mm=row[2],
            volume_m3=row[3],
            weight_kg=float(row[4]),
            pcs_per_20t=row[5],
        )
        for row in rows
    ]


def upsert_pile_catalog(db_path: str, entries: list[PileCatalogEntry]) -> tuple[int, int]:
    """Upsert по ``mark`` (безопасен для повторного запуска). Возвращает (inserted, updated)."""
    from core.kp_db_schema import ensure_schema

    ensure_schema(db_path)
    inserted = 0
    updated = 0
    conn = sqlite3.connect(db_path)
    try:
        cur = conn.cursor()
        for entry in entries:
            cur.execute("SELECT id FROM pile_catalog WHERE mark = ?", (entry.mark,))
            exists = cur.fetchone() is not None
            cur.execute(
                """
                INSERT INTO pile_catalog (mark, length_m, section_mm, volume_m3, weight_kg, pcs_per_20t)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(mark) DO UPDATE SET
                    length_m = excluded.length_m,
                    section_mm = excluded.section_mm,
                    volume_m3 = excluded.volume_m3,
                    weight_kg = excluded.weight_kg,
                    pcs_per_20t = excluded.pcs_per_20t
                """,
                (
                    entry.mark,
                    entry.length_m,
                    entry.section_mm,
                    entry.volume_m3,
                    entry.weight_kg,
                    entry.pcs_per_20t,
                ),
            )
            if exists:
                updated += 1
            else:
                inserted += 1
        conn.commit()
    finally:
        conn.close()
    return inserted, updated

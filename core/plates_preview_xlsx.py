#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""XLSX превью: ввод → распознано (имя/кол-во) → как в КП (имя/кол-во).

Двухблочная схема A/B:

**Блок A (обычные плиты, ширина <= 12 дм):**

- Строки группируются по распознанному имени плиты — дубли идут рядом.
- Колонки D–E (КП) показываются только в первой строке группы одинаковых плит;
  последующие дубли получают пустые D и E (схлопывание).
- Количество КП (колонка E) считается только по Блоку A, без вкладов Блока B.

**Блок B (широкие плиты, ширина > 12 дм):**

- Плиты идут в порядке, присланном пользователем.
- При разбиении на 2 плиты (напр. 1,5 м → 1,2 + 0,3) колонка A заполняется
  только в первой строке блока, последующие строки — пустая A.
- Количество КП считается только по Блоку B, без вкладов Блока A.

**Общее правило:** количества Блока A и Блока B не суммируются в колонке КП (кол-во),
чтобы избежать дублирования.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    Workbook = None  # type: ignore

import core.config_and_data as cfg
from core.config_and_data import LineContributionKey, make_plate_name, set_plate_lists_from_text
from core.plate_text_normalizer import get_wide_plate_lines
from core.reconciliation_xlsx import split_plate_text_lines

logger = logging.getLogger(__name__)

_HEADER_FONT = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")

# A + распознано (имя, кол-во) + КП (имя, кол-во)
_PREVIEW_COLS = 5
WIDE_WIDTH_M = 1.2
WIDE_EPS = 1e-6


def _is_wide_width(width_m: float, *, threshold_m: float = WIDE_WIDTH_M, eps: float = WIDE_EPS) -> bool:
    """True для плит шире 12 дм (> 1.2 м)."""
    return float(width_m) > (threshold_m + eps)


def _source_line_is_wide(line: str) -> bool:
    """
    Определяет, является ли исходная строка пользователя широкой (>12 дм).
    Учитывает вариант с префиксом «Плиты ...», который не всегда проходит
    через get_wide_plate_lines напрямую.
    """
    s = (line or "").strip()
    if not s:
        return False

    if get_wide_plate_lines(s):
        return True

    # Fallback: убираем префикс "Плиты " и проверяем ещё раз.
    without_plates = re.sub(r"(?i)^\s*плиты\s+", "", s).strip()
    return bool(without_plates and get_wide_plate_lines(without_plates))


def _contrib_looks_like_wide_split(keys: list[LineContributionKey]) -> bool:
    """
    Fallback-эвристика для wide-строки, если исходная строка A распознана нестабильно.
    Считаем строку wide, если в одном line_contribution есть пара вкладов по одной длине:
    1.2 м и 0.3 м (типичное разбиение 1.5 -> 1.2 + 0.3).
    """
    pairs: set[tuple[float, int, str]] = set()
    seen_12: set[tuple[float, int, str]] = set()
    seen_03: set[tuple[float, int, str]] = set()
    for lm, wm, lc, ldr in keys:
        lc_int = int(round(float(lc))) if lc is not None else 0
        marker = (round(float(lm), 3), lc_int, (ldr or "").strip())
        pairs.add(marker)
        if abs(float(wm) - 1.2) <= 1e-6:
            seen_12.add(marker)
        if abs(float(wm) - 0.3) <= 1e-6:
            seen_03.add(marker)
    return bool((seen_12 & seen_03) & pairs)



def _ldr_equal(a: str, b: str) -> bool:
    return (a or "").strip().replace(",", ".") == (b or "").strip().replace(",", ".")


def qty_for_contribution_key(
    key: LineContributionKey,
    plate_load_details: dict[tuple[float, float, int, str], int],
) -> int:
    """Количество по ключу строки; для раскола 1.5→1.2+0.3 ищет запись с шириной ~1.5 м."""
    lm, wm, lc, ldr = key
    ldr_s = (ldr or "").strip()

    def _load_ok(sload: int) -> bool:
        if lc is None:
            return True
        return cfg.load_code_for_price_match(float(sload)) == cfg.load_code_for_price_match(float(lc))

    for (sl, sw, sload, sldr), q in plate_load_details.items():
        if abs(sl - lm) >= 0.01 or abs(sw - wm) >= 0.01:
            continue
        if not _ldr_equal(ldr_s, sldr):
            continue
        if not _load_ok(sload):
            continue
        return int(q)

    if 1.14 <= wm <= 1.26 or 0.25 <= wm <= 0.34:
        for (sl, sw, sload, sldr), q in plate_load_details.items():
            if abs(sl - lm) >= 0.01:
                continue
            if not (1.45 <= sw <= 1.55):
                continue
            if not _ldr_equal(ldr_s, sldr):
                continue
            if not _load_ok(sload):
                continue
            return int(q)

    return 0


def _sort_contribution_keys(keys: list[LineContributionKey]) -> list[LineContributionKey]:
    """Сортировка по длине (м), ширине (м), затем лексикографически по марке длины в ключе."""

    def _key(k: LineContributionKey) -> tuple[float, float, str]:
        lm, wm, _lc, ldr = k
        return (lm, wm, (ldr or "").strip())

    return sorted(keys, key=_key)


def preview_row_triples_for_contributions(
    keys: list[LineContributionKey],
    line_plate_load_details: dict,
    global_plate_load_details: dict,
    *,
    max_pairs: int | None = None,
) -> list[tuple[str, int, int]]:
    """По уникальным ключам вклада: (наименование, q_line, q_global).

    Уникальные ключи сортируются по длине/ширине (см. ``_sort_contribution_keys``).
    ``max_pairs`` — необязательный лимит числа позиций; ``None`` — все уникальные ключи.
    """
    seen: set[LineContributionKey] = set()
    ordered_unique: list[LineContributionKey] = []
    for key in keys:
        if key in seen:
            continue
        seen.add(key)
        ordered_unique.append(key)

    ordered_unique = _sort_contribution_keys(ordered_unique)

    if max_pairs is not None and len(ordered_unique) > max_pairs:
        logger.warning(
            "[plates_preview] у строки %s уникальных ключей вклада — в вывод только первые %s",
            len(ordered_unique),
            max_pairs,
        )

    to_process = ordered_unique if max_pairs is None else ordered_unique[:max_pairs]

    out: list[tuple[str, int, int]] = []
    for key in to_process:
        lm, wm, lc, ldr = key
        lc_int: int | None
        if lc is not None:
            lc_int = int(round(float(lc)))
        else:
            lc_int = None
        name = make_plate_name(
            lm,
            wm,
            load_code=lc_int,
            length_dm_raw=(ldr or None) if (ldr or "").strip() else None,
        )
        q_line = qty_for_contribution_key(key, line_plate_load_details)
        q_global = qty_for_contribution_key(key, global_plate_load_details)
        out.append((name, q_line, q_global))
    return out


def preview_row_keyed_triples_for_contributions(
    keys: list[LineContributionKey],
    line_plate_load_details: dict,
    global_plate_load_details: dict,
    *,
    max_pairs: int | None = None,
) -> list[tuple[LineContributionKey, str, int, int]]:
    """По уникальным ключам вклада: (key, наименование, q_line, q_global)."""
    seen: set[LineContributionKey] = set()
    ordered_unique: list[LineContributionKey] = []
    for key in keys:
        if key in seen:
            continue
        seen.add(key)
        ordered_unique.append(key)

    ordered_unique = _sort_contribution_keys(ordered_unique)

    if max_pairs is not None and len(ordered_unique) > max_pairs:
        logger.warning(
            "[plates_preview] у строки %s уникальных ключей вклада — в вывод только первые %s",
            len(ordered_unique),
            max_pairs,
        )

    to_process = ordered_unique if max_pairs is None else ordered_unique[:max_pairs]

    out: list[tuple[LineContributionKey, str, int, int]] = []
    for key in to_process:
        lm, wm, lc, ldr = key
        lc_int: int | None
        if lc is not None:
            lc_int = int(round(float(lc)))
        else:
            lc_int = None
        name = make_plate_name(
            lm,
            wm,
            load_code=lc_int,
            length_dm_raw=(ldr or None) if (ldr or "").strip() else None,
        )
        q_line = qty_for_contribution_key(key, line_plate_load_details)
        q_global = qty_for_contribution_key(key, global_plate_load_details)
        out.append((key, name, q_line, q_global))
    return out


def name_qty_pairs_for_contributions(
    keys: list[LineContributionKey],
    plate_load_details: dict[tuple[float, float, int, str], int],
    *,
    max_pairs: int | None = None,
) -> list[tuple[str, int]]:
    """Пары (наименование, кол-во) по уникальным ключам вклада (один словарь для количества)."""
    triples = preview_row_triples_for_contributions(
        keys, plate_load_details, plate_load_details, max_pairs=max_pairs
    )
    return [(a, b) for a, b, _ in triples]


def build_plates_reconciliation_preview_xlsx(
    path: str | Path,
    *,
    plates_text: str,
    initial_user_plate_lines: list[str],
    forced_wide_line_indexes: list[int] | None = None,
) -> None:
    """
    Парсит ``plates_text``, строит лист превью с колонками A–E (см. модульный докстринг).

    Для каждой логической строки выводятся все уникальные ключи вклада (физические строки B–E);
    колонка A заполняется только в первой строке блока, далее — пусто до следующей логической строки.
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl не установлен — превью XLSX недоступно")

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    _, line_contributions, line_plate_load_details = set_plate_lists_from_text(plates_text)
    details_global = dict(cfg.PLATE_LOAD_DETAILS)

    norm_lines = split_plate_text_lines(plates_text)
    n = max(len(norm_lines), len(line_contributions), len(line_plate_load_details))
    if len(norm_lines) != len(line_contributions) or len(norm_lines) != len(line_plate_load_details):
        logger.warning(
            "[plates_preview] len(norm_lines)=%s line_contributions=%s line_plate_load_details=%s",
            len(norm_lines),
            len(line_contributions),
            len(line_plate_load_details),
        )

    nu = len(initial_user_plate_lines)
    if nu != n:
        logger.warning(
            "[plates_preview] len(initial_user_plate_lines)=%s != строк превью=%s",
            nu,
            n,
        )

    no_match_rows: list[tuple[int, str]] = []

    wide_source_lines = {
        line.strip()
        for line in initial_user_plate_lines
        if _source_line_is_wide(line)
    }
    forced_wide_idx: set[int] = {
        int(i)
        for i in (forced_wide_line_indexes or [])
        if isinstance(i, int) and i >= 0
    }

    # ── Проход 1: классификация строк и сборка блочных plate_load_details ──
    line_is_wide: list[bool] = []
    for i in range(n):
        user_cell = initial_user_plate_lines[i] if i < len(initial_user_plate_lines) else ""
        contrib = line_contributions[i] if i < len(line_contributions) else []
        source_wide = (i in forced_wide_idx) or (user_cell.strip() in wide_source_lines)
        fallback_wide = _contrib_looks_like_wide_split(contrib)
        line_is_wide.append(source_wide or fallback_wide)

    regular_details: dict[tuple, int] = {}
    wide_details: dict[tuple, int] = {}
    for i in range(n):
        line_det = line_plate_load_details[i] if i < len(line_plate_load_details) else {}
        target = wide_details if line_is_wide[i] else regular_details
        for key, qty in line_det.items():
            target[key] = target.get(key, 0) + qty

    # ── Проход 2: построение блоков с блочными КП-количествами ──
    # (col_a, name_b, qty_c, kp_name, kp_qty_int, width_m_for_sort, source_is_wide)
    row_t = tuple[str, str, int | None, str, int, float, bool]
    regular_blocks: list[list[row_t]] = []
    wide_blocks: list[list[row_t]] = []
    for i in range(n):
        user_cell = (
            initial_user_plate_lines[i]
            if i < len(initial_user_plate_lines)
            else ""
        )
        contrib = line_contributions[i] if i < len(line_contributions) else []
        line_det = line_plate_load_details[i] if i < len(line_plate_load_details) else {}
        source_is_wide = line_is_wide[i]
        block_details = wide_details if source_is_wide else regular_details
        keyed_triples = preview_row_keyed_triples_for_contributions(
            contrib, line_det, block_details
        )
        if not keyed_triples:
            no_match_rows.append((i, user_cell))
            continue

        if source_is_wide:
            keyed_triples = sorted(
                keyed_triples,
                key=lambda item: (
                    -float(item[0][1]),               # width_m desc
                    float(item[0][0]),                # length_m asc
                    (item[0][3] or "").strip(),       # length_dm_raw
                ),
            )
        block_rows: list[row_t] = []
        for idx, (key, name, q_line, q_global) in enumerate(keyed_triples):
            width_m = float(key[1])
            a_val = user_cell if idx == 0 else ""
            c_val = None if q_line <= 0 else int(q_line)
            kp_qty = int(q_global) if int(q_global) > 0 else 0
            block_rows.append((a_val, name, c_val, name, kp_qty, width_m, source_is_wide))

        block_is_wide = source_is_wide or any(_is_wide_width(row[5]) for row in block_rows)
        if block_is_wide:
            wide_blocks.append(block_rows)
        else:
            regular_blocks.append(block_rows)

    # ── Группировка Блока A: дубли рядом, схлопывание D-E ──
    regular_blocks.sort(key=lambda block: block[0][1])

    regular_physical: list[row_t] = []
    last_kp_name: str | None = None
    for block in regular_blocks:
        for row in block:
            if row[3] == last_kp_name:
                regular_physical.append((row[0], row[1], row[2], "", 0, row[5], row[6]))
            else:
                last_kp_name = row[3]
                regular_physical.append(row)

    # ── Блок B: порядок пользователя, без схлопывания ──
    wide_physical: list[row_t] = []
    for block in wide_blocks:
        wide_physical.extend(block)

    # Нераспознанные/пустые строки — после распознанных, в том же порядке блоков.
    regular_no_match: list[row_t] = []
    wide_no_match: list[row_t] = []
    for line_idx, user_cell in no_match_rows:
        source_is_wide = line_is_wide[line_idx] if 0 <= line_idx < len(line_is_wide) else (line_idx in forced_wide_idx)
        row: row_t = (user_cell, "", None, "", 0, 0.0, source_is_wide)
        if source_is_wide:
            wide_no_match.append(row)
        else:
            regular_no_match.append(row)
    regular_physical.extend(regular_no_match)
    wide_physical.extend(wide_no_match)

    physical: list[row_t] = []
    physical.extend(regular_physical)
    if regular_physical and wide_physical:
        # Явно разделяем блоки regular/wide пустой строкой.
        physical.append(("", "", None, "", 0, 0.0, False))
    physical.extend(wide_physical)

    wb = Workbook()
    ws = wb.active
    ws.title = "Превью списка"

    headers = [
        "Как прислал пользователь",
        "Распознано (наименование)",
        "Распознано (кол-во)",
        "Как в КП (наименование)",
        "КП (кол-во)",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP

    for idx, (a, b, c, kp_name, kp_qty, _width_m, _source_is_wide) in enumerate(physical):
        r = idx + 2
        ws.cell(row=r, column=1, value=a).alignment = _WRAP
        ws.cell(row=r, column=2, value=b).alignment = _WRAP
        ws.cell(row=r, column=3, value=c).alignment = _WRAP

        d_val = kp_name if kp_name else ""
        e_val = None if kp_qty <= 0 else kp_qty

        ws.cell(row=r, column=4, value=d_val).alignment = _WRAP
        ws.cell(row=r, column=5, value=e_val).alignment = _WRAP

    for col in range(1, _PREVIEW_COLS + 1):
        ws.column_dimensions[get_column_letter(col)].width = 28 if col == 1 else 22

    wb.save(str(path))

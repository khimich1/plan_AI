#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Справочник перевозчиков (SHIP-101): нормализация имён и импорт из Excel-реестра.

Импорт берёт ТОЛЬКО имя контрагента (колонка «Организация») и лист-источник;
телефоны, почта и контактные лица (ПДн) в БД не попадают.
"""

from __future__ import annotations

import re
import sqlite3
from dataclasses import dataclass, field

from openpyxl import load_workbook

CARRIER_SOURCE_SHEETS = ("Перевозчики", "Транспортные Компании")

_LEGAL_FORM_TOKENS = frozenset({"ооо", "ип", "ао", "пао", "зао", "оао"})
_QUOTE_CHARS = "«»“”\"‘’'`"
_PUNCT_RE = re.compile(r"[^\w\s]", re.UNICODE)


def normalize_carrier_name(name: str) -> str:
    """Ключ дедупликации: lower, ё→е, без кавычек/ОПФ-токенов/пунктуации, схлопнутые пробелы."""
    text = name.lower().replace("ё", "е")
    text = text.translate(str.maketrans("", "", _QUOTE_CHARS))
    text = _PUNCT_RE.sub(" ", text)
    tokens = [token for token in text.split() if token not in _LEGAL_FORM_TOKENS]
    return " ".join(tokens)


@dataclass
class CarrierImportReport:
    inserted: int = 0
    skipped_existing: int = 0
    duplicates: list[dict] = field(default_factory=list)
    per_sheet: dict[str, dict[str, int]] = field(default_factory=dict)


def import_carriers_from_xlsx(
    xlsx_path: str,
    db_path: str,
    *,
    sheets: tuple[str, ...] = CARRIER_SOURCE_SHEETS,
) -> CarrierImportReport:
    """Импорт имён контрагентов с авто-дедупом по ``name_normalized``.

    Дубликаты в БД не попадают — уходят в отчёт. Повторный запуск безопасен:
    уже существующие нормализованные имена пропускаются.
    """
    from core.kp_db_schema import ensure_schema

    ensure_schema(db_path)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    report = CarrierImportReport()
    conn = sqlite3.connect(db_path)
    try:
        cur = conn.cursor()
        seen_this_run: dict[str, str] = {}
        for sheet in sheets:
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            read = 0
            imported = 0
            for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row:
                    continue
                name = str(row[0]).strip() if row[0] is not None else ""
                if not name:
                    continue
                read += 1
                normalized = normalize_carrier_name(name)
                if not normalized:
                    continue
                if normalized in seen_this_run:
                    report.duplicates.append(
                        {
                            "name": name,
                            "sheet": sheet,
                            "row": row_no,
                            "kept": seen_this_run[normalized],
                        }
                    )
                    continue
                cur.execute(
                    "SELECT name FROM carriers WHERE name_normalized = ?",
                    (normalized,),
                )
                existing = cur.fetchone()
                if existing is not None:
                    report.skipped_existing += 1
                    seen_this_run[normalized] = str(existing[0])
                    continue
                cur.execute(
                    """
                    INSERT INTO carriers (name, name_normalized, source_sheet)
                    VALUES (?, ?, ?)
                    """,
                    (name, normalized, sheet),
                )
                seen_this_run[normalized] = name
                imported += 1
                report.inserted += 1
            report.per_sheet[sheet] = {"read": read, "imported": imported}
        conn.commit()
    finally:
        conn.close()
        wb.close()
    return report

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Модуль создания Excel-файлов формовки по дорожкам.
Использует шаблон "!КЗ ПБ Шаблон.xlsx" и заполняет его данными.
"""
import os
import shutil
from datetime import datetime
from typing import List, Dict, Optional

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[FORMOVKA] ⚠️ openpyxl не установлен. Установите: pip install openpyxl")


# Путь к шаблону по умолчанию
DEFAULT_TEMPLATE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "банк знаний",
    "!КЗ ПБ Шаблон.xlsx"
)


def create_formovka_excel(
    track_number: int,
    max_reinforcement: float,
    plates_info: List[Dict],
    output_dir: str,
    template_path: str = None,
    date_str: Optional[str] = None
) -> Optional[str]:
    """
    Создает Excel-файл формовки для дорожки на основе шаблона.
    
    Args:
        track_number: Номер дорожки
        max_reinforcement: Максимальное армирование на дорожке
        plates_info: Список словарей с информацией о плитах:
            - plate_name: название плиты (например "ПБ 80-12-10п")
            - qty: количество штук
            - length: длина плиты в метрах
            - kp_id: номер КП
            - customer: заказчик (опционально)
        output_dir: Директория для сохранения файла
        template_path: Путь к шаблону Excel (по умолчанию "банк знаний/!КЗ ПБ Шаблон.xlsx")
        date_str: Суффикс для имени файла (по умолчанию текущая дата/время)
    
    Returns:
        Путь к созданному файлу или None в случае ошибки
    """
    if not OPENPYXL_AVAILABLE:
        print("[FORMOVKA] ❌ openpyxl не установлен")
        return None
    
    # Используем шаблон по умолчанию если не указан
    if template_path is None:
        template_path = DEFAULT_TEMPLATE
    
    if not os.path.exists(template_path):
        print(f"[FORMOVKA] ❌ Шаблон не найден: {template_path}")
        return None
    
    if not plates_info:
        print(f"[FORMOVKA] ⚠️ Нет данных о плитах для дорожки {track_number}")
        return None
    
    try:
        # Создаем директорию если не существует
        os.makedirs(output_dir, exist_ok=True)
        
        # Генерируем имя выходного файла
        if date_str is None:
            date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"Формовка_Дорожка_{track_number}_{date_str}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        
        # Копируем шаблон
        shutil.copy2(template_path, output_path)
        
        # Открываем скопированный файл
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Текущая дата
        current_date = datetime.now().strftime("%d.%m.%Y")
        
        # ==================== ЗАПОЛНЯЕМ ДАННЫЕ ====================
        
        # B1 (объединено B1:B5) - Дата формовки
        ws['B1'] = current_date
        
        # D1 (объединено D1:D6) - № дорожки - НЕ ТРОГАЕМ (там заголовок)
        # Номер дорожки указывается в C10 или в другом месте? 
        # По шаблону нет отдельной ячейки для номера дорожки
        
        # B6 (объединено B6:B9) - Дата приемки ОТК (формула =B1+1, оставляем)
        # Но можно переписать если нужно
        
        # A12 - Армирование (большим шрифтом)
        ws['A12'] = max_reinforcement
        
        # ==================== ЗАПОЛНЯЕМ ПЛИТЫ (строки 13-27) ====================
        
        # Очищаем существующие данные в шаблоне (строки 13-27)
        for row in range(13, 28):
            ws.cell(row=row, column=2).value = None  # B - Заказ, №
            ws.cell(row=row, column=3).value = None  # C - Номенклатура
            ws.cell(row=row, column=4).value = None  # D - Количество
            # E - Метраж (оставляем пустым, формула будет считать)
        
        # Заполняем данные о плитах
        current_row = 13
        for plate in plates_info:
            if current_row > 27:  # Максимум 15 строк для плит
                print(f"[FORMOVKA] ⚠️ Превышен лимит строк (15), остальные плиты пропущены")
                break
            
            # B - Заказ, № (kp_id)
            kp_id_value = plate.get('kp_id', '')
            if kp_id_value is None:
                kp_id_value = ''
            cell_b = ws.cell(row=current_row, column=2)
            cell_b.value = kp_id_value
            cell_b.number_format = '0'  # Форматируем как число, чтобы не было даты
            
            # C - Номенклатура
            plate_name = plate.get('plate_name', '')
            ws.cell(row=current_row, column=3).value = plate_name
            
            # D - Количество в формовку
            qty = plate.get('qty', 0)
            ws.cell(row=current_row, column=4).value = qty
            
            # E - Метраж ПЛАН (можно оставить пустым или заполнить формулой)
            # В шаблоне колонка E для ручного ввода, G9 считает сумму
            
            current_row += 1
        
        # Сохраняем файл
        wb.save(output_path)
        wb.close()
        
        print(f"[FORMOVKA] ✅ Файл формовки создан: {output_path}")
        return output_path
        
    except Exception as e:
        print(f"[FORMOVKA] ❌ Ошибка создания файла формовки: {e}")
        import traceback
        traceback.print_exc()
        return None


def create_formovka_files_for_tracks(
    tracks_data: List[Dict],
    output_dir: str,
    start_track_number: int = 1,
    template_path: str = None,
    date_str: Optional[str] = None
) -> List[str]:
    """
    Создает Excel-файлы формовки для нескольких дорожек.
    
    Args:
        tracks_data: Список дорожек, каждая содержит:
            - track_number: номер дорожки (используется если есть)
            - max_reinforcement: максимальное армирование
            - plates_info: список плит на дорожке
        output_dir: Директория для сохранения файлов
        start_track_number: Начальный номер дорожки (если track_number не указан)
        template_path: Путь к шаблону
        date_str: Суффикс для имени файла
    
    Returns:
        Список путей к созданным файлам
    """
    created_files = []
    
    for idx, track_data in enumerate(tracks_data):
        # Используем номер дорожки из данных или вычисляем
        track_number = track_data.get('track_number', start_track_number + idx)
        max_reinforcement = track_data.get('max_reinforcement', 0)
        plates_info = track_data.get('plates_info', [])
        
        if not plates_info:
            print(f"[FORMOVKA] ⚠️ Пропускаю дорожку {track_number} - нет данных о плитах")
            continue
        
        file_path = create_formovka_excel(
            track_number=track_number,
            max_reinforcement=max_reinforcement,
            plates_info=plates_info,
            output_dir=output_dir,
            template_path=template_path,
            date_str=date_str
        )
        
        if file_path:
            created_files.append(file_path)
    
    print(f"[FORMOVKA] ✅ Создано {len(created_files)} файлов формовки")
    return created_files


# ==================== ТЕСТИРОВАНИЕ ====================

if __name__ == "__main__":
    # Тестовые данные (как на скриншоте)
    test_plates = [
        {
            'plate_name': 'Плиты ПБ 78,1-12-8п',
            'qty': 12,
            'length': 7.81,
            'kp_id': 4,
            'customer': 'Алексей'
        }
    ]
    
    # Создаем тестовый файл
    result = create_formovka_excel(
        track_number=1,
        max_reinforcement=33,
        plates_info=test_plates,
        output_dir="test_formovka"
    )
    
    if result:
        print(f"\n[OK] Тестовый файл формовки создан: {result}")
    else:
        print("\n[FAIL] Не удалось создать тестовый файл")

"""Обработчики информации о плитах ПБ в работе"""
import os
from datetime import datetime
from pathlib import Path

from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, FSInputFile, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext

# Импорты из проекта
import sys
BOT_DIR = Path(__file__).parent.parent
PROJECT_ROOT = BOT_DIR.parent
sys.path.insert(0, str(PROJECT_ROOT))

from bot.services import kp_persistence as kp_db
from ..keyboards import main_menu_kb, pb_info_kb, kp_production_details_kb
from ..bot_config import OUTPUTS_DIR_STR
from ..states import PBInfoStates

router = Router()


# Маппинг колонок для плит в производстве (kp_plates)
PRODUCTION_COLUMNS = {
    'id': '№ записи',
    'kp_id': '№ КП',
    'position_number': '№ позиции',
    'plate_name': 'Наименование плиты',
    'length_m': 'Длина (м)',
    'width_m': 'Ширина (м)',
    'load_class': 'Класс нагрузки',
    'qty': 'Количество',
    'unit_weight': 'Вес единицы (кг)',
    'total_weight': 'Общий вес (кг)',
    'discounted_price': 'Цена со скидкой',
    'customer_name': 'Заказчик',
    'execution_terms': 'Срок выполнения',
    'plate_status': 'Статус производства'
}

# Маппинг колонок для выполненных плит (completed_plates)
COMPLETED_COLUMNS = {
    'id': '№ записи',
    'kp_id': '№ КП',
    'plate_name': 'Наименование плиты',
    'length_m': 'Длина (м)',
    'width_m': 'Ширина (м)',
    'load_class': 'Класс нагрузки',
    'qty': 'Количество',
    'completed_date': 'Дата выполнения',
    'production_day': 'День производства',
    'customer_name': 'Заказчик',
    'execution_terms': 'Срок выполнения'
}


def generate_excel_from_data(data: list, columns_map: dict, filename: str) -> str:
    """
    Генерирует Excel файл из данных.
    
    Args:
        data: список словарей с данными
        columns_map: маппинг английских названий колонок на русские
        filename: имя файла (без расширения)
    
    Returns:
        Путь к созданному файлу
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise ImportError("Для работы требуется библиотека openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Плиты"
    
    # Стили
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Определяем колонки на основе маппинга
    columns = list(columns_map.keys())
    headers = list(columns_map.values())
    
    # Записываем заголовки
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_alignment
    
    # Записываем данные
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_alignment
    
    # Автоширина колонок
    for col_idx, col_name in enumerate(columns, 1):
        max_length = len(headers[col_idx - 1])
        for row in ws.iter_rows(min_row=2, max_row=len(data) + 1, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
    
    # Сохраняем файл
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(OUTPUTS_DIR_STR, f"{filename}_{timestamp}.xlsx")
    
    # Создаём папку если её нет
    os.makedirs(OUTPUTS_DIR_STR, exist_ok=True)
    
    wb.save(filepath)
    return filepath


@router.message(F.text == "Информация о ПБ в работе")
async def btn_pb_info(message: Message):
    """Обработчик кнопки 'Информация о ПБ в работе'"""
    await message.answer(
        "📊 Информация о плитах ПБ\n\n"
        "Выберите тип отчёта:",
        reply_markup=pb_info_kb()
    )


@router.callback_query(F.data == "plates_in_production")
async def export_plates_in_production(callback: CallbackQuery):
    """Экспорт плит в производстве в Excel"""
    await callback.message.answer("⏳ Загружаю данные о плитах в производстве...")
    
    try:
        # Получаем данные из БД
        db_path = PROJECT_ROOT / "plita.db"
        plates = kp_db.get_all_plates_in_production(str(db_path))
        
        if not plates:
            await callback.message.answer(
                "📭 Нет плит в производстве.\n\n"
                "Все заказы выполнены или нет КП в работе.",
                reply_markup=main_menu_kb()
            )
            await callback.answer()
            return
        
        # Генерируем Excel
        filepath = generate_excel_from_data(
            data=plates,
            columns_map=PRODUCTION_COLUMNS,
            filename="Плиты_в_производстве"
        )
        
        # Подсчитываем статистику
        total_qty = sum(p.get('qty', 0) for p in plates)
        unique_kp = len(set(p.get('kp_id') for p in plates))
        
        # Отправляем файл
        await callback.message.answer_document(
            FSInputFile(filepath),
            caption=(
                f"🏭 Плиты в производстве\n\n"
                f"📊 Статистика:\n"
                f"• Всего записей: {len(plates)}\n"
                f"• Общее кол-во плит: {total_qty}\n"
                f"• КП в работе: {unique_kp}"
            )
        )
        
        await callback.message.answer(
            "✅ Файл готов!",
            reply_markup=main_menu_kb()
        )
        
    except Exception as e:
        await callback.message.answer(
            f"❌ Ошибка при экспорте: {str(e)}",
            reply_markup=main_menu_kb()
        )
    
    await callback.answer()


@router.callback_query(F.data == "completed_plates_export")
async def export_completed_plates(callback: CallbackQuery):
    """Экспорт выполненных плит в Excel"""
    await callback.message.answer("⏳ Загружаю данные о выполненных плитах...")
    
    try:
        # Получаем данные из БД
        db_path = PROJECT_ROOT / "plita.db"
        plates = kp_db.get_all_completed_plates(str(db_path))
        
        if not plates:
            await callback.message.answer(
                "📭 Нет выполненных плит.\n\n"
                "Пока не было завершено ни одного дня производства.",
                reply_markup=main_menu_kb()
            )
            await callback.answer()
            return
        
        # Генерируем Excel
        filepath = generate_excel_from_data(
            data=plates,
            columns_map=COMPLETED_COLUMNS,
            filename="Выполненные_плиты"
        )
        
        # Подсчитываем статистику
        total_qty = sum(p.get('qty', 0) for p in plates)
        unique_kp = len(set(p.get('kp_id') for p in plates))
        unique_days = len(set(p.get('production_day') for p in plates if p.get('production_day')))
        
        # Отправляем файл
        await callback.message.answer_document(
            FSInputFile(filepath),
            caption=(
                f"✅ Выполненные плиты\n\n"
                f"📊 Статистика:\n"
                f"• Всего записей: {len(plates)}\n"
                f"• Общее кол-во плит: {total_qty}\n"
                f"• КП затронуто: {unique_kp}\n"
                f"• Дней производства: {unique_days}"
            )
        )
        
        await callback.message.answer(
            "✅ Файл готов!",
            reply_markup=main_menu_kb()
        )
        
    except Exception as e:
        await callback.message.answer(
            f"❌ Ошибка при экспорте: {str(e)}",
            reply_markup=main_menu_kb()
        )
    
    await callback.answer()


# ==================== ОБРАБОТЧИКИ ДЛЯ КП В ПРОИЗВОДСТВЕ ====================

@router.callback_query(F.data == "kp_in_production")
async def show_kp_in_production_list(callback: CallbackQuery):
    """
    Обработчик кнопки "КП в производстве".
    Показывает список всех КП в работе с процентом выполнения.
    """
    try:
        await callback.answer()
    except:
        pass
    
    # Получаем список КП со статусом "в работе"
    all_kp = kp_db.get_all_kp_list()
    production_kp = all_kp.get('in_production', [])
    
    if not production_kp:
        await callback.message.edit_text(
            "🏭 В производстве пока нет КП\n\n"
            "Чтобы отправить КП в производство, создайте его через '📝 Создать КП' "
            "и нажмите '💾 Сохранить в БД' после генерации документов.",
            reply_markup=pb_info_kb()
        )
        return
    
    # Формируем текст с информацией о КП
    text = f"🏭 КП в производстве ({len(production_kp)} шт.)\n\n"
    text += "Нажмите на КП для просмотра деталей:"
    
    # Создаём inline кнопки для каждого КП с процентом выполнения
    buttons = []
    db_path = PROJECT_ROOT / "plita.db"
    
    for kp in production_kp:
        kp_id = kp['kp_id']
        customer = kp.get('customer_name', 'Без имени')
        total = kp.get('total_amount', 0)
        execution_terms = kp.get('execution_terms', '')
        
        # Получаем процент выполнения
        completion_info = kp_db.get_kp_completion_percentage(kp_id, str(db_path))
        percentage = completion_info['percentage']
        
        # Обрезаем длинные имена клиентов
        customer_short = customer[:20] + '...' if len(customer) > 20 else customer
        
        # Формируем текст кнопки с процентом
        button_text = f"КП №{kp_id} | {customer_short} | {percentage:.0f}%"
        if execution_terms:
            button_text += f" | ⏰{execution_terms}"
        
        buttons.append([
            InlineKeyboardButton(
                text=button_text,
                callback_data=f"view_prod_kp_{kp_id}"
            )
        ])
    
    # Кнопка "Назад"
    buttons.append([
        InlineKeyboardButton(text="◀️ Назад", callback_data="pb_info_back")
    ])
    
    await callback.message.edit_text(
        text,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )


@router.callback_query(F.data.startswith("view_prod_kp_"))
async def view_kp_production_details(callback: CallbackQuery, state: FSMContext):
    """
    Обработчик просмотра детальной информации о КП в производстве.
    Показывает детальную информацию о КП с процентом выполнения.
    """
    await callback.answer()
    
    # Извлекаем kp_id из callback_data
    kp_id = int(callback.data.split("_")[-1])
    
    # Получаем информацию о КП из БД
    db_path = PROJECT_ROOT / "plita.db"
    kp_info = kp_db.get_kp_by_id(kp_id, str(db_path))
    
    if not kp_info:
        await callback.message.edit_text(
            f"❌ КП №{kp_id} не найдено в базе данных",
            reply_markup=pb_info_kb()
        )
        return
    
    # Проверяем статус КП
    status = kp_info.get('status', 'в работе')
    if status != 'в работе':
        await callback.message.edit_text(
            f"⚠️ КП №{kp_id} имеет статус '{status}'\n\n"
            f"Информация доступна только для КП в статусе 'в работе'.",
            reply_markup=pb_info_kb()
        )
        return
    
    # Получаем процент выполнения
    completion_info = kp_db.get_kp_completion_percentage(kp_id, str(db_path))
    
    # Формируем детальное описание
    customer = kp_info.get('customer_name', 'Не указан')
    total = kp_info.get('total_amount', 0)
    execution_date = kp_info.get('execution_terms', 'Не указан')
    
    text = f"📊 КП №{kp_id} | {total:,.0f}₽ | 🎯 {execution_date}\n\n"
    text += f"👤 Менеджер: {kp_info.get('manager_name', 'Не указан')}\n"
    text += f"📅 Дата создания: {kp_info.get('creation_date', 'Не указана')}\n"
    text += f"🏭 Статус: 🏭 {status}\n"
    text += f"⏰ Срок выполнения: {execution_date}\n\n"
    
    text += f"💰 Финансы:\n"
    text += f"  • Сумма без НДС: {kp_info.get('subtotal', 0):,.2f} ₽\n"
    text += f"  • НДС (22%): {kp_info.get('vat_amount', 0):,.2f} ₽\n"
    text += f"  • Итого с НДС: {total:,.2f} ₽\n"
    
    if kp_info.get('discount_percent', 0) > 0:
        text += f"  • Скидка: {kp_info['discount_percent']}%\n"
    
    text += f"\n📦 Состав заказа ({len(kp_info.get('plates', []))} позиций):\n"
    
    # Список плит (ограничиваем до 10 позиций)
    plates = kp_info.get('plates', [])
    for i, plate in enumerate(plates[:10], 1):
        text += f"  {i}. {plate['plate_name']} — {plate['qty']} шт\n"
    
    if len(plates) > 10:
        text += f"  ... и ещё {len(plates) - 10} позиций\n"
    
    # ПРОЦЕНТ ВЫПОЛНЕНИЯ
    text += f"\n🎯 Выполнено: {completion_info['completed_plates']} из {completion_info['total_plates']} плит "
    text += f"({completion_info['percentage']:.0f}%)\n"
    
    await callback.message.edit_text(
        text,
        reply_markup=kp_production_details_kb(kp_id)
    )


@router.callback_query(F.data.startswith("change_date_"))
async def change_kp_date_request(callback: CallbackQuery, state: FSMContext):
    """
    Обработчик кнопки "Изменить дату".
    Запрашивает новую дату выполнения для КП.
    """
    await callback.answer()
    
    kp_id = int(callback.data.split("_")[-1])
    
    # Получаем текущую дату
    db_path = PROJECT_ROOT / "plita.db"
    kp_info = kp_db.get_kp_by_id(kp_id, str(db_path))
    
    if not kp_info:
        await callback.message.answer(
            f"❌ КП №{kp_id} не найдено",
            reply_markup=main_menu_kb()
        )
        return
    
    current_date = kp_info.get('execution_terms', 'Не указан')
    
    await callback.message.answer(
        f"📅 Изменение срока выполнения для КП №{kp_id}\n\n"
        f"Текущий срок: {current_date}\n\n"
        f"Введите новую дату в формате:\n"
        f"• 25.03.2026 (ДД.ММ.ГГГГ)\n"
        f"• 2026-03-25 (ГГГГ-ММ-ДД)",
        reply_markup=main_menu_kb()
    )
    
    # Сохраняем kp_id в state
    await state.update_data(kp_id=kp_id, old_date=current_date)
    await state.set_state(PBInfoStates.waiting_new_date)


@router.message(PBInfoStates.waiting_new_date)
async def receive_new_date(message: Message, state: FSMContext):
    """
    Обработчик получения новой даты.
    Обновляет дату выполнения в БД.
    """
    user_input = message.text.strip()
    data = await state.get_data()
    kp_id = data.get('kp_id')
    old_date = data.get('old_date', 'неизвестно')
    
    if not kp_id:
        await message.answer(
            "❌ Ошибка: не найден номер КП",
            reply_markup=main_menu_kb()
        )
        await state.clear()
        return
    
    # Парсим дату (поддерживаем два формата)
    new_date = None
    date_formatted = None
    
    # Формат 1: ДД.ММ.ГГГГ
    try:
        parsed = datetime.strptime(user_input, '%d.%m.%Y')
        new_date = parsed.strftime('%d.%m.%Y')
        date_formatted = new_date
    except ValueError:
        pass
    
    # Формат 2: ГГГГ-ММ-ДД
    if not new_date:
        try:
            parsed = datetime.strptime(user_input, '%Y-%m-%d')
            new_date = parsed.strftime('%d.%m.%Y')
            date_formatted = new_date
        except ValueError:
            pass
    
    if not new_date:
        await message.answer(
            "❌ Неверный формат даты.\n\n"
            "Поддерживаемые форматы:\n"
            "• 25.03.2026 (ДД.ММ.ГГГГ)\n"
            "• 2026-03-25 (ГГГГ-ММ-ДД)\n\n"
            "Попробуйте снова:"
        )
        return
    
    # Обновляем дату в БД
    db_path = PROJECT_ROOT / "plita.db"
    success = kp_db.update_kp_execution_date(kp_id, new_date, str(db_path))
    
    if success:
        await message.answer(
            f"✅ Дата выполнения обновлена!\n\n"
            f"КП №{kp_id}: {old_date} → {date_formatted}\n\n"
            f"Обновлены все плиты в этом КП.",
            reply_markup=main_menu_kb()
        )
        
        # Показываем обновлённую карточку КП
        kp_info = kp_db.get_kp_by_id(kp_id, str(db_path))
        if kp_info:
            completion_info = kp_db.get_kp_completion_percentage(kp_id, str(db_path))
            
            customer = kp_info.get('customer_name', 'Не указан')
            total = kp_info.get('total_amount', 0)
            execution_date = kp_info.get('execution_terms', 'Не указан')
            status = kp_info.get('status', 'в работе')
            
            text = f"📊 КП №{kp_id} | {total:,.0f}₽ | 🎯 {execution_date}\n\n"
            text += f"👤 Менеджер: {kp_info.get('manager_name', 'Не указан')}\n"
            text += f"📅 Дата создания: {kp_info.get('creation_date', 'Не указана')}\n"
            text += f"🏭 Статус: 🏭 {status}\n"
            text += f"⏰ Срок выполнения: {execution_date}\n\n"
            
            text += f"💰 Финансы:\n"
            text += f"  • Сумма без НДС: {kp_info.get('subtotal', 0):,.2f} ₽\n"
            text += f"  • НДС (22%): {kp_info.get('vat_amount', 0):,.2f} ₽\n"
            text += f"  • Итого с НДС: {total:,.2f} ₽\n"
            
            if kp_info.get('discount_percent', 0) > 0:
                text += f"  • Скидка: {kp_info['discount_percent']}%\n"
            
            text += f"\n📦 Состав заказа ({len(kp_info.get('plates', []))} позиций):\n"
            
            plates = kp_info.get('plates', [])
            for i, plate in enumerate(plates[:10], 1):
                text += f"  {i}. {plate['plate_name']} — {plate['qty']} шт\n"
            
            if len(plates) > 10:
                text += f"  ... и ещё {len(plates) - 10} позиций\n"
            
            text += f"\n🎯 Выполнено: {completion_info['completed_plates']} из {completion_info['total_plates']} плит "
            text += f"({completion_info['percentage']:.0f}%)\n"
            
            await message.answer(
                text,
                reply_markup=kp_production_details_kb(kp_id)
            )
    else:
        await message.answer(
            f"❌ Ошибка при обновлении даты для КП №{kp_id}",
            reply_markup=main_menu_kb()
        )
    
    await state.clear()


@router.callback_query(F.data == "pb_info_back")
async def back_to_pb_info_menu(callback: CallbackQuery):
    """Возврат в меню информации о ПБ"""
    try:
        await callback.answer()
    except:
        pass
    
    await callback.message.edit_text(
        "📊 Информация о плитах ПБ\n\n"
        "Выберите тип отчёта:",
        reply_markup=pb_info_kb()
    )


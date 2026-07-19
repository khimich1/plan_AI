"""Обработчики административных команд (удаление КП и т.д.)"""
import sys
import os
import json
import shutil
from pathlib import Path

from aiogram import Router, F
from aiogram.types import Message, CallbackQuery
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext

# Добавляем корень проекта в sys.path
BOT_DIR = Path(__file__).parent.parent
PROJECT_ROOT = BOT_DIR.parent
sys.path.insert(0, str(PROJECT_ROOT))

from bot.services import kp_persistence as kp_db
from core.destructive_db_guard import DestructiveDbOperationBlocked
from bot.security.audit import log_bot_security_event
from bot.security.users import BotUser, has_role
from ..keyboards import main_menu_kb, db_management_kb, db_clear_confirm_kb
from ..states import AdminStates

router = Router()

_ADMIN_ONLY_MSG = "⛔ Только администратор может выполнить это действие."


async def _ensure_admin_message(message: Message, bot_user: BotUser, *, action: str) -> bool:
    if has_role(bot_user, "admin"):
        return True
    log_bot_security_event(
        "access_denied",
        telegram_id=bot_user.telegram_id,
        role=bot_user.role,
        action=action,
    )
    await message.answer(_ADMIN_ONLY_MSG, reply_markup=main_menu_kb(bot_user.role))
    return False


async def _ensure_admin_callback(callback: CallbackQuery, bot_user: BotUser, *, action: str) -> bool:
    if has_role(bot_user, "admin"):
        return True
    log_bot_security_event(
        "access_denied",
        telegram_id=bot_user.telegram_id,
        role=bot_user.role,
        action=action,
    )
    if callback.message:
        await callback.message.answer(_ADMIN_ONLY_MSG, reply_markup=main_menu_kb(bot_user.role))
    await callback.answer("Недостаточно прав", show_alert=True)
    return False


@router.message(Command("delete_kp"))
async def cmd_delete_kp(message: Message, state: FSMContext, bot_user: BotUser):
    """
    Удаляет КП из базы данных по номеру.
    Использование: /delete_kp 5
    """
    if not await _ensure_admin_message(message, bot_user, action="delete_kp"):
        return
    args = message.text.split()
    
    if len(args) != 2:
        await message.answer(
            "❌ Неверный формат команды!\n\n"
            "Использование: /delete_kp <номер_КП>\n"
            "Пример: /delete_kp 5",
            reply_markup=main_menu_kb(bot_user.role)
        )
        return
    
    try:
        kp_id = int(args[1])
    except ValueError:
        await message.answer(
            "❌ Номер КП должен быть числом!",
            reply_markup=main_menu_kb(bot_user.role)
        )
        return
    
    # Получаем информацию о КП перед удалением (для подтверждения)
    kp_info = kp_db.get_kp_by_id(kp_id)
    
    if not kp_info:
        await message.answer(
            f"❌ КП #{kp_id} не найдено в базе данных.",
            reply_markup=main_menu_kb(bot_user.role)
        )
        return
    
    # Сохраняем kp_id в состояние для подтверждения
    await state.update_data(delete_kp_id=kp_id)
    await state.set_state(AdminStates.waiting_delete_confirmation)
    
    # Показываем информацию о КП
    plates_count = len(kp_info.get('plates', []))
    total_amount = kp_info.get('total_amount', 0)
    
    await message.answer(
        f"⚠️ **ВНИМАНИЕ!** Вы собираетесь удалить КП #{kp_id}\n\n"
        f"📋 Информация:\n"
        f"  • Дата: {kp_info.get('creation_date', 'Не указана')}\n"
        f"  • Клиент: {kp_info.get('customer_name', 'Не указан')}\n"
        f"  • Менеджер: {kp_info.get('manager_name', 'Не указан')}\n"
        f"  • Сумма: {total_amount:,.2f} ₽ (с НДС)\n"
        f"  • Плит: {plates_count} шт\n"
        f"  • Статус: {kp_info.get('status', 'не указан')}\n\n"
        f"⚠️ **Это действие необратимо!**\n"
        f"Все данные (плиты, файлы, метаданные) будут удалены.\n\n"
        f"Для подтверждения отправьте: **ДА**\n"
        f"Для отмены отправьте: **НЕТ**",
        parse_mode="Markdown"
    )


@router.message(AdminStates.waiting_delete_confirmation)
async def confirm_delete_kp(message: Message, state: FSMContext, bot_user: BotUser):
    """Подтверждение удаления КП"""
    if not await _ensure_admin_message(message, bot_user, action="confirm_delete_kp"):
        return
    confirmation = message.text.strip().upper()
    
    if confirmation not in ['ДА', 'YES', 'Y']:
        await state.clear()
        await message.answer(
            "❌ Удаление отменено.",
            reply_markup=main_menu_kb(bot_user.role)
        )
        return
    
    # Получаем kp_id из состояния
    data = await state.get_data()
    kp_id = data.get('delete_kp_id')
    
    if not kp_id:
        await state.clear()
        await message.answer(
            "❌ Ошибка: не удалось получить номер КП.",
            reply_markup=main_menu_kb(bot_user.role)
        )
        return
    
    # Удаляем КП
    success = kp_db.delete_kp_by_id(kp_id)
    
    await state.clear()
    
    if success:
        await message.answer(
            f"✅ **КП #{kp_id} успешно удалено из базы данных!**\n\n"
            f"Все связанные данные (плиты, файлы, метаданные) также удалены.",
            parse_mode="Markdown",
            reply_markup=main_menu_kb(bot_user.role)
        )
    else:
        await message.answer(
            f"❌ Не удалось удалить КП #{kp_id}.\n"
            f"Возможно, оно уже было удалено ранее.",
            reply_markup=main_menu_kb(bot_user.role)
        )


@router.message(Command("list_kp"))
async def cmd_list_kp(message: Message, bot_user: BotUser):
    """
    Показывает список всех КП в базе данных.
    Использование: /list_kp
    """
    if not await _ensure_admin_message(message, bot_user, action="list_kp"):
        return
    try:
        # Получаем все КП со статусом "в работе"
        kp_list_active = kp_db.get_all_kp_by_status('в работе')
        kp_list_done = kp_db.get_all_kp_by_status('выполнено')
        kp_list_rejected = kp_db.get_all_kp_by_status('отклонено')
        
        response = "📋 **Список КП в базе данных:**\n\n"
        
        if kp_list_active:
            response += "🟢 **В работе:**\n"
            for kp in kp_list_active:
                response += (
                    f"  • КП #{kp['kp_id']} - {kp.get('customer_name', 'Без имени')} - "
                    f"{kp.get('total_amount', 0):,.2f} ₽\n"
                )
            response += "\n"
        
        if kp_list_done:
            response += "✅ **Выполнено:**\n"
            for kp in kp_list_done:
                response += (
                    f"  • КП #{kp['kp_id']} - {kp.get('customer_name', 'Без имени')} - "
                    f"{kp.get('total_amount', 0):,.2f} ₽\n"
                )
            response += "\n"
        
        if kp_list_rejected:
            response += "❌ **Отклонено:**\n"
            for kp in kp_list_rejected:
                response += (
                    f"  • КП #{kp['kp_id']} - {kp.get('customer_name', 'Без имени')} - "
                    f"{kp.get('total_amount', 0):,.2f} ₽\n"
                )
            response += "\n"
        
        if not (kp_list_active or kp_list_done or kp_list_rejected):
            response += "📭 База данных пуста.\n\n"
        
        response += "💡 Для удаления КП используйте: /delete_kp <номер>"
        
        await message.answer(
            response, parse_mode="Markdown", reply_markup=main_menu_kb(bot_user.role)
        )
    
    except Exception as e:
        await message.answer(
            f"❌ Ошибка при получении списка КП: {str(e)}",
            reply_markup=main_menu_kb(bot_user.role)
        )


@router.message(Command("clear_all_kp"))
async def cmd_clear_all_kp(message: Message, state: FSMContext, bot_user: BotUser):
    """
    Полностью очищает все КП из базы данных.
    Использование: /clear_all_kp
    """
    if not await _ensure_admin_message(message, bot_user, action="clear_all_kp"):
        return
    log_bot_security_event(
        "destructive_started",
        telegram_id=bot_user.telegram_id,
        role=bot_user.role,
        action="clear_all_kp",
    )
    try:
        # Подсчитываем количество КП перед удалением
        kp_list_active = kp_db.get_all_kp_by_status('в работе')
        kp_list_done = kp_db.get_all_kp_by_status('выполнено')
        kp_list_rejected = kp_db.get_all_kp_by_status('отклонено')
        kp_list_waiting = kp_db.get_all_kp_by_status('в ожидании')
        
        total_count = len(kp_list_active) + len(kp_list_done) + len(kp_list_rejected) + len(kp_list_waiting)
        
        if total_count == 0:
            await message.answer(
                "ℹ️ База данных уже пуста. Нечего удалять.",
                reply_markup=main_menu_kb(bot_user.role)
            )
            return
        
        # Сохраняем информацию в состояние
        await state.update_data(clear_all_count=total_count)
        await state.set_state(AdminStates.waiting_clear_all_confirmation)
        
        # Показываем предупреждение
        await message.answer(
            f"⚠️ **КРИТИЧЕСКОЕ ВНИМАНИЕ!**\n\n"
            f"Вы собираетесь **ПОЛНОСТЬЮ ОЧИСТИТЬ** всю базу данных с КП!\n\n"
            f"📊 Будет удалено:\n"
            f"  • КП в работе: {len(kp_list_active)}\n"
            f"  • КП выполнено: {len(kp_list_done)}\n"
            f"  • КП отклонено: {len(kp_list_rejected)}\n"
            f"  • КП в ожидании: {len(kp_list_waiting)}\n"
            f"  • **ВСЕГО: {total_count} КП**\n\n"
            f"🗑️ Будут удалены:\n"
            f"  • Все записи о КП\n"
            f"  • Все плиты в КП\n"
            f"  • Все файлы XLSX\n"
            f"  • Все метаданные\n\n"
            f"⚠️ **ЭТО ДЕЙСТВИЕ НЕОБРАТИМО!**\n"
            f"После очистки база данных будет полностью пуста.\n\n"
            f"Для подтверждения отправьте: **ДА**\n"
            f"Для отмены отправьте: **НЕТ**",
            parse_mode="Markdown"
        )
    
    except Exception as e:
        await message.answer(
            f"❌ Ошибка при проверке базы данных: {str(e)}",
            reply_markup=main_menu_kb(bot_user.role)
        )


@router.message(AdminStates.waiting_clear_all_confirmation)
async def confirm_clear_all_kp(message: Message, state: FSMContext, bot_user: BotUser):
    """Подтверждение полной очистки БД"""
    if not await _ensure_admin_message(message, bot_user, action="confirm_clear_all_kp"):
        return
    confirmation = message.text.strip().upper()
    
    if confirmation not in ['ДА', 'YES', 'Y']:
        await state.clear()
        await message.answer(
            "❌ Очистка отменена.",
            reply_markup=main_menu_kb(bot_user.role)
        )
        return
    
    # Получаем количество из состояния
    data = await state.get_data()
    expected_count = data.get('clear_all_count', 0)
    
    try:
        result = kp_db.clear_all_kp()
        await state.clear()
        log_bot_security_event(
            "destructive_completed",
            telegram_id=bot_user.telegram_id,
            role=bot_user.role,
            action="clear_all_kp",
            kp_offers=result.get("kp_offers", 0),
            kp_plates=result.get("kp_plates", 0),
        )
        await message.answer(
            f"✅ **База данных полностью очищена!**\n\n"
            f"📊 Удалено:\n"
            f"  • КП: {result.get('kp_offers', 0)}\n"
            f"  • Записей плит: {result.get('kp_plates', 0)}\n"
            f"  • Файлов: {result.get('kp_files', 0)}\n"
            f"  • Метаданных: {result.get('kp_meta', 0)}\n\n"
            f"🔄 Счётчики AUTOINCREMENT сброшены.\n"
            f"Новые КП будут начинаться с номера 1.",
            parse_mode="Markdown",
            reply_markup=main_menu_kb(bot_user.role)
        )
    except DestructiveDbOperationBlocked as exc:
        await state.clear()
        await message.answer(
            f"⛔ {exc}",
            reply_markup=main_menu_kb(bot_user.role),
        )
    except Exception as e:
        await state.clear()
        await message.answer(
            f"❌ Ошибка при очистке базы данных: {str(e)}\n\n"
            f"Попробуйте снова позже.",
            reply_markup=main_menu_kb(bot_user.role)
        )
        import logging
        logging.getLogger(__name__).exception(f"Ошибка при очистке базы данных: {e}")


# ==================== НОВОЕ МЕНЮ УПРАВЛЕНИЯ БД ====================

@router.message(F.text == "⚙️ Управление БД")
async def btn_db_management(message: Message, bot_user: BotUser):
    """Обработчик кнопки 'Управление БД' из главного меню"""
    if not await _ensure_admin_message(message, bot_user, action="db_management"):
        return
    await message.answer(
        "⚙️ **Управление базой данных**\n\n"
        "Выберите действие:",
        parse_mode="Markdown",
        reply_markup=db_management_kb(bot_user.role),
    )


@router.callback_query(F.data == "db_stats")
async def show_db_stats(callback: CallbackQuery, bot_user: BotUser):
    """Показывает статистику БД"""
    if not await _ensure_admin_callback(callback, bot_user, action="db_stats"):
        return
    try:
        stats = kp_db.get_db_stats()
        
        response = "📊 **Статистика базы данных**\n\n"
        response += f"📋 **КП:**\n"
        response += f"  • Всего: {stats['kp_total']}\n"
        response += f"  • В работе: {stats['kp_in_work']}\n"
        response += f"  • Выполнено: {stats['kp_completed']}\n\n"
        response += f"📦 **Плиты:**\n"
        response += f"  • В работе: {stats['plates_in_work']}\n"
        response += f"  • Выполнено: {stats['plates_completed']}\n\n"
        response += f"🔧 **Остатки от резки:**\n"
        response += f"  • Всего: {stats['plate_rests']}\n"
        
        await callback.message.answer(
            response,
            parse_mode="Markdown",
            reply_markup=db_management_kb(bot_user.role)
        )
        await callback.answer()
    
    except Exception as e:
        await callback.message.answer(
            f"❌ Ошибка при получении статистики: {str(e)}",
            reply_markup=db_management_kb(bot_user.role)
        )
        await callback.answer()


@router.callback_query(F.data == "db_clear_all")
async def request_db_clear(callback: CallbackQuery, bot_user: BotUser):
    """Запрос подтверждения полной очистки БД"""
    if not await _ensure_admin_callback(callback, bot_user, action="db_clear_all"):
        return
    log_bot_security_event(
        "destructive_started",
        telegram_id=bot_user.telegram_id,
        role=bot_user.role,
        action="db_clear_all",
    )
    try:
        stats = kp_db.get_db_stats()
        
        total = (stats['kp_total'] + stats['plates_in_work'] + 
                stats['plates_completed'] + stats['plate_rests'])
        
        if total == 0:
            await callback.message.answer(
                "ℹ️ База данных уже пуста. Нечего удалять.",
                reply_markup=db_management_kb(bot_user.role)
            )
            await callback.answer()
            return
        
        warning = (
            "🔴 **КРИТИЧЕСКОЕ ПРЕДУПРЕЖДЕНИЕ!**\n\n"
            "Вы собираетесь **ПОЛНОСТЬЮ УДАЛИТЬ ВСЕ ДАННЫЕ** из базы и планов!\n\n"
            f"📊 Будет удалено из БД:\n"
            f"  • КП: {stats['kp_total']} шт\n"
            f"  • Плиты в работе: {stats['plates_in_work']} шт\n"
            f"  • Выполненные плиты: {stats['plates_completed']} шт\n"
            f"  • Остатки: {stats['plate_rests']} шт\n"
            f"  • **ВСЕГО ЗАПИСЕЙ: {total}**\n\n"
            f"📅 Также будут удалены:\n"
            f"  • Текущий план производства\n"
            f"  • Весь календарь планов\n"
            f"  • Архив всех планов\n\n"
            f"⚠️ **ЭТО ДЕЙСТВИЕ НЕОБРАТИМО!**\n"
            f"Восстановить данные будет невозможно!\n\n"
            f"Вы уверены, что хотите продолжить?"
        )
        
        await callback.message.answer(
            warning,
            parse_mode="Markdown",
            reply_markup=db_clear_confirm_kb()
        )
        await callback.answer()
    
    except Exception as e:
        await callback.message.answer(
            f"❌ Ошибка при проверке БД: {str(e)}",
            reply_markup=db_management_kb(bot_user.role)
        )
        await callback.answer()


def clear_all_plans_data():
    """
    Очищает все данные планов производства (JSON файлы).
    
    Удаляет:
    - current_plan.json (текущий план)
    - plans_metadata.json (метаданные планов)
    - Все файлы планов в папке plans/
    
    Возвращает словарь с количеством удалённых файлов.
    """
    result = {
        'current_plan': 0,
        'metadata': 0,
        'plan_files': 0,
        'total': 0
    }
    
    try:
        # Путь к папке data
        data_dir = BOT_DIR / 'data'
        
        # Удаляем current_plan.json
        current_plan_file = data_dir / 'current_plan.json'
        if current_plan_file.exists():
            current_plan_file.unlink()
            result['current_plan'] = 1
        
        # Удаляем plans_metadata.json
        metadata_file = data_dir / 'plans_metadata.json'
        if metadata_file.exists():
            metadata_file.unlink()
            result['metadata'] = 1
        
        # Удаляем папку plans/ со всеми файлами
        plans_dir = data_dir / 'plans'
        if plans_dir.exists() and plans_dir.is_dir():
            # Подсчитываем файлы перед удалением
            plan_files = list(plans_dir.glob('*.json'))
            result['plan_files'] = len(plan_files)
            
            # Удаляем всю папку
            shutil.rmtree(plans_dir)
        
        result['total'] = result['current_plan'] + result['metadata'] + result['plan_files']
        
    except Exception as e:
        import logging
        logging.getLogger(__name__).exception(f"Ошибка при очистке планов: {e}")
    
    return result


@router.callback_query(F.data == "db_clear_confirmed")
async def confirm_db_clear(callback: CallbackQuery, bot_user: BotUser):
    """Подтверждение и выполнение полной очистки БД и планов"""
    if not await _ensure_admin_callback(callback, bot_user, action="db_clear_confirmed"):
        return
    log_bot_security_event(
        "destructive_started",
        telegram_id=bot_user.telegram_id,
        role=bot_user.role,
        action="db_clear_confirmed",
    )
    try:
        result_db = kp_db.clear_all_plates_data()
        result_plans = clear_all_plans_data()
        log_bot_security_event(
            "destructive_completed",
            telegram_id=bot_user.telegram_id,
            role=bot_user.role,
            action="db_clear_confirmed",
            db_total=result_db.get("total", 0),
            plan_files=result_plans.get("plan_files", 0),
        )
        
        report = (
            "✅ **БАЗА ДАННЫХ И ПЛАНЫ ПОЛНОСТЬЮ ОЧИЩЕНЫ!**\n\n"
            f"📊 Удалено из БД:\n"
            f"  • КП: {result_db['kp_offers']}\n"
            f"  • Плиты в работе: {result_db['kp_plates']}\n"
            f"  • Выполненные плиты: {result_db['completed_plates']}\n"
            f"  • Остатки: {result_db['plate_rests']}\n"
            f"  • Файлы: {result_db['kp_files']}\n"
            f"  • Метаданные: {result_db['kp_meta']}\n"
            f"  • Журнал статусов плит: {result_db['plate_status_log']}\n"
            f"  • **Итого записей БД: {result_db['total']}**\n\n"
            f"📅 Удалено файлов планов:\n"
            f"  • Текущий план: {result_plans['current_plan']}\n"
            f"  • Метаданные планов: {result_plans['metadata']}\n"
            f"  • Архивных планов: {result_plans['plan_files']}\n"
            f"  • **Итого файлов планов: {result_plans['total']}**\n\n"
            f"🔄 Счётчики сброшены. Новые КП начнутся с #1.\n"
            f"База данных и планы полностью очищены и готовы к работе!"
        )
        
        await callback.message.answer(
            report,
            parse_mode="Markdown",
            reply_markup=main_menu_kb(bot_user.role)
        )
        await callback.answer("✅ База и планы очищены!")

    except DestructiveDbOperationBlocked as exc:
        await callback.message.answer(
            f"⛔ {exc}",
            reply_markup=db_management_kb(bot_user.role),
        )
        await callback.answer("⛔ Запрещено в production")

    except Exception as e:
        await callback.message.answer(
            f"❌ Ошибка при очистке БД: {str(e)}\n\n"
            f"Попробуйте снова позже.",
            reply_markup=db_management_kb(bot_user.role)
        )
        await callback.answer("❌ Ошибка!")
        import logging
        logging.getLogger(__name__).exception(f"Ошибка при очистке БД: {e}")


@router.callback_query(F.data == "db_clear_cancel")
async def cancel_db_clear(callback: CallbackQuery, bot_user: BotUser):
    """Отмена очистки БД"""
    if not await _ensure_admin_callback(callback, bot_user, action="db_clear_cancel"):
        return
    await callback.message.answer(
        "❌ Очистка отменена. База данных не изменена.",
        reply_markup=db_management_kb(bot_user.role)
    )
    await callback.answer()


@router.callback_query(F.data == "db_back_to_menu")
async def back_to_menu(callback: CallbackQuery, bot_user: BotUser):
    """Возврат в главное меню"""
    await callback.message.answer(
        "Возврат в главное меню",
        reply_markup=main_menu_kb(bot_user.role),
    )
    await callback.answer()


@router.message(Command("recover_plates"))
async def cmd_recover_plates(message: Message, bot_user: BotUser):
    """
    Восстанавливает "застрявшие" плиты (статус 'в плане', но не в треках).
    Использование: /recover_plates
    
    Простыми словами:
    - Если плиты были добавлены в план, но не попали в tracks
    - И не были списаны через "День выполнен"
    - Они "застряли" в статусе "в плане" и недоступны для нового планирования
    - Эта команда возвращает их обратно в "в производстве"
    """
    if not await _ensure_admin_message(message, bot_user, action="recover_plates"):
        return
    try:
        # Сначала покажем, сколько плит застряло
        import sqlite3
        db_path = str(PROJECT_ROOT / "plita.db")
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        
        cur.execute('''
            SELECT COUNT(*), COALESCE(SUM(qty), 0)
            FROM kp_plates
            WHERE status = 'в плане'
        ''')
        
        result = cur.fetchone()
        records = result[0] if result else 0
        total_qty = result[1] if result else 0
        conn.close()
        
        if records == 0:
            await message.answer(
                "ℹ️ Застрявших плит не найдено.\n\n"
                "Все плиты имеют корректный статус.",
                reply_markup=main_menu_kb(bot_user.role)
            )
            return
        
        # Показываем информацию и восстанавливаем
        await message.answer(
            f"🔧 **Восстановление застрявших плит**\n\n"
            f"Найдено:\n"
            f"  • Записей: {records}\n"
            f"  • Плит: {total_qty}\n\n"
            f"⏳ Возвращаю в статус 'в производстве'...",
            parse_mode="Markdown"
        )
        
        # Выполняем восстановление
        recovered = kp_db.recover_stuck_plates(db_path)
        
        await message.answer(
            f"✅ **Восстановление завершено!**\n\n"
            f"Возвращено в производство: {recovered} записей\n\n"
            f"Теперь эти плиты снова доступны для планирования.",
            parse_mode="Markdown",
            reply_markup=main_menu_kb(bot_user.role)
        )
        
    except Exception as e:
        await message.answer(
            f"❌ Ошибка при восстановлении плит: {str(e)}",
            reply_markup=main_menu_kb(bot_user.role)
        )
        import logging
        logging.getLogger(__name__).exception(f"Ошибка при восстановлении плит: {e}")


@router.callback_query(F.data == "db_view_rests")
async def view_plate_rests(callback: CallbackQuery, bot_user: BotUser):
    """Экспорт остатков в Excel файл"""
    if not await _ensure_admin_callback(callback, bot_user, action="db_view_rests"):
        return
    try:
        await callback.message.answer("⏳ Формирую отчёт по остаткам...")
        
        # Получаем все остатки из БД
        rests = kp_db.get_all_plate_rests()
        
        if not rests:
            await callback.message.answer(
                "ℹ️ Остатков пока нет.",
                reply_markup=db_management_kb(bot_user.role)
            )
            await callback.answer()
            return
        
        # Создаём Excel файл
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from datetime import datetime
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Остатки от резки"
        
        # Заголовки
        headers = ['№', 'Ширина (мм)', 'Длина (м)', 'Количество', 'Статус', 'КП №', 'Клиент', 'Исходная плита', 'Дата создания']
        ws.append(headers)
        
        # Стиль заголовков
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Заполняем данные
        status_map = {
            'available': '✅ Доступен',
            'used': '🔧 Использован',
            'completed': '✔️ Выполнен',
            'scrapped': '❌ Списан'
        }
        
        for idx, rest in enumerate(rests, 1):
            status = status_map.get(rest['status'], rest['status'])
            customer = rest.get('customer_name', 'Не указан')
            created = rest.get('created_date', 'Не указана')
            
            # Форматируем дату
            if created and created != 'Не указана':
                try:
                    dt = datetime.fromisoformat(created)
                    created = dt.strftime('%d.%m.%Y %H:%M')
                except:
                    pass
            
            row = [
                idx,
                rest['rest_width_mm'],
                rest['length_m'],
                rest['qty'],
                status,
                rest['kp_id'],
                customer,
                rest.get('source_plate_name', 'Не указано'),
                created
            ]
            ws.append(row)
            
            # Выравнивание
            for cell in ws[idx + 1]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Настройка ширины столбцов
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 18
        
        # Сохраняем файл
        user_id = callback.from_user.id
        filename = f"остатки_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = PROJECT_ROOT / 'bot' / 'tmp' / f"{user_id}_{filename}"
        
        # Создаём папку tmp если её нет
        filepath.parent.mkdir(exist_ok=True)
        
        wb.save(str(filepath))
        
        # Отправляем файл
        from aiogram.types import FSInputFile
        file = FSInputFile(filepath)
        
        await callback.message.answer_document(
            file,
            caption=f"📋 **Отчёт по остаткам от резки**\n\n"
                    f"Всего остатков: {len(rests)} записей\n"
                    f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            parse_mode="Markdown",
            reply_markup=db_management_kb(bot_user.role)
        )
        
        # Удаляем временный файл
        filepath.unlink()
        
        await callback.answer()
        
    except Exception as e:
        await callback.message.answer(
            f"❌ Ошибка при формировании отчёта: {str(e)}",
            reply_markup=db_management_kb(bot_user.role)
        )
        await callback.answer()
